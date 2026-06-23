# -*- coding: utf-8 -*-
import argparse
import hashlib
import json
import re
import shutil
import sqlite3
import sys
import tempfile
import time
import traceback
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

PLUGIN_INFO = {
    "id": "word_excel_read_to_db_v1",
    "name": "Word/Excel读取入库V1",
    "version": "1.2.1",
    "api_version": "1.0",
    "category": "文件处理",
    "description": "按统一读取策略读取 Word/Excel，并按“每文件一表”写入数据库。",
    "input_type": "table",
    "output_type": "table",
    "danger_level": "db_write",
}
PARSE_CACHE_VERSION = "word_table_cell_canonical_v3_raw_text"

DETAIL_HEADERS = [
    "source_file",
    "source_name",
    "source_ext",
    "block_type",
    "sheet_name",
    "row_index",
    "col_index",
    "cell_address",
    "text",
    "meta_json",
    "is_merged",
    "is_merge_origin",
    "merge_origin_row",
    "merge_origin_col",
    "row_span",
    "col_span",
    "merged_range",
]

SUMMARY_HEADERS = ["source_row", "source_mode", "file_name", "file_path", "table_name", "read_rows", "write_rows", "status", "error"]

DEFAULT_READ_STRATEGY = "win32文本段落+精确表格"
READ_STRATEGY_CONVERT_XML = "转换docx后XML读取"
READ_STRATEGY_CONVERT_WIN32 = "转换docx后用win32读取"
READ_STRATEGY_CONVERT_WIN32_COMPAT = "保留旧格式转换docx后用win32读取"
READ_STRATEGIES = [
    DEFAULT_READ_STRATEGY,
    "win32快速读取",
    "win32文本反推表格",
    "win32纯文本快速读取",
    READ_STRATEGY_CONVERT_XML,
    READ_STRATEGY_CONVERT_WIN32,
    READ_STRATEGY_CONVERT_WIN32_COMPAT,
    "win32完整读取",
]

PARAMETER_UI_METADATA = {
    "read_engine": {"group": "读取设置", "order": 10},
    "word_merge_mode": {"group": "读取设置", "order": 20, "advanced": True},
    "doc_read_strategy": {"group": "读取设置", "order": 30, "advanced": True},
    "path_source": {"group": "文件来源", "order": 100},
    "path_field": {
        "group": "文件来源",
        "order": 110,
        "options_source": {"type": "preview_headers"},
        "visible_when": {"field": "path_source", "equals": "当前表字段=完整文件路径"},
        "empty_text": "当前输入表没有可选字段",
    },
    "dir_field": {
        "group": "文件来源",
        "order": 120,
        "options_source": {"type": "preview_headers"},
        "visible_when": {"field": "path_source", "equals": "当前表字段=目录路径"},
        "empty_text": "当前输入表没有可选字段",
    },
    "directory_path": {
        "group": "文件来源",
        "order": 130,
        "placeholder": "选择或输入固定目录路径",
        "visible_when": {"field": "path_source", "equals": "插件参数=固定目录路径"},
    },
    "recursive": {
        "group": "文件来源",
        "order": 140,
        "visible_when": {"field": "path_source", "in": ["当前表字段=目录路径", "插件参数=固定目录路径"]},
    },
    "file_patterns": {
        "group": "文件来源",
        "order": 150,
        "visible_when": {"field": "path_source", "in": ["当前表字段=目录路径", "插件参数=固定目录路径"]},
    },
    "table_name_mode": {"group": "写库设置", "order": 200},
    "table_prefix": {"group": "写库设置", "order": 210},
    "write_mode": {"group": "写库设置", "order": 220},
    "preview_write_db": {"group": "写库设置", "order": 230, "warning": "开启后预览也会写入数据库。"},
    "enable_cache": {"group": "缓存与失败处理", "order": 300},
    "force_refresh": {"group": "缓存与失败处理", "order": 310, "enabled_when": {"field": "enable_cache", "equals": True}},
    "cache_key_mode": {"group": "缓存与失败处理", "order": 320, "enabled_when": {"field": "enable_cache", "equals": True}},
    "error_policy": {"group": "缓存与失败处理", "order": 330},
}


def _with_parameter_ui_metadata(fields):
    result = []
    for field in fields:
        if not isinstance(field, dict):
            result.append(field)
            continue
        merged = dict(field)
        for key, value in PARAMETER_UI_METADATA.get(str(field.get("name") or ""), {}).items():
            merged.setdefault(key, value)
        result.append(merged)
    return result



def get_parameter_schema():
    return _with_parameter_ui_metadata([
        {
            "name": "read_engine",
            "label": "读取引擎",
            "type": "select",
            "choices": ["win32", "zip_xml"],
            "default": "win32",
            "help": "win32：通过 Office COM 读取；zip_xml：Word/Excel 走 ZIP+XML 解析。",
        },
        {
            "name": "word_merge_mode",
            "label": "Word合并解析",
            "type": "select",
            "choices": ["关闭", "简化", "完整"],
            "default": "关闭",
            "help": "关闭：最快，不解析 Word 合并单元格；简化：只取行列；完整：解析合并范围，复杂文档可能很慢。",
        },
        {
            "name": "doc_read_strategy",
            "label": "读取策略",
            "type": "select",
            "choices": list(READ_STRATEGIES),
            "default": DEFAULT_READ_STRATEGY,
            "help": "适用于 .doc/.docx/.docm。转换策略只使用临时 docx，不修改源文件；“保留旧格式”会保留源文档当前兼容模式。",
        },
        {
            "name": "path_source",
            "label": "文件来源",
            "type": "select",
            "choices": [
                "当前表字段=完整文件路径",
                "当前表字段=目录路径",
                "插件参数=固定目录路径",
            ],
            "default": "当前表字段=完整文件路径",
        },
        {
            "name": "path_field",
            "label": "完整路径字段",
            "type": "field_select",
            "default": "完整路径",
        },
        {
            "name": "dir_field",
            "label": "目录字段",
            "type": "field_select",
            "default": "目录",
        },
        {
            "name": "directory_path",
            "label": "固定目录路径",
            "type": "folder_path",
            "default": "",
        },
        {
            "name": "recursive",
            "label": "目录递归扫描",
            "type": "bool",
            "default": True,
        },
        {
            "name": "file_patterns",
            "label": "文件匹配",
            "type": "text",
            "default": "*.doc;*.docx;*.docm;*.xls;*.xlsx;*.xlsm",
            "help": "多个模式用 ; 或 , 分隔。",
        },
        {
            "name": "table_name_mode",
            "label": "表名模式",
            "type": "select",
            "choices": ["文件名", "文件名+短哈希"],
            "default": "文件名+短哈希",
        },
        {
            "name": "table_prefix",
            "label": "表名前缀",
            "type": "text",
            "default": "src_",
        },
        {
            "name": "write_mode",
            "label": "写表模式",
            "type": "select",
            "choices": ["replace", "timestamp", "fail"],
            "default": "replace",
        },
        {
            "name": "preview_write_db",
            "label": "预览模式仍写库",
            "type": "bool",
            "default": False,
            "help": "默认预览模式只模拟，不实际写入数据库。",
        },
        {
            "name": "enable_cache",
            "label": "启用插件缓存",
            "type": "bool",
            "default": True,
        },
        {
            "name": "force_refresh",
            "label": "强制重新处理，不使用缓存",
            "type": "bool",
            "default": False,
        },
        {
            "name": "cache_key_mode",
            "label": "缓存校验方式",
            "type": "select",
            "choices": ["快速签名", "文件Hash"],
            "default": "快速签名",
        },
        {
            "name": "error_policy",
            "label": "失败处理",
            "type": "select",
            "choices": ["继续并记录失败", "遇错停止"],
            "default": "继续并记录失败",
        },
    ])


def get_table_access_spec(params=None, context=None):
    p = dict(params or {})
    prefix = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff]+", "_", _as_text(p.get("table_prefix", "src_"))).strip()
    mode = _as_text(p.get("write_mode", "replace")) or "replace"
    permissions = {
        "read_table": False,
        "write_table": True,
        "create_table": True,
        "append_rows": mode == "append",
        "update_rows": False,
        "clear_table": False,
        "replace_table": mode == "replace",
        "alter_schema": mode == "append",
        "delete_rows": False,
        "drop_table": False,
    }
    return [{
        "role": "document_output",
        "source_type": "SQLite表",
        "table_pattern": f"{prefix}*" if prefix else "*",
        "pattern_type": "glob",
        "permissions": permissions,
        "write_mode": mode,
    }]


def get_output_schema(params=None, input_data=None, context=None):
    return {
        "type": "table",
        "headers": list(DETAIL_HEADERS),
        "rows": [],
        "meta": {
            "plugin": PLUGIN_INFO["id"],
            "lazy_schema": True,
            "summary_headers": list(SUMMARY_HEADERS),
        },
    }


def _as_text(v):
    return "" if v is None else str(v).strip()


def _safe_cell(row, idx):
    return row[idx] if idx < len(row) else ""


def _progress(context, current=None, total=None, message="", **extra):
    reporter = context.get("report_progress")
    if callable(reporter):
        try:
            reporter(current, total, message, **extra)
            return
        except Exception:
            pass
    callback = context.get("progress_callback")
    if callable(callback):
        msg = {
            "type": "node_progress",
            "node_name": context.get("node_name", "插件节点"),
            "plugin_id": context.get("plugin_id", PLUGIN_INFO["id"]),
            "current": current,
            "total": total,
            "message": message or "处理中",
        }
        msg.update(extra)
        try:
            callback(msg)
        except Exception:
            pass


def _short_detail_text(text, limit=90):
    text = re.sub(r"\s+", " ", _as_text(text))
    if len(text) > limit:
        return text[:limit] + "..."
    return text


def _record_location(rec):
    block_type = _as_text(rec.get("block_type"))
    sheet = _as_text(rec.get("sheet_name")) or ("段落" if block_type == "word_paragraph" else "")
    row = _as_text(rec.get("row_index"))
    col = _as_text(rec.get("col_index"))
    addr = _as_text(rec.get("cell_address"))
    if row and col:
        loc = f"{sheet} R{row}C{col}" if sheet else f"R{row}C{col}"
    elif addr:
        loc = f"{sheet} {addr}" if sheet else addr
    elif row:
        loc = f"{sheet}{row}" if sheet else f"行{row}"
    else:
        loc = sheet or block_type or "记录"
    return loc.strip()


def _record_detail_message(file_path, rec, action="读取中"):
    return f"{action}：{Path(file_path).name} {_record_location(rec)}：{_short_detail_text(rec.get('text', ''))}"


def _record_progress(context, current, total, file_path, rec, record_no, stage="record_read"):
    if not context:
        return
    if record_no != 1 and record_no % 50 != 0:
        return
    _progress(
        context,
        current,
        total,
        "读取中",
        stage=stage,
        object=str(file_path),
        detail_message=_record_detail_message(file_path, rec),
    )


def _check_cancel(context):
    ev = context.get("cancel_event")
    return bool(ev is not None and hasattr(ev, "is_set") and ev.is_set())


def _resolve_path(text, context):
    p = Path(_as_text(text))
    if p.is_absolute():
        return p
    app_dir = Path(context.get("app_dir", "."))
    return (app_dir / p).resolve()


def _split_patterns(text):
    raw = _as_text(text)
    if raw == "":
        return ["*.doc", "*.docx", "*.docm", "*.xls", "*.xlsx", "*.xlsm"]
    parts = re.split(r"[;,]", raw)
    cleaned = [p.strip() for p in parts if p.strip()]
    return cleaned or ["*.doc", "*.docx", "*.docm", "*.xls", "*.xlsx", "*.xlsm"]


def _collect_files(input_data, params, context):
    headers = list(input_data.get("headers", []))
    rows = [list(r) for r in input_data.get("rows", [])]
    source_mode = params.get("path_source", "当前表字段=完整文件路径")
    recursive = bool(params.get("recursive", True))
    patterns = _split_patterns(params.get("file_patterns", ""))

    def scan_dir(dir_path):
        files = []
        for pat in patterns:
            if recursive:
                files.extend([p for p in dir_path.rglob(pat) if p.is_file()])
            else:
                files.extend([p for p in dir_path.glob(pat) if p.is_file()])
        return files

    out = []
    if source_mode == "当前表字段=完整文件路径":
        field = _as_text(params.get("path_field", "完整路径"))
        idx = headers.index(field)
        for row_no, row in enumerate(rows, start=1):
            cell = _as_text(_safe_cell(row, idx))
            if cell:
                out.append({"path": _resolve_path(cell, context), "source_row": row_no, "source_mode": source_mode})
    elif source_mode == "当前表字段=目录路径":
        field = _as_text(params.get("dir_field", "目录"))
        idx = headers.index(field)
        for row_no, row in enumerate(rows, start=1):
            cell = _as_text(_safe_cell(row, idx))
            if not cell:
                continue
            folder = _resolve_path(cell, context)
            for p in scan_dir(folder):
                out.append({"path": p, "source_row": row_no, "source_mode": source_mode})
    else:
        folder = _resolve_path(_as_text(params.get("directory_path", "")), context)
        for p in scan_dir(folder):
            out.append({"path": p, "source_row": "", "source_mode": source_mode})

    dedup = {}
    for item in out:
        key = str(item["path"]).lower()
        if key not in dedup:
            dedup[key] = item
    return list(dedup.values())


def _sanitize_table_name(name):
    name = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff]+", "_", _as_text(name))
    name = name.strip("_")
    return name or "unnamed"


def _short_hash(path):
    h = hashlib.sha256(str(path).encode("utf-8")).hexdigest()
    return h[:8]


def _build_table_name(file_path, params):
    prefix = _as_text(params.get("table_prefix", "src_"))
    mode = _as_text(params.get("table_name_mode", "文件名+短哈希"))
    stem = _sanitize_table_name(file_path.stem)
    if mode == "文件名":
        return _sanitize_table_name(prefix + stem)
    return _sanitize_table_name(f"{prefix}{stem}_{_short_hash(file_path)}")


def _cache_db_path(context):
    data_dir = Path(context.get("plugin_data_dir", "."))
    data_dir.mkdir(parents=True, exist_ok=True)
    return data_dir / "cache.sqlite"


def _ensure_cache_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS plugin_cache (
            cache_key TEXT PRIMARY KEY,
            plugin_id TEXT,
            plugin_version TEXT,
            file_path TEXT,
            file_size INTEGER,
            file_mtime_ns INTEGER,
            file_hash TEXT,
            params_hash TEXT,
            result_json TEXT,
            updated_at TEXT
        )
        """
    )
    conn.commit()


def _stable_params_hash(params):
    ignore = {"enable_cache", "force_refresh", "cache_key_mode"}
    data = {k: params[k] for k in sorted(params.keys()) if k not in ignore}
    data["_parse_cache_version"] = PARSE_CACHE_VERSION
    txt = json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(txt.encode("utf-8")).hexdigest()


def _parse_params_hash(params):
    data = {
        "parse_cache_version": PARSE_CACHE_VERSION,
        "read_engine": _as_text(params.get("read_engine", "win32")) or "win32",
        "word_merge_mode": _as_text(params.get("word_merge_mode", "关闭")) or "关闭",
        "doc_read_strategy": _as_text(params.get("doc_read_strategy", DEFAULT_READ_STRATEGY)) or DEFAULT_READ_STRATEGY,
    }
    txt = json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(txt.encode("utf-8")).hexdigest()


def _file_sha256(file_path):
    h = hashlib.sha256()
    with file_path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _cache_signature(file_path, mode):
    st = file_path.stat()
    sig = {
        "file_size": int(st.st_size),
        "file_mtime_ns": int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1_000_000_000))),
        "file_hash": "",
    }
    if mode == "文件Hash":
        sig["file_hash"] = _file_sha256(file_path)
    return sig


def _cache_key(file_path, mode, params_hash, sig):
    payload = {
        "plugin_id": PLUGIN_INFO["id"],
        "plugin_version": PLUGIN_INFO["version"],
        "file_path": str(file_path.resolve()).lower(),
        "file_size": sig["file_size"],
        "file_mtime_ns": sig["file_mtime_ns"],
        "file_hash": sig.get("file_hash", ""),
        "params_hash": params_hash,
        "cache_key_mode": mode,
    }
    txt = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(txt.encode("utf-8")).hexdigest()


def _cache_load(conn, key):
    cur = conn.cursor()
    cur.execute("SELECT result_json FROM plugin_cache WHERE cache_key=?", (key,))
    row = cur.fetchone()
    if not row or not row[0]:
        return None
    try:
        data = json.loads(row[0])
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def _cache_save(conn, key, file_path, params_hash, sig, result_data):
    payload = json.dumps(result_data, ensure_ascii=False)
    conn.execute(
        """
        INSERT OR REPLACE INTO plugin_cache(
            cache_key, plugin_id, plugin_version, file_path, file_size, file_mtime_ns,
            file_hash, params_hash, result_json, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            key,
            PLUGIN_INFO["id"],
            PLUGIN_INFO["version"],
            str(file_path),
            int(sig.get("file_size", 0)),
            int(sig.get("file_mtime_ns", 0)),
            _as_text(sig.get("file_hash", "")),
            params_hash,
            payload,
            time.strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )
    conn.commit()


def _xml_texts(node, ns):
    texts = []
    for t in node.findall(".//w:t", ns):
        if t.text:
            texts.append(t.text)
    return "".join(texts).strip()


def _clean_word_text(text):
    s = _as_text(text)
    s = s.replace("\r", "").replace("\x07", "").replace("\n", "").strip()
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)
    return s


def _word_raw_visible_text(text):
    raw = "" if text is None else str(text)
    raw = raw.replace("\x00", "")
    while raw.endswith(("\r\x07", "\x07", "\r")):
        if raw.endswith("\r\x07"):
            raw = raw[:-2]
        else:
            raw = raw[:-1]
    return raw


def _split_word_cell_text(text):
    raw = "" if text is None else str(text)
    if raw == "":
        return []
    if "\x07" not in raw:
        cleaned = _clean_word_text(raw)
        return [cleaned] if cleaned else []
    parts = raw.split("\x07")
    if parts and parts[-1] == "":
        parts = parts[:-1]
    return [_clean_word_text(part) for part in parts]


def _word_range_bounds(range_obj):
    try:
        return int(range_obj.Start), int(range_obj.End)
    except Exception:
        return None


def _word_table_ranges(doc):
    ranges = []
    try:
        count = int(doc.Tables.Count)
    except Exception:
        count = 0
    for index in range(1, count + 1):
        try:
            bounds = _word_range_bounds(doc.Tables(index).Range)
        except Exception:
            bounds = None
        if not bounds:
            continue
        start, end = bounds
        if end > start:
            ranges.append((start, end))
    return ranges


def _range_overlaps_any(start, end, ranges):
    try:
        start = int(start)
        end = int(end)
    except Exception:
        return False
    for left, right in ranges or []:
        if start < right and end > left:
            return True
    return False


def _word_range_in_table(range_obj):
    try:
        return int(range_obj.Tables.Count) > 0
    except Exception:
        pass
    try:
        return bool(range_obj.Information(12))
    except Exception:
        return False


def _append_word_text_lines(rows, raw_text, file_path, context, progress_current, progress_total, meta_base=None, exclude_ranges=None):
    meta_base = dict(meta_base or {})
    exclude_ranges = list(exclude_ranges or [])
    source_text = "" if raw_text is None else str(raw_text)
    line_index = 0
    for match in re.finditer(r"[^\r\n]+", source_text):
        part = match.group(0)
        txt = _clean_word_text(part)
        if txt == "" and "\x01" not in str(part):
            continue
        range_base = int(meta_base.get("range_base") or 0)
        absolute_start = range_base + match.start()
        absolute_end = range_base + match.end()
        if exclude_ranges and _range_overlaps_any(absolute_start, absolute_end, exclude_ranges):
            continue
        line_index += 1
        meta = {
            "line_index": line_index,
            "range_start": match.start(),
            "range_end": match.end(),
            "position_kind": "word_content_range",
        }
        meta.update(meta_base)
        rec = {
            "block_type": "word_text_range",
            "sheet_name": "",
            "row_index": line_index,
            "col_index": "",
            "cell_address": f"WRANGE{absolute_start}:{absolute_end}",
            "text": txt,
            "meta_json": _row_meta(meta),
        }
        rows.append(rec)
        _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))


def _word_table_text_record(table_index, row_index, col_index, txt, inferred_by):
    merge = _merge_meta()
    return {
        "block_type": "word_table_cell",
        "sheet_name": f"table_{table_index}",
        "row_index": row_index,
        "col_index": col_index,
        "cell_address": f"R{row_index}C{col_index}",
        "text": txt,
        "meta_json": _row_meta(
            {
                "table_index": table_index,
                "row_index": row_index,
                "col_index": col_index,
                "win32_text_table_map": True,
                "position_inferred_by": inferred_by,
            },
            merge,
        ),
    }


def _timing_add(timings, label, start, **extra):
    if timings is None:
        return
    item = {"label": label, "seconds": time.perf_counter() - start}
    item.update(extra)
    timings.append(item)


def _format_timing_detail(timings):
    parts = []
    for item in timings or []:
        label = _as_text(item.get("label"))
        seconds = item.get("seconds", 0)
        try:
            seconds = float(seconds)
        except Exception:
            seconds = 0.0
        extra = []
        for key in ("count", "rows", "tables", "cells", "strategy"):
            value = item.get(key, "")
            if value != "":
                extra.append(f"{key}={value}")
        suffix = f"({','.join(extra)})" if extra else ""
        parts.append(f"{label}{suffix}={seconds:.2f}s")
    return "；".join(parts)


def _merge_meta(is_merged=False, is_merge_origin=True, row_span=1, col_span=1, merge_origin_row=None, merge_origin_col=None, merged_range=""):
    row_span = max(1, int(row_span or 1))
    col_span = max(1, int(col_span or 1))
    is_merged = bool(is_merged or row_span > 1 or col_span > 1)
    return {
        "is_merged": is_merged,
        "is_merge_origin": bool(is_merge_origin),
        "merge_origin_row": "" if merge_origin_row is None else int(merge_origin_row),
        "merge_origin_col": "" if merge_origin_col is None else int(merge_origin_col),
        "row_span": row_span,
        "col_span": col_span,
        "merged_range": _as_text(merged_range),
    }


def _row_meta(base=None, merge=None):
    data = dict(base or {})
    data.update(merge or _merge_meta())
    return json.dumps(data, ensure_ascii=False)


def _word_attr(node, ns, name, default=""):
    if node is None:
        return default
    return node.get(f"{{{ns['w']}}}{name}", node.get(name, default))


def _word_com_cell_info(cell, fallback_row=0, fallback_col=0, merge_mode="关闭"):
    mode = _as_text(merge_mode) or "关闭"
    if mode == "关闭":
        return int(fallback_row or 0), int(fallback_col or 0), _merge_meta()
    if mode == "简化":
        try:
            row_index = int(cell.RowIndex)
            col_index = int(cell.ColumnIndex)
        except Exception:
            row_index = int(fallback_row or 0)
            col_index = int(fallback_col or 0)
        return row_index, col_index, _merge_meta(merge_origin_row=row_index, merge_origin_col=col_index)

    # WdInformation: start/end row/column numbers in the containing table.
    try:
        start_row = int(cell.Range.Information(13))
        end_row = int(cell.Range.Information(14))
        start_col = int(cell.Range.Information(16))
        end_col = int(cell.Range.Information(17))
    except Exception:
        start_row = int(fallback_row or 0)
        end_row = start_row
        start_col = int(fallback_col or 0)
        end_col = start_col
    row_span = max(1, end_row - start_row + 1)
    col_span = max(1, end_col - start_col + 1)
    merge = _merge_meta(
        is_merged=(row_span > 1 or col_span > 1),
        is_merge_origin=True,
        row_span=row_span,
        col_span=col_span,
        merge_origin_row=start_row,
        merge_origin_col=start_col,
        merged_range=f"R{start_row}C{start_col}:R{end_row}C{end_col}" if row_span > 1 or col_span > 1 else "",
    )
    return start_row, start_col, merge


def _read_docx_like(file_path, word_merge_mode="关闭", context=None, progress_current=None, progress_total=None, timings=None):
    parse_start = time.perf_counter()
    parse_merge = _as_text(word_merge_mode) == "完整"
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    rows = []
    with zipfile.ZipFile(file_path, "r") as zf:
        if "word/document.xml" not in zf.namelist():
            _timing_add(timings, "解析docx XML", parse_start, rows=0)
            return rows
        root = ET.fromstring(zf.read("word/document.xml"))
        body = root.find("w:body", ns)
        if body is None:
            return rows

        para_idx = 0
        table_idx = 0
        for child in list(body):
            tag = child.tag.split("}")[-1]
            if tag == "p":
                para_idx += 1
                txt = _xml_texts(child, ns)
                if txt != "":
                    rec = {
                        "block_type": "word_paragraph",
                        "sheet_name": "",
                        "row_index": para_idx,
                        "col_index": "",
                        "cell_address": "",
                        "text": txt,
                        "meta_json": _row_meta({
                            "paragraph_index": para_idx,
                            "word_raw_text": txt,
                        }),
                    }
                    rows.append(rec)
                    _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))
            elif tag == "tbl":
                table_idx += 1
                row_idx = 0
                active_vmerges = {}
                table_cells = []
                for tr in child.findall(".//w:tr", ns):
                    row_idx += 1
                    col_idx = 0
                    next_vmerges = {}
                    incremented_origins = set()
                    for tc in tr.findall("./w:tc", ns):
                        col_idx += 1
                        tc_pr = tc.find("./w:tcPr", ns)
                        grid_span_node = tc_pr.find("./w:gridSpan", ns) if parse_merge and tc_pr is not None else None
                        grid_span = 1
                        if grid_span_node is not None:
                            try:
                                grid_span = max(1, int(_word_attr(grid_span_node, ns, "val", "1")))
                            except Exception:
                                grid_span = 1
                        vmerge_node = tc_pr.find("./w:vMerge", ns) if parse_merge and tc_pr is not None else None
                        vmerge_val = _word_attr(vmerge_node, ns, "val", "continue") if vmerge_node is not None else ""
                        if vmerge_node is not None and vmerge_val != "restart":
                            origin = active_vmerges.get(col_idx)
                            if origin is not None:
                                key = id(origin)
                                if key not in incremented_origins:
                                    origin["merge"]["row_span"] = int(origin["merge"].get("row_span", 1)) + 1
                                    origin["merge"]["is_merged"] = True
                                    incremented_origins.add(key)
                                for cc in range(col_idx, col_idx + int(origin["merge"].get("col_span", grid_span) or grid_span)):
                                    next_vmerges[cc] = origin
                            col_idx += grid_span - 1
                            continue
                        txt = _xml_texts(tc, ns)
                        merge = _merge_meta(
                            is_merged=(grid_span > 1 or vmerge_node is not None),
                            is_merge_origin=True,
                            row_span=1,
                            col_span=grid_span,
                            merge_origin_row=row_idx,
                            merge_origin_col=col_idx,
                            merged_range="",
                        )
                        rec = {"row_index": row_idx, "col_index": col_idx, "text": txt, "merge": merge}
                        table_cells.append(rec)
                        if vmerge_node is not None and vmerge_val == "restart":
                            for cc in range(col_idx, col_idx + grid_span):
                                next_vmerges[cc] = rec
                        col_idx += grid_span - 1
                    active_vmerges = next_vmerges
                for cell in table_cells:
                    txt = _as_text(cell.get("text", ""))
                    if txt == "":
                        continue
                    row_start = int(cell.get("row_index") or 0)
                    col_start = int(cell.get("col_index") or 0)
                    merge = dict(cell.get("merge") or _merge_meta())
                    row_span = max(1, int(merge.get("row_span", 1) or 1))
                    col_span = max(1, int(merge.get("col_span", 1) or 1))
                    if row_span > 1 or col_span > 1:
                        merge["is_merged"] = True
                        merge["merged_range"] = f"R{row_start}C{col_start}:R{row_start + row_span - 1}C{col_start + col_span - 1}"
                    rec = {
                        "block_type": "word_table_cell",
                        "sheet_name": f"table_{table_idx}",
                        "row_index": row_start,
                        "col_index": col_start,
                        "cell_address": f"R{row_start}C{col_start}",
                        "text": txt,
                        "meta_json": _row_meta({
                            "table_index": table_idx,
                            "row_index": row_start,
                            "col_index": col_start,
                            "word_raw_text": txt,
                        }, merge),
                    }
                    rows.append(rec)
                    _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))
    _timing_add(timings, "解析docx XML", parse_start, rows=len(rows))
    return rows


def _convert_word_to_temp_docx(file_path, preserve_compatibility=False, timings=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("转换 Word 文档需要 pywin32 + Word") from exc

    temp_docx = None
    temp_dir = None
    word = None
    doc = None
    try:
        step_start = time.perf_counter()
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        _timing_add(timings, "启动Word", step_start)

        step_start = time.perf_counter()
        doc = word.Documents.Open(str(file_path), ReadOnly=True, AddToRecentFiles=False, ConfirmConversions=False, Visible=False)
        _timing_add(timings, "打开文档", step_start)
        try:
            temp_dir = Path(tempfile.mkdtemp(prefix="word_doc_convert_"))
            temp_docx = temp_dir / f"{file_path.stem}_converted.docx"
            step_start = time.perf_counter()
            compatibility_mode = 0 if preserve_compatibility else 15
            try:
                # 12 = wdFormatXMLDocument；0 = 保留当前兼容模式；15 = 当前 Word 模式。
                doc.SaveAs2(
                    str(temp_docx),
                    FileFormat=12,
                    CompatibilityMode=compatibility_mode,
                )
            except Exception:
                if preserve_compatibility:
                    raise
                if temp_docx.exists():
                    temp_docx.unlink()
                doc.SaveAs2(str(temp_docx), FileFormat=12)
            _timing_add(
                timings,
                "转换docx",
                step_start,
                strategy="保留旧格式" if preserve_compatibility else "当前格式",
            )
        finally:
            doc.Close(False)
            doc = None

        return temp_dir, temp_docx
    except Exception:
        if temp_dir is not None and temp_dir.exists():
            shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    finally:
        step_start = time.perf_counter()
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        _timing_add(timings, "关闭Word", step_start)


def _read_word_converted(
    file_path,
    read_mode,
    word_merge_mode="关闭",
    preserve_compatibility=False,
    context=None,
    progress_current=None,
    progress_total=None,
    timings=None,
):
    temp_dir = None
    temp_docx = None
    try:
        temp_dir, temp_docx = _convert_word_to_temp_docx(
            file_path,
            preserve_compatibility=preserve_compatibility,
            timings=timings,
        )
        last_err = None
        for _ in range(5):
            try:
                if read_mode == "xml":
                    return _read_docx_like(
                        temp_docx,
                        word_merge_mode,
                        context,
                        progress_current,
                        progress_total,
                        timings,
                    )
                return _read_word_via_com(
                    temp_docx,
                    word_merge_mode,
                    context,
                    progress_current,
                    progress_total,
                    timings,
                )
            except Exception as exc:
                last_err = exc
                time.sleep(0.15)
        raise last_err if last_err else RuntimeError("读取临时 docx 失败")
    finally:
        if temp_dir is not None and temp_dir.exists():
            shutil.rmtree(temp_dir, ignore_errors=True)


def _read_doc_via_com(file_path, word_merge_mode="关闭", context=None, progress_current=None, progress_total=None, timings=None):
    return _read_word_converted(
        file_path,
        "xml",
        word_merge_mode,
        preserve_compatibility=False,
        context=context,
        progress_current=progress_current,
        progress_total=progress_total,
        timings=timings,
    )


def _read_word_text_via_com(file_path, context=None, progress_current=None, progress_total=None, timings=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("读取 Word 需要 pywin32 + Word") from exc

    word = None
    doc = None
    rows = []
    try:
        step_start = time.perf_counter()
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        _timing_add(timings, "启动Word", step_start)

        step_start = time.perf_counter()
        doc = word.Documents.Open(str(file_path), ReadOnly=True, AddToRecentFiles=False, ConfirmConversions=False, Visible=False)
        _timing_add(timings, "打开文档", step_start)

        step_start = time.perf_counter()
        raw_text = "" if doc.Content.Text is None else str(doc.Content.Text)
        _timing_add(timings, "读取全文", step_start, count=len(raw_text))

        step_start = time.perf_counter()
        _append_word_text_lines(
            rows,
            raw_text,
            file_path,
            context,
            progress_current,
            progress_total,
            {"win32_text_fast": True, "range_base": int(doc.Content.Start)},
        )
        _timing_add(timings, "拆分文本", step_start, rows=len(rows))
        return rows
    finally:
        step_start = time.perf_counter()
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        _timing_add(timings, "关闭Word", step_start)


def _read_word_text_table_map_via_com(file_path, context=None, progress_current=None, progress_total=None, timings=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("读取 Word 需要 pywin32 + Word") from exc

    word = None
    doc = None
    rows = []
    try:
        step_start = time.perf_counter()
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        _timing_add(timings, "启动Word", step_start)

        step_start = time.perf_counter()
        doc = word.Documents.Open(str(file_path), ReadOnly=True, AddToRecentFiles=False, ConfirmConversions=False, Visible=False)
        _timing_add(timings, "打开文档", step_start)

        step_start = time.perf_counter()
        raw_text = "" if doc.Content.Text is None else str(doc.Content.Text)
        _timing_add(timings, "读取全文", step_start, count=len(raw_text))

        step_start = time.perf_counter()
        table_ranges = _word_table_ranges(doc)
        _append_word_text_lines(
            rows,
            raw_text,
            file_path,
            context,
            progress_current,
            progress_total,
            {"win32_text_table_map": True, "range_base": int(doc.Content.Start)},
            exclude_ranges=table_ranges,
        )
        _timing_add(timings, "拆分文本", step_start, rows=len(rows), excluded_table_ranges=len(table_ranges))

        step_start = time.perf_counter()
        try:
            table_count = int(doc.Tables.Count)
        except Exception:
            table_count = 0
        table_rows_start = len(rows)
        cell_scan_count = 0
        fallback_count = 0
        for table_index in range(1, table_count + 1):
            table = doc.Tables(table_index)
            table_records = []
            used_row_ranges = False
            try:
                row_count = int(table.Rows.Count)
                for row_index in range(1, row_count + 1):
                    cell_texts = _split_word_cell_text(table.Rows(row_index).Range.Text)
                    cell_scan_count += len(cell_texts)
                    for col_index, txt in enumerate(cell_texts, start=1):
                        if txt == "":
                            continue
                        table_records.append(_word_table_text_record(table_index, row_index, col_index, txt, "row_range_text"))
                used_row_ranges = True
            except Exception:
                table_records = []

            if not used_row_ranges:
                fallback_count += 1
                cell_texts = _split_word_cell_text(table.Range.Text)
                cell_scan_count += len(cell_texts)
                try:
                    col_count = int(table.Columns.Count)
                except Exception:
                    col_count = 0
                for cell_index, txt in enumerate(cell_texts, start=1):
                    if col_count > 0:
                        row_index = ((cell_index - 1) // col_count) + 1
                        col_index = ((cell_index - 1) % col_count) + 1
                    else:
                        row_index = cell_index
                        col_index = 1
                    if txt == "":
                        continue
                    table_records.append(_word_table_text_record(table_index, row_index, col_index, txt, "table_range_text"))

            for rec in table_records:
                rows.append(rec)
                _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))

        _timing_add(
            timings,
            "文本反推表格",
            step_start,
            tables=table_count,
            cells=cell_scan_count,
            rows=len(rows) - table_rows_start,
            strategy=f"fallback={fallback_count}",
        )
        return rows
    finally:
        step_start = time.perf_counter()
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        _timing_add(timings, "关闭Word", step_start)


def _append_word_table_cells_via_com(doc, file_path, word_merge_mode, rows, context=None, progress_current=None, progress_total=None, timings=None):
    step_start = time.perf_counter()
    try:
        t_count = int(doc.Tables.Count)
    except Exception:
        t_count = 0
    table_rows_start = len(rows)
    cell_scan_count = 0
    for t in range(1, t_count + 1):
        table = doc.Tables(t)
        try:
            r_count = int(table.Rows.Count)
            c_count = int(table.Columns.Count)
        except Exception:
            r_count = 0
            c_count = 0
        seen_cells = set()
        if r_count > 0 and c_count > 0:
            cell_iter = []
            for r in range(1, r_count + 1):
                for c in range(1, c_count + 1):
                    cell_iter.append((r, c, None))
        else:
            cell_iter = []
            try:
                for n in range(1, int(table.Range.Cells.Count) + 1):
                    cell_iter.append((0, 0, table.Range.Cells(n)))
            except Exception:
                cell_iter = []
        for r, c, cell_obj in cell_iter:
            cell_scan_count += 1
            try:
                cell = cell_obj if cell_obj is not None else table.Cell(r, c)
                row_index, col_index, merge = _word_com_cell_info(cell, r, c, word_merge_mode)
                key = (row_index, col_index)
                if key in seen_cells:
                    continue
                seen_cells.add(key)
                raw_cell_text = cell.Range.Text
                keep_empty_marker = "\x01" in str(raw_cell_text)
                txt = _clean_word_text(raw_cell_text)
            except Exception:
                txt = ""
                keep_empty_marker = False
                row_index, col_index, merge = r, c, _merge_meta()
            if txt == "" and not keep_empty_marker:
                continue
            rec = {
                "block_type": "word_table_cell",
                "sheet_name": f"table_{t}",
                "row_index": row_index,
                "col_index": col_index,
                "cell_address": f"R{row_index}C{col_index}",
                "text": txt,
                "meta_json": _row_meta({
                    "table_index": t,
                    "row_index": row_index,
                    "col_index": col_index,
                    "word_raw_text": _word_raw_visible_text(raw_cell_text),
                }, merge),
            }
            rows.append(rec)
            _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))
    _timing_add(timings, "读取表格", step_start, tables=t_count, cells=cell_scan_count, rows=len(rows) - table_rows_start)


def _read_word_text_exact_tables_via_com(file_path, word_merge_mode="关闭", context=None, progress_current=None, progress_total=None, timings=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("读取 Word 需要 pywin32 + Word") from exc

    word = None
    doc = None
    rows = []
    try:
        step_start = time.perf_counter()
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        _timing_add(timings, "启动Word", step_start)

        step_start = time.perf_counter()
        doc = word.Documents.Open(str(file_path), ReadOnly=True, AddToRecentFiles=False, ConfirmConversions=False, Visible=False)
        _timing_add(timings, "打开文档", step_start)

        step_start = time.perf_counter()
        raw_text = "" if doc.Content.Text is None else str(doc.Content.Text)
        _timing_add(timings, "读取全文", step_start, count=len(raw_text))

        step_start = time.perf_counter()
        table_ranges = _word_table_ranges(doc)
        _append_word_text_lines(
            rows,
            raw_text,
            file_path,
            context,
            progress_current,
            progress_total,
            {"win32_text_paragraph_exact_table": True, "range_base": int(doc.Content.Start)},
            exclude_ranges=table_ranges,
        )
        _timing_add(timings, "拆分文本", step_start, rows=len(rows), excluded_table_ranges=len(table_ranges))

        _append_word_table_cells_via_com(doc, file_path, word_merge_mode, rows, context, progress_current, progress_total, timings)
        return rows
    finally:
        step_start = time.perf_counter()
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        _timing_add(timings, "关闭Word", step_start)


def _read_word_via_com(file_path, word_merge_mode="关闭", context=None, progress_current=None, progress_total=None, timings=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("读取 Word 需要 pywin32 + Word") from exc

    word = None
    doc = None
    rows = []
    try:
        step_start = time.perf_counter()
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        _timing_add(timings, "启动Word", step_start)

        step_start = time.perf_counter()
        doc = word.Documents.Open(str(file_path), ReadOnly=True, AddToRecentFiles=False, ConfirmConversions=False, Visible=False)
        _timing_add(timings, "打开文档", step_start)

        step_start = time.perf_counter()
        try:
            p_count = int(doc.Paragraphs.Count)
        except Exception:
            p_count = 0
        skipped_table_paragraphs = 0
        for i in range(1, p_count + 1):
            try:
                para = doc.Paragraphs(i)
                rng = para.Range
                if _word_range_in_table(rng):
                    skipped_table_paragraphs += 1
                    continue
                raw_paragraph_text = rng.Text
                txt = _clean_word_text(raw_paragraph_text)
            except Exception:
                txt = ""
            if txt == "":
                continue
            rec = {
                "block_type": "word_paragraph",
                "sheet_name": "",
                "row_index": i,
                "col_index": "",
                "cell_address": "",
                "text": txt,
                "meta_json": _row_meta({
                    "paragraph_index": i,
                    "word_raw_text": _word_raw_visible_text(raw_paragraph_text),
                }),
            }
            rows.append(rec)
            _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))
        _timing_add(timings, "读取段落", step_start, count=p_count, rows=len(rows), skipped_table_paragraphs=skipped_table_paragraphs)

        _append_word_table_cells_via_com(doc, file_path, word_merge_mode, rows, context, progress_current, progress_total, timings)
        return rows
    finally:
        step_start = time.perf_counter()
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        _timing_add(timings, "关闭Word", step_start)


def _read_excel_openpyxl(file_path, context=None, progress_current=None, progress_total=None):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("未安装 openpyxl") from exc

    rows = []
    wb = load_workbook(filename=str(file_path), read_only=False, data_only=False)
    try:
        for ws in wb.worksheets:
            merge_map = {}
            try:
                for merged in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged.bounds
                    row_span = max_row - min_row + 1
                    col_span = max_col - min_col + 1
                    for rr in range(min_row, max_row + 1):
                        for cc in range(min_col, max_col + 1):
                            merge_map[(rr, cc)] = _merge_meta(
                                is_merged=True,
                                is_merge_origin=(rr == min_row and cc == min_col),
                                row_span=row_span,
                                col_span=col_span,
                                merge_origin_row=min_row,
                                merge_origin_col=min_col,
                                merged_range=str(merged),
                            )
            except Exception:
                merge_map = {}
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    merge = merge_map.get((cell.row, cell.column), _merge_meta())
                    if merge.get("is_merged") and not merge.get("is_merge_origin", True):
                        continue
                    value = cell.value
                    if value is None or str(value) == "":
                        continue
                    rec = {
                        "block_type": "excel_cell",
                        "sheet_name": ws.title,
                        "row_index": cell.row,
                        "col_index": cell.column,
                        "cell_address": cell.coordinate,
                        "text": str(value),
                        "meta_json": _row_meta({"data_type": cell.data_type}, merge),
                    }
                    rows.append(rec)
                    _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))
    finally:
        wb.close()
    return rows


def _read_excel_via_com(file_path, context=None, progress_current=None, progress_total=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("读取 Excel 需要 openpyxl 或 pywin32 + Excel") from exc

    excel = None
    wb = None
    rows = []
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(file_path), ReadOnly=True, UpdateLinks=0)
        for si in range(1, wb.Worksheets.Count + 1):
            ws = wb.Worksheets(si)
            used = ws.UsedRange
            if used is None:
                continue
            sr = int(used.Row)
            sc = int(used.Column)
            rc = int(used.Rows.Count)
            cc = int(used.Columns.Count)
            for r in range(sr, sr + rc):
                for c in range(sc, sc + cc):
                    cell = ws.Cells(r, c)
                    merge = _merge_meta()
                    try:
                        if bool(cell.MergeCells):
                            area = cell.MergeArea
                            origin_row = int(area.Row)
                            origin_col = int(area.Column)
                            row_span = int(area.Rows.Count)
                            col_span = int(area.Columns.Count)
                            if r != origin_row or c != origin_col:
                                continue
                            merge = _merge_meta(
                                is_merged=True,
                                is_merge_origin=True,
                                row_span=row_span,
                                col_span=col_span,
                                merge_origin_row=origin_row,
                                merge_origin_col=origin_col,
                                merged_range=_as_text(area.Address).replace("$", ""),
                            )
                    except Exception:
                        merge = _merge_meta()
                    txt = _as_text(cell.Text)
                    if txt == "":
                        continue
                    rec = {
                        "block_type": "excel_cell",
                        "sheet_name": _as_text(ws.Name),
                        "row_index": r,
                        "col_index": c,
                        "cell_address": _as_text(cell.Address).replace("$", ""),
                        "text": txt,
                        "meta_json": _row_meta({}, merge),
                    }
                    rows.append(rec)
                    _record_progress(context, progress_current, progress_total, file_path, rec, len(rows))
        return rows
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _read_file_rows(file_path, read_engine, params=None, context=None, progress_current=None, progress_total=None, timings=None):
    ext = file_path.suffix.lower()
    engine = _as_text(read_engine).lower() or "win32"
    params = dict(params or {})
    word_merge_mode = _as_text(params.get("word_merge_mode", "关闭")) or "关闭"
    if word_merge_mode not in ("关闭", "简化", "完整"):
        word_merge_mode = "关闭"
    doc_strategy = _as_text(params.get("doc_read_strategy", DEFAULT_READ_STRATEGY)) or DEFAULT_READ_STRATEGY
    if doc_strategy not in READ_STRATEGIES:
        doc_strategy = DEFAULT_READ_STRATEGY

    if ext in (".doc", ".docx", ".docm"):
        if doc_strategy == DEFAULT_READ_STRATEGY:
            return _read_word_text_exact_tables_via_com(file_path, word_merge_mode, context, progress_current, progress_total, timings)
        if doc_strategy == "win32文本反推表格":
            return _read_word_text_table_map_via_com(file_path, context, progress_current, progress_total, timings)
        if doc_strategy == "win32纯文本快速读取":
            return _read_word_text_via_com(file_path, context, progress_current, progress_total, timings)
        if doc_strategy == READ_STRATEGY_CONVERT_XML:
            return _read_word_converted(
                file_path,
                "xml",
                word_merge_mode,
                preserve_compatibility=False,
                context=context,
                progress_current=progress_current,
                progress_total=progress_total,
                timings=timings,
            )
        if doc_strategy == READ_STRATEGY_CONVERT_WIN32:
            return _read_word_converted(
                file_path,
                "win32",
                word_merge_mode,
                preserve_compatibility=False,
                context=context,
                progress_current=progress_current,
                progress_total=progress_total,
                timings=timings,
            )
        if doc_strategy == READ_STRATEGY_CONVERT_WIN32_COMPAT:
            return _read_word_converted(
                file_path,
                "win32",
                word_merge_mode,
                preserve_compatibility=True,
                context=context,
                progress_current=progress_current,
                progress_total=progress_total,
                timings=timings,
            )
        if doc_strategy == "win32完整读取":
            return _read_word_via_com(file_path, "完整", context, progress_current, progress_total, timings)
        if doc_strategy == "win32快速读取":
            return _read_word_via_com(file_path, word_merge_mode, context, progress_current, progress_total, timings)

    if ext in (".xlsx", ".xlsm", ".xls"):
        if engine == "win32":
            try:
                return _read_excel_via_com(file_path, context, progress_current, progress_total)
            except Exception:
                return _read_excel_openpyxl(file_path, context, progress_current, progress_total)
        try:
            return _read_excel_openpyxl(file_path, context, progress_current, progress_total)
        except Exception:
            return _read_excel_via_com(file_path, context, progress_current, progress_total)

    raise ValueError(f"不支持的文件类型：{ext}")


def _write_table_with_db_api(db, table_name, headers, rows, mode):
    return db.write_table(table_name=table_name, headers=headers, rows=rows, mode=mode)


def validate_params(params, input_data, context):
    p = dict(params or {})
    headers = list(input_data.get("headers", []))
    source_mode = _as_text(p.get("path_source", "当前表字段=完整文件路径"))

    if source_mode == "当前表字段=完整文件路径":
        if _as_text(p.get("path_field", "完整路径")) not in headers:
            return False, "完整路径字段不存在"
    elif source_mode == "当前表字段=目录路径":
        if _as_text(p.get("dir_field", "目录")) not in headers:
            return False, "目录字段不存在"
    else:
        if _as_text(p.get("directory_path", "")) == "":
            return False, "固定目录路径不能为空"
    return True, ""


def run(input_data, params, context):
    p = dict(params or {})
    read_engine = _as_text(p.get("read_engine", "win32")) or "win32"
    if read_engine not in ("win32", "zip_xml"):
        read_engine = "win32"
    word_merge_mode = _as_text(p.get("word_merge_mode", "关闭")) or "关闭"
    if word_merge_mode not in ("关闭", "简化", "完整"):
        word_merge_mode = "关闭"
    doc_read_strategy = _as_text(p.get("doc_read_strategy", DEFAULT_READ_STRATEGY)) or DEFAULT_READ_STRATEGY
    if doc_read_strategy not in READ_STRATEGIES:
        doc_read_strategy = DEFAULT_READ_STRATEGY
    p["read_engine"] = read_engine
    p["word_merge_mode"] = word_merge_mode
    p["doc_read_strategy"] = doc_read_strategy
    error_policy = _as_text(p.get("error_policy", "继续并记录失败"))
    files = _collect_files(input_data, p, context)
    if not files:
        return {
            "ok": True,
            "message": "未找到待处理文件",
            "output": {"type": "table", "headers": ["提示"], "rows": [["未找到待处理文件"]], "meta": {"plugin": PLUGIN_INFO["id"]}},
            "logs": [{"level": "INFO", "message": "未找到待处理文件"}],
            "summary": {"total": 0, "success": 0, "failed": 0, "written_rows": 0, "output_rows": 0},
        }
    _progress(context, 0, len(files), f"准备读取 {len(files)} 个文件", stage="prepare")

    db = context.get("db")
    is_preview = bool(context.get("is_preview", False))
    preview_write_db = bool(p.get("preview_write_db", False))
    write_mode = _as_text(p.get("write_mode", "replace")) or "replace"
    logs = []
    enable_cache = bool(p.get("enable_cache", True))
    force_refresh = bool(p.get("force_refresh", False))
    cache_key_mode = _as_text(p.get("cache_key_mode", "快速签名")) or "快速签名"
    if cache_key_mode not in ("快速签名", "文件Hash"):
        cache_key_mode = "快速签名"
    params_hash = _parse_params_hash(p)
    legacy_params_hash = _stable_params_hash(p)
    cache_conn = None
    if enable_cache:
        try:
            cache_conn = sqlite3.connect(str(_cache_db_path(context)), timeout=10)
            cache_conn.execute("PRAGMA busy_timeout=10000")
            _ensure_cache_table(cache_conn)
        except Exception as exc:
            cache_conn = None
            logs.append({"level": "WARNING", "message": f"CACHE_ERROR：初始化缓存失败：{exc}"})

    detail_headers = list(DETAIL_HEADERS)
    summary_headers = list(SUMMARY_HEADERS)
    summary_rows = []
    output_rows = []
    cancelled = False
    cache_hit_count = 0
    cache_miss_count = 0
    cache_write_count = 0
    database_requests = []

    try:
        file_iter = enumerate(files, start=1)
    except Exception:
        file_iter = []

    for i, item in file_iter:
        if _check_cancel(context):
            cancelled = True
            logs.append({"level": "WARNING", "message": "检测到取消信号，已停止后续读取"})
            _progress(context, i - 1, len(files), "已取消，停止后续读取", stage="cancelled")
            break
        file_path = item["path"]
        table_name = _build_table_name(file_path, p)
        _progress(context, i - 1, len(files), f"读取中：{file_path.name}", stage="file_start", object=str(file_path))
        try:
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在：{file_path}")

            parsed = None
            cache_key = ""
            cache_sig = None
            cache_save_parsed = False
            cache_write_message = "CACHE_WRITE：已写入缓存"
            if enable_cache and cache_conn is not None:
                try:
                    cache_sig = _cache_signature(file_path, cache_key_mode)
                    cache_key = _cache_key(file_path, cache_key_mode, params_hash, cache_sig)
                    if force_refresh:
                        logs.append({"level": "INFO", "object": str(file_path), "message": "CACHE_REFRESH：强制重新处理"})
                    else:
                        cached = _cache_load(cache_conn, cache_key)
                        if cached and isinstance(cached.get("parsed"), list):
                            parsed = cached["parsed"]
                            cache_hit_count += 1
                            logs.append({"level": "INFO", "object": str(file_path), "message": "CACHE_HIT：文件未变化，复用缓存读取结果"})
                        else:
                            legacy_cached = None
                            if legacy_params_hash != params_hash:
                                legacy_key = _cache_key(file_path, cache_key_mode, legacy_params_hash, cache_sig)
                                legacy_cached = _cache_load(cache_conn, legacy_key)
                            if legacy_cached and isinstance(legacy_cached.get("parsed"), list):
                                parsed = legacy_cached["parsed"]
                                cache_hit_count += 1
                                cache_save_parsed = True
                                cache_write_message = "CACHE_WRITE：已迁移旧参数缓存"
                                logs.append({"level": "INFO", "object": str(file_path), "message": "CACHE_HIT：命中旧参数缓存，复用读取结果"})
                            else:
                                cache_miss_count += 1
                                logs.append({"level": "INFO", "object": str(file_path), "message": "CACHE_MISS：未命中缓存，重新读取"})
                except Exception as exc:
                    logs.append({"level": "WARNING", "object": str(file_path), "message": f"CACHE_ERROR：读取缓存失败：{exc}"})

            if parsed is None:
                read_start = time.perf_counter()
                read_timings = []
                parsed = _read_file_rows(file_path, read_engine, p, context, i - 1, len(files), read_timings)
                read_seconds = time.perf_counter() - read_start
                cache_save_parsed = True
                logs.append({
                    "level": "INFO",
                    "object": str(file_path),
                    "message": f"READ_TIMING：读取耗时 {read_seconds:.2f}s，Word合并解析={word_merge_mode}，读取策略={doc_read_strategy}",
                })
                timing_detail = _format_timing_detail(read_timings)
                if timing_detail:
                    logs.append({
                        "level": "INFO",
                        "object": str(file_path),
                        "message": f"READ_TIMING_DETAIL：{timing_detail}",
                    })
            if cache_save_parsed and enable_cache and cache_conn is not None and cache_key and cache_sig is not None:
                try:
                    _cache_save(cache_conn, cache_key, file_path, params_hash, cache_sig, {"parsed": parsed})
                    cache_write_count += 1
                    logs.append({"level": "INFO", "object": str(file_path), "message": cache_write_message})
                except Exception as exc:
                    logs.append({"level": "WARNING", "object": str(file_path), "message": f"CACHE_ERROR：写入缓存失败：{exc}"})

            detail_rows = []
            for r in parsed:
                meta_text = _as_text(r.get("meta_json", "{}"))
                try:
                    meta = json.loads(meta_text) if meta_text else {}
                    if not isinstance(meta, dict):
                        meta = {}
                except Exception:
                    meta = {}
                detail_rows.append(
                    [
                        str(file_path),
                        file_path.name,
                        file_path.suffix.lower(),
                        _as_text(r.get("block_type", "")),
                        _as_text(r.get("sheet_name", "")),
                        _as_text(r.get("row_index", "")),
                        _as_text(r.get("col_index", "")),
                        _as_text(r.get("cell_address", "")),
                        _as_text(r.get("text", "")),
                        meta_text or "{}",
                        _as_text(r.get("is_merged", meta.get("is_merged", ""))),
                        _as_text(r.get("is_merge_origin", meta.get("is_merge_origin", ""))),
                        _as_text(r.get("merge_origin_row", meta.get("merge_origin_row", ""))),
                        _as_text(r.get("merge_origin_col", meta.get("merge_origin_col", ""))),
                        _as_text(r.get("row_span", meta.get("row_span", ""))),
                        _as_text(r.get("col_span", meta.get("col_span", ""))),
                        _as_text(r.get("merged_range", meta.get("merged_range", ""))),
                    ]
                )
            output_rows.extend(detail_rows)

            written_count = 0
            status = "读取成功(预览未写库)" if (is_preview and not preview_write_db) else "成功"
            if not (is_preview and not preview_write_db):
                if db is not None and hasattr(db, "write_table"):
                    result = _write_table_with_db_api(db, table_name, detail_headers, detail_rows, write_mode)
                    table_name = _as_text(result.get("table_name", table_name)) or table_name
                elif context.get("database_access") == "managed_requests":
                    database_requests.append({
                        "operation": "write_table",
                        "table_name": table_name,
                        "headers": detail_headers,
                        "rows": detail_rows,
                        "mode": write_mode,
                    })
                else:
                    raise RuntimeError("缺少受控数据库接口：请通过 context['db'] 或 managed_requests 执行写入")
                written_count = len(detail_rows)

            summary_rows.append(
                [
                    item["source_row"],
                    item["source_mode"],
                    file_path.name,
                    str(file_path),
                    table_name,
                    len(detail_rows),
                    written_count,
                    status,
                    "",
                ]
            )
            logs.append(
                {
                    "level": "INFO",
                    "object": str(file_path),
                    "message": f"{file_path.name} -> {table_name}，引擎={read_engine}，读取 {len(detail_rows)} 行",
                }
            )
            _progress(context, i, len(files), f"已完成：{file_path.name}", stage="file_done", object=str(file_path))
        except Exception as exc:
            err = str(exc)
            summary_rows.append(
                [
                    item["source_row"],
                    item["source_mode"],
                    file_path.name,
                    str(file_path),
                    table_name,
                    0,
                    0,
                    "失败",
                    err,
                ]
            )
            logs.append(
                {
                    "level": "ERROR",
                    "object": str(file_path),
                    "message": f"{file_path.name} 失败：{err}",
                }
            )
            _progress(context, i, len(files), f"失败：{file_path.name}", stage="file_error", object=str(file_path))
            if error_policy == "遇错停止":
                raise

    ok_count = len([r for r in summary_rows if r[7].startswith("成功") or r[7].startswith("读取成功")])
    fail_count = len(summary_rows) - ok_count
    written_rows = sum(int(r[6]) for r in summary_rows if str(r[6]).isdigit())
    summary = {
        "total": len(summary_rows),
        "planned_total": len(files),
        "success": ok_count,
        "failed": fail_count,
        "written_rows": written_rows,
        "output_rows": len(output_rows),
        "read_engine": read_engine,
        "word_merge_mode": word_merge_mode,
        "doc_read_strategy": doc_read_strategy,
        "cancelled": cancelled,
        "cache_enabled": enable_cache,
        "cache_key_mode": cache_key_mode,
        "cache_hit": cache_hit_count,
        "cache_miss": cache_miss_count,
        "cache_write": cache_write_count,
    }
    if cache_conn is not None:
        try:
            cache_conn.close()
        except Exception:
            pass
    _progress(context, len(summary_rows), len(files), "读取阶段完成", stage="done", cancelled=cancelled)
    result_payload = {
        "ok": True,
        "message": f"读取完成：成功 {ok_count}，失败 {fail_count}",
        "output": {
            "type": "table",
            "headers": detail_headers,
            "rows": output_rows,
            "meta": {
                "plugin": PLUGIN_INFO["id"],
                "success_count": ok_count,
                "fail_count": fail_count,
                "summary_headers": summary_headers,
                "summary_rows": summary_rows,
            },
        },
        "logs": logs[-200:],
        "summary": summary,
    }
    if database_requests:
        result_payload["database_requests"] = database_requests
    return result_payload


def _external_progress_callback(msg):
    try:
        text = json.dumps(msg, ensure_ascii=False) + "\n"
        sys.stdout.buffer.write(text.encode("utf-8"))
        sys.stdout.buffer.flush()
    except Exception:
        pass


def _run_external_entry(input_path, output_path):
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    try:
        payload = json.loads(Path(input_path).read_text(encoding="utf-8"))
        input_data = payload.get("input_data") or {}
        params = payload.get("params") or {}
        context = payload.get("context") or {}
        context.setdefault("plugin_id", PLUGIN_INFO["id"])
        context["progress_callback"] = _external_progress_callback
        context["report_progress"] = lambda current=0, total=0, message="", **extra: _external_progress_callback({
            "type": "node_progress",
            "current": current,
            "total": total,
            "message": message,
            **extra,
        })
        result = run(input_data, params, context)
    except Exception as exc:
        result = {
            "ok": False,
            "message": str(exc),
            "output": {"type": "table", "headers": ["错误"], "rows": [[str(exc)]], "meta": {"plugin": PLUGIN_INFO["id"]}},
            "logs": [{"level": "ERROR", "message": str(exc)}, {"level": "ERROR", "message": traceback.format_exc()}],
            "summary": {"success": 0, "failed": 1},
        }
    output_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return 0 if result.get("ok", False) else 1


def main(argv=None):
    parser = argparse.ArgumentParser(description=PLUGIN_INFO["name"])
    parser.add_argument("--input", dest="input_path")
    parser.add_argument("--output", dest="output_path")
    args = parser.parse_args(argv)
    if args.input_path and args.output_path:
        return _run_external_entry(args.input_path, args.output_path)
    print("这是 DataFlowKit 插件，请在主程序插件节点中调用。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
