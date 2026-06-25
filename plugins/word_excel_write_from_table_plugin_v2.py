# -*- coding: utf-8 -*-
import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import time
import traceback
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

PLUGIN_INFO = {
    "id": "word_excel_write_from_table_v2",
    "name": "Word/Excel按数据写入V2",
    "version": "1.1.0",
    "api_version": "1.0",
    "category": "文件处理",
    "description": "按输入表数据安全写回 Word/Excel，支持范围定位、原子替换、重试及特殊对象全局替换。",
    "input_type": "table",
    "output_type": "table",
    "danger_level": "file_write",
}

WORD_MODE_PRESERVE_FORMAT = "保留原格式，仅改文字值"
WORD_MODE_OVERWRITE = "整段覆盖"
WORD_MODE_FIND_REPLACE = "按old_text查找替换"
WRITE_STRATEGY_FOLLOW_NODE = "跟随节点设置"
WRITE_STRATEGY_DIRECT = "直接定位写入"
BLOCK_WORD_TEXT_RANGE = "word_text_range"
BLOCK_WORD_GLOBAL_REPLACE = "word_global_replace"
BLOCK_WORD_SHAPE_TEXT = "word_shape_text"
BLOCK_WORD_CONTENT_CONTROL = "word_content_control"
REPLACE_SCOPE_FIRST = "替换第一次"
REPLACE_SCOPE_ALL = "替换全部"
CONFLICT_WARN = "按输入顺序执行并警告"
CONFLICT_KEEP_LAST = "保留最后一条"
CONFLICT_ERROR = "报错"
WORD_FIND_TEXT_LIMIT = 255

OUTPUT_HEADERS = [
    "source_file",
    "target_file",
    "target_file_name",
    "engine",
    "op_total",
    "applied",
    "skipped",
    "copy_status",
    "write_status",
    "status",
    "error",
]

GROUP_WRITE_ENGINE = 10
GROUP_WIN32_ADVANCED = 20
GROUP_WRITE_STRATEGY = 30
GROUP_PATH_FIELDS = 40
GROUP_TARGET_POLICY = 50
GROUP_DATA_FIELDS = 60
FIELD_SELECT_EMPTY_TEXT = "当前输入表没有可选字段"
FIELD_SELECT_INVALID_TEXT = "当前字段不在输入表字段中，仍会保留原值"


def _ui_meta(group, group_order, order, **extra):
    meta = {"group": group, "group_order": group_order, "order": order}
    meta.update(extra)
    return meta


def _field_ui_meta(group, group_order, order, **extra):
    meta = _ui_meta(
        group,
        group_order,
        order,
        options_source={"type": "preview_headers"},
        empty_text=FIELD_SELECT_EMPTY_TEXT,
        invalid_value_text=FIELD_SELECT_INVALID_TEXT,
    )
    meta.update(extra)
    return meta


PARAMETER_UI_METADATA = {
    "write_engine": _ui_meta("写入引擎", GROUP_WRITE_ENGINE, 10, refresh_on_change=["write_engine"]),
    "preview_write_files": _ui_meta("写入引擎", GROUP_WRITE_ENGINE, 20, warning="开启后预览也会写入文件。"),
    "win32_reuse_app": _ui_meta("win32高级设置", GROUP_WIN32_ADVANCED, 100, advanced=True, visible_when={"field": "write_engine", "equals": "win32"}, depends_on=["write_engine"]),
    "win32_open_retries": _ui_meta("win32高级设置", GROUP_WIN32_ADVANCED, 110, advanced=True, visible_when={"field": "write_engine", "equals": "win32"}, depends_on=["write_engine"], min=0, step=1, unit="次"),
    "win32_retry_interval_ms": _ui_meta("win32高级设置", GROUP_WIN32_ADVANCED, 120, advanced=True, visible_when={"field": "write_engine", "equals": "win32"}, depends_on=["write_engine"], min=0, step=50, unit="ms"),
    "win32_close_settle_ms": _ui_meta("win32高级设置", GROUP_WIN32_ADVANCED, 130, advanced=True, visible_when={"field": "write_engine", "equals": "win32"}, depends_on=["write_engine"], min=0, step=50, unit="ms"),
    "win32_cell_retries": _ui_meta("win32高级设置", GROUP_WIN32_ADVANCED, 140, advanced=True, visible_when={"field": "write_engine", "equals": "win32"}, depends_on=["write_engine"], min=0, step=1, unit="次"),
    "win32_save_retries": _ui_meta("win32高级设置", GROUP_WIN32_ADVANCED, 150, advanced=True, visible_when={"field": "write_engine", "equals": "win32"}, depends_on=["write_engine"], min=0, step=1, unit="次"),
    "word_text_write_mode": _ui_meta(
        "写入策略",
        GROUP_WRITE_STRATEGY,
        200,
        visible_when={"field": "write_engine", "equals": "win32"},
        depends_on=["write_engine"],
        refresh_on_change=["word_text_write_mode"],
    ),
    "scoped_replace_default": _ui_meta(
        "写入策略",
        GROUP_WRITE_STRATEGY,
        210,
        visible_when={
            "all": [
                {"field": "write_engine", "equals": "win32"},
                {"field": "word_text_write_mode", "equals": WORD_MODE_FIND_REPLACE},
            ],
        },
        depends_on=["write_engine", "word_text_write_mode"],
    ),
    "target_conflict_policy": _ui_meta("写入策略", GROUP_WRITE_STRATEGY, 220),
    "verify_after_write": _ui_meta(
        "写入策略",
        GROUP_WRITE_STRATEGY,
        230,
        visible_when={"field": "write_engine", "equals": "win32"},
        depends_on=["write_engine"],
        help="目前用于 win32 Word 范围写入后的结果校验。",
    ),
    "error_policy": _ui_meta("写入策略", GROUP_WRITE_STRATEGY, 240),
    "allow_empty_text_write": _ui_meta("写入策略", GROUP_WRITE_STRATEGY, 250),
    "path_field": _field_ui_meta("路径字段", GROUP_PATH_FIELDS, 300),
    "target_path_field": _field_ui_meta("路径字段", GROUP_PATH_FIELDS, 310),
    "target_missing_policy": _ui_meta("目标文件策略", GROUP_TARGET_POLICY, 400),
    "target_existing_policy": _ui_meta("目标文件策略", GROUP_TARGET_POLICY, 410),
    "same_path_policy": _ui_meta("目标文件策略", GROUP_TARGET_POLICY, 420),
    "create_parent_dirs": _ui_meta("目标文件策略", GROUP_TARGET_POLICY, 430),
    "backup_mode": _ui_meta("目标文件策略", GROUP_TARGET_POLICY, 440),
    "block_type_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 500),
    "sheet_name_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 510),
    "row_index_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 520),
    "col_index_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 530),
    "cell_address_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 540),
    "value_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 550),
    "old_text_field": _field_ui_meta(
        "写入数据字段",
        GROUP_DATA_FIELDS,
        560,
        depends_on=["write_engine", "word_text_write_mode", "write_strategy_field"],
        warning="仅在 Word 全局替换、按 old_text 查找替换或行级写入策略选择查找替换时使用。",
        help="指定输入表中保存原文的字段。直接定位写入不依赖此字段；查找替换和 word_global_replace 需要此字段。",
    ),
    "write_strategy_field": _field_ui_meta(
        "写入数据字段",
        GROUP_DATA_FIELDS,
        570,
        help="可选行级策略字段，支持跟随节点设置、直接定位写入、按 old_text 查找替换。",
    ),
    "replace_scope_field": _field_ui_meta(
        "写入数据字段",
        GROUP_DATA_FIELDS,
        580,
        depends_on=["write_engine", "word_text_write_mode", "write_strategy_field"],
        warning="仅在按 old_text 查找替换时决定替换第一次或全部。",
        help="指定输入表中保存替换次数策略的字段；为空时使用节点级默认替换次数。",
    ),
    "rule_old_text_field": _field_ui_meta(
        "写入数据字段",
        GROUP_DATA_FIELDS,
        590,
        depends_on=["write_engine", "word_text_write_mode"],
        warning="仅在按 old_text 查找替换时作为更精确的查找文本使用。",
        help="指定输入表中保存规则旧值的字段；可帮助 Word 查找替换避开 255 字符限制。",
    ),
    "rule_new_text_field": _field_ui_meta(
        "写入数据字段",
        GROUP_DATA_FIELDS,
        600,
        depends_on=["write_engine", "word_text_write_mode"],
        warning="仅在按 old_text 查找替换时作为更精确的替换文本使用。",
        help="指定输入表中保存规则新值的字段；通常与规则旧值字段配套使用。",
    ),
    "meta_json_field": _field_ui_meta("写入数据字段", GROUP_DATA_FIELDS, 610),
}


def _with_parameter_ui_metadata(fields):
    result = []
    for field in fields:
        if not isinstance(field, dict):
            result.append(field)
            continue
        merged = dict(field)
        for key, value in PARAMETER_UI_METADATA.get(str(field.get("name") or ""), {}).items():
            merged.setdefault(key, value)
        result.append(merged)
    return result



def get_parameter_schema():
    return _with_parameter_ui_metadata([
        {
            "name": "write_engine",
            "label": "写入引擎",
            "type": "select",
            "choices": ["win32", "zip_xml"],
            "default": "win32",
            "help": "win32：Office COM 写入；zip_xml：docx 直接 XML 写入，xlsx/xlsm 用 openpyxl 写入。",
        },
        {
            "name": "preview_write_files",
            "label": "预览模式仍写入文件",
            "type": "bool",
            "default": False,
            "help": "默认关闭：预览只模拟，不改文件。",
        },
        {
            "name": "win32_reuse_app",
            "label": "win32复用Office进程",
            "type": "bool",
            "default": True,
            "help": "开启后，本节点一次运行内复用同一个 Word/Excel 进程处理多个文件。",
        },
        {
            "name": "win32_open_retries",
            "label": "win32打开重试次数",
            "type": "number",
            "default": 5,
        },
        {
            "name": "win32_retry_interval_ms",
            "label": "win32重试间隔(ms)",
            "type": "number",
            "default": 300,
        },
        {
            "name": "win32_close_settle_ms",
            "label": "win32关闭等待(ms)",
            "type": "number",
            "default": 200,
        },
        {
            "name": "win32_cell_retries",
            "label": "win32单项写入重试次数",
            "type": "number",
            "default": 3,
        },
        {
            "name": "win32_save_retries",
            "label": "win32保存重试次数",
            "type": "number",
            "default": 3,
        },
        {
            "name": "word_text_write_mode",
            "label": "Word文字写入方式",
            "type": "select",
            "choices": [WORD_MODE_PRESERVE_FORMAT, WORD_MODE_OVERWRITE, WORD_MODE_FIND_REPLACE],
            "default": WORD_MODE_FIND_REPLACE,
            "help": "win32 写入 Word 时可按定位范围整段写入，也可在定位范围内按 old_text 查找替换。",
        },
        {
            "name": "scoped_replace_default",
            "label": "定位范围默认替换次数",
            "type": "select",
            "choices": [REPLACE_SCOPE_FIRST, REPLACE_SCOPE_ALL],
            "default": REPLACE_SCOPE_FIRST,
            "help": "仅影响段落、表格单元格、文本范围等定位对象；全文替换始终替换全部。",
        },
        {
            "name": "target_conflict_policy",
            "label": "同一位置多次写入",
            "type": "select",
            "choices": [CONFLICT_WARN, CONFLICT_KEEP_LAST, CONFLICT_ERROR],
            "default": CONFLICT_WARN,
        },
        {
            "name": "verify_after_write",
            "label": "写入后校验",
            "type": "bool",
            "default": True,
        },
        {
            "name": "error_policy",
            "label": "失败处理",
            "type": "select",
            "choices": ["继续并记录失败", "遇错停止"],
            "default": "继续并记录失败",
        },
        {
            "name": "allow_empty_text_write",
            "label": "允许写入空文本",
            "type": "bool",
            "default": False,
            "help": "关闭时，空文本记录会跳过，避免误清空内容。",
        },
        {"name": "path_field", "label": "文件路径字段", "type": "field_select", "default": "source_file"},
        {"name": "target_path_field", "label": "新文件路径字段", "type": "field_select", "default": "target_file"},
        {
            "name": "target_missing_policy",
            "label": "目标不存在时",
            "type": "select",
            "choices": ["从源文件复制", "报错"],
            "default": "从源文件复制",
        },
        {
            "name": "target_existing_policy",
            "label": "目标已存在时",
            "type": "select",
            "choices": ["直接写入", "覆盖为源文件后写入", "跳过", "报错"],
            "default": "直接写入",
        },
        {
            "name": "same_path_policy",
            "label": "新旧路径相同时",
            "type": "select",
            "choices": ["修改源文件", "跳过", "报错"],
            "default": "修改源文件",
        },
        {
            "name": "create_parent_dirs",
            "label": "自动创建目标目录",
            "type": "bool",
            "default": True,
        },
        {
            "name": "backup_mode",
            "label": "文件备份策略",
            "type": "select",
            "choices": ["失败时恢复原文件", "写入前保留备份", "不备份"],
            "default": "失败时恢复原文件",
            "help": "实际写入始终先在同目录临时副本完成；保留备份会额外保存原目标文件。",
        },
        {"name": "block_type_field", "label": "类型字段", "type": "field_select", "default": "block_type"},
        {"name": "sheet_name_field", "label": "sheet/表名字段", "type": "field_select", "default": "sheet_name"},
        {"name": "row_index_field", "label": "行号字段", "type": "field_select", "default": "row_index"},
        {"name": "col_index_field", "label": "列号字段", "type": "field_select", "default": "col_index"},
        {"name": "cell_address_field", "label": "地址字段", "type": "field_select", "default": "cell_address"},
        {"name": "value_field", "label": "写入值字段", "type": "field_select", "default": "text"},
        {"name": "old_text_field", "label": "原文匹配字段", "type": "field_select", "default": "old_text"},
        {"name": "write_strategy_field", "label": "单行写入策略字段", "type": "field_select", "default": "write_strategy"},
        {"name": "replace_scope_field", "label": "替换次数策略字段", "type": "field_select", "default": "replace_scope"},
        {"name": "rule_old_text_field", "label": "规则旧值字段", "type": "field_select", "default": "rule_old_text"},
        {"name": "rule_new_text_field", "label": "规则新值字段", "type": "field_select", "default": "rule_new_text"},
        {"name": "meta_json_field", "label": "meta字段", "type": "field_select", "default": "meta_json"},
    ])


def get_output_schema(params=None, input_data=None, context=None):
    return {
        "type": "table",
        "headers": list(OUTPUT_HEADERS),
        "rows": [],
        "meta": {"plugin": PLUGIN_INFO["id"], "lazy_schema": True},
    }


def _as_text(v):
    return "" if v is None else str(v).strip()


def _value_text(v):
    return "" if v is None else str(v)


def _op_meta(op):
    raw = (op or {}).get("meta_json", "")
    if isinstance(raw, dict):
        return dict(raw)
    try:
        value = json.loads(raw) if _as_text(raw) else {}
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def _default_word_text_write_mode(write_engine="win32"):
    engine = _as_text(write_engine).lower() or "win32"
    return WORD_MODE_FIND_REPLACE if engine == "win32" else WORD_MODE_PRESERVE_FORMAT


def _word_text_mode_from_params(params):
    params = params or {}
    write_engine = _as_text(params.get("write_engine", "win32")).lower() or "win32"
    mode = _as_text(params.get("word_text_write_mode", "")) or _default_word_text_write_mode(write_engine)
    if mode not in (WORD_MODE_PRESERVE_FORMAT, WORD_MODE_OVERWRITE, WORD_MODE_FIND_REPLACE):
        return _default_word_text_write_mode(write_engine)
    return mode


def _safe_cell(row, idx):
    return row[idx] if idx < len(row) else ""


def _resolve_path(path_text, context):
    p = Path(path_text)
    if p.is_absolute():
        return p
    app_dir = Path(context.get("app_dir", "."))
    return (app_dir / p).resolve()


def _same_file_path(a, b):
    try:
        return str(Path(a).resolve()).lower() == str(Path(b).resolve()).lower()
    except Exception:
        return str(a).strip().lower() == str(b).strip().lower()


def _to_int_or_none(v):
    s = _as_text(v)
    if s == "":
        return None
    try:
        return int(s, 10)
    except Exception:
        try:
            return int(s, 0)
        except Exception:
            return None


def _to_int_with_default(v, default, min_value=None):
    n = _to_int_or_none(v)
    if n is None:
        n = int(default)
    if min_value is not None and n < min_value:
        n = int(min_value)
    return n


def _stop_on_error(context):
    return _as_text(((context or {}).get("params") or {}).get("error_policy", "")) == "遇错停止"


def _retry_count(context, key, default):
    return _to_int_with_default(((context or {}).get("params") or {}).get(key, default), default, min_value=1)


def _retry_interval(context):
    milliseconds = _to_int_with_default(
        ((context or {}).get("params") or {}).get("win32_retry_interval_ms", 300),
        300,
        min_value=0,
    )
    return milliseconds / 1000.0


def _run_with_retry(action, attempts, interval, label):
    last_exc = None
    for attempt in range(1, max(1, int(attempts or 1)) + 1):
        try:
            return action()
        except Exception as exc:
            last_exc = exc
            if attempt < attempts and interval > 0:
                time.sleep(interval * attempt)
    raise RuntimeError(f"{label}失败，已重试 {attempts} 次：{last_exc}") from last_exc


def _parse_table_index(sheet_name, meta_json_text):
    m = re.search(r"table_(\d+)", _as_text(sheet_name), flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    try:
        meta = json.loads(meta_json_text) if _as_text(meta_json_text) else {}
        v = meta.get("table_index")
        vi = _to_int_or_none(v)
        if vi and vi > 0:
            return vi
    except Exception:
        pass
    return None


def _parse_rc_from_address(addr_text):
    s = _as_text(addr_text).upper().replace("$", "")
    if not s:
        return None, None
    m = re.match(r"^R(\d+)C(\d+)$", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^([A-Z]+)(\d+)$", s)
    if not m:
        return None, None
    letters = m.group(1)
    row = int(m.group(2))
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


def _excel_addr_from_rc(row, col):
    if not row or not col:
        return ""
    n = int(col)
    out = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out.append(chr(ord("A") + rem))
    return "".join(reversed(out)) + str(int(row))


def _log(level, message, obj="", tb=""):
    return {"level": str(level).upper(), "object": _as_text(obj), "message": _as_text(message), "traceback": _as_text(tb)}


def _progress(context, current=None, total=None, message="", **extra):
    reporter = context.get("report_progress")
    if callable(reporter):
        try:
            reporter(current, total, message, **extra)
            return
        except Exception:
            pass
    callback = context.get("progress_callback")
    if callable(callback):
        msg = {
            "type": "node_progress",
            "node_name": context.get("node_name", "插件节点"),
            "plugin_id": context.get("plugin_id", PLUGIN_INFO["id"]),
            "current": current,
            "total": total,
            "message": message or "处理中",
        }
        msg.update(extra)
        try:
            callback(msg)
        except Exception:
            pass


def _short_detail_text(text, limit=90):
    text = re.sub(r"\s+", " ", _as_text(text))
    if len(text) > limit:
        return text[:limit] + "..."
    return text


def _visible_log_text(text, limit=180):
    raw = _value_text(text)
    raw = (
        raw.replace("\\", "\\\\")
        .replace("\r", "\\r")
        .replace("\n", "\\n")
        .replace("\t", "\\t")
        .replace("\x07", "\\x07")
    )
    raw = re.sub(
        r"[\x00-\x06\x08\x0b\x0c\x0e-\x1f]",
        lambda match: f"\\x{ord(match.group(0)):02x}",
        raw,
    )
    if len(raw) > limit:
        return raw[:limit] + "..."
    return raw


def _op_location(op):
    sheet = _as_text(op.get("sheet_name"))
    row = _as_text(op.get("row_index"))
    col = _as_text(op.get("col_index"))
    addr = _as_text(op.get("cell_address"))
    if row and col:
        loc = f"{sheet} R{row}C{col}" if sheet else f"R{row}C{col}"
    elif addr:
        loc = f"{sheet} {addr}" if sheet else addr
    elif row:
        loc = f"{sheet} 段落{row}" if sheet else f"段落{row}"
    else:
        loc = sheet or _as_text(op.get("block_type")) or "写入位置"
    return loc.strip()


def _op_detail_message(file_path, op):
    return f"写入中：{Path(file_path).name} {_op_location(op)}：{_short_detail_text(op.get('value', ''))}"


def _op_progress(context, current, total, file_path, op, op_no):
    if not context:
        return
    if op_no != 1 and op_no % 20 != 0:
        return
    _progress(
        context,
        current,
        total,
        "写入中",
        stage="op_write",
        object=str(file_path),
        detail_message=_op_detail_message(file_path, op),
    )


def _check_cancel(context):
    ev = context.get("cancel_event")
    return bool(ev is not None and hasattr(ev, "is_set") and ev.is_set())


def validate_params(params, input_data, context):
    p = dict(params or {})
    headers = list(input_data.get("headers", []))

    path_field = _as_text(p.get("path_field", "source_file")) or "source_file"
    target_path_field = _as_text(p.get("target_path_field", "target_file")) or "target_file"
    value_field = _as_text(p.get("value_field", "text")) or "text"
    old_text_field = _as_text(p.get("old_text_field", "old_text")) or "old_text"
    block_type_field = _as_text(p.get("block_type_field", "block_type")) or "block_type"
    write_strategy_field = _as_text(p.get("write_strategy_field", "write_strategy")) or "write_strategy"
    write_engine = _as_text(p.get("write_engine", "win32")).lower() or "win32"
    word_text_write_mode = _word_text_mode_from_params(p)
    block_types = set()
    requires_old_text = False
    requires_win32_find = False
    if block_type_field in headers:
        block_index = headers.index(block_type_field)
        strategy_index = headers.index(write_strategy_field) if write_strategy_field in headers else None
        for row in input_data.get("rows", []) or []:
            block_type = _as_text(_safe_cell(row, block_index)).lower()
            block_types.add(block_type)
            strategy = _op_write_strategy({
                "write_strategy": _safe_cell(row, strategy_index) if strategy_index is not None else "",
            })
            is_word_scoped = block_type in {
                "word_paragraph",
                "word_table_cell",
                BLOCK_WORD_TEXT_RANGE,
                BLOCK_WORD_SHAPE_TEXT,
                BLOCK_WORD_CONTENT_CONTROL,
            }
            row_uses_find = (
                block_type in {BLOCK_WORD_TEXT_RANGE, BLOCK_WORD_GLOBAL_REPLACE}
                or strategy == WORD_MODE_FIND_REPLACE
                or (
                    is_word_scoped
                    and strategy == WRITE_STRATEGY_FOLLOW_NODE
                    and word_text_write_mode == WORD_MODE_FIND_REPLACE
                )
            )
            requires_old_text = requires_old_text or row_uses_find
            requires_win32_find = requires_win32_find or (
                row_uses_find and block_type != BLOCK_WORD_GLOBAL_REPLACE
            )

    if path_field not in headers:
        return False, f"文件路径字段不存在：{path_field}"
    if target_path_field and target_path_field not in headers:
        # 兼容旧工作流：默认 target_file 不存在时自动回退到 source_file。
        if target_path_field != "target_file":
            return False, f"新文件路径字段不存在：{target_path_field}"
    if value_field not in headers:
        return False, f"写入值字段不存在：{value_field}"
    if requires_old_text and old_text_field not in headers:
        return False, f"原文匹配字段不存在：{old_text_field}"
    if requires_win32_find and write_engine != "win32":
        return False, "按old_text查找替换仅支持 win32 写入引擎"
    if write_engine not in ("win32", "zip_xml"):
        return False, f"不支持的写入引擎：{write_engine}"
    com_only_blocks = block_types.intersection({
        BLOCK_WORD_TEXT_RANGE,
        BLOCK_WORD_SHAPE_TEXT,
        BLOCK_WORD_CONTENT_CONTROL,
    })
    if write_engine == "zip_xml" and com_only_blocks:
        return False, f"{sorted(com_only_blocks)[0]} 依赖 Word COM，请使用 win32 写入引擎"
    return True, ""


def _op_target_key(op):
    block_type = _as_text((op or {}).get("block_type")).lower()
    if block_type == BLOCK_WORD_GLOBAL_REPLACE:
        return None
    meta = _op_meta(op)
    if block_type == BLOCK_WORD_TEXT_RANGE:
        return (
            block_type,
            _to_int_or_none(meta.get("range_base")) or 0,
            _to_int_or_none(meta.get("range_start")),
            _to_int_or_none(meta.get("range_end")),
            _as_text((op or {}).get("cell_address")),
        )
    if block_type == BLOCK_WORD_SHAPE_TEXT:
        return (
            block_type,
            _as_text(meta.get("shape_scope")),
            _to_int_or_none(meta.get("section_index")),
            _to_int_or_none(meta.get("header_footer_type")),
            _to_int_or_none(meta.get("shape_index")),
        )
    if block_type == BLOCK_WORD_CONTENT_CONTROL:
        return (block_type, _to_int_or_none(meta.get("content_control_index")) or _to_int_or_none((op or {}).get("row_index")))
    return (
        block_type,
        _as_text((op or {}).get("sheet_name")).lower(),
        _as_text((op or {}).get("row_index")),
        _as_text((op or {}).get("col_index")),
        _as_text((op or {}).get("cell_address")).upper(),
    )


def _collect_ops(input_data, params, context):
    headers = list(input_data.get("headers", []))
    rows = [list(r) for r in input_data.get("rows", [])]
    idx = {name: headers.index(name) for name in headers}

    path_field = _as_text(params.get("path_field", "source_file")) or "source_file"
    target_path_field = _as_text(params.get("target_path_field", "target_file")) or "target_file"
    block_type_field = _as_text(params.get("block_type_field", "block_type")) or "block_type"
    sheet_name_field = _as_text(params.get("sheet_name_field", "sheet_name")) or "sheet_name"
    row_index_field = _as_text(params.get("row_index_field", "row_index")) or "row_index"
    col_index_field = _as_text(params.get("col_index_field", "col_index")) or "col_index"
    cell_address_field = _as_text(params.get("cell_address_field", "cell_address")) or "cell_address"
    value_field = _as_text(params.get("value_field", "text")) or "text"
    old_text_field = _as_text(params.get("old_text_field", "old_text")) or "old_text"
    write_strategy_field = _as_text(params.get("write_strategy_field", "write_strategy")) or "write_strategy"
    replace_scope_field = _as_text(params.get("replace_scope_field", "replace_scope")) or "replace_scope"
    rule_old_text_field = _as_text(params.get("rule_old_text_field", "rule_old_text")) or "rule_old_text"
    rule_new_text_field = _as_text(params.get("rule_new_text_field", "rule_new_text")) or "rule_new_text"
    meta_json_field = _as_text(params.get("meta_json_field", "meta_json")) or "meta_json"
    allow_empty = bool(params.get("allow_empty_text_write", False))
    conflict_policy = _as_text(params.get("target_conflict_policy", CONFLICT_WARN)) or CONFLICT_WARN

    def cell_by_field(row, field_name):
        if field_name in idx:
            return _safe_cell(row, idx[field_name])
        return ""

    grouped = {}
    skipped_no_path = 0
    skipped_empty_value = 0
    skipped_duplicate_global = 0
    skipped_duplicate_target = 0
    skipped_conflict_replaced = 0
    target_conflicts = 0

    for row_no, row in enumerate(rows, start=1):
        raw_path = _as_text(cell_by_field(row, path_field))
        if raw_path == "":
            skipped_no_path += 1
            continue
        raw_target = _as_text(cell_by_field(row, target_path_field))
        if raw_target == "":
            raw_target = raw_path
        value_text = _value_text(cell_by_field(row, value_field))
        if (not allow_empty) and (value_text == ""):
            skipped_empty_value += 1
            continue

        source_path = _resolve_path(raw_path, context)
        target_path = _resolve_path(raw_target, context)
        op = {
            "source_row": row_no,
            "source_path": source_path,
            "target_path": target_path,
            "block_type": _as_text(cell_by_field(row, block_type_field)).lower(),
            "sheet_name": _as_text(cell_by_field(row, sheet_name_field)),
            "row_index": _as_text(cell_by_field(row, row_index_field)),
            "col_index": _as_text(cell_by_field(row, col_index_field)),
            "cell_address": _as_text(cell_by_field(row, cell_address_field)),
            "value": value_text,
            "old_text": _value_text(cell_by_field(row, old_text_field)),
            "write_strategy": _as_text(cell_by_field(row, write_strategy_field)),
            "replace_scope": _as_text(cell_by_field(row, replace_scope_field)),
            "rule_old_text": _value_text(cell_by_field(row, rule_old_text_field)),
            "rule_new_text": _value_text(cell_by_field(row, rule_new_text_field)),
            "meta_json": _as_text(cell_by_field(row, meta_json_field)),
        }
        key = str(target_path).lower()
        group = grouped.setdefault(key, {"source_path": source_path, "target_path": target_path, "ops": [], "source_conflict": False})
        if not _same_file_path(group["source_path"], source_path):
            group["source_conflict"] = True
        if op["block_type"] == BLOCK_WORD_GLOBAL_REPLACE:
            global_key = (op["old_text"], op["value"])
            seen_global = group.setdefault("seen_global_ops", set())
            if global_key in seen_global:
                skipped_duplicate_global += 1
                continue
            seen_global.add(global_key)
        target_key = _op_target_key(op)
        if target_key is not None:
            target_positions = group.setdefault("target_positions", {})
            previous_index = target_positions.get(target_key)
            if previous_index is not None:
                previous = group["ops"][previous_index]
                same_write = (
                    _value_text(previous.get("value")) == _value_text(op.get("value"))
                    and _value_text(previous.get("old_text")) == _value_text(op.get("old_text"))
                    and _op_write_strategy(previous) == _op_write_strategy(op)
                )
                if same_write:
                    skipped_duplicate_target += 1
                    continue
                target_conflicts += 1
                group.setdefault("target_conflicts", []).append(
                    f"{_op_location(op)}：源行{previous.get('source_row')}与源行{op.get('source_row')}写入值不同"
                )
                if conflict_policy == CONFLICT_KEEP_LAST:
                    group["ops"][previous_index] = op
                    skipped_conflict_replaced += 1
                    continue
                if conflict_policy == CONFLICT_ERROR:
                    group["conflict_error"] = True
            else:
                target_positions[target_key] = len(group["ops"])
        group["ops"].append(op)

    return grouped, {
        "input_rows": len(rows),
        "skipped_no_path": skipped_no_path,
        "skipped_empty_value": skipped_empty_value,
        "skipped_duplicate_global": skipped_duplicate_global,
        "skipped_duplicate_target": skipped_duplicate_target,
        "skipped_conflict_replaced": skipped_conflict_replaced,
        "target_conflicts": target_conflicts,
    }


def _old_text_length(value):
    return len(_value_text(value))


def _sort_global_replace_ops_by_old_text_length(ops):
    indexed = list(enumerate(ops or []))
    indexed.sort(key=lambda item: (-_old_text_length(item[1].get("old_text", "")), item[0]))
    return [op for _index, op in indexed]


def _sort_global_replace_runs_by_old_text_length(ops):
    result = []
    run = []

    def flush_run():
        if not run:
            return
        result.extend(_sort_global_replace_ops_by_old_text_length(run))
        run.clear()

    for op in ops or []:
        if _as_text((op or {}).get("block_type", "")).lower() == BLOCK_WORD_GLOBAL_REPLACE:
            run.append(op)
            continue
        flush_run()
        result.append(op)
    flush_run()
    return result


def _sort_global_replacements_by_old_text_length(replacements):
    indexed = list(enumerate(replacements or []))
    indexed.sort(key=lambda item: (-_old_text_length(item[1][0] if item[1] else ""), item[0]))
    return [replacement for _index, replacement in indexed]


def _word_text_write_mode(context):
    params = (context or {}).get("params") or {}
    return _word_text_mode_from_params(params)


def _word_preserve_format_enabled(context):
    return _word_text_write_mode(context) == WORD_MODE_PRESERVE_FORMAT


def _word_find_replace_enabled(context):
    return _word_text_write_mode(context) == WORD_MODE_FIND_REPLACE


def _verification_enabled(context):
    return bool(((context or {}).get("params") or {}).get("verify_after_write", True))


def _op_write_strategy(op):
    text = _as_text((op or {}).get("write_strategy"))
    if text in ("", WRITE_STRATEGY_FOLLOW_NODE, "follow", "node"):
        return WRITE_STRATEGY_FOLLOW_NODE
    if text in (WRITE_STRATEGY_DIRECT, "direct", "direct_write", "直接写入", "定位写入"):
        return WRITE_STRATEGY_DIRECT
    if text in (WORD_MODE_FIND_REPLACE, "find_replace", "old_text", "按 old_text 查找替换"):
        return WORD_MODE_FIND_REPLACE
    return text


def _op_replace_all(op, context=None):
    if _as_text((op or {}).get("block_type")).lower() == BLOCK_WORD_GLOBAL_REPLACE:
        return True
    scope = _as_text((op or {}).get("replace_scope"))
    if scope in (REPLACE_SCOPE_ALL, "all", "全部", "0"):
        return True
    if scope in (REPLACE_SCOPE_FIRST, "first", "第一次", "1"):
        return False
    default_scope = _as_text(
        ((context or {}).get("params") or {}).get("scoped_replace_default", REPLACE_SCOPE_FIRST)
    ) or REPLACE_SCOPE_FIRST
    return default_scope == REPLACE_SCOPE_ALL


def _op_failure_detail(op):
    op = op or {}
    old_text = _value_text(op.get("old_text"))
    rule_old_text = _value_text(op.get("rule_old_text"))
    rule_new_text = _value_text(op.get("rule_new_text"))
    value = _value_text(op.get("value"))
    return "；".join(
        [
            f"block_type={_as_text(op.get('block_type')) or '-'}",
            f"位置={_op_location(op)}",
            f"cell_address={_as_text(op.get('cell_address')) or '-'}",
            f"write_strategy={_op_write_strategy(op)}",
            f"old_text长度={len(old_text)}",
            f"old_text={_visible_log_text(old_text)}",
            f"规则旧值={_visible_log_text(rule_old_text)}",
            f"规则新值={_visible_log_text(rule_new_text)}",
            f"写入值长度={len(value)}",
            f"写入值={_visible_log_text(value)}",
        ]
    )


def _operation_failure_message(prefix, op, exc):
    return f"{prefix}(源行{(op or {}).get('source_row')})：{exc}；{_op_failure_detail(op)}"


def _word_body_range(range_obj):
    start = int(range_obj.Start)
    end = int(range_obj.End)
    if end > start:
        text = str(range_obj.Text)
        while end > start and text.endswith(("\r\x07", "\r", "\x07")):
            end -= 1
            text = text[:-1]
    return range_obj.Document.Range(start, max(start, end))


def _word_visible_body_range(range_obj):
    text = str(range_obj.Text)
    if text.endswith("\r\x07"):
        try:
            body = range_obj.Duplicate
            if int(body.End) > int(body.Start):
                body.End = int(body.End) - 1
            return body
        except Exception:
            start = int(range_obj.Start)
            end = int(range_obj.End)
            return range_obj.Document.Range(start, max(start, end - 1))
    return _word_body_range(range_obj)


def _word_capture_font(range_obj):
    try:
        body = _word_visible_body_range(range_obj)
        if int(body.End) > int(body.Start):
            sample = body.Document.Range(int(body.Start), int(body.Start) + 1)
        else:
            sample = body
        font = sample.Font
        return {
            "Name": font.Name,
            "NameFarEast": getattr(font, "NameFarEast", ""),
            "Size": font.Size,
            "Bold": font.Bold,
            "Italic": font.Italic,
            "Underline": font.Underline,
            "Color": font.Color,
        }
    except Exception:
        return {}


def _word_apply_font(range_obj, font_info):
    if not font_info:
        return
    try:
        font = range_obj.Font
        for key, value in font_info.items():
            if value is None or value == 9999999:
                continue
            try:
                setattr(font, key, value)
            except Exception:
                pass
    except Exception:
        pass


def _word_write_text_preserve_format(range_obj, value):
    text = str(value if value is not None else "")
    font_info = _word_capture_font(range_obj)
    body = _word_visible_body_range(range_obj)
    start = int(body.Start)
    body.Text = text
    if text:
        new_body = range_obj.Document.Range(start, start + len(text))
        _word_apply_font(new_body, font_info)


def _word_write_visible_text(range_obj, value, preserve_format=True):
    if preserve_format:
        _word_write_text_preserve_format(range_obj, value)
    else:
        body = _word_visible_body_range(range_obj)
        body.Text = str(value if value is not None else "")


def _normalized_word_text(value):
    text = unicodedata.normalize("NFKC", _value_text(value))
    text = text.replace("\x00", "").replace("\x07", "")
    text = text.replace("\r", "").replace("\n", "").replace("\v", "")
    text = re.sub(r"[ \t\u00a0\u3000]+", " ", text)
    return text.strip()


def _minimal_text_change(old_text, new_text):
    old_text = _value_text(old_text)
    new_text = _value_text(new_text)
    if old_text == new_text:
        return "", ""
    prefix = 0
    max_prefix = min(len(old_text), len(new_text))
    while prefix < max_prefix and old_text[prefix] == new_text[prefix]:
        prefix += 1
    suffix = 0
    max_suffix = min(len(old_text) - prefix, len(new_text) - prefix)
    while suffix < max_suffix and old_text[len(old_text) - 1 - suffix] == new_text[len(new_text) - 1 - suffix]:
        suffix += 1
    old_end = len(old_text) - suffix if suffix else len(old_text)
    new_end = len(new_text) - suffix if suffix else len(new_text)
    return prefix, old_end, new_end


def _replacement_reaches_target(old_text, new_text, search_text, replacement_text):
    old_normalized = _normalized_word_text(old_text)
    new_normalized = _normalized_word_text(new_text)
    search_normalized = _normalized_word_text(search_text)
    replacement_normalized = _normalized_word_text(replacement_text)
    if not search_normalized or old_normalized.count(search_normalized) != 1:
        return False
    return old_normalized.replace(search_normalized, replacement_normalized, 1) == new_normalized


def _unique_context_text_change(old_text, new_text, current_text=""):
    old_text = _value_text(old_text)
    new_text = _value_text(new_text)
    if old_text == new_text:
        return "", ""
    prefix, old_end, new_end = _minimal_text_change(old_text, new_text)
    left = prefix
    right_old = old_end
    right_new = new_end
    comparison = _normalized_word_text(current_text or old_text)
    while True:
        old_part = old_text[left:right_old]
        new_part = new_text[left:right_new]
        normalized_part = _normalized_word_text(old_part)
        if (
            normalized_part
            and comparison.count(normalized_part) == 1
            and _replacement_reaches_target(old_text, new_text, old_part, new_part)
        ):
            return old_part, new_part
        if left <= 0 and right_old >= len(old_text):
            return old_text, new_text
        if left > 0:
            left -= 1
        if right_old < len(old_text):
            right_old += 1
            right_new += 1


def _word_replace_pair(old_text, new_text, current_text="", rule_old_text="", rule_new_text=""):
    rule_old_text = _value_text(rule_old_text)
    rule_new_text = _value_text(rule_new_text)
    current_normalized = _normalized_word_text(current_text or old_text)
    rule_normalized = _normalized_word_text(rule_old_text)
    if (
        rule_old_text
        and len(rule_old_text) <= WORD_FIND_TEXT_LIMIT
        and len(rule_new_text) <= WORD_FIND_TEXT_LIMIT
        and current_normalized.count(rule_normalized) == 1
        and _replacement_reaches_target(old_text, new_text, rule_old_text, rule_new_text)
    ):
        return rule_old_text, rule_new_text, "plan_rule"
    diff_old, diff_new = _unique_context_text_change(old_text, new_text, current_text)
    if diff_old and len(diff_old) <= WORD_FIND_TEXT_LIMIT and len(diff_new) <= WORD_FIND_TEXT_LIMIT:
        return diff_old, diff_new, "unique_context"
    return "", "", "full_value"


def _word_visible_text(range_obj):
    try:
        return _value_text(_word_visible_body_range(range_obj).Text)
    except Exception:
        return ""


def _verify_word_range_value(range_obj, expected):
    actual = _word_visible_text(range_obj)
    if _normalized_word_text(actual) != _normalized_word_text(expected):
        raise ValueError(
            "写入后校验失败"
            f"；期望={_visible_log_text(expected)}"
            f"；实际={_visible_log_text(actual)}"
        )


def _word_find_replace_visible_text(
    range_obj,
    old_text,
    value,
    rule_old_text="",
    rule_new_text="",
    raw_old_text="",
    replace_all=False,
    preserve_format=True,
    verify=True,
):
    old_text = str(old_text if old_text is not None else "")
    new_text = str(value if value is not None else "")
    if old_text == "":
        raise ValueError("缺少 old_text，无法执行查找替换")
    body = _word_visible_body_range(range_obj)
    if int(body.End) <= int(body.Start):
        raise ValueError("定位范围为空，无法执行查找替换")
    range_text = _word_visible_text(range_obj)
    if _normalized_word_text(range_text) == _normalized_word_text(new_text):
        return "already_target"

    search_text, replacement_text, pair_source = _word_replace_pair(
        old_text,
        new_text,
        current_text=range_text,
        rule_old_text=rule_old_text,
        rule_new_text=rule_new_text,
    )
    if search_text:
        finder = body.Find
        try:
            finder.ClearFormatting()
        except Exception:
            pass
        try:
            finder.Replacement.ClearFormatting()
        except Exception:
            pass
        find_error = None
        try:
            replaced = finder.Execute(
                FindText=search_text,
                MatchCase=False,
                MatchWholeWord=False,
                MatchWildcards=False,
                MatchSoundsLike=False,
                MatchAllWordForms=False,
                Forward=True,
                Wrap=0,
                Format=False,
                ReplaceWith=replacement_text,
                Replace=2 if replace_all else 1,
            )
        except Exception as exc:
            if _is_retryable_operation_error(exc):
                raise
            find_error = exc
            replaced = False
        current = _word_visible_text(range_obj)
        if replaced and _normalized_word_text(current) == _normalized_word_text(new_text):
            return f"find_{pair_source}"
        if current != range_text:
            _word_write_visible_text(range_obj, range_text, preserve_format)
            current = _word_visible_text(range_obj)
            if _normalized_word_text(current) != _normalized_word_text(range_text):
                raise ValueError("Word Find 校验失败后无法恢复原范围文本")
        if search_text in range_text:
            fallback_text = range_text.replace(search_text, replacement_text, -1 if replace_all else 1)
            _word_write_visible_text(range_obj, fallback_text, preserve_format)
            if verify:
                _verify_word_range_value(range_obj, new_text)
            return f"fallback_{pair_source}"
        if find_error is not None:
            range_text = _word_visible_text(range_obj)

    old_candidates = [old_text, raw_old_text]
    if any(
        candidate and _normalized_word_text(range_text) == _normalized_word_text(candidate)
        for candidate in old_candidates
    ):
        fallback_text = new_text
        _word_write_visible_text(range_obj, fallback_text, preserve_format)
        if verify:
            _verify_word_range_value(range_obj, new_text)
        return "fallback_full_value"
    raise ValueError(
        "定位范围内未找到 old_text，且规范化文本不一致"
        f"；old_text长度={len(old_text)}"
        f"；old_text={_visible_log_text(old_text)}"
        f"；规则旧值={_visible_log_text(search_text)}"
        f"；范围文本长度={len(range_text)}"
        f"；范围文本={_visible_log_text(range_text)}"
    )


def _word_write_range_by_mode(range_obj, op, context=None):
    strategy = _op_write_strategy(op)
    if strategy == WRITE_STRATEGY_DIRECT:
        _word_write_visible_text(range_obj, op.get("value", ""), _word_preserve_format_enabled(context))
        if _verification_enabled(context):
            _verify_word_range_value(range_obj, op.get("value", ""))
    elif strategy == WORD_MODE_FIND_REPLACE or (strategy == WRITE_STRATEGY_FOLLOW_NODE and _word_find_replace_enabled(context)):
        _word_find_replace_visible_text(
            range_obj,
            op.get("old_text", ""),
            op.get("value", ""),
            rule_old_text=op.get("rule_old_text", ""),
            rule_new_text=op.get("rule_new_text", ""),
            raw_old_text=_op_meta(op).get("word_raw_text", ""),
            replace_all=_op_replace_all(op, context),
            preserve_format=_word_preserve_format_enabled(context),
            verify=_verification_enabled(context),
        )
    else:
        _word_write_visible_text(range_obj, op.get("value", ""), _word_preserve_format_enabled(context))
        if _verification_enabled(context):
            _verify_word_range_value(range_obj, op.get("value", ""))


def _word_write_range_transactional(range_obj, op, context=None):
    original_text = _word_visible_text(range_obj)
    preserve_format = _word_preserve_format_enabled(context)
    try:
        return _word_write_range_by_mode(range_obj, op, context)
    except Exception as exc:
        current_text = _word_visible_text(range_obj)
        if current_text != original_text:
            try:
                _word_write_visible_text(range_obj, original_text, preserve_format)
                restored_text = _word_visible_text(range_obj)
                if _normalized_word_text(restored_text) != _normalized_word_text(original_text):
                    raise ValueError("恢复后文本仍不一致")
            except Exception as restore_exc:
                raise RuntimeError(f"{exc}；单项回滚失败：{restore_exc}") from exc
        raise


def _word_content_range(doc, op):
    meta = _op_meta(op)
    start = _to_int_or_none(meta.get("range_start"))
    end = _to_int_or_none(meta.get("range_end"))
    base = _to_int_or_none(meta.get("range_base")) or 0
    address_match = re.fullmatch(r"WRANGE(\d+):(\d+)", _as_text(op.get("cell_address", "")), flags=re.I)
    if (start is None or end is None) and address_match:
        start = int(address_match.group(1))
        end = int(address_match.group(2))
        base = 0
    if start is None or end is None or end <= start:
        raise ValueError("缺少有效的 Word 文本范围(range_start/range_end)")
    return doc.Range(int(base + start), int(base + end))


def _word_story_range(doc, meta):
    story_type = _to_int_or_none((meta or {}).get("story_type"))
    if not story_type:
        return None
    try:
        story = doc.StoryRanges(int(story_type))
    except Exception:
        try:
            story = doc.StoryRanges.Item(int(story_type))
        except Exception as exc:
            raise ValueError(f"无法定位 Word story_type={story_type}") from exc
    story_seq = _to_int_or_none((meta or {}).get("story_seq")) or 1
    for _index in range(1, story_seq):
        try:
            story = story.NextStoryRange
        except Exception:
            story = None
        if story is None:
            raise ValueError(f"Word story_seq 越界：{story_seq}")
    return story


def _word_shape_from_meta(doc, meta):
    shape_index = _to_int_or_none((meta or {}).get("shape_index"))
    if not shape_index or shape_index <= 0:
        raise ValueError("缺少 shape_index")
    scope = _as_text((meta or {}).get("shape_scope")).lower() or "document"
    if scope in ("document", "正文", "doc"):
        collection = doc.Shapes
    else:
        section_index = _to_int_or_none((meta or {}).get("section_index")) or 1
        area_type = _to_int_or_none((meta or {}).get("header_footer_type")) or 1
        try:
            section = doc.Sections(int(section_index))
        except Exception:
            section = doc.Sections.Item(int(section_index))
        collection_name = "Footers" if scope in ("footer", "页脚") else "Headers"
        collection = getattr(section, collection_name)(int(area_type)).Shapes
    try:
        return collection(int(shape_index))
    except Exception:
        return collection.Item(int(shape_index))


def _word_write_shape(shape, op, context=None):
    try:
        if shape.HasTextFrame and shape.TextFrame.HasText:
            _word_write_range_by_mode(shape.TextFrame.TextRange, op, context)
            return
    except Exception:
        pass
    old_text = _value_text(op.get("old_text"))
    new_text = _value_text(op.get("value"))
    try:
        current = _value_text(shape.TextEffect.Text)
        if _op_write_strategy(op) == WRITE_STRATEGY_DIRECT:
            shape.TextEffect.Text = new_text
        elif old_text and old_text in current:
            shape.TextEffect.Text = current.replace(old_text, new_text, -1 if _op_replace_all(op, context) else 1)
        else:
            raise ValueError("形状文本中未找到 old_text")
        return
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("形状不包含可写文字") from exc


def _word_replace_on_range(range_obj, old_text, new_text):
    old_text = _value_text(old_text)
    if old_text == "":
        raise ValueError("缺少 old_text，无法执行 Word 全局替换")
    finder = range_obj.Find
    try:
        finder.ClearFormatting()
        finder.Replacement.ClearFormatting()
    except Exception:
        pass
    return bool(finder.Execute(
        FindText=old_text,
        MatchCase=False,
        MatchWholeWord=False,
        MatchWildcards=False,
        MatchSoundsLike=False,
        MatchAllWordForms=False,
        Forward=True,
        Wrap=0,
        Format=False,
        ReplaceWith=_value_text(new_text),
        Replace=2,
    ))


def _word_collection_items(collection):
    try:
        count = int(collection.Count)
    except Exception:
        return []
    items = []
    for index in range(1, count + 1):
        try:
            items.append(collection(index))
        except Exception:
            try:
                items.append(collection.Item(index))
            except Exception:
                pass
    return items


def _word_shape_key(shape, scope=""):
    try:
        return (
            scope,
            _as_text(getattr(shape, "Name", "")),
            int(getattr(getattr(shape, "Anchor", None), "Start", -1)),
            int(getattr(shape, "Type", -1)),
        )
    except Exception:
        return (scope, id(shape))


def _word_replace_in_shape(shape, old_text, new_text, visited=None, scope=""):
    visited = visited if visited is not None else set()
    shape_key = _word_shape_key(shape, scope)
    if shape_key in visited:
        return 0
    visited.add(shape_key)
    replaced = 0
    try:
        text = _value_text(shape.TextEffect.Text)
        if old_text in text:
            shape.TextEffect.Text = text.replace(old_text, new_text)
            replaced += 1
    except Exception:
        pass
    try:
        if shape.HasTextFrame and shape.TextFrame.HasText:
            if _word_replace_on_range(shape.TextFrame.TextRange, old_text, new_text):
                replaced += 1
    except Exception:
        pass
    try:
        text_range = shape.TextFrame2.TextRange
        text = _value_text(text_range.Text)
        if old_text in text:
            text_range.Text = text.replace(old_text, new_text)
            replaced += 1
    except Exception:
        pass
    try:
        if int(shape.Type) == 6:
            for child in _word_collection_items(shape.GroupItems):
                replaced += _word_replace_in_shape(child, old_text, new_text, visited, f"{scope}/group")
    except Exception:
        pass
    return replaced


def _word_global_replace(doc, op):
    old_text = _value_text(op.get("old_text"))
    new_text = _value_text(op.get("value"))
    if old_text == "":
        raise ValueError("word_global_replace 缺少 old_text")
    if old_text == new_text:
        return 1
    replaced = 0
    visited_shapes = set()

    def range_key(range_obj):
        try:
            return (int(range_obj.StoryType), int(range_obj.Start), int(range_obj.End))
        except Exception:
            return ("range", id(range_obj))

    content_key = range_key(doc.Content)

    def replace_range(range_obj):
        nonlocal replaced
        if _word_replace_on_range(range_obj, old_text, new_text):
            replaced += 1

    replace_range(doc.Content)

    try:
        stories = list(doc.StoryRanges)
    except Exception:
        stories = []
    for first_story in stories:
        story = first_story
        seen_story_objects = set()
        story_steps = 0
        while story is not None:
            story_steps += 1
            object_key = id(story)
            if object_key in seen_story_objects or story_steps > 10000:
                break
            seen_story_objects.add(object_key)
            if range_key(story) != content_key:
                replace_range(story)
            try:
                story = story.NextStoryRange
            except Exception:
                break

    for shape in _word_collection_items(doc.Shapes):
        replaced += _word_replace_in_shape(shape, old_text, new_text, visited_shapes, "document")
    for inline_shape in _word_collection_items(doc.InlineShapes):
        replaced += _word_replace_in_shape(inline_shape, old_text, new_text, visited_shapes, "document-inline")

    try:
        section_count = int(doc.Sections.Count)
    except Exception:
        section_count = 0
    for section_index in range(1, section_count + 1):
        section = doc.Sections(section_index)
        for header_footer_type in (1, 2, 3):
            for collection_name in ("Headers", "Footers"):
                try:
                    area = getattr(section, collection_name)(header_footer_type)
                    for shape in _word_collection_items(area.Shapes):
                        replaced += _word_replace_in_shape(
                            shape,
                            old_text,
                            new_text,
                            visited_shapes,
                            f"{collection_name}:{section_index}:{header_footer_type}",
                        )
                except Exception:
                    pass
    if replaced <= 0:
        raise ValueError(
            "Word全文及特殊对象中未找到 old_text"
            f"；old_text长度={len(old_text)}"
            f"；old_text={_visible_log_text(old_text)}"
        )
    return replaced


def _word_op_range_start(op):
    if _as_text((op or {}).get("block_type")).lower() != BLOCK_WORD_TEXT_RANGE:
        return None
    meta = _op_meta(op)
    start = _to_int_or_none(meta.get("range_start"))
    base = _to_int_or_none(meta.get("range_base")) or 0
    address_match = re.fullmatch(r"WRANGE(\d+):(\d+)", _as_text((op or {}).get("cell_address", "")), flags=re.I)
    if start is None and address_match:
        start = int(address_match.group(1))
        base = 0
    return None if start is None else base + start


def _ordered_word_ops(ops):
    indexed = list(enumerate(ops or []))
    range_ops = [(index, op) for index, op in indexed if _word_op_range_start(op) is not None]
    other_ops = [(index, op) for index, op in indexed if _word_op_range_start(op) is None]
    range_ops.sort(key=lambda item: (_word_op_range_start(item[1]), item[0]), reverse=True)
    ordered_range_ops = [op for _index, op in range_ops]
    ordered_other_ops = _sort_global_replace_runs_by_old_text_length([op for _index, op in other_ops])
    return ordered_range_ops + ordered_other_ops


def _apply_word_com_op(doc, op, context=None):
    bt = _as_text(op.get("block_type", "")).lower()
    if bt == "word_paragraph":
        paragraph_index = _to_int_or_none(op.get("row_index"))
        if not paragraph_index or paragraph_index <= 0:
            raise ValueError("缺少 paragraph 索引(row_index)")
        story = _word_story_range(doc, _op_meta(op))
        paragraphs = story.Paragraphs if story is not None else doc.Paragraphs
        _word_write_range_transactional(paragraphs(int(paragraph_index)).Range, op, context)
        return
    if bt == BLOCK_WORD_TEXT_RANGE:
        range_obj = _word_content_range(doc, op)
        old_text = _value_text(op.get("old_text"))
        if old_text == "":
            raise ValueError("word_text_range 必须提供 old_text 以避免范围漂移误写")
        _word_write_range_transactional(range_obj, op, context)
        return
    if bt == "word_table_cell":
        table_index = _parse_table_index(op.get("sheet_name", ""), op.get("meta_json", ""))
        if not table_index:
            raise ValueError("无法确定表索引（sheet_name 或 meta_json.table_index）")
        row_index = _to_int_or_none(op.get("row_index"))
        col_index = _to_int_or_none(op.get("col_index"))
        if not row_index or not col_index:
            parsed_row, parsed_col = _parse_rc_from_address(op.get("cell_address", ""))
            row_index = row_index or parsed_row
            col_index = col_index or parsed_col
        meta = _op_meta(op)
        row_index = _to_int_or_none(meta.get("merge_origin_row")) or row_index
        col_index = _to_int_or_none(meta.get("merge_origin_col")) or col_index
        if not row_index or not col_index:
            raise ValueError("缺少单元格行列索引(row_index/col_index/cell_address)")
        table = doc.Tables(int(table_index))
        cell_index = _to_int_or_none(meta.get("cell_index"))
        if cell_index:
            try:
                cell = table.Range.Cells(int(cell_index))
            except Exception:
                cell = table.Range.Cells.Item(int(cell_index))
        else:
            cell = table.Cell(int(row_index), int(col_index))
        _word_write_range_transactional(cell.Range, op, context)
        return
    if bt == BLOCK_WORD_CONTENT_CONTROL:
        control_index = _to_int_or_none(_op_meta(op).get("content_control_index")) or _to_int_or_none(op.get("row_index"))
        if not control_index or control_index <= 0:
            raise ValueError("缺少 content_control_index")
        try:
            control = doc.ContentControls(int(control_index))
        except Exception:
            control = doc.ContentControls.Item(int(control_index))
        _word_write_range_transactional(control.Range, op, context)
        return
    if bt == BLOCK_WORD_SHAPE_TEXT:
        _word_write_shape(_word_shape_from_meta(doc, _op_meta(op)), op, context)
        return
    if bt == BLOCK_WORD_GLOBAL_REPLACE:
        _word_global_replace(doc, op)
        return
    raise ValueError(f"不支持的 Word block_type：{bt}")


def _excel_values_equal(expected, actual):
    if expected is None and actual is None:
        return True
    return _value_text(expected) == _value_text(actual)


def _excel_com_target(worksheet, op):
    meta = _op_meta(op)
    row_index = _to_int_or_none(meta.get("merge_origin_row")) or _to_int_or_none(op.get("row_index"))
    col_index = _to_int_or_none(meta.get("merge_origin_col")) or _to_int_or_none(op.get("col_index"))
    address = _as_text(op.get("cell_address", "")).replace("$", "")
    if row_index and col_index:
        target = worksheet.Cells(int(row_index), int(col_index))
    elif address:
        target = worksheet.Range(address)
    else:
        raise ValueError("缺少 Excel 定位(row/col/cell_address)")
    try:
        if bool(target.MergeCells):
            target = target.MergeArea.Cells(1, 1)
    except Exception:
        pass
    return target


def _apply_excel_com_op(workbook, op, context=None):
    sheet_name = _as_text(op.get("sheet_name", ""))
    if sheet_name:
        try:
            worksheet = workbook.Worksheets(sheet_name)
        except Exception as exc:
            raise ValueError(f"Excel 工作表不存在：{sheet_name}") from exc
    else:
        worksheet = workbook.Worksheets(1)
    target = _excel_com_target(worksheet, op)
    target.Value = op.get("value", "")
    if _verification_enabled(context) and not _excel_values_equal(op.get("value", ""), target.Value):
        raise ValueError(
            f"Excel写入后校验失败：期望={_visible_log_text(op.get('value', ''))}"
            f"；实际={_visible_log_text(target.Value)}"
        )


def _is_retryable_operation_error(exc):
    if isinstance(exc, (ValueError, KeyError, IndexError, TypeError)):
        return False
    text = _value_text(exc).lower()
    if any(item in text for item in (
        "字符串参量过长",
        "未找到 old_text",
        "校验失败",
        "索引越界",
        "缺少 ",
        "不支持",
    )):
        return False
    return any(item in text for item in (
        "-2147418111",
        "rpc_e_call_rejected",
        "call was rejected by callee",
        "被呼叫方拒绝接收呼叫",
        "server busy",
        "服务器忙",
        "应用程序正忙",
        "rpc server is unavailable",
        "rpc 服务器不可用",
    ))


def _apply_operation_with_retry(action, op, context, label):
    block_type = _as_text((op or {}).get("block_type", "")).lower()
    attempts = 1 if block_type == BLOCK_WORD_GLOBAL_REPLACE else _retry_count(context, "win32_cell_retries", 3)
    interval = _retry_interval(context)
    last_exc = None
    for attempt in range(1, attempts + 1):
        try:
            return action()
        except Exception as exc:
            last_exc = exc
            if attempt >= attempts or not _is_retryable_operation_error(exc):
                raise
            if interval > 0:
                time.sleep(interval * attempt)
    raise RuntimeError(f"{label}失败，已重试 {attempts} 次：{last_exc}") from last_exc


def _save_with_retry(action, context, label):
    return _run_with_retry(
        action,
        _retry_count(context, "win32_save_retries", 3),
        _retry_interval(context),
        label,
    )


def _word_file_format(file_path):
    return {".doc": 0, ".docx": 12, ".docm": 13}.get(Path(file_path).suffix.lower())


def _save_word_with_fallback(doc, file_path, context):
    try:
        return _save_with_retry(doc.Save, context, "Word保存")
    except Exception as save_exc:
        file_format = _word_file_format(file_path)

        def save_as():
            kwargs = {"FileFormat": file_format} if file_format is not None else {}
            try:
                return doc.SaveAs2(str(file_path), **kwargs)
            except AttributeError:
                return doc.SaveAs(str(file_path), **kwargs)

        try:
            return _save_with_retry(save_as, context, "Word SaveAs兜底保存")
        except Exception as save_as_exc:
            raise RuntimeError(f"Word保存及SaveAs兜底均失败：{save_exc}；{save_as_exc}") from save_as_exc


def _record_operation_failure(logs, message, file_path, context):
    if _stop_on_error(context):
        raise RuntimeError(message)
    logs.append(_log("ERROR", message, str(file_path)))


def _write_word_via_com(file_path, ops, context=None, progress_current=None, progress_total=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("win32 写入 Word 需要 pywin32 + Word") from exc

    word = None
    doc = None
    applied = 0
    skipped = 0
    logs = []
    try:
        pythoncom.CoInitialize()
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0
        doc = _run_with_retry(
            lambda: word.Documents.Open(
                str(file_path),
                ReadOnly=False,
                AddToRecentFiles=False,
                ConfirmConversions=False,
                Visible=False,
            ),
            _retry_count(context, "win32_open_retries", 5),
            _retry_interval(context),
            "Word打开",
        )

        ordered_ops = _ordered_word_ops(ops)
        for op_no, op in enumerate(ordered_ops, start=1):
            _op_progress(context, progress_current, progress_total, file_path, op, op_no)
            try:
                _apply_operation_with_retry(
                    lambda: _apply_word_com_op(doc, op, context),
                    op,
                    context,
                    f"Word写入(源行{op.get('source_row')})",
                )
                applied += 1
            except Exception as exc:
                skipped += 1
                _record_operation_failure(
                    logs,
                    _operation_failure_message("写入失败", op, exc),
                    file_path,
                    context,
                )
        _save_word_with_fallback(doc, file_path, context)
        return applied, skipped, logs
    finally:
        if doc is not None:
            try:
                doc.Close(False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _write_excel_via_com(file_path, ops, context=None, progress_current=None, progress_total=None):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError("win32 写入 Excel 需要 pywin32 + Excel") from exc

    excel = None
    wb = None
    applied = 0
    skipped = 0
    logs = []
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = _run_with_retry(
            lambda: excel.Workbooks.Open(str(file_path), ReadOnly=False, UpdateLinks=0),
            _retry_count(context, "win32_open_retries", 5),
            _retry_interval(context),
            "Excel打开",
        )

        for op_no, op in enumerate(ops, start=1):
            _op_progress(context, progress_current, progress_total, file_path, op, op_no)
            try:
                _apply_operation_with_retry(
                    lambda: _apply_excel_com_op(wb, op, context),
                    op,
                    context,
                    f"Excel写入(源行{op.get('source_row')})",
                )
                applied += 1
            except Exception as exc:
                skipped += 1
                _record_operation_failure(
                    logs,
                    _operation_failure_message("Excel写入失败", op, exc),
                    file_path,
                    context,
                )

        _save_with_retry(wb.Save, context, "Excel保存")
        return applied, skipped, logs
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


class _Win32OfficeSession:
    def __init__(self, open_retries=5, retry_interval_ms=300, close_settle_ms=200):
        self.open_retries = max(1, int(open_retries or 1))
        self.retry_interval = max(0, int(retry_interval_ms or 0)) / 1000.0
        self.close_settle = max(0, int(close_settle_ms or 0)) / 1000.0
        self.pythoncom = None
        self.win32com_client = None
        self.com_initialized = False
        self.word = None
        self.excel = None

    def _ensure_com(self):
        if self.com_initialized:
            return
        try:
            import pythoncom
            import win32com.client
        except Exception as exc:
            raise RuntimeError("win32 写入需要 pywin32 + Office") from exc
        pythoncom.CoInitialize()
        self.pythoncom = pythoncom
        self.win32com_client = win32com.client
        self.com_initialized = True

    def _word_app(self):
        self._ensure_com()
        if self.word is None:
            self.word = self.win32com_client.DispatchEx("Word.Application")
            self.word.Visible = False
            self.word.DisplayAlerts = 0
        return self.word

    def _excel_app(self):
        self._ensure_com()
        if self.excel is None:
            self.excel = self.win32com_client.DispatchEx("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
        return self.excel

    def _open_with_retry(self, opener, file_path):
        last_exc = None
        for attempt in range(1, self.open_retries + 1):
            try:
                return opener()
            except Exception as exc:
                last_exc = exc
                if attempt < self.open_retries and self.retry_interval > 0:
                    time.sleep(self.retry_interval)
        raise RuntimeError(f"打开文件失败，已重试 {self.open_retries} 次：{file_path}；{last_exc}")

    def write_word(self, file_path, ops, context=None, progress_current=None, progress_total=None):
        doc = None
        saved = False
        applied = 0
        skipped = 0
        logs = []
        try:
            word = self._word_app()
            doc = self._open_with_retry(
                lambda: word.Documents.Open(str(file_path), ReadOnly=False, AddToRecentFiles=False, ConfirmConversions=False, Visible=False),
                file_path,
            )
            ordered_ops = _ordered_word_ops(ops)
            for op_no, op in enumerate(ordered_ops, start=1):
                _op_progress(context, progress_current, progress_total, file_path, op, op_no)
                try:
                    _apply_operation_with_retry(
                        lambda: _apply_word_com_op(doc, op, context),
                        op,
                        context,
                        f"Word写入(源行{op.get('source_row')})",
                    )
                    applied += 1
                except Exception as exc:
                    skipped += 1
                    _record_operation_failure(
                        logs,
                        _operation_failure_message("写入失败", op, exc),
                        file_path,
                        context,
                    )
            _save_word_with_fallback(doc, file_path, context)
            saved = True
            return applied, skipped, logs
        finally:
            if doc is not None:
                try:
                    doc.Close(SaveChanges=False)
                except Exception:
                    try:
                        doc.Close(False)
                    except Exception:
                        pass
            if saved and self.close_settle > 0:
                time.sleep(self.close_settle)

    def write_excel(self, file_path, ops, context=None, progress_current=None, progress_total=None):
        wb = None
        saved = False
        applied = 0
        skipped = 0
        logs = []
        try:
            excel = self._excel_app()
            wb = self._open_with_retry(
                lambda: excel.Workbooks.Open(str(file_path), ReadOnly=False, UpdateLinks=0),
                file_path,
            )
            for op_no, op in enumerate(ops, start=1):
                _op_progress(context, progress_current, progress_total, file_path, op, op_no)
                try:
                    _apply_operation_with_retry(
                        lambda: _apply_excel_com_op(wb, op, context),
                        op,
                        context,
                        f"Excel写入(源行{op.get('source_row')})",
                    )
                    applied += 1
                except Exception as exc:
                    skipped += 1
                    _record_operation_failure(
                        logs,
                        _operation_failure_message("Excel写入失败", op, exc),
                        file_path,
                        context,
                    )
            _save_with_retry(wb.Save, context, "Excel保存")
            saved = True
            return applied, skipped, logs
        finally:
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if saved and self.close_settle > 0:
                time.sleep(self.close_settle)

    def write_file(self, file_path, ops, context=None, progress_current=None, progress_total=None):
        ext = file_path.suffix.lower()
        if ext in (".doc", ".docx", ".docm"):
            return self.write_word(file_path, ops, context, progress_current, progress_total)
        if ext in (".xls", ".xlsx", ".xlsm"):
            return self.write_excel(file_path, ops, context, progress_current, progress_total)
        raise ValueError(f"win32 不支持的文件类型：{ext}")

    def close(self):
        if self.word is not None:
            try:
                self.word.Quit()
            except Exception:
                pass
            self.word = None
        if self.excel is not None:
            try:
                self.excel.Quit()
            except Exception:
                pass
            self.excel = None
        if self.com_initialized and self.pythoncom is not None:
            try:
                self.pythoncom.CoUninitialize()
            except Exception:
                pass
        self.com_initialized = False


def _set_xml_text_node(text_node, text):
    txt = str(text or "")
    if txt.startswith(" ") or txt.endswith(" "):
        text_node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    else:
        text_node.attrib.pop("{http://www.w3.org/XML/1998/namespace}space", None)
    text_node.text = txt


def _set_word_paragraph_text(p_node, text, ns_w, preserve_format=True):
    text_nodes = list(p_node.iter(f"{{{ns_w}}}t"))
    if preserve_format and text_nodes:
        _set_xml_text_node(text_nodes[0], text)
        for text_node in text_nodes[1:]:
            _set_xml_text_node(text_node, "")
        return

    paragraph_properties = p_node.find(f"{{{ns_w}}}pPr")
    for child in list(p_node):
        if child is not paragraph_properties:
            p_node.remove(child)
    run = ET.SubElement(p_node, f"{{{ns_w}}}r")
    text_node = ET.SubElement(run, f"{{{ns_w}}}t")
    _set_xml_text_node(text_node, text)


def _clear_word_paragraph_text(p_node, ns_w):
    for text_node in p_node.iter(f"{{{ns_w}}}t"):
        _set_xml_text_node(text_node, "")


def _set_word_cell_text(cell_node, text, ns, preserve_format=True):
    paragraphs = cell_node.findall("./w:p", ns)
    if not paragraphs:
        paragraphs = [ET.SubElement(cell_node, f"{{{ns['w']}}}p")]
    _set_word_paragraph_text(paragraphs[0], text, ns["w"], preserve_format=preserve_format)
    for paragraph in paragraphs[1:]:
        _clear_word_paragraph_text(paragraph, ns["w"])


def _xml_grid_span(cell_node, ns):
    span_node = cell_node.find("./w:tcPr/w:gridSpan", ns)
    if span_node is None:
        return 1
    raw = span_node.get(f"{{{ns['w']}}}val", span_node.get("val", "1"))
    try:
        return max(1, int(raw or 1))
    except Exception:
        return 1


def _xml_cell_for_logical_column(row_node, logical_column, ns):
    current_column = 1
    for cell_node in row_node.findall("./w:tc", ns):
        span = _xml_grid_span(cell_node, ns)
        if current_column <= logical_column < current_column + span:
            return cell_node, current_column
        current_column += span
    return None, current_column


def _replace_xml_text(root, old_text, new_text):
    replaced = 0
    for node in root.iter():
        local_name = node.tag.split("}")[-1]
        if local_name in {"t", "instrText", "textpath"}:
            current = node.text or ""
            if old_text in current:
                node.text = current.replace(old_text, new_text)
                replaced += 1
        for attribute_name, attribute_value in list(node.attrib.items()):
            if attribute_name.split("}")[-1] not in {"string", "text"}:
                continue
            if old_text in attribute_value:
                node.set(attribute_name, attribute_value.replace(old_text, new_text))
                replaced += 1
    return replaced


def _apply_docx_global_replacements(payload, replacements):
    if not replacements:
        return payload, []
    updated = dict(payload)
    replacement_counts = [0 for _item in replacements]
    for name, data in payload.items():
        if not name.lower().endswith(".xml") or not name.lower().startswith("word/"):
            continue
        try:
            root = ET.fromstring(data)
        except Exception:
            continue
        file_replaced = 0
        for replacement_index, (old_text, new_text) in enumerate(replacements):
            count = _replace_xml_text(root, old_text, new_text)
            replacement_counts[replacement_index] += count
            file_replaced += count
        if file_replaced:
            updated[name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return updated, replacement_counts


def _write_docx_zip_xml(file_path, ops, context=None, progress_current=None, progress_total=None):
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    ns = {"w": ns_w}
    applied = 0
    skipped = 0
    logs = []

    with zipfile.ZipFile(file_path, "r") as zf:
        names = zf.namelist()
        if "word/document.xml" not in names:
            raise RuntimeError("docx 缺少 word/document.xml")
        payload = {name: zf.read(name) for name in names}

    root = ET.fromstring(payload["word/document.xml"])
    body = root.find("w:body", ns)
    if body is None:
        raise RuntimeError("docx 文档 body 不存在")

    body_children = list(body)
    paragraphs = [x for x in body_children if x.tag.split("}")[-1] == "p"]
    tables = [x for x in body_children if x.tag.split("}")[-1] == "tbl"]
    global_replacements = []
    preserve_format = _word_preserve_format_enabled(context)

    for op_no, op in enumerate(ops, start=1):
        _op_progress(context, progress_current, progress_total, file_path, op, op_no)
        bt = _as_text(op.get("block_type", "")).lower()
        try:
            if bt == "word_paragraph":
                pi = _to_int_or_none(op.get("row_index"))
                if not pi or pi <= 0 or pi > len(paragraphs):
                    raise ValueError(f"paragraph 索引越界：{pi}")
                _set_word_paragraph_text(
                    paragraphs[pi - 1],
                    op.get("value", ""),
                    ns_w,
                    preserve_format=preserve_format,
                )
                applied += 1
            elif bt == "word_table_cell":
                table_idx = _parse_table_index(op.get("sheet_name", ""), op.get("meta_json", ""))
                if not table_idx or table_idx <= 0 or table_idx > len(tables):
                    raise ValueError(f"table 索引越界：{table_idx}")
                row_i = _to_int_or_none(op.get("row_index"))
                col_i = _to_int_or_none(op.get("col_index"))
                if not row_i or not col_i:
                    rr, cc = _parse_rc_from_address(op.get("cell_address", ""))
                    row_i = row_i or rr
                    col_i = col_i or cc
                if not row_i or not col_i:
                    raise ValueError("缺少单元格行列索引")
                meta = _op_meta(op)
                row_i = _to_int_or_none(meta.get("merge_origin_row")) or row_i
                col_i = _to_int_or_none(meta.get("merge_origin_col")) or col_i

                tbl = tables[table_idx - 1]
                trs = tbl.findall("./w:tr", ns)
                if row_i <= 0 or row_i > len(trs):
                    raise ValueError(f"行索引越界：{row_i}")
                tc, cell_start_column = _xml_cell_for_logical_column(trs[row_i - 1], col_i, ns)
                if tc is None:
                    raise ValueError(f"列索引越界：{col_i}")
                if cell_start_column != col_i:
                    raise ValueError(f"列 {col_i} 位于合并单元格内部，合并起点为列 {cell_start_column}")
                _set_word_cell_text(
                    tc,
                    op.get("value", ""),
                    ns,
                    preserve_format=preserve_format,
                )
                applied += 1
            elif bt == BLOCK_WORD_GLOBAL_REPLACE:
                old_text = _value_text(op.get("old_text"))
                if old_text == "":
                    raise ValueError("word_global_replace 缺少 old_text")
                global_replacements.append((old_text, _value_text(op.get("value"))))
                applied += 1
            else:
                raise ValueError(f"zip_xml 不支持的 Word block_type：{bt}")
        except Exception as exc:
            skipped += 1
            _record_operation_failure(
                logs,
                _operation_failure_message("写入失败", op, exc),
                file_path,
                context,
            )

    if applied > 0:
        payload["word/document.xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        global_replacements = _sort_global_replacements_by_old_text_length(global_replacements)
        payload, replacement_counts = _apply_docx_global_replacements(payload, global_replacements)
        for (old_text, _new_text), replace_count in zip(global_replacements, replacement_counts):
            if replace_count > 0:
                continue
            message = (
                "Word全文及特殊对象中未找到 old_text"
                f"；old_text长度={len(_value_text(old_text))}"
                f"；old_text={_visible_log_text(old_text)}"
            )
            if _stop_on_error(context):
                raise RuntimeError(message)
            logs.append(_log("ERROR", message, str(file_path)))
            skipped += 1
            applied -= 1
        tmp_fd = tempfile.NamedTemporaryFile(delete=False, suffix=".docx", dir=str(file_path.parent))
        tmp_path = Path(tmp_fd.name)
        tmp_fd.close()
        try:
            with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zf:
                for name, data in payload.items():
                    out_zf.writestr(name, data)
            tmp_path.replace(file_path)
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

    return applied, skipped, logs


def _openpyxl_target_address(worksheet, op):
    meta = _op_meta(op)
    row_i = _to_int_or_none(meta.get("merge_origin_row")) or _to_int_or_none(op.get("row_index"))
    col_i = _to_int_or_none(meta.get("merge_origin_col")) or _to_int_or_none(op.get("col_index"))
    addr = _as_text(op.get("cell_address", "")).replace("$", "")
    if row_i and col_i:
        addr = _excel_addr_from_rc(row_i, col_i)
    if not addr:
        rr, cc = _parse_rc_from_address(op.get("cell_address", ""))
        if rr and cc:
            addr = _excel_addr_from_rc(rr, cc)
    if not addr:
        raise ValueError("缺少 Excel 定位(row/col/cell_address)")
    for merged_range in worksheet.merged_cells.ranges:
        if addr in merged_range:
            return merged_range.start_cell.coordinate
    return addr


def _write_excel_openpyxl(file_path, ops, context=None, progress_current=None, progress_total=None):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("zip_xml 写入 Excel 需要 openpyxl") from exc

    ext = file_path.suffix.lower()
    keep_vba = ext in (".xlsm", ".xltm")
    wb = load_workbook(filename=str(file_path), keep_vba=keep_vba)
    applied = 0
    skipped = 0
    logs = []
    try:
        for op_no, op in enumerate(ops, start=1):
            _op_progress(context, progress_current, progress_total, file_path, op, op_no)
            try:
                sheet_name = _as_text(op.get("sheet_name", ""))
                if sheet_name and sheet_name not in wb.sheetnames:
                    raise ValueError(f"Excel 工作表不存在：{sheet_name}")
                ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
                addr = _openpyxl_target_address(ws, op)
                ws[addr] = op.get("value", "")
                if _verification_enabled(context) and not _excel_values_equal(op.get("value", ""), ws[addr].value):
                    raise ValueError(
                        f"Excel写入后校验失败：期望={_visible_log_text(op.get('value', ''))}"
                        f"；实际={_visible_log_text(ws[addr].value)}"
                    )
                applied += 1
            except Exception as exc:
                skipped += 1
                _record_operation_failure(
                    logs,
                    _operation_failure_message("Excel写入失败", op, exc),
                    file_path,
                    context,
                )
        wb.save(str(file_path))
        return applied, skipped, logs
    finally:
        wb.close()


def _write_file(file_path, ops, engine, context=None, progress_current=None, progress_total=None):
    ext = file_path.suffix.lower()
    if engine == "win32":
        if ext in (".doc", ".docx", ".docm"):
            return _write_word_via_com(file_path, ops, context, progress_current, progress_total)
        if ext in (".xls", ".xlsx", ".xlsm"):
            return _write_excel_via_com(file_path, ops, context, progress_current, progress_total)
        raise ValueError(f"win32 不支持的文件类型：{ext}")

    # zip_xml
    if ext == ".doc":
        raise ValueError("zip_xml 不支持 .doc，请改用 win32")
    if ext in (".docx", ".docm"):
        return _write_docx_zip_xml(file_path, ops, context, progress_current, progress_total)
    if ext in (".xlsx", ".xlsm"):
        return _write_excel_openpyxl(file_path, ops, context, progress_current, progress_total)
    if ext == ".xls":
        raise ValueError("zip_xml 不支持 .xls，请改用 win32")
    raise ValueError(f"zip_xml 不支持的文件类型：{ext}")


def _prepare_target_file(source_path, target_path, params, preview_protected):
    missing_policy = _as_text(params.get("target_missing_policy", "从源文件复制")) or "从源文件复制"
    existing_policy = _as_text(params.get("target_existing_policy", "直接写入")) or "直接写入"
    same_policy = _as_text(params.get("same_path_policy", "修改源文件")) or "修改源文件"
    create_dirs = bool(params.get("create_parent_dirs", True))

    same_path = _same_file_path(source_path, target_path)
    source_exists = source_path.exists()
    target_exists = target_path.exists()

    if same_path:
        if same_policy == "跳过":
            return "skip", "新旧路径相同，按设置跳过"
        if same_policy == "报错":
            raise RuntimeError("新旧路径相同，按设置报错")
        if not source_exists:
            raise FileNotFoundError(f"文件不存在：{source_path}")
        return "原地写入", ""

    if not source_exists:
        raise FileNotFoundError(f"源文件不存在：{source_path}")

    if create_dirs and target_path.parent:
        if not preview_protected:
            target_path.parent.mkdir(parents=True, exist_ok=True)

    if target_exists:
        if existing_policy == "跳过":
            return "skip", "目标文件已存在，按设置跳过"
        if existing_policy == "报错":
            raise RuntimeError(f"目标文件已存在：{target_path}")
        if existing_policy == "覆盖为源文件后写入":
            if preview_protected:
                return "预览覆盖复制", ""
            shutil.copy2(str(source_path), str(target_path))
            return "已覆盖复制", ""
        return "目标已存在直接写入", ""

    if missing_policy == "报错":
        raise FileNotFoundError(f"目标文件不存在：{target_path}")
    if preview_protected:
        return "预览复制", ""
    shutil.copy2(str(source_path), str(target_path))
    return "已复制", ""


def _neighbor_temp_path(target_path, label):
    file_descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{target_path.stem}.{label}.",
        suffix=target_path.suffix,
        dir=str(target_path.parent),
    )
    os.close(file_descriptor)
    return Path(temp_name)


def _snapshot_target(target_path, preview_protected):
    state = {
        "existed": bool(target_path.exists()),
        "snapshot": None,
    }
    if preview_protected or not state["existed"]:
        return state
    snapshot = _neighbor_temp_path(target_path, "original")
    shutil.copy2(str(target_path), str(snapshot))
    state["snapshot"] = snapshot
    return state


def _cleanup_path(path):
    if not path:
        return
    try:
        Path(path).unlink()
    except FileNotFoundError:
        pass
    except Exception:
        pass


def _rollback_target(target_path, state):
    snapshot = (state or {}).get("snapshot")
    if (state or {}).get("existed"):
        if snapshot and Path(snapshot).exists():
            shutil.copy2(str(snapshot), str(target_path))
    else:
        _cleanup_path(target_path)


def _permanent_backup_path(target_path):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return target_path.with_name(f"{target_path.stem}_backup_{stamp}{target_path.suffix}")


def _prepare_working_copy(target_path):
    working_path = _neighbor_temp_path(target_path, "working")
    shutil.copy2(str(target_path), str(working_path))
    return working_path


def _commit_working_copy(working_path, target_path, state, backup_mode):
    backup_path = None
    snapshot = (state or {}).get("snapshot")
    if backup_mode == "写入前保留备份" and (state or {}).get("existed") and snapshot and Path(snapshot).exists():
        backup_path = _permanent_backup_path(target_path)
        shutil.copy2(str(snapshot), str(backup_path))
    os.replace(str(working_path), str(target_path))
    return backup_path


def _error_messages(logs):
    return [
        _as_text(item.get("message"))
        for item in (logs or [])
        if isinstance(item, dict) and _as_text(item.get("level")).upper() == "ERROR"
    ]


def _run_impl(input_data, params, context, _session_holder=None):
    p = dict(params or {})
    context = dict(context or {})
    context["params"] = p
    engine = _as_text(p.get("write_engine", "win32")).lower() or "win32"
    error_policy = _as_text(p.get("error_policy", "继续并记录失败"))
    preview_write = bool(p.get("preview_write_files", False))
    is_preview = bool(context.get("is_preview", False))
    win32_reuse_app = bool(p.get("win32_reuse_app", True))
    win32_open_retries = _to_int_with_default(p.get("win32_open_retries", 5), 5, min_value=1)
    win32_retry_interval_ms = _to_int_with_default(p.get("win32_retry_interval_ms", 300), 300, min_value=0)
    win32_close_settle_ms = _to_int_with_default(p.get("win32_close_settle_ms", 200), 200, min_value=0)
    backup_mode = _as_text(p.get("backup_mode", "失败时恢复原文件")) or "失败时恢复原文件"

    grouped, prep = _collect_ops(input_data, p, context)
    total_files = len(grouped)
    total_ops = sum(len(item["ops"]) for item in grouped.values())
    _progress(context, 0, total_files if total_files > 0 else 1, f"准备写入：{total_files} 个文件", stage="prepare")

    logs = []
    if prep["skipped_no_path"] > 0:
        logs.append(_log("WARNING", f"有 {prep['skipped_no_path']} 行缺少路径，已跳过"))
    if prep["skipped_empty_value"] > 0:
        logs.append(_log("INFO", f"有 {prep['skipped_empty_value']} 行空文本，按配置已跳过"))
    if prep["skipped_duplicate_global"] > 0:
        logs.append(_log("INFO", f"有 {prep['skipped_duplicate_global']} 条重复全文替换操作，已合并"))
    if prep["skipped_duplicate_target"] > 0:
        logs.append(_log("INFO", f"有 {prep['skipped_duplicate_target']} 条同位置同值操作，已合并"))
    if prep["skipped_conflict_replaced"] > 0:
        logs.append(_log("WARNING", f"有 {prep['skipped_conflict_replaced']} 条同位置冲突操作，按配置保留最后一条"))
    if prep["target_conflicts"] > 0 and _as_text(p.get("target_conflict_policy", CONFLICT_WARN)) == CONFLICT_WARN:
        logs.append(_log("WARNING", f"检测到 {prep['target_conflicts']} 处同位置多值写入，将按输入顺序执行"))

    if total_ops <= 0:
        return {
            "ok": True,
            "message": "没有可写入的数据",
            "output": {
                "type": "table",
                "headers": ["提示"],
                "rows": [["没有可写入的数据"]],
                "meta": {"plugin": PLUGIN_INFO["id"]},
            },
            "logs": logs,
            "summary": {
                "input_rows": prep["input_rows"],
                "total_files": 0,
                "total_ops": 0,
                "applied": 0,
                "skipped": (
                    prep["skipped_no_path"]
                    + prep["skipped_empty_value"]
                    + prep["skipped_duplicate_global"]
                    + prep["skipped_duplicate_target"]
                    + prep["skipped_conflict_replaced"]
                ),
                "failed_files": 0,
                "preview_protected": bool(is_preview and not preview_write),
            },
        }

    out_headers = list(OUTPUT_HEADERS)
    out_rows = []
    total_applied = 0
    total_skipped = (
        prep["skipped_no_path"]
        + prep["skipped_empty_value"]
        + prep["skipped_duplicate_global"]
        + prep["skipped_duplicate_target"]
        + prep["skipped_conflict_replaced"]
    )
    failed_files = 0
    partial_files = 0
    successful_files = 0
    skipped_files = 0
    preview_protected = bool(is_preview and not preview_write)
    processed_files = 0
    cancelled = False
    win32_session = None
    if engine == "win32" and win32_reuse_app and not preview_protected:
        win32_session = _Win32OfficeSession(
            open_retries=win32_open_retries,
            retry_interval_ms=win32_retry_interval_ms,
            close_settle_ms=win32_close_settle_ms,
        )
        if isinstance(_session_holder, dict):
            _session_holder["session"] = win32_session
        logs.append(_log("INFO", "WIN32_REUSE：本节点将复用同一个 Word/Excel 进程处理多个文件"))

    for i, item in enumerate(grouped.values(), start=1):
        if _check_cancel(context):
            cancelled = True
            logs.append(_log("WARNING", "检测到取消信号，已停止后续写入"))
            _progress(context, i - 1, total_files, "已取消，停止后续写入", stage="cancelled")
            break
        source_path = item["source_path"]
        target_path = item["target_path"]
        ops = item["ops"]
        _progress(context, i - 1, total_files, f"写入中：{target_path.name}", stage="file_start", object=str(target_path))
        if item.get("source_conflict"):
            err = f"同一个目标文件对应多个源文件：{target_path}"
            failed_files += 1
            total_skipped += len(ops)
            processed_files += 1
            out_rows.append([str(source_path), str(target_path), target_path.name, engine, len(ops), 0, len(ops), "", "失败", "失败", err])
            logs.append(_log("ERROR", err, str(target_path)))
            _progress(context, i, total_files, f"失败：{target_path.name}", stage="file_error", object=str(target_path))
            if error_policy == "遇错停止":
                if win32_session is not None:
                    win32_session.close()
                    win32_session = None
                raise RuntimeError(err)
            continue
        if item.get("conflict_error"):
            conflict_detail = "；".join((item.get("target_conflicts") or [])[:5])
            err = f"同一目标位置存在多条不同写入：{conflict_detail}"
            failed_files += 1
            total_skipped += len(ops)
            processed_files += 1
            out_rows.append([str(source_path), str(target_path), target_path.name, engine, len(ops), 0, len(ops), "", "失败", "失败", err])
            logs.append(_log("ERROR", err, str(target_path)))
            if error_policy == "遇错停止":
                if win32_session is not None:
                    win32_session.close()
                    win32_session = None
                    if isinstance(_session_holder, dict):
                        _session_holder["session"] = None
                raise RuntimeError(err)
            continue
        for conflict_message in (item.get("target_conflicts") or []):
            logs.append(_log("WARNING", f"同位置写入冲突：{conflict_message}", str(target_path)))

        copy_status = ""
        transaction_state = _snapshot_target(target_path, preview_protected)
        try:
            copy_status, prepare_msg = _prepare_target_file(source_path, target_path, p, preview_protected)
            if copy_status == "skip":
                out_rows.append([str(source_path), str(target_path), target_path.name, engine, len(ops), 0, len(ops), copy_status, "跳过", "跳过", prepare_msg])
                logs.append(_log("INFO", f"{target_path.name} 已跳过：{prepare_msg}", str(target_path)))
                total_skipped += len(ops)
                skipped_files += 1
                processed_files += 1
                _progress(context, i, total_files, f"跳过：{target_path.name}", stage="file_skipped", object=str(target_path))
                _cleanup_path(transaction_state.get("snapshot"))
                continue
        except Exception as exc:
            _rollback_target(target_path, transaction_state)
            _cleanup_path(transaction_state.get("snapshot"))
            failed_files += 1
            err = str(exc)
            out_rows.append([str(source_path), str(target_path), target_path.name, engine, len(ops), 0, len(ops), copy_status, "失败", "失败", err])
            logs.append(_log("ERROR", f"{target_path.name} 准备目标文件失败：{err}", str(target_path)))
            total_skipped += len(ops)
            processed_files += 1
            _progress(context, i, total_files, f"失败：{target_path.name}", stage="file_error", object=str(target_path))
            if error_policy == "遇错停止":
                if win32_session is not None:
                    win32_session.close()
                    win32_session = None
                    if isinstance(_session_holder, dict):
                        _session_holder["session"] = None
                raise
            continue

        if preview_protected:
            out_rows.append([str(source_path), str(target_path), target_path.name, engine, len(ops), 0, len(ops), copy_status, "预览保护(未写入)", "预览保护(未写入)", ""])
            logs.append(_log("INFO", f"预览模式已保护，未写入：{target_path.name}", str(target_path)))
            total_skipped += len(ops)
            processed_files += 1
            _progress(context, i, total_files, f"预览保护：{target_path.name}", stage="file_skipped", object=str(target_path))
            _cleanup_path(transaction_state.get("snapshot"))
            continue

        working_path = None
        try:
            if not target_path.exists():
                raise FileNotFoundError(f"目标文件不存在：{target_path}")
            working_path = _prepare_working_copy(target_path)
            if win32_session is not None:
                applied, skipped, write_logs = win32_session.write_file(working_path, ops, context, i - 1, total_files)
            else:
                applied, skipped, write_logs = _write_file(working_path, ops, engine, context, i - 1, total_files)
            error_messages = _error_messages(write_logs)
            logs.extend(write_logs)
            if applied <= 0:
                raise RuntimeError(error_messages[0] if error_messages else "没有任何写入操作成功")
            backup_path = _commit_working_copy(working_path, target_path, transaction_state, backup_mode)
            working_path = None
            total_applied += applied
            total_skipped += skipped
            if skipped > 0 or error_messages:
                partial_files += 1
                file_status = "部分成功"
                file_error = "；".join(error_messages[:3])
            else:
                successful_files += 1
                file_status = "成功"
                file_error = ""
            if backup_path is not None:
                copy_status = f"{copy_status}；已备份:{backup_path.name}"
            out_rows.append([
                str(source_path),
                str(target_path),
                target_path.name,
                engine,
                len(ops),
                applied,
                skipped,
                copy_status,
                file_status,
                file_status,
                file_error,
            ])
            logs.append(_log("INFO", f"{target_path.name} 写入完成：状态 {file_status}，应用 {applied}，跳过 {skipped}", str(target_path)))
            processed_files += 1
            _progress(context, i, total_files, f"已完成：{target_path.name}", stage="file_done", object=str(target_path))
        except Exception as exc:
            _cleanup_path(working_path)
            _rollback_target(target_path, transaction_state)
            failed_files += 1
            err = str(exc)
            out_rows.append([str(source_path), str(target_path), target_path.name, engine, len(ops), 0, len(ops), copy_status, "失败", "失败", err])
            logs.append(_log("ERROR", f"{target_path.name} 写入失败：{err}", str(target_path)))
            total_skipped += len(ops)
            processed_files += 1
            _progress(context, i, total_files, f"失败：{target_path.name}", stage="file_error", object=str(target_path))
            if error_policy == "遇错停止":
                if win32_session is not None:
                    win32_session.close()
                    win32_session = None
                    if isinstance(_session_holder, dict):
                        _session_holder["session"] = None
                raise
        finally:
            _cleanup_path(working_path)
            _cleanup_path(transaction_state.get("snapshot") if transaction_state else None)

    if win32_session is not None:
        win32_session.close()
        if isinstance(_session_holder, dict):
            _session_holder["session"] = None
        logs.append(_log("INFO", "WIN32_REUSE：Office 进程已统一退出"))

    ok_files = successful_files + partial_files
    if cancelled:
        msg = (
            f"写入已取消：已处理文件 {processed_files}/{total_files}；"
            f"成功 {successful_files}，部分成功 {partial_files}，跳过 {skipped_files}，失败 {failed_files}；"
            f"应用 {total_applied}，跳过操作 {total_skipped}"
        )
    else:
        msg = (
            f"写入完成：文件成功 {successful_files}，部分成功 {partial_files}，"
            f"跳过 {skipped_files}，失败 {failed_files}；"
            f"操作应用 {total_applied}，跳过 {total_skipped}"
        )
    _progress(context, processed_files, total_files if total_files > 0 else 1, "写入阶段完成", stage="done", cancelled=cancelled)
    return {
        "ok": bool(total_files == 0 or ok_files > 0 or skipped_files > 0 or preview_protected),
        "message": msg,
        "output": {
            "type": "table",
            "headers": out_headers,
            "rows": out_rows,
            "meta": {"plugin": PLUGIN_INFO["id"]},
        },
        "logs": logs[-500:],
        "summary": {
            "input_rows": prep["input_rows"],
            "total_files": total_files,
            "total_ops": total_ops,
            "applied": total_applied,
            "skipped": total_skipped,
            "failed_files": failed_files,
            "successful_files": successful_files,
            "partial_files": partial_files,
            "skipped_files": skipped_files,
            "processed_files": processed_files,
            "preview_protected": preview_protected,
            "engine": engine,
            "cancelled": cancelled,
            "win32_reuse_app": bool(engine == "win32" and win32_reuse_app),
            "win32_open_retries": win32_open_retries,
            "win32_retry_interval_ms": win32_retry_interval_ms,
            "win32_close_settle_ms": win32_close_settle_ms,
            "target_path_field": _as_text(p.get("target_path_field", "target_file")) or "target_file",
            "word_text_write_mode": _word_text_mode_from_params(p),
            "old_text_field": _as_text(p.get("old_text_field", "old_text")) or "old_text",
            "write_strategy_field": _as_text(p.get("write_strategy_field", "write_strategy")) or "write_strategy",
            "replace_scope_field": _as_text(p.get("replace_scope_field", "replace_scope")) or "replace_scope",
            "rule_old_text_field": _as_text(p.get("rule_old_text_field", "rule_old_text")) or "rule_old_text",
            "rule_new_text_field": _as_text(p.get("rule_new_text_field", "rule_new_text")) or "rule_new_text",
            "scoped_replace_default": _as_text(p.get("scoped_replace_default", REPLACE_SCOPE_FIRST)) or REPLACE_SCOPE_FIRST,
            "target_conflict_policy": _as_text(p.get("target_conflict_policy", CONFLICT_WARN)) or CONFLICT_WARN,
            "target_conflicts": prep["target_conflicts"],
            "verify_after_write": bool(p.get("verify_after_write", True)),
            "target_missing_policy": _as_text(p.get("target_missing_policy", "从源文件复制")) or "从源文件复制",
            "target_existing_policy": _as_text(p.get("target_existing_policy", "直接写入")) or "直接写入",
            "same_path_policy": _as_text(p.get("same_path_policy", "修改源文件")) or "修改源文件",
            "backup_mode": backup_mode,
        },
    }


def run(input_data, params, context):
    session_holder = {}
    try:
        return _run_impl(
            input_data,
            params,
            context,
            _session_holder=session_holder,
        )
    finally:
        win32_session = session_holder.get("session")
        if win32_session is not None:
            win32_session.close()
            session_holder["session"] = None


def _external_progress_callback(message):
    try:
        text = json.dumps(message, ensure_ascii=False) + "\n"
        sys.stdout.buffer.write(text.encode("utf-8"))
        sys.stdout.buffer.flush()
    except Exception:
        pass


def _run_external_entry(input_path, output_path):
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    try:
        payload = json.loads(Path(input_path).read_text(encoding="utf-8"))
        input_data = payload.get("input_data") or {}
        params = payload.get("params") or {}
        context = payload.get("context") or {}
        context.setdefault("plugin_id", PLUGIN_INFO["id"])
        context["progress_callback"] = _external_progress_callback
        context["report_progress"] = lambda current=0, total=0, message="", **extra: _external_progress_callback({
            "type": "node_progress",
            "current": current,
            "total": total,
            "message": message,
            **extra,
        })
        result = run(input_data, params, context)
    except Exception as exc:
        result = {
            "ok": False,
            "message": str(exc),
            "output": {
                "type": "table",
                "headers": ["错误"],
                "rows": [[str(exc)]],
                "meta": {"plugin": PLUGIN_INFO["id"]},
            },
            "logs": [
                {"level": "ERROR", "message": str(exc)},
                {"level": "ERROR", "message": traceback.format_exc()},
            ],
            "summary": {"success": 0, "failed": 1},
        }
    output_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return 0 if result.get("ok", False) else 1


def main(argv=None):
    parser = argparse.ArgumentParser(description=PLUGIN_INFO["name"])
    parser.add_argument("--input", dest="input_path")
    parser.add_argument("--output", dest="output_path")
    args = parser.parse_args(argv)
    if args.input_path and args.output_path:
        return _run_external_entry(args.input_path, args.output_path)
    print("这是 DataFlowKit 插件，请在主程序插件节点中调用。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
