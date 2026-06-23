# -*- coding: utf-8 -*-
import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from plugins import word_excel_read_to_db_plugin_v1 as reader
from plugins import word_excel_write_from_table_plugin_v2 as writer


NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def make_docx(path, document_xml, extra_payload=None):
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '</Types>'
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("word/document.xml", document_xml)
        for name, data in (extra_payload or {}).items():
            archive.writestr(name, data)


def read_zip_text(path, name="word/document.xml"):
    with zipfile.ZipFile(path, "r") as archive:
        return archive.read(name).decode("utf-8")


def base_params(**overrides):
    params = {
        "write_engine": "zip_xml",
        "word_text_write_mode": writer.WORD_MODE_PRESERVE_FORMAT,
        "path_field": "source_file",
        "target_path_field": "target_file",
        "block_type_field": "block_type",
        "sheet_name_field": "sheet_name",
        "row_index_field": "row_index",
        "col_index_field": "col_index",
        "cell_address_field": "cell_address",
        "value_field": "text",
        "old_text_field": "old_text",
        "meta_json_field": "meta_json",
        "same_path_policy": "修改源文件",
        "error_policy": "继续并记录失败",
        "backup_mode": "失败时恢复原文件",
    }
    params.update(overrides)
    return params


class WordExcelWriterTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory(dir=Path.cwd())
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def simple_docx(self, name="simple.docx", text="old"):
        path = self.root / name
        document_xml = (
            f'<?xml version="1.0" encoding="UTF-8"?>'
            f'<w:document xmlns:w="{NS_W}"><w:body>'
            f'<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
            f'<w:r><w:rPr><w:b/></w:rPr><w:t>{text}</w:t></w:r></w:p>'
            f'<w:sectPr/></w:body></w:document>'
        )
        make_docx(path, document_xml)
        return path

    def test_reused_win32_session_closes_when_progress_callback_raises(self):
        path = self.simple_docx()
        state = {"closed": 0}

        class FakeSession:
            def __init__(self, **_kwargs):
                pass

            def close(self):
                state["closed"] += 1

        original_progress = writer._progress

        def fail_on_file_start(context, current=None, total=None, message="", **extra):
            if extra.get("stage") == "file_start":
                raise RuntimeError("forced progress failure")
            return original_progress(context, current, total, message, **extra)

        with mock.patch.object(writer, "_Win32OfficeSession", FakeSession), mock.patch.object(
            writer,
            "_progress",
            side_effect=fail_on_file_start,
        ):
            with self.assertRaisesRegex(RuntimeError, "forced progress failure"):
                writer.run(
                    {
                        "type": "table",
                        "headers": ["source_file", "block_type", "row_index", "text"],
                        "rows": [[str(path), "word_paragraph", 1, "new"]],
                    },
                    base_params(write_engine="win32", win32_reuse_app=True),
                    {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
                )
        self.assertEqual(state["closed"], 1)

    def test_numeric_zero_is_not_treated_as_empty(self):
        grouped, prep = writer._collect_ops(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "text"],
                "rows": [["demo.docx", "word_paragraph", 0]],
            },
            base_params(),
            {"app_dir": str(self.root)},
        )
        self.assertEqual(sum(len(item["ops"]) for item in grouped.values()), 1)
        self.assertEqual(prep["skipped_empty_value"], 0)
        self.assertEqual(next(iter(grouped.values()))["ops"][0]["value"], "0")

    def test_duplicate_global_replacements_are_merged(self):
        path = self.root / "demo.docx"
        grouped, prep = writer._collect_ops(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "old_text", "text"],
                "rows": [
                    [str(path), writer.BLOCK_WORD_GLOBAL_REPLACE, "(2)", "(3)"],
                    [str(path), writer.BLOCK_WORD_GLOBAL_REPLACE, "(2)", "(3)"],
                ],
            },
            base_params(),
            {"app_dir": str(self.root)},
        )
        self.assertEqual(sum(len(item["ops"]) for item in grouped.values()), 1)
        self.assertEqual(prep["skipped_duplicate_global"], 1)

    def test_validate_excel_only_does_not_require_old_text(self):
        ok, message = writer.validate_params(
            base_params(write_engine="win32", word_text_write_mode=writer.WORD_MODE_FIND_REPLACE),
            {
                "headers": ["source_file", "block_type", "sheet_name", "cell_address", "text"],
                "rows": [["demo.xlsx", "excel_cell", "Sheet1", "A1", "new"]],
            },
            {},
        )
        self.assertTrue(ok, message)

    def test_validate_direct_word_write_does_not_require_old_text(self):
        ok, message = writer.validate_params(
            base_params(write_engine="win32", word_text_write_mode=writer.WORD_MODE_FIND_REPLACE),
            {
                "headers": ["source_file", "block_type", "row_index", "text", "write_strategy"],
                "rows": [["demo.docx", "word_paragraph", 1, "new", writer.WRITE_STRATEGY_DIRECT]],
            },
            {},
        )
        self.assertTrue(ok, message)

    def test_collect_ops_detects_same_target_conflict(self):
        grouped, prep = writer._collect_ops(
            {
                "headers": ["source_file", "block_type", "sheet_name", "cell_address", "text"],
                "rows": [
                    ["demo.xlsx", "excel_cell", "Sheet1", "A1", "first"],
                    ["demo.xlsx", "excel_cell", "Sheet1", "A1", "second"],
                ],
            },
            base_params(target_conflict_policy=writer.CONFLICT_WARN),
            {"app_dir": str(self.root)},
        )
        group = next(iter(grouped.values()))
        self.assertEqual(len(group["ops"]), 2)
        self.assertEqual(prep["target_conflicts"], 1)
        self.assertIn("写入值不同", group["target_conflicts"][0])

    def test_find_replace_falls_back_to_literal_first_replace(self):
        class FakeReplacement:
            def ClearFormatting(self):
                return None

        class FakeFind:
            def __init__(self):
                self.Replacement = FakeReplacement()

            def ClearFormatting(self):
                return None

            def Execute(self, **_kwargs):
                return False

        class FakeRange:
            def __init__(self, text):
                self.Start = 0
                self.End = len(text)
                self.Text = text
                self.Document = self
                self.Find = FakeFind()

            def Range(self, _start, _end):
                return self

        target = FakeRange("old + old")
        result = writer._word_find_replace_visible_text(
            target,
            "old + old",
            "new + old",
            rule_old_text="old",
            rule_new_text="new",
            replace_all=False,
            preserve_format=False,
            verify=True,
        )
        self.assertEqual(result, "fallback_unique_context")
        self.assertEqual(target.Text, "new + old")

    def test_find_replace_short_diff_preserves_internal_word_line_breaks(self):
        class FakeReplacement:
            def ClearFormatting(self):
                return None

        class FakeFind:
            def __init__(self, owner):
                self.owner = owner
                self.Replacement = FakeReplacement()
                self.find_texts = []

            def ClearFormatting(self):
                return None

            def Execute(self, **kwargs):
                self.find_texts.append(kwargs["FindText"])
                if kwargs["FindText"] not in self.owner.Text:
                    return False
                self.owner.Text = self.owner.Text.replace(kwargs["FindText"], kwargs["ReplaceWith"], 1)
                return True

        class FakeRange:
            def __init__(self, text):
                self.Start = 0
                self.End = len(text)
                self.Text = text
                self.Document = self
                self.Find = FakeFind(self)

            def Range(self, _start, _end):
                return self

        target = FakeRange("HYBP2435TK-150RD\r电脑板")
        result = writer._word_find_replace_visible_text(
            target,
            "HYBP2435TK-150RD电脑板",
            "HYBP2435TK-51RD电脑板",
            preserve_format=False,
        )

        self.assertEqual(result, "find_unique_context")
        self.assertEqual(target.Find.find_texts, ["150"])
        self.assertEqual(target.Text, "HYBP2435TK-51RD\r电脑板")

    def test_find_replace_already_target_is_idempotent(self):
        class NeverFind:
            def Execute(self, **_kwargs):
                raise AssertionError("already-target value must not call Word Find")

        class FakeRange:
            def __init__(self):
                self.Start = 0
                self.End = 23
                self.Text = "HYBP2435TK-51RD\r电脑板"
                self.Document = self
                self.Find = NeverFind()

            def Range(self, _start, _end):
                return self

        result = writer._word_find_replace_visible_text(
            FakeRange(),
            "HYBP2435TK-150RD电脑板",
            "HYBP2435TK-51RD电脑板",
        )
        self.assertEqual(result, "already_target")

    def test_find_replace_com_false_success_uses_stable_fallback(self):
        class FakeReplacement:
            def ClearFormatting(self):
                return None

        class FakeFind:
            def __init__(self):
                self.Replacement = FakeReplacement()

            def ClearFormatting(self):
                return None

            def Execute(self, **_kwargs):
                return True

        class FakeRange:
            def __init__(self):
                self.Start = 0
                self.End = 16
                self.Text = "HYBP2435TK-150RD"
                self.Document = self
                self.Find = FakeFind()

            def Range(self, _start, _end):
                return self

        target = FakeRange()
        result = writer._word_find_replace_visible_text(
            target,
            "HYBP2435TK-150RD",
            "HYBP2435TK-51RD",
            preserve_format=False,
        )
        self.assertEqual(result, "fallback_unique_context")
        self.assertEqual(target.Text, "HYBP2435TK-51RD")

    def test_long_block_uses_short_diff_instead_of_long_word_find(self):
        prefix = "产 品 使 用 说 明" + ("说明文字" * 50)
        old_text = prefix + "HYBP2435TK-150RD" + ("结尾" * 40)
        new_text = prefix + "HYBP2435TK-51RD" + ("结尾" * 40)

        class FakeReplacement:
            def ClearFormatting(self):
                return None

        class FakeFind:
            def __init__(self, owner):
                self.owner = owner
                self.Replacement = FakeReplacement()
                self.find_text = ""

            def ClearFormatting(self):
                return None

            def Execute(self, **kwargs):
                self.find_text = kwargs["FindText"]
                self.owner.Text = self.owner.Text.replace(kwargs["FindText"], kwargs["ReplaceWith"], 1)
                return True

        class FakeRange:
            def __init__(self, text):
                self.Start = 0
                self.End = len(text)
                self.Text = text
                self.Document = self
                self.Find = FakeFind(self)

            def Range(self, _start, _end):
                return self

        target = FakeRange(old_text)
        result = writer._word_find_replace_visible_text(target, old_text, new_text, preserve_format=False)
        self.assertEqual(result, "find_unique_context")
        self.assertLessEqual(len(target.Find.find_text), writer.WORD_FIND_TEXT_LIMIT)
        self.assertEqual(target.Find.find_text, "150")
        self.assertEqual(target.Text, new_text)

    def test_unique_context_avoids_replacing_unrelated_150(self):
        old_text = "名称：Q款变频天花机150冷暖电脑板型号：HYBP2435TK-150RD"
        new_text = "名称：Q款变频天花机150冷暖电脑板型号：HYBP2435TK-51RD"

        search_text, replacement_text, source = writer._word_replace_pair(
            old_text,
            new_text,
            current_text=old_text,
            rule_old_text="150",
            rule_new_text="51",
        )

        self.assertEqual(source, "unique_context")
        self.assertEqual(old_text.count(search_text), 1)
        self.assertIn("-150R", search_text)
        self.assertEqual(old_text.replace(search_text, replacement_text, 1), new_text)

    def test_unique_context_avoids_single_digit_model_corruption(self):
        old_text = "设计文件图样HYBP2435TK-150RD电脑板"
        new_text = "设计文件图样HYBP2435TK-120RD电脑板"

        search_text, replacement_text, source = writer._word_replace_pair(
            old_text,
            new_text,
            current_text=old_text,
            rule_old_text="5",
            rule_new_text="2",
        )

        self.assertEqual(source, "unique_context")
        self.assertNotEqual(search_text, "5")
        self.assertEqual(search_text, "150")
        self.assertEqual(replacement_text, "120")

    def test_false_success_restores_original_before_fallback(self):
        class FakeReplacement:
            def ClearFormatting(self):
                return None

        class FakeFind:
            def __init__(self, owner):
                self.owner = owner
                self.Replacement = FakeReplacement()

            def ClearFormatting(self):
                return None

            def Execute(self, **_kwargs):
                self.owner.Text = "WRONG"
                return True

        class FakeRange:
            def __init__(self):
                self.Start = 0
                self.End = 3
                self._text = "old"
                self.text_history = []
                self.Document = self
                self.Find = FakeFind(self)

            @property
            def Text(self):
                return self._text

            @Text.setter
            def Text(self, value):
                self.text_history.append(value)
                self._text = value

            def Range(self, _start, _end):
                return self

        target = FakeRange()
        result = writer._word_find_replace_visible_text(
            target,
            "old",
            "new",
            rule_old_text="missing",
            rule_new_text="new",
            preserve_format=False,
        )
        self.assertEqual(result, "fallback_unique_context")
        self.assertEqual(target.Text, "new")
        self.assertEqual(target.text_history, ["WRONG", "old", "new"])

    def test_transactional_range_write_restores_text_after_failure(self):
        class FakeRange:
            def __init__(self):
                self.Start = 0
                self.End = 8
                self.Text = "original"
                self.Document = self

            def Range(self, _start, _end):
                return self

        target = FakeRange()

        def broken_write(range_obj, _op, _context):
            range_obj.Text = "corrupted"
            raise ValueError("写入后校验失败")

        with mock.patch.object(writer, "_word_write_range_by_mode", side_effect=broken_write):
            with self.assertRaisesRegex(ValueError, "写入后校验失败"):
                writer._word_write_range_transactional(
                    target,
                    {"block_type": "word_table_cell"},
                    {"params": {"preserve_text_format": False}},
                )

        self.assertEqual(target.Text, "original")

    def test_deterministic_operation_failure_is_not_retried(self):
        calls = []

        def action():
            calls.append(1)
            raise ValueError("定位范围内未找到 old_text")

        with self.assertRaises(ValueError):
            writer._apply_operation_with_retry(
                action,
                {"block_type": "word_table_cell"},
                {"params": {"win32_cell_retries": 3, "win32_retry_interval_ms": 0}},
                "Word写入",
            )
        self.assertEqual(len(calls), 1)

    def test_transient_com_operation_failure_is_retried(self):
        calls = []

        def action():
            calls.append(1)
            if len(calls) < 3:
                raise RuntimeError("-2147418111 call was rejected by callee")
            return "ok"

        result = writer._apply_operation_with_retry(
            action,
            {"block_type": "word_table_cell"},
            {"params": {"win32_cell_retries": 3, "win32_retry_interval_ms": 0}},
            "Word写入",
        )
        self.assertEqual(result, "ok")
        self.assertEqual(len(calls), 3)

    def test_global_replace_deduplicates_same_story_range(self):
        class FakeReplacement:
            def ClearFormatting(self):
                return None

        class FakeFind:
            def __init__(self, owner):
                self.owner = owner
                self.Replacement = FakeReplacement()

            def ClearFormatting(self):
                return None

            def Execute(self, **kwargs):
                self.owner.calls += 1
                old_text = kwargs["FindText"]
                if old_text not in self.owner.Text:
                    return False
                self.owner.Text = self.owner.Text.replace(old_text, kwargs["ReplaceWith"])
                return True

        class FakeRange:
            StoryType = 1
            Start = 0
            End = 1
            NextStoryRange = None

            def __init__(self):
                self.Text = "A"
                self.calls = 0
                self.Find = FakeFind(self)

        class EmptyCollection:
            Count = 0

        class FakeDoc:
            def __init__(self):
                self.Content = FakeRange()
                self.StoryRanges = [self.Content]
                self.Shapes = EmptyCollection()
                self.InlineShapes = EmptyCollection()
                self.Sections = EmptyCollection()

        doc = FakeDoc()
        writer._word_global_replace(doc, {"old_text": "A", "value": "AA"})
        self.assertEqual(doc.Content.Text, "AA")
        self.assertEqual(doc.Content.calls, 1)

    def test_word_save_falls_back_to_save_as(self):
        class FakeDoc:
            def __init__(self):
                self.save_as_calls = []

            def Save(self):
                raise RuntimeError("locked")

            def SaveAs2(self, path, **kwargs):
                self.save_as_calls.append((path, kwargs))

        doc = FakeDoc()
        target = self.root / "fallback.docx"
        writer._save_word_with_fallback(
            doc,
            target,
            {"params": {"win32_save_retries": 1, "win32_retry_interval_ms": 0}},
        )
        self.assertEqual(doc.save_as_calls[0][0], str(target))
        self.assertEqual(doc.save_as_calls[0][1]["FileFormat"], 12)

    def test_global_replace_ops_sort_long_old_text_first_within_runs(self):
        def global_op(old_text):
            return {"block_type": writer.BLOCK_WORD_GLOBAL_REPLACE, "old_text": old_text, "value": old_text.lower()}

        normal_op = {"block_type": "word_table_cell", "old_text": "normal", "value": "normal"}
        ordered = writer._ordered_word_ops([
            global_op("123"),
            global_op("123569"),
            normal_op,
            global_op("xx"),
            global_op("abcd"),
        ])

        self.assertEqual(ordered[0]["old_text"], "123569")
        self.assertEqual(ordered[1]["old_text"], "123")
        self.assertIs(ordered[2], normal_op)
        self.assertEqual(ordered[3]["old_text"], "abcd")
        self.assertEqual(ordered[4]["old_text"], "xx")

    def test_operation_failure_message_includes_write_context(self):
        op = {
            "source_row": 7,
            "block_type": "word_table_cell",
            "sheet_name": "table_5",
            "row_index": 6,
            "col_index": 5,
            "cell_address": "R6C5",
            "old_text": "HYBP2435TK-150RD\r_V1.0",
            "value": "1.01.01.00592",
            "write_strategy": "",
        }
        message = writer._operation_failure_message("写入失败", op, ValueError("定位范围内未找到 old_text"))

        self.assertIn("源行7", message)
        self.assertIn("block_type=word_table_cell", message)
        self.assertIn("位置=table_5 R6C5", message)
        self.assertIn("cell_address=R6C5", message)
        self.assertIn("write_strategy=跟随节点设置", message)
        self.assertIn("old_text长度=22", message)
        self.assertIn("HYBP2435TK-150RD\\r_V1.0", message)
        self.assertIn("写入值=1.01.01.00592", message)

    def test_preserve_format_table_cell_writes_body_range_only(self):
        class FakeFont:
            Name = "Arial"
            NameFarEast = ""
            Size = 10
            Bold = 0
            Italic = 0
            Underline = 0
            Color = 0

        class FakeDocument:
            def __init__(self):
                self.ranges = []

            def Range(self, start, end):
                rng = FakeRange(self, start, end, "")
                self.ranges.append(rng)
                return rng

        class FakeRange:
            def __init__(self, document, start, end, text):
                self.Document = document
                self.Start = start
                self.End = end
                self._text = text
                self.assigned_text = None
                self.Font = FakeFont()

            @property
            def Text(self):
                return self._text

            @Text.setter
            def Text(self, value):
                self.assigned_text = value
                self._text = value

            @property
            def Duplicate(self):
                rng = FakeRange(self.Document, self.Start, self.End, self.Text)
                self.Document.ranges.append(rng)
                return rng

        doc = FakeDocument()
        cell_range = FakeRange(doc, 0, 5, "old\r\x07")

        writer._word_write_text_preserve_format(cell_range, "new")

        self.assertIsNone(cell_range.assigned_text)
        written = [rng for rng in doc.ranges if rng.assigned_text == "new"]
        self.assertEqual(len(written), 1)
        self.assertEqual((written[0].Start, written[0].End), (0, 4))
        self.assertNotIn("new\r\x07", [rng.assigned_text for rng in doc.ranges])

    def test_all_operations_failed_returns_failed_file(self):
        path = self.simple_docx()
        before = path.read_bytes()
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "text"],
                "rows": [[str(path), "unsupported_block", "new"]],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        self.assertFalse(result["ok"])
        self.assertEqual(result["summary"]["failed_files"], 1)
        self.assertEqual(result["output"]["rows"][0][8], "失败")
        self.assertEqual(path.read_bytes(), before)

    def test_error_policy_stop_raises_on_operation_error(self):
        path = self.simple_docx()
        with self.assertRaises(RuntimeError):
            writer.run(
                {
                    "type": "table",
                    "headers": ["source_file", "block_type", "text"],
                    "rows": [[str(path), "unsupported_block", "new"]],
                },
                base_params(error_policy="遇错停止"),
                {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
            )

    def test_reader_emits_word_range_contract(self):
        rows = []
        reader._append_word_text_lines(
            rows,
            "alpha\rbravo",
            self.root / "source.doc",
            {},
            0,
            1,
            {"range_base": 10},
        )
        self.assertEqual(rows[0]["block_type"], writer.BLOCK_WORD_TEXT_RANGE)
        self.assertEqual(rows[0]["cell_address"], "WRANGE10:15")
        meta = json.loads(rows[0]["meta_json"])
        self.assertEqual(meta["range_start"], 0)
        self.assertEqual(meta["range_end"], 5)

    def test_reader_preserves_internal_word_line_breaks_in_meta(self):
        self.assertEqual(
            reader._word_raw_visible_text("HYBP2435TK-150RD\r电脑板\r\x07"),
            "HYBP2435TK-150RD\r电脑板",
        )

    def test_reader_excludes_table_ranges_from_word_text_lines(self):
        rows = []
        reader._append_word_text_lines(
            rows,
            "body-before\rtable-cell-text\rbody-after",
            self.root / "source.doc",
            {},
            0,
            1,
            {"range_base": 100},
            exclude_ranges=[(112, 127)],
        )

        self.assertEqual([row["text"] for row in rows], ["body-before", "body-after"])
        self.assertEqual([row["row_index"] for row in rows], [1, 2])
        self.assertEqual(rows[0]["cell_address"], "WRANGE100:111")
        self.assertEqual(rows[1]["cell_address"], "WRANGE128:138")

    def test_writer_recovers_word_range_from_cell_address(self):
        calls = []

        class FakeDocument:
            def Range(self, start, end):
                calls.append((start, end))
                return (start, end)

        result = writer._word_content_range(
            FakeDocument(),
            {"cell_address": "WRANGE12:34", "meta_json": ""},
        )
        self.assertEqual(result, (12, 34))
        self.assertEqual(calls, [(12, 34)])

    def test_zip_xml_preserves_paragraph_and_run_format(self):
        path = self.simple_docx()
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "row_index", "text"],
                "rows": [[str(path), "word_paragraph", 1, "new"]],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        after = read_zip_text(path)
        self.assertTrue(result["ok"])
        self.assertIn("pPr", after)
        self.assertIn("rPr", after)
        self.assertIn("new", after)
        self.assertNotIn(">old<", after)

    def test_zip_xml_uses_logical_column_after_grid_span(self):
        path = self.root / "merged.docx"
        document_xml = (
            f'<?xml version="1.0" encoding="UTF-8"?>'
            f'<w:document xmlns:w="{NS_W}"><w:body><w:tbl><w:tr>'
            f'<w:tc><w:tcPr><w:gridSpan w:val="2"/></w:tcPr><w:p><w:r><w:t>A</w:t></w:r></w:p></w:tc>'
            f'<w:tc><w:p><w:r><w:t>B</w:t></w:r></w:p></w:tc>'
            f'</w:tr></w:tbl><w:sectPr/></w:body></w:document>'
        )
        make_docx(path, document_xml)
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "sheet_name", "row_index", "col_index", "text"],
                "rows": [[str(path), "word_table_cell", "table_1", 1, 3, "C"]],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        after = read_zip_text(path)
        self.assertTrue(result["ok"])
        self.assertIn(">A<", after)
        self.assertIn(">C<", after)
        self.assertNotIn(">B<", after)

    def test_openpyxl_missing_sheet_is_an_error(self):
        path = self.root / "book.xlsx"
        workbook = Workbook()
        workbook.save(path)
        before = path.read_bytes()
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "sheet_name", "cell_address", "text"],
                "rows": [[str(path), "excel_cell", "Missing", "A1", "new"]],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        self.assertFalse(result["ok"])
        self.assertIn("工作表不存在", result["output"]["rows"][0][-1])
        self.assertEqual(path.read_bytes(), before)

    def test_openpyxl_writes_merged_cell_origin(self):
        path = self.root / "merged.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.merge_cells("A1:B2")
        ws["A1"] = "old"
        wb.save(path)
        wb.close()

        applied, skipped, logs = writer._write_excel_openpyxl(
            path,
            [{
                "source_row": 1,
                "block_type": "excel_cell",
                "sheet_name": "Sheet1",
                "row_index": "2",
                "col_index": "2",
                "cell_address": "B2",
                "value": "new",
                "meta_json": "",
            }],
            {"params": {"verify_after_write": True}},
        )

        from openpyxl import load_workbook
        check = load_workbook(path)
        try:
            self.assertEqual(check["Sheet1"]["A1"].value, "new")
        finally:
            check.close()
        self.assertEqual((applied, skipped, logs), (1, 0, []))

    def test_failed_writer_restores_original_target(self):
        path = self.simple_docx()
        before = path.read_bytes()
        original_write_file = writer._write_file

        def fail_after_mutation(file_path, *args, **kwargs):
            Path(file_path).write_bytes(b"partial")
            raise RuntimeError("forced failure")

        writer._write_file = fail_after_mutation
        try:
            result = writer.run(
                {
                    "type": "table",
                    "headers": ["source_file", "block_type", "row_index", "text"],
                    "rows": [[str(path), "word_paragraph", 1, "new"]],
                },
                base_params(),
                {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
            )
        finally:
            writer._write_file = original_write_file
        self.assertFalse(result["ok"])
        self.assertEqual(path.read_bytes(), before)
        self.assertEqual(list(self.root.glob(".*.working.*")), [])

    def test_partial_operation_failure_reports_partial_success(self):
        path = self.simple_docx()
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "row_index", "text"],
                "rows": [
                    [str(path), "word_paragraph", 1, "new"],
                    [str(path), "unsupported_block", "", "ignored"],
                ],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        self.assertTrue(result["ok"])
        self.assertEqual(result["summary"]["partial_files"], 1)
        self.assertEqual(result["output"]["rows"][0][8], "部分成功")
        self.assertIn("new", read_zip_text(path))

    def test_backup_mode_keeps_original_file(self):
        path = self.simple_docx()
        original = path.read_bytes()
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "row_index", "text"],
                "rows": [[str(path), "word_paragraph", 1, "new"]],
            },
            base_params(backup_mode="写入前保留备份"),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        backups = list(self.root.glob("simple_backup_*.docx"))
        self.assertTrue(result["ok"])
        self.assertEqual(len(backups), 1)
        self.assertEqual(backups[0].read_bytes(), original)

    def test_docx_global_replace_updates_header_and_wordart_attribute(self):
        path = self.simple_docx(text="body")
        header_xml = (
            f'<?xml version="1.0" encoding="UTF-8"?>'
            f'<w:hdr xmlns:w="{NS_W}" xmlns:v="urn:schemas-microsoft-com:vml">'
            f'<w:p><w:r><w:t>version (2)</w:t></w:r></w:p>'
            f'<v:textpath string="art (2)"/>'
            f'</w:hdr>'
        )
        document_xml = read_zip_text(path)
        make_docx(path, document_xml, {"word/header1.xml": header_xml})
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "old_text", "text"],
                "rows": [[str(path), writer.BLOCK_WORD_GLOBAL_REPLACE, "(2)", "(3)"]],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        header_after = read_zip_text(path, "word/header1.xml")
        self.assertTrue(result["ok"])
        self.assertIn("version (3)", header_after)
        self.assertIn("art (3)", header_after)
        self.assertNotIn("(2)", header_after)

    def test_docx_global_replace_applies_long_old_text_first(self):
        path = self.simple_docx(text="123569 123")
        result = writer.run(
            {
                "type": "table",
                "headers": ["source_file", "block_type", "old_text", "text"],
                "rows": [
                    [str(path), writer.BLOCK_WORD_GLOBAL_REPLACE, "123", "B"],
                    [str(path), writer.BLOCK_WORD_GLOBAL_REPLACE, "123569", "A"],
                ],
            },
            base_params(),
            {"app_dir": str(self.root), "is_preview": False, "execute_actions": True},
        )
        after = read_zip_text(path)
        self.assertTrue(result["ok"], result)
        self.assertIn(">A B<", after)
        self.assertNotIn("B569", after)

    def test_external_entry_writes_standard_output_json(self):
        path = self.simple_docx()
        input_path = self.root / "input.json"
        output_path = self.root / "output.json"
        payload = {
            "input_data": {
                "type": "table",
                "headers": ["source_file", "block_type", "row_index", "text"],
                "rows": [[str(path), "word_paragraph", 1, "external"]],
            },
            "params": base_params(),
            "context": {
                "app_dir": str(self.root),
                "is_preview": False,
                "execute_actions": True,
            },
        }
        input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        exit_code = writer._run_external_entry(input_path, output_path)
        result = json.loads(output_path.read_text(encoding="utf-8"))
        self.assertEqual(exit_code, 0)
        self.assertTrue(result["ok"])
        self.assertIn("external", read_zip_text(path))


class WordReaderStrategyTests(unittest.TestCase):
    def test_schema_uses_generic_read_strategy_label(self):
        item = next(
            field
            for field in reader.get_parameter_schema()
            if field.get("name") == "doc_read_strategy"
        )
        self.assertEqual(item["label"], "读取策略")
        self.assertIn(reader.READ_STRATEGY_CONVERT_WIN32, item["choices"])
        self.assertIn(reader.READ_STRATEGY_CONVERT_WIN32_COMPAT, item["choices"])

    def test_selected_strategy_applies_to_all_word_extensions(self):
        original = reader._read_word_text_via_com
        calls = []

        def fake_reader(file_path, *args, **kwargs):
            calls.append(Path(file_path).suffix.lower())
            return [{"text": "ok"}]

        reader._read_word_text_via_com = fake_reader
        try:
            for extension in (".doc", ".docx", ".docm"):
                rows = reader._read_file_rows(
                    Path(f"sample{extension}"),
                    "zip_xml",
                    {"doc_read_strategy": "win32纯文本快速读取"},
                )
                self.assertEqual(rows, [{"text": "ok"}])
        finally:
            reader._read_word_text_via_com = original
        self.assertEqual(calls, [".doc", ".docx", ".docm"])

    def test_conversion_win32_strategies_pass_compatibility_mode(self):
        original = reader._read_word_converted
        calls = []

        def fake_converted(file_path, read_mode, word_merge_mode="关闭", preserve_compatibility=False, **kwargs):
            calls.append((Path(file_path).suffix.lower(), read_mode, preserve_compatibility))
            return [{"text": "converted"}]

        reader._read_word_converted = fake_converted
        try:
            normal = reader._read_file_rows(
                Path("sample.docx"),
                "zip_xml",
                {"doc_read_strategy": reader.READ_STRATEGY_CONVERT_WIN32},
            )
            compatible = reader._read_file_rows(
                Path("sample.docm"),
                "zip_xml",
                {"doc_read_strategy": reader.READ_STRATEGY_CONVERT_WIN32_COMPAT},
            )
        finally:
            reader._read_word_converted = original
        self.assertEqual(normal, [{"text": "converted"}])
        self.assertEqual(compatible, [{"text": "converted"}])
        self.assertEqual(
            calls,
            [
                (".docx", "win32", False),
                (".docm", "win32", True),
            ],
        )


if __name__ == "__main__":
    unittest.main()
