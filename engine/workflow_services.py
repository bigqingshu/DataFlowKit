# -*- coding: utf-8 -*-
"""UI-free side-effect services used by workflow execution and output."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from core.text_utils import make_sql_columns, sanitize_sql_name
from db.table_manager import TableAccessManager
from engine.models import TableData


@dataclass
class WorkflowServices:
    """Minimal service boundary for outputs that write outside the frontend."""

    db_path: str = ""

    def write_table(self, name, table, mode="timestamp", backup=True, db_path=None):
        target_db = str(db_path or self.db_path or "").strip()
        if not target_db:
            raise ValueError("SQLite 数据库路径为空。")
        target = Path(target_db)
        target.parent.mkdir(parents=True, exist_ok=True)

        table_data = TableData.from_payload(table or {})
        table_name = sanitize_sql_name(name, "计划结果")
        headers = make_sql_columns(table_data.headers)
        if not headers:
            raise ValueError("没有可写入的字段。")
        rows = _normalize_rows(table_data.rows, len(headers))

        manager = TableAccessManager(str(target), node_type="WorkflowServices")
        backup_name = ""
        if TableAccessManager.normalize_write_mode(mode) == "replace_table":
            if backup and manager.table_exists(table_name):
                backup_name = manager.backup_table(table_name)
        info = manager.write_table(table_name, headers, rows, mode=mode)
        return {
            "ok": True,
            "db_path": str(target),
            "table_name": info.get("table_name", table_name),
            "rows": info.get("rows", len(rows)),
            "columns": info.get("columns", len(headers)),
            "mode": info.get("mode", mode),
            "write_mode": info.get("write_mode", mode),
            "backup_table": backup_name,
        }

    def export_xlsx(self, path, table, sheet_name="结果"):
        target = Path(str(path or "").strip())
        if not str(target):
            raise ValueError("xlsx 导出路径为空。")
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(parents=True, exist_ok=True)

        table_data = TableData.from_payload(table or {})
        _export_xlsx_with_openpyxl(target, table_data.headers, table_data.rows, sheet_name=sheet_name)
        return {
            "ok": True,
            "path": str(target),
            "rows": len(table_data.rows),
            "columns": len(table_data.headers),
            "sheet_name": _normalize_sheet_title(sheet_name),
        }


def _normalize_rows(rows, width):
    result = []
    for row in rows or []:
        fixed = list(row)
        if len(fixed) < width:
            fixed += [""] * (width - len(fixed))
        if len(fixed) > width:
            fixed = fixed[:width]
        result.append(["" if value is None else str(value) for value in fixed])
    return result


def _normalize_sheet_title(value):
    title = str(value or "结果").strip() or "结果"
    for char in "[]:*?/\\":  # Excel sheet names cannot contain these characters.
        title = title.replace(char, "_")
    return title[:31] or "结果"


def _export_xlsx_with_openpyxl(path, headers, rows, sheet_name="结果"):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    headers = [str(header) for header in (headers or [])]
    rows = [list(row) for row in (rows or [])]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _normalize_sheet_title(sheet_name)
    sheet.append(headers)
    for row in rows:
        fixed = list(row)
        if len(fixed) < len(headers):
            fixed += [""] * (len(headers) - len(fixed))
        if len(fixed) > len(headers):
            fixed = fixed[:len(headers)]
        sheet.append(["" if value is None else str(value) for value in fixed])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    sheet.freeze_panes = "A2"
    if headers:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 1)}"
    for col_idx, header in enumerate(headers, start=1):
        max_width = _display_width(header)
        for row in rows[:3000]:
            if col_idx - 1 < len(row):
                max_width = max(max_width, _display_width(row[col_idx - 1]))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_width + 2, 10), 40)
    workbook.save(str(path))


def _display_width(value):
    text = "" if value is None else str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width
