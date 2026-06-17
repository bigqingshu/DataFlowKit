# -*- coding: utf-8 -*-
"""
剪贴板表格解析器 - SQLite保存版 + 高级筛选/数据匹配窗口

功能概览：
1. 从 Windows 剪贴板读取 Excel/WPS/网页表格数据。
2. 在 Tkinter GUI 中预览、编辑、保存到 SQLite。
3. 下拉选择 SQLite 表后，可自动加载数据库表数据。
4. 新增“高级筛选 / 数据匹配”窗口：
   - 支持选择一个或多个 SQLite 表作为数据源。
   - 支持多条件筛选：等于、不等于、包含、大于、小于、为空等。
   - 支持多表匹配规则：字段相等、字段包含等。
   - 支持选择输出字段。
   - 支持预览筛选结果。
   - 支持保存筛选结果为新表。
   - 支持保存/载入筛选模板 JSON。
5. 新增“批量替换 / 数据处理”窗口：
   - 支持按字段进行局部字符串替换或整格替换。
   - 支持替换前预览、执行替换、撤销上一次替换。
   - 支持保存/载入替换规则模板 JSON。
6. 新增主界面“导出为 xlsx”按钮，可导出当前预览数据。
7. 新增“数据提取 / 字段生成”窗口：
   - 支持 Python 正则提取、固定位置提取、按分隔符提取、关键字之间提取等。
   - 支持预览、执行、撤销、生成新字段、覆盖源字段、保存/载入规则模板。
8. 新增“合并列 / 生成新列”窗口：
   - 支持从字段池添加字段到合并顺序列表。
   - 支持上移、下移、删除、清空字段顺序。
   - 支持每两列之间设置不同连接符，也支持自定义连接符和 {换行符}/{制表符} 等特殊占位符。
   - 支持预览、执行、撤销、保存/载入合并模板。
9. 新增“计划 / 工作流处理”窗口：
   - 支持把批量替换、数据提取、合并列、高级筛选、删除列、移动列组成顺序节点。
   - 上一步输出可直接作为下一步输入。
   - 支持预览到当前节点、预览完整计划、输出到主界面、保存/覆盖SQLite表、导出xlsx。
10. 新增文件工作流节点：获取文件列表、批量重命名，可与数据提取/替换/合并列组合生成新文件名。
11. 新增表格编辑类工作流节点：复制列、复制行、删除行、填充值、序列填充、区域填充。
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import csv
import io
import re
import os
import sys
import json
import traceback
import copy
import threading
import queue
import time
import subprocess
import uuid
from datetime import datetime

from core.data_utils import (
    make_unique_headers_for_append as core_make_unique_headers_for_append,
    normalize_rows as core_normalize_rows,
    safe_cell as core_safe_cell,
)
from core.text_utils import (
    make_sql_columns as core_make_sql_columns,
    quote_ident as core_quote_ident,
    sanitize_sql_name as core_sanitize_sql_name,
)
from db import PluginDatabaseAPI, TableAccessManager
from plugin_runtime.progress import handle_plugin_stdout_line
from plugin_runtime.scanner import scan_plugins
from shared.atomic_json_utils import atomic_write_json, load_json_with_backup
from workflow.nodes.data_nodes import (
    apply_area_fill_node as workflow_apply_area_fill_node,
    apply_copy_column_node as workflow_apply_copy_column_node,
    apply_copy_row_node as workflow_apply_copy_row_node,
    apply_current_datetime_column_node as workflow_apply_current_datetime_column_node,
    apply_delete_columns_node as workflow_apply_delete_columns_node,
    apply_delete_rows_node as workflow_apply_delete_rows_node,
    apply_dedupe_node as workflow_apply_dedupe_node,
    apply_extract_node as workflow_apply_extract_node,
    apply_filter_node as workflow_apply_filter_node,
    apply_format_datetime_node as workflow_apply_format_datetime_node,
    apply_fill_value_node as workflow_apply_fill_value_node,
    apply_merge_node as workflow_apply_merge_node,
    apply_move_columns_node as workflow_apply_move_columns_node,
    apply_new_columns_node as workflow_apply_new_columns_node,
    apply_match_value_output_field_name_node as workflow_apply_match_value_output_field_name_node,
    apply_numeric_column_node as workflow_apply_numeric_column_node,
    apply_replace_node as workflow_apply_replace_node,
    apply_rename_columns_node as workflow_apply_rename_columns_node,
    apply_row_data_mapping_node as workflow_apply_row_data_mapping_node,
    apply_sequence_fill_node as workflow_apply_sequence_fill_node,
    apply_unmatched_format_value as workflow_apply_unmatched_format_value,
    apply_unmatched_extract as workflow_apply_unmatched_extract,
    add_plan_filter_required_field as workflow_add_plan_filter_required_field,
    build_plan_filter_right_index as workflow_build_plan_filter_right_index,
    build_filter_config_probe_result as workflow_build_filter_config_probe_result,
    build_filter_runtime_plan as workflow_build_filter_runtime_plan,
    build_date_parts as workflow_build_date_parts,
    build_format_component_columns as workflow_build_format_component_columns,
    build_time_parts as workflow_build_time_parts,
    complete_format_year as workflow_complete_format_year,
    collect_plan_filter_required_fields as workflow_collect_plan_filter_required_fields,
    ensure_column_count as workflow_ensure_column_count,
    ensure_field_exists as workflow_ensure_field_exists,
    ensure_row_count as workflow_ensure_row_count,
    ensure_target_cell_limit as workflow_ensure_target_cell_limit,
    eval_plan_condition_record as workflow_eval_plan_condition_record,
    eval_plan_join_rule_record as workflow_eval_plan_join_rule_record,
    extract_one_value as workflow_extract_one_value,
    format_output_value as workflow_format_output_value,
    format_numeric_column_result as workflow_format_numeric_column_result,
    format_sequence_value as workflow_format_sequence_value,
    get_plan_filter_config_warnings as workflow_get_plan_filter_config_warnings,
    get_plan_filter_field_owner as workflow_get_plan_filter_field_owner,
    get_plan_filter_hash_join_availability as workflow_get_plan_filter_hash_join_availability,
    get_plan_filter_hash_join_rules as workflow_get_plan_filter_hash_join_rules,
    get_plan_filter_output_base_headers as workflow_get_plan_filter_output_base_headers,
    get_plan_filter_output_header_conflicts as workflow_get_plan_filter_output_header_conflicts,
    get_plan_filter_output_headers as workflow_get_plan_filter_output_headers,
    get_required_columns_for_plan_table as workflow_get_required_columns_for_plan_table,
    get_datetime_parse_warning as workflow_get_datetime_parse_warning,
    get_config_cell_value as workflow_get_config_cell_value,
    get_cycle_source_values_by_config as workflow_get_cycle_source_values_by_config,
    get_fill_targets as workflow_get_fill_targets,
    get_source_area_values_by_config as workflow_get_source_area_values_by_config,
    get_source_column_values_by_config as workflow_get_source_column_values_by_config,
    get_source_row_multi_field_values_by_config as workflow_get_source_row_multi_field_values_by_config,
    get_numeric_node_row_indexes as workflow_get_numeric_node_row_indexes,
    get_row_mapping_end_index as workflow_get_row_mapping_end_index,
    iter_plan_filter_join_candidates as workflow_iter_plan_filter_join_candidates,
    last_non_empty_row_index_by_field as workflow_last_non_empty_row_index_by_field,
    make_current_table_records as workflow_make_current_table_records,
    make_unique_plan_headers as workflow_make_unique_plan_headers,
    match_value_output_column_match as workflow_match_value_output_column_match,
    normalize_datetime_source_text as workflow_normalize_datetime_source_text,
    normalize_filter_condition_value_source as workflow_normalize_filter_condition_value_source,
    normalize_plan_filter_config_field_references as workflow_normalize_plan_filter_config_field_references,
    normalize_plan_filter_field_reference as workflow_normalize_plan_filter_field_reference,
    numeric_node_fallback_value as workflow_numeric_node_fallback_value,
    plan_filter_condition_dependencies as workflow_plan_filter_condition_dependencies,
    plan_filter_field_belongs_to_table as workflow_plan_filter_field_belongs_to_table,
    parse_new_columns_specs as workflow_parse_new_columns_specs,
    parse_numeric_value_for_column_op as workflow_parse_numeric_value_for_column_op,
    parse_date_auto_common as workflow_parse_date_auto_common,
    parse_date_delimited as workflow_parse_date_delimited,
    parse_date_fixed as workflow_parse_date_fixed,
    parse_format_datetime_value as workflow_parse_format_datetime_value,
    parse_format_int as workflow_parse_format_int,
    parse_int as workflow_parse_int,
    parse_row_spec_to_indexes as workflow_parse_row_spec_to_indexes,
    parse_time_auto_common as workflow_parse_time_auto_common,
    parse_time_delimited as workflow_parse_time_delimited,
    parse_time_fixed as workflow_parse_time_fixed,
    post_extract_result as workflow_post_extract_result,
    render_current_datetime_template as workflow_render_current_datetime_template,
    render_format_template as workflow_render_format_template,
    record_passes_plan_conditions as workflow_record_passes_plan_conditions,
    record_passes_plan_join_rules as workflow_record_passes_plan_join_rules,
    record_survives_available_plan_conditions as workflow_record_survives_available_plan_conditions,
    resolve_plan_condition_value as workflow_resolve_plan_condition_value,
    resolve_area_end_row_index as workflow_resolve_area_end_row_index,
    resolve_sequence_count_by_source as workflow_resolve_sequence_count_by_source,
    resolve_start_row_index_by_mode as workflow_resolve_start_row_index_by_mode,
    row_is_empty as workflow_row_is_empty,
    should_write_cell as workflow_should_write_cell,
    slice_by_position as workflow_slice_by_position,
    split_by_config_delimiter as workflow_split_by_config_delimiter,
)
from workflow.nodes.file_nodes import (
    BATCH_RENAME_LOG_HEADERS,
    apply_batch_rename_node as workflow_apply_batch_rename_node,
    apply_file_list_node as workflow_apply_file_list_node,
    is_hidden_path as workflow_is_hidden_path,
    make_numbered_path as workflow_make_numbered_path,
    parse_extensions_filter as workflow_parse_extensions_filter,
)
from workflow.nodes.group_nodes import (
    add_group_inner_node as workflow_add_group_inner_node,
    apply_group_inner_node_list_action as workflow_apply_group_inner_node_list_action,
    apply_group_mapping as workflow_apply_group_mapping,
    apply_inferred_group_inputs as workflow_apply_inferred_group_inputs,
    apply_group_template_config as workflow_apply_group_template_config,
    auto_group_mapping_by_name as workflow_auto_group_mapping_by_name,
    build_group_input_table as workflow_build_group_input_table,
    build_group_output_config_state as workflow_build_group_output_config_state,
    ensure_group_config_defaults as workflow_ensure_group_config_defaults,
    group_input_fields_text as workflow_group_input_fields_text,
    group_inner_node_type_values as workflow_group_inner_node_type_values,
    group_infer_input_apply_decision as workflow_group_infer_input_apply_decision,
    group_mapping_detail as workflow_group_mapping_detail,
    group_mapping_rows as workflow_group_mapping_rows,
    group_mapping_selection_detail as workflow_group_mapping_selection_detail,
    group_node_label as workflow_group_node_label,
    make_group_inner_node as workflow_make_group_inner_node,
    group_selected_input_state as workflow_group_selected_input_state,
    group_source_field_combo_state as workflow_group_source_field_combo_state,
    group_source_headers_for_mapping as workflow_group_source_headers_for_mapping,
    make_group_child_context as workflow_make_group_child_context,
    normalize_group_sqlite_mode as workflow_normalize_group_sqlite_mode,
    normalize_group_transit_conflict_mode as workflow_normalize_group_transit_conflict_mode,
    parse_group_inner_node_json as workflow_parse_group_inner_node_json,
    parse_group_input_fields as workflow_parse_group_input_fields,
    update_group_input_fields_config as workflow_update_group_input_fields_config,
    use_source_headers_as_group_inputs as workflow_use_source_headers_as_group_inputs,
    unique_keep_order as workflow_unique_keep_order,
)
from workflow.nodes.loop_nodes import (
    apply_loop_judge_to_state as workflow_apply_loop_judge_to_state,
    build_loop_judge_output as workflow_build_loop_judge_output,
    build_loop_start_output as workflow_build_loop_start_output,
    evaluate_loop_condition as workflow_evaluate_loop_condition,
    find_loop_judge_index as workflow_find_loop_judge_index,
    find_loop_start_index as workflow_find_loop_start_index,
    init_loop_state_from_source as workflow_init_loop_state_from_source,
    loop_last_non_empty_row_index as workflow_loop_last_non_empty_row_index,
    take_next_loop_item as workflow_take_next_loop_item,
)
from workflow.nodes.plugin_nodes import (
    build_plugin_failure_output as workflow_build_plugin_failure_output,
    build_plugin_final_output as workflow_build_plugin_final_output,
    build_plugin_probe_final_output as workflow_build_plugin_probe_final_output,
    build_plugin_probe_stat as workflow_build_plugin_probe_stat,
    build_plugin_status_text as workflow_build_plugin_status_text,
    get_plugin_output_schema_table as workflow_get_plugin_output_schema_table,
    is_external_plugin_mode as workflow_is_external_plugin_mode,
    make_plugin_input_data as workflow_make_plugin_input_data,
    merge_plugin_output_fields_to_current as workflow_merge_plugin_output_fields_to_current,
    normalize_plugin_logs as workflow_normalize_plugin_logs,
    normalize_plugin_output_schema as workflow_normalize_plugin_output_schema,
    normalize_plugin_run_result as workflow_normalize_plugin_run_result,
    plugin_log_items_to_table as workflow_plugin_log_items_to_table,
    should_save_plugin_output_as_transit as workflow_should_save_plugin_output_as_transit,
)
from workflow.default_configs import default_config_for_type as workflow_default_config_for_type
from workflow.plugin_config_helpers import (
    apply_plugin_custom_config_result as workflow_apply_plugin_custom_config_result,
    build_plugin_dynamic_control_state as workflow_build_plugin_dynamic_control_state,
    build_plugin_dynamic_select_choices as workflow_build_plugin_dynamic_select_choices,
    build_plugin_field_select_initial_value as workflow_build_plugin_field_select_initial_value,
    build_plugin_input_spec as workflow_build_plugin_input_spec,
    build_plugin_load_status_state as workflow_build_plugin_load_status_state,
    build_plugin_select_initial_value as workflow_build_plugin_select_initial_value,
    default_plugin_input_spec as workflow_default_plugin_input_spec,
    ensure_plugin_input_specs as workflow_ensure_plugin_input_specs,
    format_plugin_input_spec as workflow_format_plugin_input_spec,
    get_plugin_field_choices_for_table_param as workflow_get_plugin_field_choices_for_table_param,
    get_plugin_input_table_alias_choices as workflow_get_plugin_input_table_alias_choices,
    get_plugin_static_parameter_choices as workflow_get_plugin_static_parameter_choices,
    normalize_plugin_dynamic_parameter_choices as workflow_normalize_plugin_dynamic_parameter_choices,
    normalize_plugin_run_mode as workflow_normalize_plugin_run_mode,
    plugin_config_transit_reuse_note as workflow_plugin_config_transit_reuse_note,
)
from workflow.plugin_schema_config_ui import (
    build_plugin_output_and_log_section as workflow_build_plugin_output_and_log_section_ui,
    build_plugin_schema_parameter_controls as workflow_build_plugin_schema_parameter_controls_ui,
)
from workflow.plugin_config_ui import (
    build_plugin_node_config as workflow_build_plugin_node_config_ui,
)
from workflow.basic_data_config_ui import (
    build_current_datetime_column_config as workflow_build_current_datetime_column_config_ui,
    build_extract_config as workflow_build_extract_config_ui,
    build_format_datetime_config as workflow_build_format_datetime_config_ui,
    build_new_columns_config as workflow_build_new_columns_config_ui,
    build_replace_config as workflow_build_replace_config_ui,
)
from workflow.rename_columns_config_ui import (
    build_rename_columns_config as workflow_build_rename_columns_config_ui,
)
from workflow.match_value_output_config_ui import (
    build_match_value_output_field_name_config as workflow_build_match_value_output_field_name_config_ui,
)
from workflow.merge_config_ui import (
    build_merge_config as workflow_build_merge_config_ui,
    display_to_sep_value as workflow_display_to_sep_value,
    ensure_separator_count as workflow_ensure_separator_count_ui,
    preview_plan_separator as workflow_preview_plan_separator_ui,
    refresh_merge_separator_ui as workflow_refresh_merge_separator_ui,
    sep_value_to_display as workflow_sep_value_to_display,
    separator_to_input_text as workflow_separator_to_input_text,
)
from workflow.numeric_column_config_ui import (
    build_numeric_column_config as workflow_build_numeric_column_config_ui,
)
from workflow.dedupe_config_ui import (
    build_dedupe_config as workflow_build_dedupe_config_ui,
)
from workflow.file_config_ui import (
    build_batch_rename_config as workflow_build_batch_rename_config_ui,
    build_file_list_config as workflow_build_file_list_config_ui,
)
from workflow.table_edit_config_ui import (
    build_area_fill_config as workflow_build_area_fill_config_ui,
    build_copy_column_config as workflow_build_copy_column_config_ui,
    build_copy_row_config as workflow_build_copy_row_config_ui,
    build_delete_columns_config as workflow_build_delete_columns_config_ui,
    build_delete_rows_config as workflow_build_delete_rows_config_ui,
    build_fill_value_config as workflow_build_fill_value_config_ui,
    build_move_columns_config as workflow_build_move_columns_config_ui,
    build_sequence_fill_config as workflow_build_sequence_fill_config_ui,
)
from workflow.control_flow_config_ui import (
    anchor_id_from_choice as workflow_anchor_id_from_choice_ui,
    build_condition_check_config as workflow_build_condition_check_config_ui,
    build_conditional_jump_config as workflow_build_conditional_jump_config_ui,
    build_jump_anchor_config as workflow_build_jump_anchor_config_ui,
    build_loop_judge_config as workflow_build_loop_judge_config_ui,
    build_loop_start_config as workflow_build_loop_start_config_ui,
    build_unconditional_jump_config as workflow_build_unconditional_jump_config_ui,
    jump_anchor_choices as workflow_jump_anchor_choices_ui,
    set_anchor_var_to_config as workflow_set_anchor_var_to_config_ui,
)
from workflow.filter_config_ui import (
    build_filter_condition_action_buttons as workflow_build_filter_condition_action_buttons_ui,
    build_filter_condition_section as workflow_build_filter_condition_section_ui,
    build_filter_config as workflow_build_filter_config_ui,
    build_filter_header_risk_section as workflow_build_filter_header_risk_section_ui,
    build_filter_join_action_buttons as workflow_build_filter_join_action_buttons_ui,
    build_filter_join_section as workflow_build_filter_join_section_ui,
    build_filter_output_action_buttons as workflow_build_filter_output_action_buttons_ui,
    build_filter_output_section as workflow_build_filter_output_section_ui,
    build_filter_source_table_section as workflow_build_filter_source_table_section_ui,
    edit_filter_condition_cell as workflow_edit_filter_condition_cell_ui,
    filter_tree_rows as workflow_filter_tree_rows_ui,
    invert_output_fields as workflow_invert_output_fields_ui,
    refresh_filter_actual_output_text as workflow_refresh_filter_actual_output_text_ui,
    refresh_filter_condition_value_input as workflow_refresh_filter_condition_value_input_ui,
    refresh_filter_field_sources as workflow_refresh_filter_field_sources_ui,
    refresh_filter_risk_text as workflow_refresh_filter_risk_text_ui,
    replace_filter_tree_rows as workflow_replace_filter_tree_rows_ui,
    select_all_output_fields as workflow_select_all_output_fields_ui,
    select_current_table_output_fields as workflow_select_current_table_output_fields_ui,
    sync_filter_output_fields as workflow_sync_filter_output_fields_ui,
)
from workflow import group_config_ui as workflow_group_config_ui
from workflow import group_runtime as workflow_group_runtime
from workflow import group_template_ui as workflow_group_template_ui
from workflow import jump_runtime as workflow_jump_runtime
from workflow import run_plan_context as workflow_run_plan_context
from workflow import run_plan_dispatch as workflow_run_plan_dispatch
from workflow import run_plan_step as workflow_run_plan_step
from workflow.row_data_mapping_config_ui import (
    build_row_data_mapping_config as workflow_build_row_data_mapping_config_ui,
)
from workflow.save_transit_config_ui import (
    build_save_transit_config as workflow_build_save_transit_config_ui,
)
from workflow.writeback_config_ui import (
    build_writeback_config as workflow_build_writeback_config_ui,
)
from workflow.selected_columns_write_config_ui import (
    build_selected_columns_write_config as workflow_build_selected_columns_write_config_ui,
)
from workflow.nodes.selected_columns_nodes import (
    apply_selected_columns_to_memory_table as workflow_apply_selected_columns_to_memory_table,
    build_selected_columns_write_payload as workflow_build_selected_columns_write_payload,
    build_selected_columns_write_preview_rows as workflow_build_selected_columns_write_preview_rows,
    get_selected_columns_write_skip_stat as workflow_get_selected_columns_write_skip_stat,
    get_selected_columns_write_selected_fields as workflow_get_selected_columns_write_selected_fields,
    make_selected_columns_target_fields as workflow_make_selected_columns_target_fields,
    normalize_selected_columns_write_mode as workflow_normalize_selected_columns_write_mode,
    resolve_selected_columns_write_target as workflow_resolve_selected_columns_write_target,
    selected_columns_should_write as workflow_selected_columns_should_write,
)
from workflow.nodes.transit_nodes import (
    append_headers_rows as workflow_append_headers_rows,
    apply_save_transit_node as workflow_apply_save_transit_node,
    make_unique_transit_name as workflow_make_unique_transit_name,
)
from workflow.nodes.writeback_nodes import (
    apply_external_table_to_current_node as workflow_apply_external_table_to_current_node,
    build_writeback_execute_stat as workflow_build_writeback_execute_stat,
    build_writeback_actions as workflow_build_writeback_actions,
    build_writeback_full_structure_rows_for_sqlite as workflow_build_writeback_full_structure_rows_for_sqlite,
    build_writeback_full_structure_execute_stat as workflow_build_writeback_full_structure_execute_stat,
    build_writeback_preview_rows as workflow_build_writeback_preview_rows,
    build_writeback_preview_stat as workflow_build_writeback_preview_stat,
    count_writeback_actions as workflow_count_writeback_actions,
    compare_writeback_values as workflow_compare_writeback_values,
    finish_writeback_node_output as workflow_finish_writeback_node_output,
    get_writeback_non_execute_suffix as workflow_get_writeback_non_execute_suffix,
    get_writeback_target_fields as workflow_get_writeback_target_fields,
    should_execute_writeback_update as workflow_should_execute_writeback_update,
)
from workflow.table_access_precheck import (
    evaluate_node_table_access_precheck as workflow_evaluate_node_table_access_precheck,
    evaluate_workflow_output_precheck as workflow_evaluate_workflow_output_precheck,
    find_table_access_field_rule as workflow_find_table_access_field_rule,
    find_matching_table_access_entry as workflow_find_matching_table_access_entry,
    iter_nodes_for_table_access_precheck as workflow_iter_nodes_for_table_access_precheck,
    make_table_access_precheck_issue as workflow_make_table_access_precheck_issue,
    normalize_precheck_transit_name as workflow_normalize_precheck_transit_name,
    table_access_entry_match_score as workflow_table_access_entry_match_score,
    table_access_field_items as workflow_table_access_field_items,
    table_access_entry_status as workflow_table_access_entry_status,
    table_access_entry_table_label as workflow_table_access_entry_table_label,
    table_access_operation_summary as workflow_table_access_operation_summary,
    table_access_precheck_actionable as workflow_table_access_precheck_actionable,
    table_access_precheck_blocking as workflow_table_access_precheck_blocking,
    table_access_precheck_sort_key as workflow_table_access_precheck_sort_key,
    table_access_precheck_summary_text as workflow_table_access_precheck_summary_text,
)
from workflow.table_access_defaults import (
    build_default_table_access_for_node as workflow_build_default_table_access_for_node,
)
from workflow.table_access_window_ui import (
    add_table_access_entry as workflow_add_table_access_entry,
    apply_auto_field_mapping_by_name as workflow_apply_auto_field_mapping_by_name,
    apply_auto_field_mapping_by_order as workflow_apply_auto_field_mapping_by_order,
    build_table_access_impact_preview as workflow_build_table_access_impact_preview,
    build_table_access_permission_check as workflow_build_table_access_permission_check,
    clear_field_mapping as workflow_clear_field_mapping,
    delete_table_access_entry as workflow_delete_table_access_entry,
    delete_field_mapping_entry as workflow_delete_field_mapping_entry,
    table_access_field_mapping_mode_choices as workflow_table_access_field_mapping_mode_choices,
    table_access_field_tree_columns as workflow_table_access_field_tree_columns,
    field_mapping_item as workflow_field_mapping_item,
    field_mapping_mode_display as workflow_field_mapping_mode_display,
    field_mapping_mode_value as workflow_field_mapping_mode_value,
    load_field_form as workflow_load_field_form,
    make_table_access_field_key as workflow_make_table_access_field_key,
    table_access_node_tree_columns as workflow_table_access_node_tree_columns,
    rebuild_table_access as workflow_rebuild_table_access,
    render_field_mapping_tree as workflow_render_field_mapping_tree,
    render_table_access_tree as workflow_render_table_access_tree,
    reset_field_form as workflow_reset_field_form,
    table_access_preset_choices as workflow_table_access_preset_choices,
    table_access_role_choices as workflow_table_access_role_choices,
    save_table_access_entry as workflow_save_table_access_entry,
    selected_field_key as workflow_selected_field_key,
    table_access_source_type_choices as workflow_table_access_source_type_choices,
    table_access_table_tree_columns as workflow_table_access_table_tree_columns,
    table_access_preset_config as workflow_table_access_preset_config,
    upsert_field_mapping_entry as workflow_upsert_field_mapping_entry,
)


def get_app_dir():
    """
    返回程序真实工作目录。

    - 直接运行 .py：使用 .py 文件所在目录。
    - PyInstaller 打包为 exe 后：使用 exe 所在目录。

    这样 plan / logs / export / 默认数据库等目录不会被创建到
    PyInstaller 单文件模式的 C 盘临时解压目录 _MEIxxxxx 中。
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def load_json_file_with_recovery(path, parent=None):
    data, info = load_json_with_backup(path)
    warning = info.get("warning", "")
    if warning:
        messagebox.showwarning("配置已从备份恢复", warning, parent=parent)
    return data





class ClipboardTableApp:
    def __init__(self, root):
        self.root = root
        self.root.title("剪贴板表格解析器 - SQLite保存版")
        self.root.geometry("1420x760")

        self.raw_data = ""
        self.headers = []
        self.rows = []

        self.edit_mode = False
        self.edit_entry = None

        # 主界面搜索状态
        self.search_var = tk.StringVar(value="")
        self.search_matches = []
        self.search_index = -1

        # 程序真实目录：兼容直接运行 .py 和 PyInstaller 单文件 exe。
        # 所有需要长期保留的文件都应基于此目录，避免写到 _MEI 临时目录。
        self.app_dir = get_app_dir()

        self.db_path_var = tk.StringVar(value=os.path.join(self.app_dir, "clipboard_tables.db"))
        self.table_name_var = tk.StringVar(value="paste_table")
        self.first_row_header_var = tk.BooleanVar(value=True)
        self.recreate_table_var = tk.BooleanVar(value=True)
        self.edit_btn_text = tk.StringVar(value="修改模式:关")

        self.build_ui()

    def build_ui(self):
        top_frame = ttk.Frame(self.root, padding=8)
        top_frame.pack(fill=tk.X)

        ttk.Button(
            top_frame,
            text="读取剪贴板并解析",
            command=self.load_clipboard
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="清空预览",
            command=self.clear_preview
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="删除字段名，并用下一行作为字段名",
            command=self.delete_header_and_promote_next_row
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            textvariable=self.edit_btn_text,
            command=self.toggle_edit_mode
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="计划 / 工作流处理",
            command=self.open_plan_workflow
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="批量替换 / 数据处理",
            command=self.open_batch_replace
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="数据提取 / 字段生成",
            command=self.open_data_extract
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="合并列 / 生成新列",
            command=self.open_merge_columns
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="高级筛选 / 数据匹配",
            command=self.open_advanced_filter
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="导出为 xlsx",
            command=self.export_current_preview_to_xlsx
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="保存到 SQLite",
            command=self.save_to_sqlite
        ).pack(side=tk.LEFT, padx=4)

        ttk.Button(
            top_frame,
            text="删除当前表",
            command=self.delete_current_sqlite_table
        ).pack(side=tk.LEFT, padx=4)

        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)

        # 主界面选项区拆成独立行，避免不同 row 共用同一个 grid 列宽互相影响。
        # 之前搜索按钮通过较大的 padx 放在 option_frame 的 column=1，
        # 会把数据库路径输入框所在列撑宽，导致“选择 / 刷新表名”整体右移。
        option_frame = ttk.Frame(self.root, padding=8)
        option_frame.pack(fill=tk.X)

        # 第1行：数据库路径设置
        db_frame = ttk.Frame(option_frame)
        db_frame.pack(fill=tk.X, anchor=tk.W)

        ttk.Label(db_frame, text="数据库：").pack(side=tk.LEFT, padx=(4, 4))

        ttk.Entry(
            db_frame,
            textvariable=self.db_path_var,
            width=80
        ).pack(side=tk.LEFT, padx=(4, 4))

        ttk.Button(
            db_frame,
            text="选择",
            command=self.choose_db
        ).pack(side=tk.LEFT, padx=(4, 4))

        ttk.Button(
            db_frame,
            text="刷新表名",
            command=self.refresh_table_list
        ).pack(side=tk.LEFT, padx=(4, 4))

        # 第2行：表名与保存选项
        table_option_frame = ttk.Frame(option_frame)
        table_option_frame.pack(fill=tk.X, anchor=tk.W, pady=(6, 0))

        ttk.Label(table_option_frame, text="表名：").pack(side=tk.LEFT, padx=(4, 4))

        self.table_combo = ttk.Combobox(
            table_option_frame,
            textvariable=self.table_name_var,
            width=32,
            state="normal"
        )
        self.table_combo.pack(side=tk.LEFT, padx=(4, 18))

        self.table_combo.configure(postcommand=self.refresh_table_list)
        self.table_combo.bind("<<ComboboxSelected>>", self.on_table_selected)

        ttk.Checkbutton(
            table_option_frame,
            text="第一行作为字段名",
            variable=self.first_row_header_var,
            command=self.reparse_current_raw
        ).pack(side=tk.LEFT, padx=(12, 12))

        ttk.Checkbutton(
            table_option_frame,
            text="保存时重建同名表",
            variable=self.recreate_table_var
        ).pack(side=tk.LEFT, padx=(12, 12))

        # 第3行：搜索区。搜索按钮保留你指定的 padx=330，但只影响 search_frame 自身，
        # 不再影响上方数据库行和表名行的布局。
        search_frame = ttk.Frame(option_frame)
        search_frame.pack(fill=tk.X, anchor=tk.W, pady=(6, 0))

        ttk.Label(search_frame, text="搜索：").grid(row=0, column=0, sticky=tk.W, padx=(4, 4), pady=4)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=38)
        search_entry.grid(row=0, column=1, sticky=tk.W, padx=(4, 4), pady=4)
        search_entry.bind("<Return>", lambda e: self.search_main_preview(reset=True))
        ttk.Button(search_frame, text="搜索", command=lambda: self.search_main_preview(reset=True)).grid(row=0, column=2, sticky=tk.W, padx=(12, 8), pady=4)
        ttk.Button(search_frame, text="上一个", command=self.search_main_prev).grid(row=0, column=3, sticky=tk.W, padx=(12, 8), pady=4)
        ttk.Button(search_frame, text="下一个", command=self.search_main_next).grid(row=0, column=4, sticky=tk.W, padx=(12, 8), pady=4)

        self.info_var = tk.StringVar(value="等待读取剪贴板数据。")
        ttk.Label(self.root, textvariable=self.info_var, padding=8).pack(fill=tk.X)

        table_frame = ttk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.tree = ttk.Treeview(table_frame, show="headings")

        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)

        self.tree.configure(
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )

        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # 程序启动时尝试刷新表名
        self.refresh_table_list()

    def open_plan_workflow(self):
        if not self.headers:
            messagebox.showwarning("提示", "当前没有可处理的数据，请先读取剪贴板或加载数据库表。")
            return

        PlanWorkflowWindow(self)

    def open_advanced_filter(self):
        db_path = self.db_path_var.get().strip()
        if not db_path:
            messagebox.showwarning("提示", "请先设置 SQLite 数据库路径。")
            return

        if not os.path.exists(db_path):
            messagebox.showwarning("提示", "当前 SQLite 数据库不存在，请先保存数据或选择已有数据库。")
            return

        AdvancedFilterWindow(self)


    def open_batch_replace(self):
        if not self.headers:
            messagebox.showwarning("提示", "当前没有可处理的数据，请先读取剪贴板或加载数据库表。")
            return

        BatchReplaceWindow(self)

    def open_data_extract(self):
        if not self.headers:
            messagebox.showwarning("提示", "当前没有可处理的数据，请先读取剪贴板或加载数据库表。")
            return

        DataExtractWindow(self)

    def open_merge_columns(self):
        if not self.headers:
            messagebox.showwarning("提示", "当前没有可处理的数据，请先读取剪贴板或加载数据库表。")
            return

        MergeColumnsWindow(self)

    def normalize_sheet_title(self, name):
        name = str(name or "导出数据").strip() or "导出数据"
        name = re.sub(r"[\\/*?:\[\]]", "_", name)
        return name[:31] or "导出数据"

    def column_letter(self, index):
        result = ""
        while index > 0:
            index, rem = divmod(index - 1, 26)
            result = chr(65 + rem) + result
        return result or "A"

    def calc_display_width(self, value):
        text = str(value or "")
        width = 0
        for ch in text:
            width += 2 if ord(ch) > 127 else 1
        return width

    def export_current_preview_to_xlsx(self, headers=None, rows=None, table_name=None, title="导出为 xlsx"):
        headers = list(self.headers if headers is None else headers)
        rows = [list(row) for row in (self.rows if rows is None else rows)]
        table_name = self.table_name_var.get() if table_name is None else table_name

        if not headers:
            messagebox.showwarning("提示", "当前没有可导出的表格字段。")
            return

        default_base = self.sanitize_sql_name(table_name, "导出数据")
        default_name = f"{default_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            title=title,
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")]
        )

        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            try:
                self.export_xlsx_with_openpyxl(path, headers=headers, rows=rows, table_name=table_name)
                engine = "openpyxl"
            except ModuleNotFoundError:
                self.export_xlsx_minimal(path, headers=headers, rows=rows, table_name=table_name)
                engine = "内置简易导出器"

            self.info_var.set(f"导出成功：{path}")
            messagebox.showinfo(
                "导出成功",
                f"已导出当前预览数据。\n\n文件：{path}\n行数：{len(rows)}\n列数：{len(headers)}\n导出方式：{engine}"
            )
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def export_xlsx_with_openpyxl(self, path, headers=None, rows=None, table_name=None):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        headers = [str(h) for h in (self.headers if headers is None else headers)]
        rows = [list(row) for row in (self.rows if rows is None else rows)]
        table_name = self.table_name_var.get() if table_name is None else table_name

        wb = Workbook()
        ws = wb.active
        ws.title = self.normalize_sheet_title(table_name)

        ws.append(headers)

        for row in rows:
            fixed = list(row)
            if len(fixed) < len(headers):
                fixed += [""] * (len(headers) - len(fixed))
            if len(fixed) > len(headers):
                fixed = fixed[:len(headers)]
            ws.append(["" if value is None else str(value) for value in fixed])

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        ws.freeze_panes = "A2"
        if headers:
            last_col = self.column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 1)}"

        for col_idx, header in enumerate(headers, start=1):
            max_width = self.calc_display_width(header)
            for row in rows[:3000]:
                if col_idx - 1 < len(row):
                    max_width = max(max_width, self.calc_display_width(row[col_idx - 1]))
            ws.column_dimensions[self.column_letter(col_idx)].width = min(max(max_width + 2, 10), 40)

        wb.save(path)

    def export_xlsx_minimal(self, path, headers=None, rows=None, table_name=None):
        import zipfile
        from xml.sax.saxutils import escape

        headers = [str(h) for h in (self.headers if headers is None else headers)]
        rows = [list(row) for row in (self.rows if rows is None else rows)]
        sheet_rows = [headers]
        for row in rows:
            fixed = list(row)
            if len(fixed) < len(headers):
                fixed += [""] * (len(headers) - len(fixed))
            if len(fixed) > len(headers):
                fixed = fixed[:len(headers)]
            sheet_rows.append(["" if value is None else str(value) for value in fixed])

        def cell_xml(row_idx, col_idx, value, style_id="0"):
            ref = f"{self.column_letter(col_idx)}{row_idx}"
            value = escape(str(value))
            return f'<c r="{ref}" t="inlineStr" s="{style_id}"><is><t>{value}</t></is></c>'

        col_xml = []
        for col_idx, header in enumerate(headers, start=1):
            max_width = self.calc_display_width(header)
            for row in rows[:3000]:
                if col_idx - 1 < len(row):
                    max_width = max(max_width, self.calc_display_width(row[col_idx - 1]))
            width = min(max(max_width + 2, 10), 40)
            col_xml.append(f'<col min="{col_idx}" max="{col_idx}" width="{width}" customWidth="1"/>')

        row_xml_list = []
        for r_idx, row in enumerate(sheet_rows, start=1):
            style_id = "1" if r_idx == 1 else "0"
            cells = "".join(cell_xml(r_idx, c_idx, value, style_id) for c_idx, value in enumerate(row, start=1))
            row_xml_list.append(f'<row r="{r_idx}">{cells}</row>')

        last_col = self.column_letter(len(headers) if headers else 1)
        last_row = max(len(sheet_rows), 1)
        auto_filter = f'<autoFilter ref="A1:{last_col}{last_row}"/>' if headers else ""

        sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheetViews>
    <sheetView workbookViewId="0">
      <pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>
      <selection pane="bottomLeft"/>
    </sheetView>
  </sheetViews>
  <cols>{''.join(col_xml)}</cols>
  <sheetData>{''.join(row_xml_list)}</sheetData>
  {auto_filter}
</worksheet>'''

        styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FFD9EAF7"/></patternFill></fill></fills>
  <borders count="2"><border/><border><left style="thin"><color rgb="FFCCCCCC"/></left><right style="thin"><color rgb="FFCCCCCC"/></right><top style="thin"><color rgb="FFCCCCCC"/></top><bottom style="thin"><color rgb="FFCCCCCC"/></bottom></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"/><xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/></cellXfs>
</styleSheet>'''

        content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>'''

        rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>'''

        workbook_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="{escape(self.normalize_sheet_title(self.table_name_var.get()))}" sheetId="1" r:id="rId1"/></sheets>
</workbook>'''

        workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''

        now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        core_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>ClipboardTableTool</dc:creator><cp:lastModifiedBy>ClipboardTableTool</cp:lastModifiedBy><dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created><dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>
</cp:coreProperties>'''

        app_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>ClipboardTableTool</Application></Properties>'''

        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", content_types)
            zf.writestr("_rels/.rels", rels)
            zf.writestr("xl/workbook.xml", workbook_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
            zf.writestr("xl/styles.xml", styles_xml)
            zf.writestr("docProps/core.xml", core_xml)
            zf.writestr("docProps/app.xml", app_xml)

    def choose_db(self):
        path = filedialog.asksaveasfilename(
            title="选择 SQLite 数据库",
            defaultextension=".db",
            filetypes=[
                ("SQLite 数据库", "*.db"),
                ("SQLite 数据库", "*.sqlite"),
                ("所有文件", "*.*")
            ]
        )
        if path:
            self.db_path_var.set(path)
            self.refresh_table_list()

    def get_db_path(self):
        return self.db_path_var.get().strip()

    def get_table_names(self):
        db_path = self.get_db_path()

        if not db_path or not os.path.exists(db_path):
            return []
        return TableAccessManager(db_path, node_type="主界面").list_tables()

    def get_table_columns(self, table_name):
        db_path = self.get_db_path()

        return TableAccessManager(db_path, node_type="主界面").get_columns(table_name)

    def refresh_table_list(self):
        try:
            tables = self.get_table_names()
            self.table_combo["values"] = tables

            if tables:
                self.info_var.set(f"已读取当前数据库表：{len(tables)} 个。")
            else:
                self.info_var.set("当前数据库中没有普通数据表。")
        except Exception as e:
            self.table_combo["values"] = []
            self.info_var.set(f"读取数据库表失败：{e}")

    def on_table_selected(self, event=None):
        table_name = self.table_name_var.get().strip()

        if not table_name:
            return

        self.load_table_from_sqlite(table_name)

    def toggle_edit_mode(self):
        self.edit_mode = not self.edit_mode

        if self.edit_mode:
            self.edit_btn_text.set("修改模式:开")
            self.info_var.set("修改模式已开启：双击预览表格中的单元格即可修改。")
        else:
            self.edit_btn_text.set("修改模式:关")
            self.info_var.set("修改模式已关闭。")

            if self.edit_entry is not None:
                self.edit_entry.destroy()
                self.edit_entry = None

    def on_tree_double_click(self, event):
        if not self.edit_mode:
            return

        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)

        if not row_id or not col_id:
            return

        try:
            col_index = int(col_id.replace("#", "")) - 1
            row_index = self.tree.index(row_id)
        except Exception:
            return

        if row_index < 0 or row_index >= len(self.rows):
            return

        if col_index < 0 or col_index >= len(self.headers):
            return

        bbox = self.tree.bbox(row_id, col_id)
        if not bbox:
            return

        x, y, width, height = bbox

        old_value = ""
        if col_index < len(self.rows[row_index]):
            old_value = self.rows[row_index][col_index]

        if self.edit_entry is not None:
            self.edit_entry.destroy()
            self.edit_entry = None

        entry = ttk.Entry(self.tree)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, old_value)
        entry.select_range(0, tk.END)
        entry.focus()

        closed = {"done": False}

        def close_editor(save=True):
            if closed["done"]:
                return

            closed["done"] = True

            if save:
                new_value = entry.get()

                while len(self.rows[row_index]) < len(self.headers):
                    self.rows[row_index].append("")

                self.rows[row_index][col_index] = new_value

                values = list(self.tree.item(row_id, "values"))

                while len(values) < len(self.headers):
                    values.append("")

                values[col_index] = new_value
                self.tree.item(row_id, values=values)

                self.info_var.set(f"已修改：第 {row_index + 1} 行，第 {col_index + 1} 列。")

            entry.destroy()
            self.edit_entry = None

        entry.bind("<Return>", lambda e: close_editor(save=True))
        entry.bind("<FocusOut>", lambda e: close_editor(save=True))
        entry.bind("<Escape>", lambda e: close_editor(save=False))

        self.edit_entry = entry

    def load_clipboard(self):
        try:
            data = self.root.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("提示", "剪贴板中没有可读取的文本数据。")
            return

        if not data.strip():
            messagebox.showwarning("提示", "剪贴板内容为空。")
            return

        self.raw_data = data
        self.parse_data(data)

    def reparse_current_raw(self):
        if self.raw_data:
            self.parse_data(self.raw_data)

    def parse_data(self, data):
        data = data.replace("\r\n", "\n").replace("\r", "\n")

        delimiter = "\t"
        if "\t" not in data and "," in data:
            delimiter = ","

        reader = csv.reader(io.StringIO(data), delimiter=delimiter)
        parsed_rows = []

        for row in reader:
            if not row:
                continue

            cleaned_row = [cell.strip() for cell in row]

            if all(cell == "" for cell in cleaned_row):
                continue

            parsed_rows.append(cleaned_row)

        if not parsed_rows:
            messagebox.showwarning("提示", "没有解析到有效表格数据。")
            return

        max_cols = max(len(row) for row in parsed_rows)

        normalized_rows = []
        for row in parsed_rows:
            row = row + [""] * (max_cols - len(row))
            normalized_rows.append(row)

        if self.first_row_header_var.get() and len(normalized_rows) >= 2:
            raw_headers = normalized_rows[0]
            data_rows = normalized_rows[1:]
        else:
            raw_headers = [f"列{i + 1}" for i in range(max_cols)]
            data_rows = normalized_rows

        self.headers = self.make_display_headers(raw_headers)
        self.rows = data_rows

        self.refresh_tree()

        self.info_var.set(
            f"解析完成：{len(self.rows)} 行 × {len(self.headers)} 列。"
            f" 分隔符：{'TAB制表符' if delimiter == chr(9) else '逗号'}"
        )

    def make_display_headers(self, headers):
        result = []
        used = {}

        for index, header in enumerate(headers, start=1):
            name = str(header).strip()
            if not name:
                name = f"列{index}"

            if name in used:
                used[name] += 1
                name = f"{name}_{used[name]}"
            else:
                used[name] = 1

            result.append(name)

        return result

    def refresh_tree(self):
        self.search_matches = []
        self.search_index = -1
        self.tree.delete(*self.tree.get_children())

        self.tree["columns"] = self.headers

        for col in self.headers:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140, minwidth=80, anchor=tk.W, stretch=False)

        self.tree.tag_configure("search_match", background="#fff7cc")
        self.tree.tag_configure("search_current", background="#ffd580")

        for row in self.rows:
            fixed = list(row)
            if len(fixed) < len(self.headers):
                fixed += [""] * (len(self.headers) - len(fixed))
            if len(fixed) > len(self.headers):
                fixed = fixed[:len(self.headers)]
            self.tree.insert("", tk.END, values=fixed)

    def clear_main_search_marks(self):
        for iid in self.tree.get_children():
            self.tree.item(iid, tags=())
        self.search_matches = []
        self.search_index = -1

    def search_main_preview(self, reset=True):
        keyword = self.search_var.get().strip()
        if not keyword:
            messagebox.showwarning("提示", "请输入搜索关键词。")
            return

        keyword_lower = keyword.lower()
        self.clear_main_search_marks()

        for iid in self.tree.get_children():
            values = self.tree.item(iid, "values")
            row_text = "\t".join(str(v) for v in values)
            if keyword_lower in row_text.lower():
                self.search_matches.append(iid)
                self.tree.item(iid, tags=("search_match",))

        if not self.search_matches:
            self.info_var.set(f"搜索完成：未找到包含『{keyword}』的行。")
            return

        self.search_index = 0 if reset else max(self.search_index, 0)
        self.goto_main_search_result()
        self.info_var.set(f"搜索完成：找到 {len(self.search_matches)} 行匹配『{keyword}』。")

    def goto_main_search_result(self):
        if not self.search_matches:
            return
        self.search_index %= len(self.search_matches)
        current_iid = self.search_matches[self.search_index]
        for iid in self.search_matches:
            self.tree.item(iid, tags=("search_match",))
        self.tree.item(current_iid, tags=("search_current",))
        self.tree.selection_set(current_iid)
        self.tree.focus(current_iid)
        self.tree.see(current_iid)
        self.info_var.set(f"当前搜索结果：{self.search_index + 1}/{len(self.search_matches)}")

    def search_main_next(self):
        if not self.search_matches:
            self.search_main_preview(reset=True)
            return
        self.search_index += 1
        self.goto_main_search_result()

    def search_main_prev(self):
        if not self.search_matches:
            self.search_main_preview(reset=True)
            return
        self.search_index -= 1
        self.goto_main_search_result()

    def clear_preview(self):
        self.raw_data = ""
        self.headers = []
        self.rows = []

        if self.edit_entry is not None:
            self.edit_entry.destroy()
            self.edit_entry = None

        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.info_var.set("已清空预览。")

    def delete_header_and_promote_next_row(self):
        if not self.headers:
            messagebox.showwarning("提示", "当前没有字段名，请先读取剪贴板数据。")
            return

        if not self.rows:
            messagebox.showwarning("提示", "当前没有下一行数据，无法提升为字段名。")
            return

        new_headers_raw = self.rows[0]
        new_rows = self.rows[1:]

        self.headers = self.make_display_headers(new_headers_raw)
        self.rows = new_rows

        self.refresh_tree()

        self.info_var.set(
            f"已删除原字段名，并使用下一行作为新字段名："
            f"{len(self.rows)} 行 × {len(self.headers)} 列。"
        )

    def sanitize_sql_name(self, name, default_name):
        return core_sanitize_sql_name(name, default_name)

    def make_sql_columns(self, headers):
        return core_make_sql_columns(headers)

    def quote_ident(self, name):
        return core_quote_ident(name)

    def table_exists(self, conn, table_name):
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,)
        )
        return cur.fetchone() is not None

    def get_available_table_name(self, conn, base_name):
        if not self.table_exists(conn, base_name):
            return base_name

        suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{base_name}_{suffix}"

        counter = 2
        while self.table_exists(conn, new_name):
            new_name = f"{base_name}_{suffix}_{counter}"
            counter += 1

        return new_name

    def format_db_value(self, value):
        if value is None:
            return ""

        if isinstance(value, bytes):
            return f"<BLOB {len(value)} bytes>"

        return str(value)

    def load_table_from_sqlite(self, table_name):
        db_path = self.db_path_var.get().strip()

        if not db_path:
            messagebox.showwarning("提示", "请先选择 SQLite 数据库。")
            return

        if not os.path.exists(db_path):
            messagebox.showwarning("提示", "当前数据库文件不存在。")
            return

        try:
            manager = TableAccessManager(db_path, node_type="主界面读取")
            if not manager.table_exists(table_name):
                messagebox.showwarning("提示", f"表不存在：{table_name}")
                return

            data = manager.read_table(table_name)
            headers = list(data.get("headers", []))

            if not headers:
                messagebox.showwarning("提示", f"表没有字段：{table_name}")
                return

            if self.edit_entry is not None:
                self.edit_entry.destroy()
                self.edit_entry = None

            self.raw_data = ""

            self.headers = self.make_display_headers(headers)
            self.rows = [list(row) for row in data.get("rows", [])]

            self.refresh_tree()

            self.info_var.set(
                f"已加载数据库表：{table_name}，"
                f"{len(self.rows)} 行 × {len(self.headers)} 列。"
            )

        except Exception as e:
            messagebox.showerror("读取表失败", str(e))

    def save_rows_to_sqlite_table(self, table_name_raw, headers, rows, recreate=True):
        db_path = self.get_db_path()
        if not db_path:
            raise ValueError("数据库路径为空。")

        table_name = self.sanitize_sql_name(table_name_raw, "result_table")
        sql_columns = self.make_sql_columns(headers)

        normalized_rows = []
        for row in rows:
            fixed_row = list(row)
            if len(fixed_row) < len(sql_columns):
                fixed_row += [""] * (len(sql_columns) - len(fixed_row))
            if len(fixed_row) > len(sql_columns):
                fixed_row = fixed_row[:len(sql_columns)]
            normalized_rows.append(fixed_row)

        mode = "replace" if recreate else "timestamp"
        info = TableAccessManager(db_path, node_type="主界面保存").write_table(
            table_name,
            sql_columns,
            normalized_rows,
            mode=mode,
        )
        self.refresh_table_list()

        return info.get("table_name", table_name), len(normalized_rows)

    def save_to_sqlite(self):
        if not self.headers or not self.rows:
            messagebox.showwarning("提示", "当前没有可保存的数据，请先读取剪贴板。")
            return

        db_path = self.db_path_var.get().strip()
        if not db_path:
            messagebox.showwarning("提示", "请填写 SQLite 数据库路径。")
            return

        table_name_raw = self.table_name_var.get().strip()

        try:
            table_name, row_count = self.save_rows_to_sqlite_table(
                table_name_raw=table_name_raw,
                headers=self.headers,
                rows=self.rows,
                recreate=self.recreate_table_var.get()
            )

            self.info_var.set(
                f"保存成功：数据库 {db_path}，表 {table_name}，共 {row_count} 行。"
            )

            messagebox.showinfo(
                "保存成功",
                f"已保存到 SQLite。\n\n数据库：{db_path}\n表名：{table_name}\n行数：{row_count}"
            )

        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def make_table_backup_name(self, conn, table_name):
        """生成当前 SQLite 表的备份表名，避免覆盖已有备份。"""
        base_name = f"{table_name}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        backup_name = base_name
        counter = 2
        while self.table_exists(conn, backup_name):
            backup_name = f"{base_name}_{counter}"
            counter += 1
        return backup_name

    def backup_sqlite_table_before_delete(self, conn, table_name):
        """删除前复制当前表到同库备份表，返回备份表名。"""
        backup_name = self.make_table_backup_name(conn, table_name)
        conn.execute(
            f"CREATE TABLE {self.quote_ident(backup_name)} AS "
            f"SELECT * FROM {self.quote_ident(table_name)}"
        )
        return backup_name

    def delete_current_sqlite_table(self):
        """主页删除当前下拉框选中的 SQLite 表。

        安全规则：
        1. 必须先开启修改模式。
        2. 只允许删除当前数据库中已存在、且当前下拉框选中的普通表。
        3. 删除前询问是否备份，备份失败则不删除。
        4. 删除前进行二次确认。
        """
        if not self.edit_mode:
            messagebox.showwarning(
                "禁止删除",
                "删除 SQLite 表属于高风险操作。\n\n请先开启“修改模式:开”，再点击“删除当前表”。"
            )
            return

        db_path = self.get_db_path()
        if not db_path:
            messagebox.showwarning("提示", "请先选择 SQLite 数据库。")
            return

        if not os.path.exists(db_path):
            messagebox.showwarning("提示", "当前 SQLite 数据库文件不存在。")
            return

        table_name = self.table_name_var.get().strip()
        if not table_name:
            messagebox.showwarning("提示", "请先在“表名”下拉框选择要删除的 SQLite 表。")
            return

        try:
            tables = self.get_table_names()
        except Exception as e:
            messagebox.showerror("读取表失败", str(e))
            return

        if table_name not in tables:
            messagebox.showwarning(
                "禁止删除",
                "只能删除当前 SQLite 数据库中已存在、并且从表名下拉框选中的普通表。\n\n"
                f"当前表名：{table_name}"
            )
            return

        if table_name.lower().startswith("sqlite_"):
            messagebox.showwarning("禁止删除", "不能删除 SQLite 系统内部表。")
            return

        backup_choice = messagebox.askyesnocancel(
            "删除当前表",
            "即将删除当前 SQLite 表：\n\n"
            f"数据库：{db_path}\n"
            f"表名：{table_name}\n\n"
            "是否先备份后删除？\n\n"
            "是：先复制为备份表，再删除当前表。\n"
            "否：不备份，直接删除当前表。\n"
            "取消：放弃删除。"
        )

        if backup_choice is None:
            self.info_var.set("已取消删除当前表。")
            return

        confirm_text = (
            "请再次确认删除操作。\n\n"
            f"将删除 SQLite 表：{table_name}\n"
        )
        if backup_choice:
            confirm_text += "删除前会先在当前数据库中创建备份表。\n\n"
        else:
            confirm_text += "本次选择不备份，删除后只能依靠你自己的数据库备份恢复。\n\n"
        confirm_text += "确认继续删除吗？"

        if not messagebox.askyesno("二次确认删除", confirm_text):
            self.info_var.set("已取消删除当前表。")
            return

        try:
            backup_name = TableAccessManager(db_path, node_type="主界面删除").drop_table(
                table_name,
                backup=bool(backup_choice),
            )

            self.refresh_table_list()

            if backup_name:
                self.table_name_var.set(backup_name)
                try:
                    self.load_table_from_sqlite(backup_name)
                except Exception:
                    self.clear_preview()
                msg = f"已备份并删除当前表。备份表：{backup_name}"
            else:
                self.table_name_var.set("")
                self.clear_preview()
                msg = f"已删除当前表：{table_name}"

            self.info_var.set(msg)
            messagebox.showinfo(
                "删除完成",
                f"已删除 SQLite 表：{table_name}"
                + (f"\n\n备份表：{backup_name}" if backup_name else "\n\n本次未创建备份表。")
            )

        except Exception as e:
            messagebox.showerror("删除失败", str(e))
            self.info_var.set(f"删除失败：{e}")



class DataExtractWindow:
    # 数据提取 / 字段生成窗口

    METHODS = [
        "正则提取",
        "固定位置提取",
        "从左取N位",
        "从右取N位",
        "按分隔符提取",
        "前后关键字之间提取",
        "指定字符前提取",
        "指定字符后提取",
        "删除前缀",
        "删除后缀",
    ]
    OUTPUT_MODES = ["生成新字段", "覆盖源字段"]
    UNMATCHED_MODES = ["留空", "保留原值", "填写固定值", "跳过该行"]
    POSITION_BASES = ["从1开始", "从0开始"]
    FIND_MODES = ["第一次出现", "最后一次出现"]

    def __init__(self, app):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("数据提取 / 字段生成")
        self.window.geometry("1320x780")
        self.window.transient(app.root)

        self.preview_results = []
        self.last_backup = None

        self.source_field_var = tk.StringVar(value=app.headers[0] if app.headers else "")
        self.method_var = tk.StringVar(value="正则提取")
        self.output_mode_var = tk.StringVar(value="生成新字段")
        self.new_field_var = tk.StringVar(value="提取结果")
        self.unmatched_mode_var = tk.StringVar(value="留空")
        self.unmatched_fixed_var = tk.StringVar(value="未匹配")
        self.result_limit_var = tk.StringVar(value="1000")
        self.case_sensitive_var = tk.BooleanVar(value=True)
        self.strip_result_var = tk.BooleanVar(value=True)

        # 正则提取
        self.regex_pattern_var = tk.StringVar()
        self.regex_group_var = tk.StringVar(value="0")
        self.regex_find_all_var = tk.BooleanVar(value=False)
        self.regex_joiner_var = tk.StringVar(value=";")

        # 固定位置提取
        self.start_pos_var = tk.StringVar(value="1")
        self.extract_len_var = tk.StringVar(value="1")
        self.position_base_var = tk.StringVar(value="从1开始")

        # 左/右取N位
        self.n_chars_var = tk.StringVar(value="1")

        # 分隔符提取
        self.delimiter_var = tk.StringVar(value="-")
        self.part_index_var = tk.StringVar(value="1")
        self.ignore_empty_part_var = tk.BooleanVar(value=False)

        # 前后关键字之间提取
        self.before_key_var = tk.StringVar()
        self.after_key_var = tk.StringVar()
        self.between_occurrence_var = tk.StringVar(value="1")

        # 指定字符前/后提取
        self.marker_var = tk.StringVar(value="-")
        self.find_mode_var = tk.StringVar(value="第一次出现")

        # 删除前缀/后缀
        self.prefix_var = tk.StringVar()
        self.suffix_var = tk.StringVar()

        self.build_ui()
        self.update_param_ui()

    def build_ui(self):
        main = ttk.Frame(self.window, padding=8)
        main.pack(fill=tk.BOTH, expand=True)

        top = ttk.LabelFrame(main, text="1. 数据源与提取方式", padding=8)
        top.pack(fill=tk.X)

        ttk.Label(top, text="源字段：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        self.source_combo = ttk.Combobox(top, textvariable=self.source_field_var, values=self.app.headers, width=28, state="readonly")
        self.source_combo.grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)

        ttk.Label(top, text="提取方式：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
        method_combo = ttk.Combobox(top, textvariable=self.method_var, values=self.METHODS, width=22, state="readonly")
        method_combo.grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)
        method_combo.bind("<<ComboboxSelected>>", lambda event: self.update_param_ui())

        ttk.Checkbutton(top, text="区分大小写", variable=self.case_sensitive_var).grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(top, text="提取结果去除首尾空格", variable=self.strip_result_var).grid(row=0, column=5, sticky=tk.W, padx=4, pady=4)

        self.param_frame = ttk.LabelFrame(main, text="2. 提取参数", padding=8)
        self.param_frame.pack(fill=tk.X, pady=8)

        output = ttk.LabelFrame(main, text="3. 输出设置", padding=8)
        output.pack(fill=tk.X)

        ttk.Label(output, text="输出方式：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        output_combo = ttk.Combobox(output, textvariable=self.output_mode_var, values=self.OUTPUT_MODES, width=16, state="readonly")
        output_combo.grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
        output_combo.bind("<<ComboboxSelected>>", lambda event: self.update_output_state())

        ttk.Label(output, text="新字段名：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
        self.new_field_entry = ttk.Entry(output, textvariable=self.new_field_var, width=28)
        self.new_field_entry.grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)

        ttk.Label(output, text="未匹配时：").grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(output, textvariable=self.unmatched_mode_var, values=self.UNMATCHED_MODES, width=14, state="readonly").grid(row=0, column=5, sticky=tk.W, padx=4, pady=4)

        ttk.Label(output, text="固定值：").grid(row=0, column=6, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(output, textvariable=self.unmatched_fixed_var, width=18).grid(row=0, column=7, sticky=tk.W, padx=4, pady=4)

        center = ttk.LabelFrame(main, text="4. 提取结果预览", padding=6)
        center.pack(fill=tk.BOTH, expand=True, pady=8)

        self.preview_tree = ttk.Treeview(
            center,
            columns=("行号", "原内容", "提取结果", "状态"),
            show="headings",
            height=16
        )
        for col, width in [("行号", 70), ("原内容", 420), ("提取结果", 420), ("状态", 140)]:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=width, anchor=tk.W, stretch=False)

        y_scroll = ttk.Scrollbar(center, orient=tk.VERTICAL, command=self.preview_tree.yview)
        x_scroll = ttk.Scrollbar(center, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        center.rowconfigure(0, weight=1)
        center.columnconfigure(0, weight=1)

        bottom = ttk.Frame(main)
        bottom.pack(fill=tk.X)

        ttk.Label(bottom, text="预览最大显示行数：").pack(side=tk.LEFT, padx=4)
        ttk.Entry(bottom, textvariable=self.result_limit_var, width=8).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="预览提取结果", command=self.preview_extract).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="执行提取", command=self.execute_extract).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="撤销上一次提取", command=self.undo_last_extract).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="保存当前结果为新表", command=self.save_current_result_to_new_table).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="保存规则模板", command=self.save_template).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="载入规则模板", command=self.load_template).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="关闭", command=self.window.destroy).pack(side=tk.RIGHT, padx=4)

        self.status_var = tk.StringVar(value="提示：正则提取直接使用 Python re 规则。分组 0 表示完整匹配，分组 1 表示第一个括号内容。")
        ttk.Label(main, textvariable=self.status_var, padding=(0, 6)).pack(fill=tk.X)
        self.update_output_state()

    def clear_param_frame(self):
        for child in self.param_frame.winfo_children():
            child.destroy()

    def update_param_ui(self):
        self.clear_param_frame()
        method = self.method_var.get()

        if method == "正则提取":
            ttk.Label(self.param_frame, text="Python正则：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.regex_pattern_var, width=60).grid(row=0, column=1, columnspan=4, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="提取分组：").grid(row=0, column=5, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.regex_group_var, width=8).grid(row=0, column=6, sticky=tk.W, padx=4, pady=4)
            ttk.Checkbutton(self.param_frame, text="提取全部匹配", variable=self.regex_find_all_var).grid(row=1, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="全部匹配连接符：").grid(row=1, column=2, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.regex_joiner_var, width=12).grid(row=1, column=3, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：BP\\d+GK 或 客码[:：]([A-Za-z0-9_-]+)").grid(row=1, column=4, columnspan=4, sticky=tk.W, padx=4, pady=4)

        elif method == "固定位置提取":
            ttk.Label(self.param_frame, text="起始位置：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.start_pos_var, width=10).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="提取长度：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.extract_len_var, width=10).grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="位置规则：").grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)
            ttk.Combobox(self.param_frame, textvariable=self.position_base_var, values=self.POSITION_BASES, width=12, state="readonly").grid(row=0, column=5, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：123456789，起始3、长度4 → 3456（从1开始）").grid(row=1, column=0, columnspan=6, sticky=tk.W, padx=4, pady=4)

        elif method in ["从左取N位", "从右取N位"]:
            ttk.Label(self.param_frame, text="N：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.n_chars_var, width=10).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：ABC123456，取3位 → 左取ABC / 右取456").grid(row=0, column=2, columnspan=5, sticky=tk.W, padx=4, pady=4)

        elif method == "按分隔符提取":
            ttk.Label(self.param_frame, text="分隔符：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.delimiter_var, width=16).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="取第几段：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.part_index_var, width=10).grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)
            ttk.Checkbutton(self.param_frame, text="忽略空段", variable=self.ignore_empty_part_var).grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="段序号从1开始；可填 -1 表示最后一段，-2 表示倒数第2段。").grid(row=1, column=0, columnspan=6, sticky=tk.W, padx=4, pady=4)

        elif method == "前后关键字之间提取":
            ttk.Label(self.param_frame, text="开始关键字：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.before_key_var, width=24).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="结束关键字：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.after_key_var, width=24).grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="第几个匹配：").grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.between_occurrence_var, width=8).grid(row=0, column=5, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：型号[BP2526GK]，开始 型号[，结束 ] → BP2526GK").grid(row=1, column=0, columnspan=6, sticky=tk.W, padx=4, pady=4)

        elif method in ["指定字符前提取", "指定字符后提取"]:
            ttk.Label(self.param_frame, text="指定字符/字符串：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.marker_var, width=20).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="查找位置：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
            ttk.Combobox(self.param_frame, textvariable=self.find_mode_var, values=self.FIND_MODES, width=12, state="readonly").grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：BP2526GK-35RD-01，指定 -，前提取 → BP2526GK").grid(row=1, column=0, columnspan=6, sticky=tk.W, padx=4, pady=4)

        elif method == "删除前缀":
            ttk.Label(self.param_frame, text="要删除的前缀：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.prefix_var, width=30).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：HYBP2526GK 删除 HY → BP2526GK").grid(row=0, column=2, columnspan=5, sticky=tk.W, padx=4, pady=4)

        elif method == "删除后缀":
            ttk.Label(self.param_frame, text="要删除的后缀：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
            ttk.Entry(self.param_frame, textvariable=self.suffix_var, width=30).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
            ttk.Label(self.param_frame, text="示例：BP2526GK_TEMP 删除 _TEMP → BP2526GK").grid(row=0, column=2, columnspan=5, sticky=tk.W, padx=4, pady=4)

    def update_output_state(self):
        if self.output_mode_var.get() == "生成新字段":
            self.new_field_entry.configure(state="normal")
        else:
            self.new_field_entry.configure(state="disabled")

    def get_source_index(self):
        field = self.source_field_var.get().strip()
        if field not in self.app.headers:
            raise ValueError("请选择有效的源字段。")
        return self.app.headers.index(field)

    def parse_int(self, value, name):
        try:
            return int(str(value).strip())
        except Exception:
            raise ValueError(f"{name} 必须是整数。")

    def normalize_case(self, text):
        return text if self.case_sensitive_var.get() else text.lower()

    def find_marker_index(self, text, marker):
        if marker == "":
            raise ValueError("指定字符/字符串不能为空。")
        search_text = self.normalize_case(text)
        search_marker = self.normalize_case(marker)
        if self.find_mode_var.get() == "最后一次出现":
            return search_text.rfind(search_marker)
        return search_text.find(search_marker)

    def apply_unmatched(self, original, status):
        mode = self.unmatched_mode_var.get()
        if mode == "留空":
            return "", status
        if mode == "保留原值":
            return original, status
        if mode == "填写固定值":
            return self.unmatched_fixed_var.get(), status
        if mode == "跳过该行":
            return "", "跳过"
        return "", status

    def post_process_result(self, result):
        if result is None:
            result = ""
        result = str(result)
        if self.strip_result_var.get():
            result = result.strip()
        return result

    def extract_one(self, original):
        text = "" if original is None else str(original)
        method = self.method_var.get()

        try:
            if method == "正则提取":
                pattern = self.regex_pattern_var.get()
                if not pattern:
                    raise ValueError("正则表达式不能为空。")
                flags = 0 if self.case_sensitive_var.get() else re.IGNORECASE
                group_index = self.parse_int(self.regex_group_var.get(), "提取分组")

                if self.regex_find_all_var.get():
                    results = []
                    for m in re.finditer(pattern, text, flags):
                        try:
                            results.append(m.group(group_index))
                        except IndexError:
                            return self.apply_unmatched(text, "分组不存在")
                    if not results:
                        return self.apply_unmatched(text, "未匹配")
                    return self.post_process_result(self.regex_joiner_var.get().join(results)), "成功"

                m = re.search(pattern, text, flags)
                if not m:
                    return self.apply_unmatched(text, "未匹配")
                try:
                    return self.post_process_result(m.group(group_index)), "成功"
                except IndexError:
                    return self.apply_unmatched(text, "分组不存在")

            if method == "固定位置提取":
                start = self.parse_int(self.start_pos_var.get(), "起始位置")
                length = self.parse_int(self.extract_len_var.get(), "提取长度")
                if length < 0:
                    raise ValueError("提取长度不能小于0。")
                start_idx = start - 1 if self.position_base_var.get() == "从1开始" else start
                if start_idx < 0 or start_idx >= len(text):
                    return self.apply_unmatched(text, "越界")
                return self.post_process_result(text[start_idx:start_idx + length]), "成功"

            if method == "从左取N位":
                n = self.parse_int(self.n_chars_var.get(), "N")
                if n < 0:
                    raise ValueError("N不能小于0。")
                return self.post_process_result(text[:n]), "成功"

            if method == "从右取N位":
                n = self.parse_int(self.n_chars_var.get(), "N")
                if n < 0:
                    raise ValueError("N不能小于0。")
                return self.post_process_result(text[-n:] if n else ""), "成功"

            if method == "按分隔符提取":
                delimiter = self.delimiter_var.get()
                if delimiter == "":
                    raise ValueError("分隔符不能为空。")
                parts = text.split(delimiter)
                if self.ignore_empty_part_var.get():
                    parts = [p for p in parts if p != ""]
                part_index = self.parse_int(self.part_index_var.get(), "取第几段")
                if part_index == 0:
                    raise ValueError("段序号不能为0。正数从1开始，负数表示倒数。")
                idx = part_index - 1 if part_index > 0 else part_index
                if idx < -len(parts) or idx >= len(parts):
                    return self.apply_unmatched(text, "越界")
                return self.post_process_result(parts[idx]), "成功"

            if method == "前后关键字之间提取":
                start_key = self.before_key_var.get()
                end_key = self.after_key_var.get()
                if start_key == "" or end_key == "":
                    raise ValueError("开始关键字和结束关键字不能为空。")
                occurrence = self.parse_int(self.between_occurrence_var.get(), "第几个匹配")
                if occurrence <= 0:
                    raise ValueError("第几个匹配必须大于0。")

                search_text = self.normalize_case(text)
                search_start = self.normalize_case(start_key)
                search_end = self.normalize_case(end_key)
                pos = 0
                found = None
                for _ in range(occurrence):
                    s = search_text.find(search_start, pos)
                    if s < 0:
                        return self.apply_unmatched(text, "未匹配")
                    content_start = s + len(start_key)
                    e = search_text.find(search_end, content_start)
                    if e < 0:
                        return self.apply_unmatched(text, "未匹配")
                    found = text[content_start:e]
                    pos = e + len(end_key)
                return self.post_process_result(found), "成功"

            if method == "指定字符前提取":
                marker = self.marker_var.get()
                idx = self.find_marker_index(text, marker)
                if idx < 0:
                    return self.apply_unmatched(text, "未匹配")
                return self.post_process_result(text[:idx]), "成功"

            if method == "指定字符后提取":
                marker = self.marker_var.get()
                idx = self.find_marker_index(text, marker)
                if idx < 0:
                    return self.apply_unmatched(text, "未匹配")
                return self.post_process_result(text[idx + len(marker):]), "成功"

            if method == "删除前缀":
                prefix = self.prefix_var.get()
                if prefix == "":
                    raise ValueError("前缀不能为空。")
                if self.normalize_case(text).startswith(self.normalize_case(prefix)):
                    return self.post_process_result(text[len(prefix):]), "成功"
                return self.apply_unmatched(text, "未匹配")

            if method == "删除后缀":
                suffix = self.suffix_var.get()
                if suffix == "":
                    raise ValueError("后缀不能为空。")
                if self.normalize_case(text).endswith(self.normalize_case(suffix)):
                    return self.post_process_result(text[:-len(suffix)]), "成功"
                return self.apply_unmatched(text, "未匹配")

            raise ValueError(f"未知提取方式：{method}")

        except re.error as e:
            raise ValueError(f"正则错误：{e}")

    def build_preview_results(self):
        source_idx = self.get_source_index()
        results = []
        for row_index, row in enumerate(self.app.rows):
            original = ""
            if source_idx < len(row):
                original = row[source_idx]
            extracted, status = self.extract_one(original)
            results.append({
                "row_index": row_index,
                "original": "" if original is None else str(original),
                "extracted": "" if extracted is None else str(extracted),
                "status": status
            })
        return results

    def get_preview_limit(self):
        try:
            limit = int(self.result_limit_var.get().strip())
            return max(limit, 1)
        except Exception:
            return 1000

    def refresh_preview_tree(self, results):
        self.preview_tree.delete(*self.preview_tree.get_children())
        limit = self.get_preview_limit()
        for item in results[:limit]:
            self.preview_tree.insert("", tk.END, values=(
                item["row_index"] + 1,
                item["original"],
                item["extracted"],
                item["status"]
            ))

    def preview_extract(self):
        try:
            results = self.build_preview_results()
        except Exception as e:
            messagebox.showwarning("预览失败", str(e))
            return

        self.preview_results = results
        self.refresh_preview_tree(results)
        success_count = sum(1 for r in results if r.get("status") == "成功")
        skip_count = sum(1 for r in results if r.get("status") == "跳过")
        self.status_var.set(
            f"预览完成：共 {len(results)} 行，成功 {success_count} 行，跳过 {skip_count} 行，"
            f"当前显示前 {min(self.get_preview_limit(), len(results))} 行。"
        )

    def get_unique_header(self, base_name, headers):
        name = str(base_name or "提取结果").strip() or "提取结果"
        if name not in headers:
            return name
        counter = 2
        while f"{name}_{counter}" in headers:
            counter += 1
        return f"{name}_{counter}"

    def execute_extract(self):
        try:
            results = self.build_preview_results()
            source_idx = self.get_source_index()
        except Exception as e:
            messagebox.showwarning("执行失败", str(e))
            return

        if self.output_mode_var.get() == "覆盖源字段":
            ok = messagebox.askyesno("确认覆盖", "覆盖源字段会直接修改当前预览数据，是否继续？")
            if not ok:
                return

        self.last_backup = {
            "headers": list(self.app.headers),
            "rows": [list(row) for row in self.app.rows]
        }

        changed = 0
        skipped = 0
        if self.output_mode_var.get() == "生成新字段":
            new_field = self.get_unique_header(self.new_field_var.get(), self.app.headers)
            self.app.headers.append(new_field)
            for item, row in zip(results, self.app.rows):
                if item["status"] == "跳过":
                    skipped += 1
                    row.append("")
                    continue
                row.append(item["extracted"])
                changed += 1
        else:
            for item, row in zip(results, self.app.rows):
                if item["status"] == "跳过":
                    skipped += 1
                    continue
                while len(row) < len(self.app.headers):
                    row.append("")
                if source_idx < len(row):
                    row[source_idx] = item["extracted"]
                    changed += 1

        self.preview_results = results
        self.refresh_preview_tree(results)
        self.app.refresh_tree()
        self.app.info_var.set(f"数据提取完成：修改/写入 {changed} 行，跳过 {skipped} 行。")
        self.status_var.set(f"执行完成：修改/写入 {changed} 行，跳过 {skipped} 行。可点击“撤销上一次提取”恢复。")

    def undo_last_extract(self):
        if not self.last_backup:
            messagebox.showwarning("提示", "没有可撤销的提取操作。")
            return

        self.app.headers = list(self.last_backup["headers"])
        self.app.rows = [list(row) for row in self.last_backup["rows"]]
        self.last_backup = None
        self.source_combo.configure(values=self.app.headers)
        if self.app.headers:
            self.source_field_var.set(self.app.headers[0])
        self.app.refresh_tree()
        self.app.info_var.set("已撤销上一次数据提取。")
        self.status_var.set("已撤销上一次数据提取。")

    def save_current_result_to_new_table(self):
        if not self.app.headers:
            messagebox.showwarning("提示", "当前没有可保存的数据。")
            return
        default_name = f"提取结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        name = simpledialog.askstring("保存为新表", "请输入新表名：", initialvalue=default_name, parent=self.window)
        if not name:
            return
        try:
            table_name, row_count = self.app.save_rows_to_sqlite_table(
                table_name_raw=name,
                headers=self.app.headers,
                rows=self.app.rows,
                recreate=False
            )
            self.app.table_name_var.set(table_name)
            self.app.info_var.set(f"数据提取结果已保存为新表：{table_name}，共 {row_count} 行。")
            messagebox.showinfo("保存成功", f"已保存为新表：{table_name}\n行数：{row_count}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def collect_template(self):
        return {
            "source_field": self.source_field_var.get(),
            "method": self.method_var.get(),
            "output_mode": self.output_mode_var.get(),
            "new_field": self.new_field_var.get(),
            "unmatched_mode": self.unmatched_mode_var.get(),
            "unmatched_fixed": self.unmatched_fixed_var.get(),
            "case_sensitive": bool(self.case_sensitive_var.get()),
            "strip_result": bool(self.strip_result_var.get()),
            "regex_pattern": self.regex_pattern_var.get(),
            "regex_group": self.regex_group_var.get(),
            "regex_find_all": bool(self.regex_find_all_var.get()),
            "regex_joiner": self.regex_joiner_var.get(),
            "start_pos": self.start_pos_var.get(),
            "extract_len": self.extract_len_var.get(),
            "position_base": self.position_base_var.get(),
            "n_chars": self.n_chars_var.get(),
            "delimiter": self.delimiter_var.get(),
            "part_index": self.part_index_var.get(),
            "ignore_empty_part": bool(self.ignore_empty_part_var.get()),
            "before_key": self.before_key_var.get(),
            "after_key": self.after_key_var.get(),
            "between_occurrence": self.between_occurrence_var.get(),
            "marker": self.marker_var.get(),
            "find_mode": self.find_mode_var.get(),
            "prefix": self.prefix_var.get(),
            "suffix": self.suffix_var.get(),
        }

    def apply_template(self, data):
        def set_if(name, var):
            if name in data:
                var.set(data[name])

        set_if("source_field", self.source_field_var)
        set_if("method", self.method_var)
        set_if("output_mode", self.output_mode_var)
        set_if("new_field", self.new_field_var)
        set_if("unmatched_mode", self.unmatched_mode_var)
        set_if("unmatched_fixed", self.unmatched_fixed_var)
        if "case_sensitive" in data:
            self.case_sensitive_var.set(bool(data["case_sensitive"]))
        if "strip_result" in data:
            self.strip_result_var.set(bool(data["strip_result"]))
        set_if("regex_pattern", self.regex_pattern_var)
        set_if("regex_group", self.regex_group_var)
        if "regex_find_all" in data:
            self.regex_find_all_var.set(bool(data["regex_find_all"]))
        set_if("regex_joiner", self.regex_joiner_var)
        set_if("start_pos", self.start_pos_var)
        set_if("extract_len", self.extract_len_var)
        set_if("position_base", self.position_base_var)
        set_if("n_chars", self.n_chars_var)
        set_if("delimiter", self.delimiter_var)
        set_if("part_index", self.part_index_var)
        if "ignore_empty_part" in data:
            self.ignore_empty_part_var.set(bool(data["ignore_empty_part"]))
        set_if("before_key", self.before_key_var)
        set_if("after_key", self.after_key_var)
        set_if("between_occurrence", self.between_occurrence_var)
        set_if("marker", self.marker_var)
        set_if("find_mode", self.find_mode_var)
        set_if("prefix", self.prefix_var)
        set_if("suffix", self.suffix_var)
        self.update_param_ui()
        self.update_output_state()

    def save_template(self):
        path = filedialog.asksaveasfilename(
            title="保存数据提取规则模板",
            defaultextension=".json",
            filetypes=[("JSON 模板", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return
        try:
            atomic_write_json(path, self.collect_template())
            self.status_var.set(f"已保存模板：{path}")
        except Exception as e:
            messagebox.showerror("保存模板失败", str(e))

    def load_template(self):
        path = filedialog.askopenfilename(
            title="载入数据提取规则模板",
            filetypes=[("JSON 模板", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return
        try:
            data = load_json_file_with_recovery(path, parent=self.window)
            self.apply_template(data)
            self.status_var.set(f"已载入模板：{path}")
        except Exception as e:
            messagebox.showerror("载入模板失败", str(e))



class MergeColumnsWindow:
    """
    合并列 / 生成新列窗口。

    作用：
    - 从当前主界面预览数据中选择多个字段；
    - 通过“合并顺序列表”明确字段拼接顺序；
    - 支持每两列之间设置不同连接符，也支持自定义连接符；
    - 生成一个新的字段列；
    - 支持预览、执行、撤销、保存/载入模板。
    """

    SEPARATOR_OPTIONS = [
        "空字符", "空格", "换行", "Windows换行", "制表符",
        "-", "_", "/", "\\", "|", ",", ";", ":", ".", "+", "自定义"
    ]

    SEPARATOR_MAP = {
        "空字符": "",
        "空格": " ",
        "换行": "\n",
        "Windows换行": "\r\n",
        "制表符": "\t",
        "-": "-",
        "_": "_",
        "/": "/",
        "\\": "\\",
        "|": "|",
        ",": ",",
        ";": ";",
        ":": ":",
        ".": ".",
        "+": "+",
    }

    def __init__(self, app):
        self.app = app
        self.last_snapshot = None
        self.separator_rows = []

        self.window = tk.Toplevel(app.root)
        self.window.title("合并列 / 生成新列")
        self.window.geometry("1120x760")
        self.window.transient(app.root)

        self.new_field_var = tk.StringVar(value="合并结果")
        self.default_separator_var = tk.StringVar(value="空字符")
        self.default_separator_custom_var = tk.StringVar(value="")
        self.skip_empty_var = tk.BooleanVar(value=False)
        self.trim_value_var = tk.BooleanVar(value=False)
        self.empty_placeholder_var = tk.StringVar(value="")
        self.preview_limit_var = tk.IntVar(value=500)
        self.status_var = tk.StringVar(value="请从左侧字段池添加字段到右侧合并顺序列表。")

        self.build_ui()

    def build_ui(self):
        main = ttk.Frame(self.window, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        source_frame = ttk.LabelFrame(main, text="1. 字段选择与合并顺序", padding=8)
        source_frame.pack(fill=tk.X)

        # 左侧：全部字段池
        left = ttk.Frame(source_frame)
        left.grid(row=0, column=0, sticky="nsw", padx=(0, 8))

        ttk.Label(left, text="可选字段：").pack(anchor=tk.W)

        available_frame = ttk.Frame(left)
        available_frame.pack(fill=tk.Y, pady=4)

        self.available_listbox = tk.Listbox(
            available_frame,
            selectmode=tk.EXTENDED,
            height=12,
            width=32,
            exportselection=False
        )
        available_scroll = ttk.Scrollbar(available_frame, orient=tk.VERTICAL, command=self.available_listbox.yview)
        self.available_listbox.configure(yscrollcommand=available_scroll.set)
        self.available_listbox.pack(side=tk.LEFT, fill=tk.Y)
        available_scroll.pack(side=tk.LEFT, fill=tk.Y)

        self.refresh_available_fields()
        self.available_listbox.bind("<Double-1>", lambda event: self.add_selected_fields())

        left_btns = ttk.Frame(left)
        left_btns.pack(fill=tk.X, pady=4)
        ttk.Button(left_btns, text="全选", command=lambda: self.available_listbox.select_set(0, tk.END)).pack(side=tk.LEFT, padx=2)
        ttk.Button(left_btns, text="清空选择", command=lambda: self.available_listbox.selection_clear(0, tk.END)).pack(side=tk.LEFT, padx=2)

        # 中间：添加/删除按钮
        middle = ttk.Frame(source_frame)
        middle.grid(row=0, column=1, sticky="n", padx=8, pady=28)

        ttk.Button(middle, text="添加 →", command=self.add_selected_fields).pack(fill=tk.X, pady=3)
        ttk.Button(middle, text="← 删除", command=self.remove_order_fields).pack(fill=tk.X, pady=3)
        ttk.Separator(middle, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)
        ttk.Button(middle, text="上移", command=self.move_order_up).pack(fill=tk.X, pady=3)
        ttk.Button(middle, text="下移", command=self.move_order_down).pack(fill=tk.X, pady=3)
        ttk.Button(middle, text="清空", command=self.clear_order_fields).pack(fill=tk.X, pady=3)

        # 右侧：合并顺序列表
        right = ttk.Frame(source_frame)
        right.grid(row=0, column=2, sticky="nsw", padx=(8, 16))

        ttk.Label(right, text="合并顺序：").pack(anchor=tk.W)

        order_frame = ttk.Frame(right)
        order_frame.pack(fill=tk.Y, pady=4)

        self.order_listbox = tk.Listbox(
            order_frame,
            selectmode=tk.EXTENDED,
            height=12,
            width=34,
            exportselection=False
        )
        order_scroll = ttk.Scrollbar(order_frame, orient=tk.VERTICAL, command=self.order_listbox.yview)
        self.order_listbox.configure(yscrollcommand=order_scroll.set)
        self.order_listbox.pack(side=tk.LEFT, fill=tk.Y)
        order_scroll.pack(side=tk.LEFT, fill=tk.Y)
        self.order_listbox.bind("<Delete>", lambda event: self.remove_order_fields())

        # 右侧设置区
        setting = ttk.Frame(source_frame)
        setting.grid(row=0, column=3, sticky="nsew", padx=(8, 0))
        source_frame.columnconfigure(3, weight=1)

        row = 0
        ttk.Label(setting, text="新字段名：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(setting, textvariable=self.new_field_var, width=32).grid(row=row, column=1, sticky=tk.W, padx=4, pady=4)

        row += 1
        ttk.Checkbutton(
            setting,
            text="合并前去除每个字段首尾空格",
            variable=self.trim_value_var
        ).grid(row=row, column=0, columnspan=3, sticky=tk.W, padx=4, pady=4)

        row += 1
        ttk.Checkbutton(
            setting,
            text="跳过空值字段，不参与拼接",
            variable=self.skip_empty_var
        ).grid(row=row, column=0, columnspan=3, sticky=tk.W, padx=4, pady=4)

        row += 1
        ttk.Label(setting, text="空值占位符：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(setting, textvariable=self.empty_placeholder_var, width=32).grid(row=row, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Label(setting, text="不跳过空值时可用，例如填 NA").grid(row=row, column=2, sticky=tk.W, padx=4, pady=4)

        row += 1
        ttk.Label(setting, text="预览最大行数：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Spinbox(setting, from_=10, to=100000, textvariable=self.preview_limit_var, width=12).grid(row=row, column=1, sticky=tk.W, padx=4, pady=4)

        # 列间隔符设置区
        sep_frame = ttk.LabelFrame(main, text="2. 列间隔符设置：每两列之间可使用不同连接符", padding=8)
        sep_frame.pack(fill=tk.X, pady=8)

        sep_top = ttk.Frame(sep_frame)
        sep_top.pack(fill=tk.X)

        ttk.Label(sep_top, text="批量设为：").pack(side=tk.LEFT, padx=4)
        self.default_separator_combo = ttk.Combobox(
            sep_top,
            textvariable=self.default_separator_var,
            values=self.SEPARATOR_OPTIONS,
            width=12,
            state="readonly"
        )
        self.default_separator_combo.pack(side=tk.LEFT, padx=4)
        self.default_separator_combo.bind("<<ComboboxSelected>>", lambda event: self.update_default_custom_state())

        self.default_separator_custom_entry = ttk.Entry(sep_top, textvariable=self.default_separator_custom_var, width=16)
        self.default_separator_custom_entry.pack(side=tk.LEFT, padx=4)
        ttk.Button(sep_top, text="应用到全部间隔符", command=self.apply_default_separator_to_all).pack(side=tk.LEFT, padx=4)
        ttk.Label(sep_top, text="常用：空字符、空格、换行、制表符、-、_，也可选择“自定义”。").pack(side=tk.LEFT, padx=8)
        ttk.Label(
            sep_frame,
            text="提示：自定义连接符支持 {换行符}、{制表符}、{空格}、{空字符}，也兼容 \\n、\\t，可组合普通文字，如 {换行符}客码:",
            foreground="gray",
            wraplength=1060
        ).pack(anchor=tk.W, padx=4, pady=(6, 0))
        self.update_default_custom_state()

        sep_body = ttk.Frame(sep_frame)
        sep_body.pack(fill=tk.X, pady=6)

        self.sep_canvas = tk.Canvas(sep_body, height=150, highlightthickness=0)
        sep_scroll = ttk.Scrollbar(sep_body, orient=tk.VERTICAL, command=self.sep_canvas.yview)
        self.sep_canvas.configure(yscrollcommand=sep_scroll.set)
        self.sep_canvas.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sep_scroll.pack(side=tk.LEFT, fill=tk.Y)

        self.sep_inner = ttk.Frame(self.sep_canvas)
        self.sep_window_id = self.sep_canvas.create_window((0, 0), window=self.sep_inner, anchor="nw")
        self.sep_inner.bind("<Configure>", self.on_separator_inner_configure)
        self.sep_canvas.bind("<Configure>", self.on_separator_canvas_configure)

        # 操作区
        action = ttk.LabelFrame(main, text="3. 操作", padding=8)
        action.pack(fill=tk.X, pady=8)

        ttk.Button(action, text="预览合并结果", command=self.preview_merge).pack(side=tk.LEFT, padx=4)
        ttk.Button(action, text="执行合并到新列", command=self.apply_merge).pack(side=tk.LEFT, padx=4)
        ttk.Button(action, text="撤销上一次合并", command=self.undo_merge).pack(side=tk.LEFT, padx=4)
        ttk.Separator(action, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(action, text="保存合并模板", command=self.save_template).pack(side=tk.LEFT, padx=4)
        ttk.Button(action, text="载入合并模板", command=self.load_template).pack(side=tk.LEFT, padx=4)
        ttk.Button(action, text="关闭", command=self.window.destroy).pack(side=tk.RIGHT, padx=4)

        ttk.Label(main, textvariable=self.status_var).pack(fill=tk.X, pady=4)

        preview_frame = ttk.LabelFrame(main, text="4. 合并结果预览", padding=8)
        preview_frame.pack(fill=tk.BOTH, expand=True)

        self.preview_tree = ttk.Treeview(preview_frame, show="headings")
        y_scroll = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_tree.yview)
        x_scroll = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        self.rebuild_separator_ui()

    def on_separator_inner_configure(self, event=None):
        self.sep_canvas.configure(scrollregion=self.sep_canvas.bbox("all"))

    def on_separator_canvas_configure(self, event=None):
        if event:
            self.sep_canvas.itemconfigure(self.sep_window_id, width=event.width)

    def refresh_available_fields(self):
        if not hasattr(self, "available_listbox"):
            return
        self.available_listbox.delete(0, tk.END)
        for header in self.app.headers:
            self.available_listbox.insert(tk.END, header)

    def add_selected_fields(self):
        selections = list(self.available_listbox.curselection())
        if not selections:
            messagebox.showinfo("提示", "请先在左侧可选字段中选择字段。")
            return

        existing = set(self.get_order_headers())
        added = 0
        for index in selections:
            header = self.app.headers[index]
            if header in existing:
                continue
            self.order_listbox.insert(tk.END, header)
            existing.add(header)
            added += 1

        if added == 0:
            self.status_var.set("所选字段已经在合并顺序列表中。")
        else:
            self.status_var.set(f"已添加 {added} 个字段到合并顺序列表。")

        self.rebuild_separator_ui()

    def remove_order_fields(self):
        selections = list(self.order_listbox.curselection())
        if not selections:
            return
        for index in reversed(selections):
            self.order_listbox.delete(index)
        self.rebuild_separator_ui()
        self.status_var.set("已从合并顺序中删除选中字段。")

    def clear_order_fields(self):
        self.order_listbox.delete(0, tk.END)
        self.rebuild_separator_ui()
        self.status_var.set("已清空合并顺序。")

    def move_order_up(self):
        selections = list(self.order_listbox.curselection())
        if not selections:
            return

        for index in selections:
            if index <= 0:
                continue
            value = self.order_listbox.get(index)
            self.order_listbox.delete(index)
            self.order_listbox.insert(index - 1, value)

        self.order_listbox.selection_clear(0, tk.END)
        for index in selections:
            self.order_listbox.selection_set(max(0, index - 1))

        self.rebuild_separator_ui()

    def move_order_down(self):
        selections = list(self.order_listbox.curselection())
        if not selections:
            return

        size = self.order_listbox.size()
        for index in reversed(selections):
            if index >= size - 1:
                continue
            value = self.order_listbox.get(index)
            self.order_listbox.delete(index)
            self.order_listbox.insert(index + 1, value)

        self.order_listbox.selection_clear(0, tk.END)
        for index in selections:
            self.order_listbox.selection_set(min(size - 1, index + 1))

        self.rebuild_separator_ui()

    def get_order_headers(self):
        return [self.order_listbox.get(i) for i in range(self.order_listbox.size())]

    def get_order_indices(self):
        indices = []
        missing = []
        for header in self.get_order_headers():
            try:
                indices.append(self.app.headers.index(header))
            except ValueError:
                missing.append(header)
        if missing:
            raise ValueError("以下字段在当前表格中不存在：" + ", ".join(missing))
        return indices

    def parse_separator_text(self, text):
        """把用户可读的特殊分隔符写法转换成真实字符。"""
        value = "" if text is None else str(text)
        replacements = [
            ("{Windows换行}", "\r\n"),
            ("{windows换行}", "\r\n"),
            ("{换行符}", "\n"),
            ("{换行}", "\n"),
            ("{newline}", "\n"),
            ("{NEWLINE}", "\n"),
            ("{制表符}", "\t"),
            ("{tab}", "\t"),
            ("{TAB}", "\t"),
            ("{空格}", " "),
            ("{space}", " "),
            ("{SPACE}", " "),
            ("{空字符}", ""),
            ("{empty}", ""),
            ("{EMPTY}", ""),
        ]
        for key, real in replacements:
            value = value.replace(key, real)
        # 兼容高级用户直接输入的转义写法。
        value = value.replace("\\r\\n", "\r\n")
        value = value.replace("\\n", "\n")
        value = value.replace("\\t", "\t")
        return value

    def separator_to_input_text(self, text):
        """把真实换行/制表符转换成输入框里更容易识别的占位符。"""
        value = "" if text is None else str(text)
        value = value.replace("\r\n", "{Windows换行}")
        value = value.replace("\n", "{换行符}")
        value = value.replace("\t", "{制表符}")
        return value

    def display_to_separator(self, option, custom_value=""):
        if option == "自定义":
            return self.parse_separator_text(custom_value)
        return self.SEPARATOR_MAP.get(option, "")

    def separator_to_display(self, sep):
        for display, value in self.SEPARATOR_MAP.items():
            if value == sep:
                return display, ""
        return "自定义", self.separator_to_input_text(sep)

    def get_separator_raw_text(self, index):
        if index < 0 or index >= len(self.separator_rows):
            return ""
        item = self.separator_rows[index]
        option = item["option_var"].get()
        if option == "自定义":
            return item["custom_var"].get()
        return option

    def preview_separator_pair(self, index, left_name, right_name):
        raw_text = self.get_separator_raw_text(index)
        sep = self.get_current_separators()[index] if index < len(self.get_current_separators()) else ""

        win = tk.Toplevel(self.window)
        win.title("连接符效果预览")
        win.geometry("520x360")
        win.transient(self.window)

        frame = ttk.Frame(win, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=f"模拟列数据：{left_name}=A，{right_name}=B").pack(anchor=tk.W, pady=(0, 6))
        ttk.Label(frame, text="用户输入：").pack(anchor=tk.W)
        raw_box = tk.Text(frame, height=4, wrap=tk.WORD)
        raw_box.pack(fill=tk.X, pady=4)
        raw_box.insert("1.0", raw_text)
        raw_box.configure(state="disabled")

        ttk.Label(frame, text="实际合并效果：").pack(anchor=tk.W, pady=(8, 0))
        effect_box = tk.Text(frame, height=7, wrap=tk.WORD)
        effect_box.pack(fill=tk.BOTH, expand=True, pady=4)
        effect_box.insert("1.0", "A" + sep + "B")
        effect_box.configure(state="disabled")

        ttk.Label(frame, text="支持：{换行符}、{制表符}、{空格}、{空字符}，也兼容 \\n、\\t。", foreground="gray").pack(anchor=tk.W, pady=(4, 0))
        ttk.Button(frame, text="关闭", command=win.destroy).pack(anchor=tk.E, pady=(8, 0))

    def get_current_separators(self):
        result = []
        for item in self.separator_rows:
            option = item["option_var"].get()
            custom = item["custom_var"].get()
            result.append(self.display_to_separator(option, custom))
        return result

    def update_default_custom_state(self):
        if not hasattr(self, "default_separator_custom_entry"):
            return
        if self.default_separator_var.get() == "自定义":
            self.default_separator_custom_entry.configure(state="normal")
        else:
            self.default_separator_custom_entry.configure(state="disabled")

    def update_custom_entry_state(self, index):
        if index < 0 or index >= len(self.separator_rows):
            return
        item = self.separator_rows[index]
        if item["option_var"].get() == "自定义":
            item["custom_entry"].configure(state="normal")
        else:
            item["custom_entry"].configure(state="disabled")

    def apply_default_separator_to_all(self):
        option = self.default_separator_var.get()
        custom = self.default_separator_custom_var.get()
        for i, item in enumerate(self.separator_rows):
            item["option_var"].set(option)
            item["custom_var"].set(custom)
            self.update_custom_entry_state(i)
        self.status_var.set("已将批量连接符应用到全部列间隔符。")

    def rebuild_separator_ui(self):
        old_separators = self.get_current_separators() if self.separator_rows else []
        headers = self.get_order_headers() if hasattr(self, "order_listbox") else []

        for child in self.sep_inner.winfo_children():
            child.destroy()
        self.separator_rows = []

        if len(headers) < 2:
            ttk.Label(
                self.sep_inner,
                text="合并顺序少于 2 个字段时，不需要设置列间隔符。"
            ).grid(row=0, column=0, sticky=tk.W, padx=4, pady=6)
            self.sep_canvas.configure(scrollregion=self.sep_canvas.bbox("all"))
            return

        for i in range(len(headers) - 1):
            sep_value = old_separators[i] if i < len(old_separators) else self.display_to_separator(
                self.default_separator_var.get(),
                self.default_separator_custom_var.get()
            )
            option, custom = self.separator_to_display(sep_value)

            option_var = tk.StringVar(value=option)
            custom_var = tk.StringVar(value=custom)

            ttk.Label(
                self.sep_inner,
                text=f"{i + 1}. {headers[i]} 和 {headers[i + 1]} 之间："
            ).grid(row=i, column=0, sticky=tk.W, padx=4, pady=3)

            combo = ttk.Combobox(
                self.sep_inner,
                textvariable=option_var,
                values=self.SEPARATOR_OPTIONS,
                width=12,
                state="readonly"
            )
            combo.grid(row=i, column=1, sticky=tk.W, padx=4, pady=3)

            entry = ttk.Entry(self.sep_inner, textvariable=custom_var, width=24)
            entry.grid(row=i, column=2, sticky=tk.W, padx=4, pady=3)

            preview_btn = ttk.Button(
                self.sep_inner,
                text="预览",
                command=lambda idx=i, left=headers[i], right=headers[i + 1]: self.preview_separator_pair(idx, left, right)
            )
            preview_btn.grid(row=i, column=3, sticky=tk.W, padx=4, pady=3)

            self.separator_rows.append({
                "option_var": option_var,
                "custom_var": custom_var,
                "custom_entry": entry,
            })

            combo.bind("<<ComboboxSelected>>", lambda event, idx=i: self.update_custom_entry_state(idx))
            self.update_custom_entry_state(i)

        self.sep_canvas.configure(scrollregion=self.sep_canvas.bbox("all"))

    def make_unique_header(self, base_name):
        base_name = str(base_name or "合并结果").strip() or "合并结果"
        existing = set(self.app.headers)
        if base_name not in existing:
            return base_name

        i = 2
        while f"{base_name}_{i}" in existing:
            i += 1
        return f"{base_name}_{i}"

    def get_cell_value(self, row, index):
        if index < len(row):
            value = row[index]
        else:
            value = ""

        value = "" if value is None else str(value)

        if self.trim_value_var.get():
            value = value.strip()

        return value

    def build_merged_value_and_status(self, row, selected_indices, separators):
        placeholder = self.empty_placeholder_var.get()
        skip_empty = self.skip_empty_var.get()
        values = [self.get_cell_value(row, index) for index in selected_indices]
        empty_count = sum(1 for value in values if value == "")

        if not values:
            return "", "无字段"

        if skip_empty:
            result = ""
            has_value = False
            for i, value in enumerate(values):
                if value == "":
                    continue
                if not has_value:
                    result = value
                    has_value = True
                else:
                    sep = separators[i - 1] if i - 1 < len(separators) else ""
                    result += sep + value
        else:
            parts = []
            for value in values:
                if value == "":
                    value = placeholder
                parts.append(value)

            result = ""
            for i, part in enumerate(parts):
                if i == 0:
                    result = part
                else:
                    sep = separators[i - 1] if i - 1 < len(separators) else ""
                    result += sep + part

        if empty_count == len(values):
            status = "全部为空"
        elif empty_count > 0:
            status = "部分字段为空"
        else:
            status = "成功"

        return result, status

    def build_merged_value(self, row, selected_indices, separators):
        merged, _status = self.build_merged_value_and_status(row, selected_indices, separators)
        return merged

    def collect_preview_rows(self):
        selected_indices = self.get_order_indices()
        if not selected_indices:
            raise ValueError("请先添加需要合并的字段到右侧合并顺序列表。")

        if not self.app.rows:
            raise ValueError("当前没有可合并的数据行。")

        try:
            limit = int(self.preview_limit_var.get())
        except Exception:
            limit = 500
        if limit <= 0:
            limit = 500

        selected_headers = self.get_order_headers()
        separators = self.get_current_separators()
        if len(separators) < max(0, len(selected_indices) - 1):
            separators += [""] * (len(selected_indices) - 1 - len(separators))

        preview_rows = []
        for row_no, row in enumerate(self.app.rows[:limit], start=1):
            source_values = [self.get_cell_value(row, idx) for idx in selected_indices]
            merged, status = self.build_merged_value_and_status(row, selected_indices, separators)
            preview_rows.append([row_no] + source_values + [merged, status])

        return selected_headers, preview_rows

    def preview_merge(self):
        try:
            selected_headers, preview_rows = self.collect_preview_rows()
        except Exception as e:
            messagebox.showwarning("提示", str(e))
            return

        columns = ["行号"] + selected_headers + ["合并结果", "状态"]
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = columns

        for col in columns:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=150, minwidth=80, anchor=tk.W, stretch=False)

        for row in preview_rows:
            self.preview_tree.insert("", tk.END, values=row)

        self.status_var.set(f"已预览 {len(preview_rows)} 行。合并顺序：{' → '.join(selected_headers)}")

    def apply_merge(self):
        try:
            selected_indices = self.get_order_indices()
        except Exception as e:
            messagebox.showwarning("提示", str(e))
            return

        if not selected_indices:
            messagebox.showwarning("提示", "请先添加需要合并的字段到右侧合并顺序列表。")
            return

        if not self.app.rows:
            messagebox.showwarning("提示", "当前没有可合并的数据行。")
            return

        selected_headers = self.get_order_headers()
        separators = self.get_current_separators()
        if len(separators) < max(0, len(selected_indices) - 1):
            separators += [""] * (len(selected_indices) - 1 - len(separators))

        new_header = self.make_unique_header(self.new_field_var.get())

        confirm = messagebox.askyesno(
            "确认合并",
            "将按以下顺序合并字段：\n\n"
            + " → ".join(selected_headers)
            + f"\n\n生成新字段：{new_header}\n\n是否继续？"
        )
        if not confirm:
            return

        self.last_snapshot = (
            list(self.app.headers),
            [list(row) for row in self.app.rows]
        )

        self.app.headers.append(new_header)

        for row in self.app.rows:
            while len(row) < len(self.app.headers) - 1:
                row.append("")
            row.append(self.build_merged_value(row, selected_indices, separators))

        self.app.refresh_tree()
        self.app.info_var.set(f"合并列完成：已生成新字段 {new_header}，共处理 {len(self.app.rows)} 行。")
        self.status_var.set(f"合并完成：已生成新字段 {new_header}。")

        # 主界面字段变化后，刷新左侧字段池；合并顺序保持原字段不变。
        self.refresh_available_fields()
        self.preview_merge()

    def undo_merge(self):
        if not self.last_snapshot:
            messagebox.showinfo("提示", "没有可撤销的合并操作。")
            return

        headers, rows = self.last_snapshot
        self.app.headers = headers
        self.app.rows = rows
        self.app.refresh_tree()
        self.app.info_var.set("已撤销上一次列合并。")
        self.status_var.set("已撤销上一次列合并。")
        self.last_snapshot = None

        self.refresh_available_fields()
        # 移除顺序列表中已不存在的字段。
        existing = set(self.app.headers)
        current = [h for h in self.get_order_headers() if h in existing]
        self.order_listbox.delete(0, tk.END)
        for header in current:
            self.order_listbox.insert(tk.END, header)
        self.rebuild_separator_ui()

        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = []

    def collect_template(self):
        return {
            "output_field": self.new_field_var.get(),
            "fields": self.get_order_headers(),
            "separators": self.get_current_separators(),
            "skip_empty": self.skip_empty_var.get(),
            "trim_value": self.trim_value_var.get(),
            "empty_placeholder": self.empty_placeholder_var.get(),
            "preview_limit": self.preview_limit_var.get(),
        }

    def apply_template(self, data):
        self.new_field_var.set(data.get("output_field", "合并结果"))
        self.skip_empty_var.set(bool(data.get("skip_empty", False)))
        self.trim_value_var.set(bool(data.get("trim_value", False)))
        self.empty_placeholder_var.set(data.get("empty_placeholder", ""))
        try:
            self.preview_limit_var.set(int(data.get("preview_limit", 500)))
        except Exception:
            self.preview_limit_var.set(500)

        fields = data.get("fields", [])
        existing = set(self.app.headers)
        missing = [field for field in fields if field not in existing]
        valid_fields = [field for field in fields if field in existing]

        self.order_listbox.delete(0, tk.END)
        for field in valid_fields:
            self.order_listbox.insert(tk.END, field)

        # 先重建，再写入模板中的连接符。
        self.rebuild_separator_ui()
        separators = data.get("separators", [])
        for i, sep in enumerate(separators[:len(self.separator_rows)]):
            option, custom = self.separator_to_display(sep)
            self.separator_rows[i]["option_var"].set(option)
            self.separator_rows[i]["custom_var"].set(custom)
            self.update_custom_entry_state(i)

        if missing:
            self.status_var.set("模板已载入，但以下字段不存在，已跳过：" + ", ".join(missing))
        else:
            self.status_var.set("合并模板已载入。")

    def save_template(self):
        path = filedialog.asksaveasfilename(
            title="保存合并规则模板",
            defaultextension=".json",
            filetypes=[("JSON 模板", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return
        try:
            atomic_write_json(path, self.collect_template())
            self.status_var.set(f"已保存合并模板：{path}")
        except Exception as e:
            messagebox.showerror("保存模板失败", str(e))

    def load_template(self):
        path = filedialog.askopenfilename(
            title="载入合并规则模板",
            filetypes=[("JSON 模板", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return
        try:
            data = load_json_file_with_recovery(path, parent=self.window)
            self.apply_template(data)
            self.status_var.set(f"已载入合并模板：{path}")
        except Exception as e:
            messagebox.showerror("载入模板失败", str(e))


class BatchReplaceWindow:
    # 批量替换 / 数据处理窗口

    OPERATORS = ["包含", "不包含", "完全相等", "不等于", "开头是", "结尾是", "为空", "不为空", "正则匹配"]
    REPLACE_MODES = ["局部替换匹配字符串", "整格替换为新值"]
    SCOPES = ["全部行", "当前选中行"]

    def __init__(self, app):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("批量替换 / 数据处理")
        self.window.geometry("1280x760")
        self.window.transient(app.root)

        self.rules = []
        self.preview_changes = []
        self.preview_final_rows = None
        self.last_backup = None

        self.field_var = tk.StringVar(value=app.headers[0] if app.headers else "")
        self.operator_var = tk.StringVar(value="包含")
        self.match_value_var = tk.StringVar()
        self.replace_value_var = tk.StringVar()
        self.replace_mode_var = tk.StringVar(value="局部替换匹配字符串")
        self.scope_var = tk.StringVar(value="全部行")
        self.case_sensitive_var = tk.BooleanVar(value=False)
        self.replace_first_only_var = tk.BooleanVar(value=False)
        self.result_limit_var = tk.StringVar(value="1000")

        self.build_ui()

    def build_ui(self):
        main = ttk.Frame(self.window, padding=8)
        main.pack(fill=tk.BOTH, expand=True)

        rule_frame = ttk.LabelFrame(main, text="1. 替换规则设置", padding=8)
        rule_frame.pack(fill=tk.X)

        ttk.Label(rule_frame, text="目标字段：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        self.field_combo = ttk.Combobox(rule_frame, textvariable=self.field_var, values=self.app.headers, width=24, state="readonly")
        self.field_combo.grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)

        ttk.Label(rule_frame, text="匹配方式：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(rule_frame, textvariable=self.operator_var, values=self.OPERATORS, width=14, state="readonly").grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)

        ttk.Label(rule_frame, text="匹配值：").grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(rule_frame, textvariable=self.match_value_var, width=28).grid(row=0, column=5, sticky=tk.W, padx=4, pady=4)

        ttk.Label(rule_frame, text="替换值：").grid(row=1, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(rule_frame, textvariable=self.replace_value_var, width=28).grid(row=1, column=1, sticky=tk.W, padx=4, pady=4)

        ttk.Label(rule_frame, text="替换方式：").grid(row=1, column=2, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(rule_frame, textvariable=self.replace_mode_var, values=self.REPLACE_MODES, width=22, state="readonly").grid(row=1, column=3, sticky=tk.W, padx=4, pady=4)

        ttk.Label(rule_frame, text="作用范围：").grid(row=1, column=4, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(rule_frame, textvariable=self.scope_var, values=self.SCOPES, width=14, state="readonly").grid(row=1, column=5, sticky=tk.W, padx=4, pady=4)

        ttk.Checkbutton(rule_frame, text="区分大小写", variable=self.case_sensitive_var).grid(row=2, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(rule_frame, text="只替换第一次出现", variable=self.replace_first_only_var).grid(row=2, column=3, sticky=tk.W, padx=4, pady=4)

        btns = ttk.Frame(rule_frame)
        btns.grid(row=2, column=5, sticky=tk.E, padx=4, pady=4)
        ttk.Button(btns, text="添加当前规则", command=self.add_rule).pack(side=tk.LEFT, padx=3)
        ttk.Button(btns, text="删除选中规则", command=self.delete_selected_rule).pack(side=tk.LEFT, padx=3)
        ttk.Button(btns, text="清空规则", command=self.clear_rules).pack(side=tk.LEFT, padx=3)

        center = ttk.PanedWindow(main, orient=tk.HORIZONTAL)
        center.pack(fill=tk.BOTH, expand=True, pady=8)

        rules_frame = ttk.LabelFrame(center, text="2. 规则列表（为空时，预览/执行会使用上方当前输入规则）", padding=6)
        center.add(rules_frame, weight=1)

        self.rules_tree = ttk.Treeview(
            rules_frame,
            columns=("序号", "字段", "匹配方式", "匹配值", "替换值", "替换方式", "范围", "选项"),
            show="headings",
            height=12
        )
        for col, width in [
            ("序号", 50), ("字段", 120), ("匹配方式", 90), ("匹配值", 150),
            ("替换值", 150), ("替换方式", 150), ("范围", 90), ("选项", 150)
        ]:
            self.rules_tree.heading(col, text=col)
            self.rules_tree.column(col, width=width, anchor=tk.W, stretch=False)

        rules_y = ttk.Scrollbar(rules_frame, orient=tk.VERTICAL, command=self.rules_tree.yview)
        rules_x = ttk.Scrollbar(rules_frame, orient=tk.HORIZONTAL, command=self.rules_tree.xview)
        self.rules_tree.configure(yscrollcommand=rules_y.set, xscrollcommand=rules_x.set)
        self.rules_tree.grid(row=0, column=0, sticky="nsew")
        rules_y.grid(row=0, column=1, sticky="ns")
        rules_x.grid(row=1, column=0, sticky="ew")
        rules_frame.rowconfigure(0, weight=1)
        rules_frame.columnconfigure(0, weight=1)

        preview_frame = ttk.LabelFrame(center, text="3. 替换结果预览", padding=6)
        center.add(preview_frame, weight=2)

        self.preview_tree = ttk.Treeview(
            preview_frame,
            columns=("行号", "字段", "原内容", "新内容", "规则"),
            show="headings",
            height=12
        )
        for col, width in [("行号", 70), ("字段", 120), ("原内容", 260), ("新内容", 260), ("规则", 180)]:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=width, anchor=tk.W, stretch=False)

        prev_y = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_tree.yview)
        prev_x = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=prev_y.set, xscrollcommand=prev_x.set)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        prev_y.grid(row=0, column=1, sticky="ns")
        prev_x.grid(row=1, column=0, sticky="ew")
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        bottom = ttk.Frame(main)
        bottom.pack(fill=tk.X)

        ttk.Label(bottom, text="预览最大显示行数：").pack(side=tk.LEFT, padx=4)
        ttk.Entry(bottom, textvariable=self.result_limit_var, width=8).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="预览替换结果", command=self.preview_replace).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="执行替换", command=self.execute_replace).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="撤销上一次替换", command=self.undo_last_replace).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="保存规则模板", command=self.save_template).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="载入规则模板", command=self.load_template).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="关闭", command=self.window.destroy).pack(side=tk.RIGHT, padx=4)

        self.status_var = tk.StringVar(value="提示：局部替换会把字段内部匹配到的字符串替换掉，例如 123456 中 45 → 54，结果为 123546。")
        ttk.Label(main, textvariable=self.status_var, padding=(0, 6)).pack(fill=tk.X)

    def normalize_rule(self):
        field = self.field_var.get().strip()
        if field not in self.app.headers:
            raise ValueError("请选择有效的目标字段。")

        operator = self.operator_var.get().strip()
        match_value = self.match_value_var.get()
        replace_value = self.replace_value_var.get()
        replace_mode = self.replace_mode_var.get().strip()

        if operator not in ["为空", "不为空"] and match_value == "":
            raise ValueError("匹配值不能为空。若要判断空值，请选择“为空”或“不为空”。")

        if replace_mode == "局部替换匹配字符串" and operator in ["为空", "不为空"]:
            raise ValueError("“为空/不为空”建议使用“整格替换为新值”，局部替换没有可替换的匹配字符串。")

        return {
            "field": field,
            "operator": operator,
            "match_value": match_value,
            "replace_value": replace_value,
            "replace_mode": replace_mode,
            "scope": self.scope_var.get().strip() or "全部行",
            "case_sensitive": bool(self.case_sensitive_var.get()),
            "replace_first_only": bool(self.replace_first_only_var.get())
        }

    def add_rule(self):
        try:
            rule = self.normalize_rule()
        except Exception as e:
            messagebox.showwarning("规则无效", str(e))
            return

        self.rules.append(rule)
        self.refresh_rules_tree()
        self.status_var.set(f"已添加规则：{len(self.rules)} 条。")

    def delete_selected_rule(self):
        selected = self.rules_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的规则。")
            return

        indices = sorted([self.rules_tree.index(item) for item in selected], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.rules):
                self.rules.pop(idx)
        self.refresh_rules_tree()
        self.status_var.set(f"已删除选中规则，剩余 {len(self.rules)} 条。")

    def clear_rules(self):
        self.rules.clear()
        self.refresh_rules_tree()
        self.status_var.set("已清空规则列表。")

    def refresh_rules_tree(self):
        self.rules_tree.delete(*self.rules_tree.get_children())
        for i, rule in enumerate(self.rules, start=1):
            opts = []
            opts.append("区分大小写" if rule.get("case_sensitive") else "忽略大小写")
            opts.append("只替换第一次" if rule.get("replace_first_only") else "替换所有")
            self.rules_tree.insert("", tk.END, values=(
                i,
                rule.get("field", ""),
                rule.get("operator", ""),
                rule.get("match_value", ""),
                rule.get("replace_value", ""),
                rule.get("replace_mode", ""),
                rule.get("scope", "全部行"),
                "；".join(opts)
            ))

    def get_rules_for_action(self):
        if self.rules:
            return list(self.rules)
        return [self.normalize_rule()]

    def get_target_indices(self, scope):
        if scope == "当前选中行":
            selected = self.app.tree.selection()
            if not selected:
                return []
            return sorted({self.app.tree.index(item) for item in selected})
        return list(range(len(self.app.rows)))

    def compare_text(self, text, pattern, rule):
        operator = rule.get("operator", "包含")
        case_sensitive = rule.get("case_sensitive", False)

        text = "" if text is None else str(text)
        pattern = "" if pattern is None else str(pattern)

        if operator == "为空":
            return text == ""
        if operator == "不为空":
            return text != ""

        if operator == "正则匹配":
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                return re.search(pattern, text, flags) is not None
            except re.error as e:
                raise ValueError(f"正则表达式错误：{e}")

        cmp_text = text if case_sensitive else text.lower()
        cmp_pattern = pattern if case_sensitive else pattern.lower()

        if operator == "包含":
            return cmp_pattern in cmp_text
        if operator == "不包含":
            return cmp_pattern not in cmp_text
        if operator == "完全相等":
            return cmp_text == cmp_pattern
        if operator == "不等于":
            return cmp_text != cmp_pattern
        if operator == "开头是":
            return cmp_text.startswith(cmp_pattern)
        if operator == "结尾是":
            return cmp_text.endswith(cmp_pattern)

        return False

    def build_replaced_text(self, text, rule):
        text = "" if text is None else str(text)
        match_value = str(rule.get("match_value", ""))
        replace_value = str(rule.get("replace_value", ""))
        replace_mode = rule.get("replace_mode", "局部替换匹配字符串")
        operator = rule.get("operator", "包含")
        case_sensitive = rule.get("case_sensitive", False)
        count = 1 if rule.get("replace_first_only", False) else 0

        if replace_mode == "整格替换为新值":
            return replace_value

        if operator == "正则匹配":
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                return re.sub(match_value, replace_value, text, count=count, flags=flags)
            except re.error as e:
                raise ValueError(f"正则表达式错误：{e}")

        if match_value == "":
            return text

        if case_sensitive:
            return text.replace(match_value, replace_value, count if count else -1)

        return re.sub(re.escape(match_value), replace_value, text, count=count, flags=re.IGNORECASE)

    def normalize_rows_copy(self):
        normalized = []
        col_count = len(self.app.headers)
        for row in self.app.rows:
            fixed = list(row)
            if len(fixed) < col_count:
                fixed += [""] * (col_count - len(fixed))
            if len(fixed) > col_count:
                fixed = fixed[:col_count]
            normalized.append(fixed)
        return normalized

    def compute_changes(self):
        rules = self.get_rules_for_action()
        final_rows = self.normalize_rows_copy()
        changes = []

        for rule_index, rule in enumerate(rules, start=1):
            field = rule.get("field")
            if field not in self.app.headers:
                continue

            col_idx = self.app.headers.index(field)
            target_indices = self.get_target_indices(rule.get("scope", "全部行"))

            for row_idx in target_indices:
                if row_idx < 0 or row_idx >= len(final_rows):
                    continue

                old_value = final_rows[row_idx][col_idx]
                if self.compare_text(old_value, rule.get("match_value", ""), rule):
                    new_value = self.build_replaced_text(old_value, rule)
                    if new_value != old_value:
                        changes.append({
                            "row_index": row_idx,
                            "field": field,
                            "old": old_value,
                            "new": new_value,
                            "rule_index": rule_index,
                            "rule": rule
                        })
                        final_rows[row_idx][col_idx] = new_value

        return changes, final_rows

    def get_preview_limit(self):
        try:
            value = int(self.result_limit_var.get().strip())
            return max(value, 1)
        except Exception:
            return 1000

    def preview_replace(self):
        try:
            changes, final_rows = self.compute_changes()
        except Exception as e:
            messagebox.showerror("预览失败", str(e))
            return

        self.preview_changes = changes
        self.preview_final_rows = final_rows
        self.preview_tree.delete(*self.preview_tree.get_children())

        limit = self.get_preview_limit()
        for change in changes[:limit]:
            self.preview_tree.insert("", tk.END, values=(
                change["row_index"] + 1,
                change["field"],
                change["old"],
                change["new"],
                f"规则{change['rule_index']}"
            ))

        if not changes:
            self.status_var.set("没有找到可替换的数据。")
        else:
            suffix = f"，仅显示前 {limit} 条" if len(changes) > limit else ""
            self.status_var.set(f"预览完成：共 {len(changes)} 处变更{suffix}。")

    def execute_replace(self):
        try:
            changes, final_rows = self.compute_changes()
        except Exception as e:
            messagebox.showerror("执行失败", str(e))
            return

        if not changes:
            messagebox.showinfo("提示", "没有找到可替换的数据。")
            return

        ok = messagebox.askyesno("确认替换", f"本次将修改 {len(changes)} 处内容。\n是否继续？")
        if not ok:
            return

        self.last_backup = [list(row) for row in self.app.rows]
        self.app.rows = final_rows
        self.app.refresh_tree()

        self.preview_changes = changes
        self.preview_final_rows = final_rows
        self.preview_tree.delete(*self.preview_tree.get_children())
        limit = self.get_preview_limit()
        for change in changes[:limit]:
            self.preview_tree.insert("", tk.END, values=(
                change["row_index"] + 1,
                change["field"],
                change["old"],
                change["new"],
                f"规则{change['rule_index']}"
            ))

        self.app.info_var.set(f"批量替换完成：共修改 {len(changes)} 处内容。")
        self.status_var.set(f"执行完成：共修改 {len(changes)} 处内容。可点击“撤销上一次替换”恢复。")

    def undo_last_replace(self):
        if self.last_backup is None:
            messagebox.showwarning("提示", "当前没有可撤销的替换操作。")
            return

        self.app.rows = [list(row) for row in self.last_backup]
        self.app.refresh_tree()
        self.last_backup = None
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.status_var.set("已撤销上一次替换操作。")
        self.app.info_var.set("已撤销上一次批量替换。")

    def save_template(self):
        rules = self.rules
        if not rules:
            try:
                rules = [self.normalize_rule()]
            except Exception as e:
                messagebox.showwarning("规则无效", str(e))
                return

        path = filedialog.asksaveasfilename(
            title="保存替换规则模板",
            defaultextension=".json",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return

        data = {
            "version": 1,
            "type": "batch_replace_template",
            "rules": rules
        }
        try:
            atomic_write_json(path, data)
            self.status_var.set(f"已保存替换规则模板：{path}")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def load_template(self):
        path = filedialog.askopenfilename(
            title="载入替换规则模板",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return

        try:
            data = load_json_file_with_recovery(path, parent=self.window)

            rules = data.get("rules", data if isinstance(data, list) else [])
            valid_rules = []
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                if rule.get("field") in self.app.headers:
                    valid_rules.append({
                        "field": rule.get("field", ""),
                        "operator": rule.get("operator", "包含"),
                        "match_value": rule.get("match_value", ""),
                        "replace_value": rule.get("replace_value", ""),
                        "replace_mode": rule.get("replace_mode", "局部替换匹配字符串"),
                        "scope": rule.get("scope", "全部行"),
                        "case_sensitive": bool(rule.get("case_sensitive", False)),
                        "replace_first_only": bool(rule.get("replace_first_only", False))
                    })

            self.rules = valid_rules
            self.refresh_rules_tree()
            self.status_var.set(f"已载入模板：{path}，有效规则 {len(valid_rules)} 条。")
        except Exception as e:
            messagebox.showerror("载入失败", str(e))

class AdvancedFilterWindow:
    def __init__(self, app):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("高级筛选 / 数据匹配")
        self.window.geometry("1380x820")
        self.window.transient(app.root)

        self.tables_cache = []
        self.columns_cache = {}
        self.field_display_cache = []

        self.conditions = []
        self.join_rules = []
        self.output_fields = []
        self.preview_headers = []
        self.preview_rows = []

        self.main_table_var = tk.StringVar()
        self.add_table_var = tk.StringVar()

        self.filter_field_var = tk.StringVar()
        self.filter_operator_var = tk.StringVar(value="包含")
        self.filter_value_var = tk.StringVar()
        self.logic_var = tk.StringVar(value="AND")

        self.join_left_var = tk.StringVar()
        self.join_operator_var = tk.StringVar(value="等于")
        self.join_right_var = tk.StringVar()
        self.join_logic_var = tk.StringVar(value="AND")

        self.result_limit_var = tk.StringVar(value="5000")
        self.max_intermediate_var = tk.StringVar(value="200000")
        self.save_table_var = tk.StringVar(
            value="筛选结果_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        )

        self.status_var = tk.StringVar(value="请选择数据源。")

        self.build_ui()
        self.refresh_tables()

    def build_ui(self):
        main = ttk.Frame(self.window, padding=8)
        main.pack(fill=tk.BOTH, expand=True)

        top = ttk.Frame(main)
        top.pack(fill=tk.X)

        ttk.Label(top, text="数据库：").pack(side=tk.LEFT)
        ttk.Label(top, text=self.app.get_db_path()).pack(side=tk.LEFT, padx=4)

        ttk.Button(top, text="刷新表/字段", command=self.refresh_tables).pack(side=tk.RIGHT, padx=4)
        ttk.Button(top, text="保存筛选模板", command=self.save_template).pack(side=tk.RIGHT, padx=4)
        ttk.Button(top, text="载入筛选模板", command=self.load_template).pack(side=tk.RIGHT, padx=4)

        body = ttk.Panedwindow(main, orient=tk.HORIZONTAL)
        body.pack(fill=tk.BOTH, expand=True, pady=8)

        left_panel = ttk.Frame(body, padding=4)
        right_panel = ttk.Frame(body, padding=4)

        body.add(left_panel, weight=1)
        body.add(right_panel, weight=2)

        self.build_left_panel(left_panel)
        self.build_right_panel(right_panel)

        ttk.Label(main, textvariable=self.status_var, padding=4).pack(fill=tk.X)

    def build_left_panel(self, parent):
        source_frame = ttk.LabelFrame(parent, text="1. 数据源选择", padding=6)
        source_frame.pack(fill=tk.X, pady=4)

        row1 = ttk.Frame(source_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="主表：", width=8).pack(side=tk.LEFT)
        self.main_table_combo = ttk.Combobox(row1, textvariable=self.main_table_var, state="readonly", width=30)
        self.main_table_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.main_table_combo.bind("<<ComboboxSelected>>", self.on_main_table_selected)

        row2 = ttk.Frame(source_frame)
        row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="添加表：", width=8).pack(side=tk.LEFT)
        self.add_table_combo = ttk.Combobox(row2, textvariable=self.add_table_var, state="readonly", width=30)
        self.add_table_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(row2, text="添加", command=self.add_selected_table).pack(side=tk.LEFT, padx=4)

        row3 = ttk.Frame(source_frame)
        row3.pack(fill=tk.BOTH, expand=True, pady=2)

        self.selected_tables_listbox = tk.Listbox(row3, height=5, exportselection=False)
        self.selected_tables_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        table_scroll = ttk.Scrollbar(row3, orient=tk.VERTICAL, command=self.selected_tables_listbox.yview)
        table_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.selected_tables_listbox.configure(yscrollcommand=table_scroll.set)

        row4 = ttk.Frame(source_frame)
        row4.pack(fill=tk.X, pady=2)
        ttk.Button(row4, text="移除选中表", command=self.remove_selected_table).pack(side=tk.LEFT, padx=2)
        ttk.Button(row4, text="刷新字段列表", command=self.refresh_fields).pack(side=tk.LEFT, padx=2)
        ttk.Button(row4, text="预览选中表格", command=self.preview_selected_source_table).pack(side=tk.LEFT, padx=2)

        filter_frame = ttk.LabelFrame(parent, text="2. 条件筛选", padding=6)
        filter_frame.pack(fill=tk.BOTH, expand=True, pady=4)

        cond_add = ttk.Frame(filter_frame)
        cond_add.pack(fill=tk.X, pady=2)

        ttk.Label(cond_add, text="字段").grid(row=0, column=0, sticky=tk.W)
        ttk.Label(cond_add, text="操作").grid(row=0, column=1, sticky=tk.W)
        ttk.Label(cond_add, text="值").grid(row=0, column=2, sticky=tk.W)

        self.filter_field_combo = ttk.Combobox(cond_add, textvariable=self.filter_field_var, state="readonly", width=24)
        self.filter_field_combo.grid(row=1, column=0, padx=2, pady=2)

        self.filter_operator_combo = ttk.Combobox(
            cond_add,
            textvariable=self.filter_operator_var,
            state="readonly",
            width=12,
            values=[
                "等于", "不等于", "包含", "不包含",
                "开头是", "结尾是",
                "大于", "小于", "大于等于", "小于等于",
                "为空", "不为空",
                "忽略大小写等于", "忽略大小写包含"
            ]
        )
        self.filter_operator_combo.grid(row=1, column=1, padx=2, pady=2)

        ttk.Entry(cond_add, textvariable=self.filter_value_var, width=18).grid(row=1, column=2, padx=2, pady=2)
        ttk.Button(cond_add, text="添加条件", command=self.add_condition).grid(row=1, column=3, padx=2, pady=2)

        logic_row = ttk.Frame(filter_frame)
        logic_row.pack(fill=tk.X, pady=2)
        ttk.Label(logic_row, text="多条件关系：").pack(side=tk.LEFT)
        ttk.Combobox(
            logic_row,
            textvariable=self.logic_var,
            state="readonly",
            width=8,
            values=["AND", "OR"]
        ).pack(side=tk.LEFT)

        self.conditions_tree = ttk.Treeview(
            filter_frame,
            columns=("field", "op", "value"),
            show="headings",
            height=8
        )
        self.conditions_tree.heading("field", text="字段")
        self.conditions_tree.heading("op", text="操作")
        self.conditions_tree.heading("value", text="值")
        self.conditions_tree.column("field", width=170, stretch=False)
        self.conditions_tree.column("op", width=90, stretch=False)
        self.conditions_tree.column("value", width=130, stretch=False)
        self.conditions_tree.pack(fill=tk.BOTH, expand=True, pady=2)

        cond_buttons = ttk.Frame(filter_frame)
        cond_buttons.pack(fill=tk.X, pady=2)
        ttk.Button(cond_buttons, text="删除选中条件", command=self.delete_selected_condition).pack(side=tk.LEFT, padx=2)
        ttk.Button(cond_buttons, text="清空条件", command=self.clear_conditions).pack(side=tk.LEFT, padx=2)

        join_frame = ttk.LabelFrame(parent, text="3. 多表匹配规则", padding=6)
        join_frame.pack(fill=tk.BOTH, expand=True, pady=4)

        join_add = ttk.Frame(join_frame)
        join_add.pack(fill=tk.X, pady=2)

        # 匹配关系放在“左字段”上方同一行，避免单独占用下方空间
        ttk.Label(join_add, text="匹配关系").grid(row=0, column=0, sticky=tk.W)
        ttk.Label(join_add, text="左字段").grid(row=0, column=1, sticky=tk.W)
        ttk.Label(join_add, text="规则").grid(row=0, column=2, sticky=tk.W)
        ttk.Label(join_add, text="右字段").grid(row=0, column=3, sticky=tk.W)

        ttk.Combobox(
            join_add,
            textvariable=self.join_logic_var,
            state="readonly",
            width=8,
            values=["AND", "OR"]
        ).grid(row=1, column=0, padx=2, pady=2)

        self.join_left_combo = ttk.Combobox(join_add, textvariable=self.join_left_var, state="readonly", width=22)
        self.join_left_combo.grid(row=1, column=1, padx=2, pady=2)

        self.join_operator_combo = ttk.Combobox(
            join_add,
            textvariable=self.join_operator_var,
            state="readonly",
            width=14,
            values=["等于", "不等于", "左包含右", "右包含左", "双向包含"]
        )
        self.join_operator_combo.grid(row=1, column=2, padx=2, pady=2)

        self.join_right_combo = ttk.Combobox(join_add, textvariable=self.join_right_var, state="readonly", width=22)
        self.join_right_combo.grid(row=1, column=3, padx=2, pady=2)

        ttk.Button(join_add, text="添加匹配", command=self.add_join_rule).grid(row=1, column=4, padx=2, pady=2)
        ttk.Label(join_add, text="AND=所有匹配规则都满足；OR=任意一条匹配规则满足。", foreground="gray").grid(row=2, column=0, columnspan=5, sticky=tk.W, padx=2, pady=(0, 2))

        self.join_tree = ttk.Treeview(
            join_frame,
            columns=("left", "op", "right"),
            show="headings",
            height=6
        )
        self.join_tree.heading("left", text="左字段")
        self.join_tree.heading("op", text="规则")
        self.join_tree.heading("right", text="右字段")
        self.join_tree.column("left", width=155, stretch=False)
        self.join_tree.column("op", width=85, stretch=False)
        self.join_tree.column("right", width=155, stretch=False)
        self.join_tree.pack(fill=tk.BOTH, expand=True, pady=2)

        join_buttons = ttk.Frame(join_frame)
        join_buttons.pack(fill=tk.X, pady=2)
        ttk.Button(join_buttons, text="删除选中匹配", command=self.delete_selected_join_rule).pack(side=tk.LEFT, padx=2)
        ttk.Button(join_buttons, text="清空匹配规则", command=self.clear_join_rules).pack(side=tk.LEFT, padx=2)

    def build_right_panel(self, parent):
        output_frame = ttk.LabelFrame(parent, text="4. 输出字段选择", padding=6)
        output_frame.pack(fill=tk.X, pady=4)

        output_body = ttk.Frame(output_frame)
        output_body.pack(fill=tk.BOTH, expand=True)

        left_box = ttk.Frame(output_body)
        left_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)

        ttk.Label(left_box, text="可用字段").pack(anchor=tk.W)
        self.available_fields_listbox = tk.Listbox(left_box, selectmode=tk.EXTENDED, height=8, exportselection=False)
        self.available_fields_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        available_scroll = ttk.Scrollbar(left_box, orient=tk.VERTICAL, command=self.available_fields_listbox.yview)
        available_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.available_fields_listbox.configure(yscrollcommand=available_scroll.set)

        mid_buttons = ttk.Frame(output_body)
        mid_buttons.pack(side=tk.LEFT, fill=tk.Y, padx=6)
        ttk.Button(mid_buttons, text="添加 >", command=self.add_output_fields).pack(pady=3)
        ttk.Button(mid_buttons, text="全部添加 >>", command=self.add_all_output_fields).pack(pady=3)
        ttk.Button(mid_buttons, text="< 删除", command=self.remove_output_fields).pack(pady=3)
        ttk.Button(mid_buttons, text="清空", command=self.clear_output_fields).pack(pady=3)

        right_box = ttk.Frame(output_body)
        right_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)

        ttk.Label(right_box, text="输出字段").pack(anchor=tk.W)
        self.output_fields_listbox = tk.Listbox(right_box, selectmode=tk.EXTENDED, height=8, exportselection=False)
        self.output_fields_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        output_scroll = ttk.Scrollbar(right_box, orient=tk.VERTICAL, command=self.output_fields_listbox.yview)
        output_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.output_fields_listbox.configure(yscrollcommand=output_scroll.set)

        setting_frame = ttk.LabelFrame(parent, text="5. 预览与保存", padding=6)
        setting_frame.pack(fill=tk.X, pady=4)

        row1 = ttk.Frame(setting_frame)
        row1.pack(fill=tk.X, pady=2)

        ttk.Label(row1, text="预览最大行数：").pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=self.result_limit_var, width=10).pack(side=tk.LEFT, padx=2)

        ttk.Label(row1, text="中间组合上限：").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(row1, textvariable=self.max_intermediate_var, width=12).pack(side=tk.LEFT, padx=2)

        ttk.Button(row1, text="预览结果", command=self.preview_result).pack(side=tk.LEFT, padx=8)
        ttk.Button(row1, text="去除重复内容", command=self.remove_duplicate_preview_rows).pack(side=tk.LEFT, padx=4)
        ttk.Button(row1, text="载入主界面预览", command=self.load_preview_to_main).pack(side=tk.LEFT, padx=4)

        row2 = ttk.Frame(setting_frame)
        row2.pack(fill=tk.X, pady=2)

        ttk.Label(row2, text="保存为新表：").pack(side=tk.LEFT)
        ttk.Entry(row2, textvariable=self.save_table_var, width=35).pack(side=tk.LEFT, padx=2)
        ttk.Button(row2, text="保存结果到新表", command=self.save_result_to_table).pack(side=tk.LEFT, padx=8)

        preview_frame = ttk.LabelFrame(parent, text="6. 筛选结果预览", padding=6)
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=4)

        self.preview_tree = ttk.Treeview(preview_frame, show="headings")
        y_scroll = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_tree.yview)
        x_scroll = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

    def refresh_tables(self):
        try:
            self.tables_cache = self.app.get_table_names()

            self.main_table_combo["values"] = self.tables_cache
            self.add_table_combo["values"] = self.tables_cache

            if self.tables_cache and not self.main_table_var.get():
                self.main_table_var.set(self.tables_cache[0])
                self.reset_selected_tables_to_main()

            self.columns_cache = {}
            for table in self.tables_cache:
                try:
                    self.columns_cache[table] = self.app.get_table_columns(table)
                except Exception:
                    self.columns_cache[table] = []

            self.refresh_fields()
            self.status_var.set(f"已读取数据库表：{len(self.tables_cache)} 个。")

        except Exception as e:
            messagebox.showerror("刷新失败", str(e))

    def on_main_table_selected(self, event=None):
        self.reset_selected_tables_to_main()
        self.refresh_fields()

    def reset_selected_tables_to_main(self):
        table = self.main_table_var.get().strip()
        self.selected_tables_listbox.delete(0, tk.END)
        if table:
            self.selected_tables_listbox.insert(tk.END, table)

    def get_selected_tables(self):
        return list(self.selected_tables_listbox.get(0, tk.END))

    def get_current_selected_source_table(self):
        """
        获取“1. 数据源选择”区域中当前要预览的表。
        优先使用列表框中选中的表；如果没有选中，则使用主表。
        """
        selections = list(self.selected_tables_listbox.curselection())
        if selections:
            return self.selected_tables_listbox.get(selections[0])

        table = self.main_table_var.get().strip()
        if table:
            return table

        table = self.add_table_var.get().strip()
        if table:
            return table

        return ""

    def preview_selected_source_table(self):
        """
        在右侧“筛选结果预览”区域预览数据源列表中当前选中的表。
        这个预览不会改变筛选条件，也不会执行多表匹配，只是快速查看原表内容。
        """
        table_name = self.get_current_selected_source_table()

        if not table_name:
            messagebox.showwarning("提示", "请先选择一个需要预览的数据表。")
            return

        try:
            columns = self.columns_cache.get(table_name)
            if columns is None:
                columns = self.app.get_table_columns(table_name)
                self.columns_cache[table_name] = columns

            if not columns:
                messagebox.showwarning("提示", f"表没有字段：{table_name}")
                return

            limit = self.get_int_setting(self.result_limit_var, 5000)
            data = TableAccessManager(
                self.app.get_db_path(),
                node_type="高级筛选窗口预览",
            ).read_table(
                table_name,
                limit=limit,
            )

            self.preview_headers = list(data.get("headers", columns))
            self.preview_rows = [list(row) for row in data.get("rows", [])]

            self.refresh_preview_tree()

            self.status_var.set(
                f"已预览选中表格：{table_name}，"
                f"{len(self.preview_rows)} 行 × {len(self.preview_headers)} 列。"
                f" 当前预览行数受“预览最大行数”限制。"
            )

        except Exception as e:
            messagebox.showerror("预览表格失败", str(e))

    def add_selected_table(self):
        table = self.add_table_var.get().strip()
        if not table:
            return

        current = self.get_selected_tables()
        if table not in current:
            self.selected_tables_listbox.insert(tk.END, table)

        self.refresh_fields()

    def remove_selected_table(self):
        selections = list(self.selected_tables_listbox.curselection())
        if not selections:
            return

        main_table = self.main_table_var.get().strip()

        for index in reversed(selections):
            value = self.selected_tables_listbox.get(index)
            if value == main_table:
                messagebox.showwarning("提示", "主表不能从数据源列表中移除。")
                continue
            self.selected_tables_listbox.delete(index)

        self.remove_invalid_rules_and_outputs()
        self.refresh_fields()

    def refresh_fields(self):
        selected_tables = self.get_selected_tables()

        fields = []
        for table in selected_tables:
            columns = self.columns_cache.get(table)
            if columns is None:
                try:
                    columns = self.app.get_table_columns(table)
                    self.columns_cache[table] = columns
                except Exception:
                    columns = []

            for col in columns:
                fields.append(f"{table}.{col}")

        self.field_display_cache = fields

        for combo in [
            self.filter_field_combo,
            self.join_left_combo,
            self.join_right_combo
        ]:
            combo["values"] = fields

        self.available_fields_listbox.delete(0, tk.END)
        for field in fields:
            self.available_fields_listbox.insert(tk.END, field)

        if fields:
            if not self.filter_field_var.get() or self.filter_field_var.get() not in fields:
                self.filter_field_var.set(fields[0])
            if not self.join_left_var.get() or self.join_left_var.get() not in fields:
                self.join_left_var.set(fields[0])
            if not self.join_right_var.get() or self.join_right_var.get() not in fields:
                self.join_right_var.set(fields[min(1, len(fields) - 1)])

        self.remove_invalid_rules_and_outputs()

    def remove_invalid_rules_and_outputs(self):
        valid = set(self.field_display_cache)

        self.conditions = [
            cond for cond in self.conditions
            if cond["field"] in valid
        ]

        self.join_rules = [
            rule for rule in self.join_rules
            if rule["left"] in valid and rule["right"] in valid
        ]

        self.output_fields = [
            field for field in self.output_fields
            if field in valid
        ]

        self.refresh_conditions_tree()
        self.refresh_join_tree()
        self.refresh_output_fields_listbox()

    def add_condition(self):
        field = self.filter_field_var.get().strip()
        op = self.filter_operator_var.get().strip()
        value = self.filter_value_var.get()

        if not field:
            messagebox.showwarning("提示", "请选择筛选字段。")
            return

        if op not in ["为空", "不为空"] and value == "":
            if not messagebox.askyesno("确认", "当前条件值为空，是否继续添加？"):
                return

        self.conditions.append({
            "field": field,
            "op": op,
            "value": value
        })

        self.refresh_conditions_tree()
        self.filter_value_var.set("")

    def delete_selected_condition(self):
        selections = list(self.conditions_tree.selection())
        if not selections:
            return

        indexes = sorted([self.conditions_tree.index(item) for item in selections], reverse=True)
        for index in indexes:
            if 0 <= index < len(self.conditions):
                self.conditions.pop(index)

        self.refresh_conditions_tree()

    def clear_conditions(self):
        self.conditions = []
        self.refresh_conditions_tree()

    def refresh_conditions_tree(self):
        self.conditions_tree.delete(*self.conditions_tree.get_children())
        for cond in self.conditions:
            self.conditions_tree.insert(
                "",
                tk.END,
                values=(cond["field"], cond["op"], cond["value"])
            )

    def add_join_rule(self):
        left = self.join_left_var.get().strip()
        op = self.join_operator_var.get().strip()
        right = self.join_right_var.get().strip()

        if not left or not right:
            messagebox.showwarning("提示", "请选择左右匹配字段。")
            return

        if left == right:
            if not messagebox.askyesno("确认", "左右字段相同，是否仍然添加？"):
                return

        self.join_rules.append({
            "left": left,
            "op": op,
            "right": right
        })

        self.refresh_join_tree()

    def delete_selected_join_rule(self):
        selections = list(self.join_tree.selection())
        if not selections:
            return

        indexes = sorted([self.join_tree.index(item) for item in selections], reverse=True)
        for index in indexes:
            if 0 <= index < len(self.join_rules):
                self.join_rules.pop(index)

        self.refresh_join_tree()

    def clear_join_rules(self):
        self.join_rules = []
        self.refresh_join_tree()

    def refresh_join_tree(self):
        self.join_tree.delete(*self.join_tree.get_children())
        for rule in self.join_rules:
            self.join_tree.insert(
                "",
                tk.END,
                values=(rule["left"], rule["op"], rule["right"])
            )

    def add_output_fields(self):
        selections = list(self.available_fields_listbox.curselection())
        if not selections:
            return

        for index in selections:
            field = self.available_fields_listbox.get(index)
            if field not in self.output_fields:
                self.output_fields.append(field)

        self.refresh_output_fields_listbox()

    def add_all_output_fields(self):
        for field in self.field_display_cache:
            if field not in self.output_fields:
                self.output_fields.append(field)

        self.refresh_output_fields_listbox()

    def remove_output_fields(self):
        selections = list(self.output_fields_listbox.curselection())
        if not selections:
            return

        for index in reversed(selections):
            if 0 <= index < len(self.output_fields):
                self.output_fields.pop(index)

        self.refresh_output_fields_listbox()

    def clear_output_fields(self):
        self.output_fields = []
        self.refresh_output_fields_listbox()

    def refresh_output_fields_listbox(self):
        self.output_fields_listbox.delete(0, tk.END)
        for field in self.output_fields:
            self.output_fields_listbox.insert(tk.END, field)

    def format_db_value(self, value):
        return self.app.format_db_value(value)

    def load_table_records(self, table_name):
        columns = self.columns_cache.get(table_name)
        if columns is None:
            columns = self.app.get_table_columns(table_name)
            self.columns_cache[table_name] = columns

        data = TableAccessManager(
            self.app.get_db_path(),
            node_type="高级筛选窗口读取",
        ).read_table(table_name)
        rows = [list(row) for row in data.get("rows", [])]

        records = []
        for row in rows:
            record = {}
            for idx, col in enumerate(columns):
                key = f"{table_name}.{col}"
                value = row[idx] if idx < len(row) else ""
                record[key] = value
            records.append(record)

        return records

    def parse_number(self, value):
        text = str(value).strip()
        if text == "":
            return None

        # 去掉常见千分位逗号
        text = text.replace(",", "")
        return float(text)

    def eval_condition(self, record, cond):
        field = cond["field"]
        op = cond["op"]
        target = cond.get("value", "")

        value = record.get(field, "")
        value_text = "" if value is None else str(value)
        target_text = "" if target is None else str(target)

        if op == "等于":
            return value_text == target_text
        if op == "不等于":
            return value_text != target_text
        if op == "包含":
            return target_text in value_text
        if op == "不包含":
            return target_text not in value_text
        if op == "开头是":
            return value_text.startswith(target_text)
        if op == "结尾是":
            return value_text.endswith(target_text)
        if op == "为空":
            return value_text.strip() == ""
        if op == "不为空":
            return value_text.strip() != ""
        if op == "忽略大小写等于":
            return value_text.lower() == target_text.lower()
        if op == "忽略大小写包含":
            return target_text.lower() in value_text.lower()

        if op in ["大于", "小于", "大于等于", "小于等于"]:
            try:
                left = self.parse_number(value_text)
                right = self.parse_number(target_text)
                if left is None or right is None:
                    return False

                if op == "大于":
                    return left > right
                if op == "小于":
                    return left < right
                if op == "大于等于":
                    return left >= right
                if op == "小于等于":
                    return left <= right
            except Exception:
                return False

        return False

    def eval_join_rule(self, record, rule):
        left_value = str(record.get(rule["left"], ""))
        right_value = str(record.get(rule["right"], ""))
        op = rule["op"]

        if op == "等于":
            return left_value == right_value
        if op == "不等于":
            return left_value != right_value
        if op == "左包含右":
            if right_value == "":
                return False
            return right_value in left_value
        if op == "右包含左":
            if left_value == "":
                return False
            return left_value in right_value
        if op == "双向包含":
            if left_value == "" or right_value == "":
                return False
            return left_value in right_value or right_value in left_value

        return False

    def eval_conditions(self, record):
        if not self.conditions:
            return True

        results = [self.eval_condition(record, cond) for cond in self.conditions]

        if self.logic_var.get() == "OR":
            return any(results)

        return all(results)

    def eval_join_rules(self, record):
        if not self.join_rules:
            return True

        checks = []
        for rule in self.join_rules:
            # 规则引用的字段还没组合进当前中间记录时，暂时不参与本轮判断，等后续表组合后再生效。
            if rule["left"] in record and rule["right"] in record:
                checks.append(self.eval_join_rule(record, rule))

        if not checks:
            return True
        return any(checks) if self.join_logic_var.get() == "OR" else all(checks)

    def get_int_setting(self, var, default_value):
        try:
            value = int(str(var.get()).strip())
            if value <= 0:
                return default_value
            return value
        except Exception:
            return default_value

    def build_result_records(self):
        selected_tables = self.get_selected_tables()

        if not selected_tables:
            raise ValueError("请至少选择一个数据表。")

        result_limit = self.get_int_setting(self.result_limit_var, 5000)
        max_intermediate = self.get_int_setting(self.max_intermediate_var, 200000)

        table_records_map = {}
        for table in selected_tables:
            table_records_map[table] = self.load_table_records(table)

        if len(selected_tables) == 1:
            records = table_records_map[selected_tables[0]]
            filtered = []
            for record in records:
                if self.eval_conditions(record):
                    filtered.append(record)
                    if len(filtered) >= result_limit:
                        break
            return filtered

        combined_records = table_records_map[selected_tables[0]]

        for table in selected_tables[1:]:
            new_records = []
            right_records = table_records_map[table]

            for left_record in combined_records:
                for right_record in right_records:
                    merged = {}
                    merged.update(left_record)
                    merged.update(right_record)

                    if self.eval_join_rules(merged):
                        new_records.append(merged)

                        if len(new_records) > max_intermediate:
                            raise RuntimeError(
                                f"中间结果超过上限 {max_intermediate} 行。"
                                "请增加匹配规则或筛选条件，避免笛卡尔组合过大。"
                            )

            combined_records = new_records

            if not combined_records:
                break

        filtered = []
        for record in combined_records:
            if self.eval_conditions(record):
                filtered.append(record)
                if len(filtered) >= result_limit:
                    break

        return filtered

    def get_output_fields(self):
        if self.output_fields:
            return self.output_fields

        return self.field_display_cache

    def preview_result(self):
        try:
            fields = self.get_output_fields()
            if not fields:
                messagebox.showwarning("提示", "没有可输出字段，请先选择数据源。")
                return

            records = self.build_result_records()

            self.preview_headers = fields
            self.preview_rows = []

            for record in records:
                self.preview_rows.append([record.get(field, "") for field in fields])

            self.refresh_preview_tree()

            self.status_var.set(
                f"预览完成：{len(self.preview_rows)} 行 × {len(self.preview_headers)} 列。"
                f" 当前预览行数受“预览最大行数”限制。"
            )

        except Exception as e:
            messagebox.showerror("预览失败", str(e))

    def refresh_preview_tree(self):
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = self.preview_headers

        for col in self.preview_headers:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=150, minwidth=80, anchor=tk.W, stretch=False)

        for row in self.preview_rows:
            self.preview_tree.insert("", tk.END, values=row)

    def remove_duplicate_preview_rows(self):
        if not self.preview_headers:
            self.preview_result()
        if not self.preview_headers:
            return

        seen = set()
        new_rows = []
        removed = 0
        for row in self.preview_rows:
            key = tuple("" if value is None else str(value) for value in row)
            if key in seen:
                removed += 1
                continue
            seen.add(key)
            new_rows.append(list(row))

        self.preview_rows = new_rows
        self.refresh_preview_tree()
        self.status_var.set(
            f"已去除重复内容：删除 {removed} 行，剩余 {len(self.preview_rows)} 行。"
            " 判断规则：按当前预览输出整行内容去重，保留第一条。"
        )

    def load_preview_to_main(self):
        if not self.preview_headers:
            messagebox.showwarning("提示", "请先预览结果。")
            return

        self.app.headers = self.preview_headers[:]
        self.app.rows = [row[:] for row in self.preview_rows]
        self.app.raw_data = ""
        self.app.refresh_tree()
        self.app.info_var.set(
            f"已从高级筛选载入预览结果：{len(self.app.rows)} 行 × {len(self.app.headers)} 列。"
        )

    def save_result_to_table(self):
        if not self.preview_headers:
            self.preview_result()

        if not self.preview_headers:
            return

        save_name = self.save_table_var.get().strip()
        if not save_name:
            messagebox.showwarning("提示", "请填写保存的新表名。")
            return

        try:
            table_name, row_count = self.app.save_rows_to_sqlite_table(
                table_name_raw=save_name,
                headers=self.preview_headers,
                rows=self.preview_rows,
                recreate=False
            )

            self.status_var.set(f"保存成功：{table_name}，{row_count} 行。")
            messagebox.showinfo(
                "保存成功",
                f"筛选结果已保存到新表。\n\n表名：{table_name}\n行数：{row_count}"
            )

            self.refresh_tables()

        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def export_template_data(self):
        return {
            "main_table": self.main_table_var.get(),
            "selected_tables": self.get_selected_tables(),
            "conditions": self.conditions,
            "logic": self.logic_var.get(),
            "join_logic": self.join_logic_var.get(),
            "join_rules": self.join_rules,
            "output_fields": self.output_fields,
            "result_limit": self.result_limit_var.get(),
            "max_intermediate": self.max_intermediate_var.get(),
            "save_table": self.save_table_var.get()
        }

    def apply_template_data(self, data):
        main_table = data.get("main_table", "")
        selected_tables = data.get("selected_tables", [])
        conditions = data.get("conditions", [])
        join_rules = data.get("join_rules", [])
        output_fields = data.get("output_fields", [])

        if main_table:
            self.main_table_var.set(main_table)

        self.selected_tables_listbox.delete(0, tk.END)
        for table in selected_tables:
            if table in self.tables_cache:
                self.selected_tables_listbox.insert(tk.END, table)

        if self.selected_tables_listbox.size() == 0 and main_table in self.tables_cache:
            self.selected_tables_listbox.insert(tk.END, main_table)

        self.refresh_fields()

        valid_fields = set(self.field_display_cache)

        self.conditions = [
            cond for cond in conditions
            if cond.get("field") in valid_fields
        ]

        self.join_rules = [
            rule for rule in join_rules
            if rule.get("left") in valid_fields and rule.get("right") in valid_fields
        ]

        self.output_fields = [
            field for field in output_fields
            if field in valid_fields
        ]

        self.logic_var.set(data.get("logic", "AND"))
        self.join_logic_var.set(data.get("join_logic", "AND"))
        self.result_limit_var.set(str(data.get("result_limit", "5000")))
        self.max_intermediate_var.set(str(data.get("max_intermediate", "200000")))
        self.save_table_var.set(str(data.get("save_table", self.save_table_var.get())))

        self.refresh_conditions_tree()
        self.refresh_join_tree()
        self.refresh_output_fields_listbox()

    def save_template(self):
        path = filedialog.asksaveasfilename(
            title="保存筛选模板",
            defaultextension=".json",
            filetypes=[
                ("JSON 文件", "*.json"),
                ("所有文件", "*.*")
            ]
        )
        if not path:
            return

        try:
            data = self.export_template_data()
            atomic_write_json(path, data)

            self.status_var.set(f"筛选模板已保存：{path}")

        except Exception as e:
            messagebox.showerror("保存模板失败", str(e))

    def load_template(self):
        path = filedialog.askopenfilename(
            title="载入筛选模板",
            filetypes=[
                ("JSON 文件", "*.json"),
                ("所有文件", "*.*")
            ]
        )
        if not path:
            return

        try:
            data = load_json_file_with_recovery(path, parent=self.window)

            self.apply_template_data(data)
            self.status_var.set(f"筛选模板已载入：{path}")

        except Exception as e:
            messagebox.showerror("载入模板失败", str(e))


class PlanWorkflowWindow:
    """
    计划 / 工作流处理窗口。

    设计目标：
    1. 把批量替换、数据提取、合并列、高级筛选、删除列、移动列作为节点串联。
    2. 每个节点都接收 headers / rows，输出新的 headers / rows。
    3. 支持预览到当前节点、预览完整计划、输出到主界面或保存到 SQLite。

    说明：
    - 计划内的“高级筛选”支持以上一步结果作为“当前表”，再选择数据库中的其他表进行多表匹配。
    """

    NODE_TYPES = ["获取文件列表", "节点组 / 子工作流", "循环执行起点", "跳转锚点节点", "无条件跳转节点", "条件判断节点", "条件跳转节点", "批量替换", "数据提取", "格式规范化 / 日期时间解析", "新建日期时间列", "新建列", "合并列", "批量更改列名", "去重 / 重复数据处理", "列数字运算", "匹配值输出列名", "复制列", "复制行", "删除行", "填充值", "序列填充", "区域填充", "行数据映射填充", "保存中转数据", "选定列写入指定表", "字段映射写入表", "高级筛选", "删除列", "移动列", "批量重命名", "循环判断回跳"]
    TABLE_ACCESS_POLICY_CHOICES = ["只审计", "预检确认", "强制拦截"]
    MAX_EXPANDED_ROWS = 200000
    MAX_TARGET_CELLS = 1000000
    TABLE_ACCESS_POLICY_DISPLAY = {
        "audit": "只审计",
        "prompt": "预检确认",
        "strict": "强制拦截",
        "off": "关闭",
    }
    STANDARD_WRITE_MODE_CHOICES = [
        "",
        "current_table_default",
        "create_new",
        "append",
        "overlay_by_order",
        "update_by_key",
        "upsert_by_key",
        "clear_keep_schema",
        "keep_schema_insert",
        "replace_table",
        "timestamp_new",
        "fail_if_exists",
        "write_fields_only",
        "fill_blank_fields",
    ]
    LOGIC_TYPES = ["AND", "OR"]
    FILTER_OPS = ["等于", "不等于", "包含", "不包含", "开头是", "结尾是", "大于", "小于", "大于等于", "小于等于", "为空", "不为空", "正则匹配"]
    FILTER_VALUE_SOURCES = ["固定值", "字段值"]
    REPLACE_MATCH_MODES = ["包含", "完全相等", "开头是", "结尾是", "正则匹配", "为空", "不为空"]
    REPLACE_MODES = ["局部替换匹配字符串", "整格替换为新值"]
    REPLACE_VALUE_SOURCES = ["手动输入", "列字段"]
    REPLACE_ROW_POLICIES = ["当前行", "第一行", "固定行号", "按匹配行号", "按命中序号"]
    EXTRACT_METHODS = [
        "正则提取", "固定位置提取", "从左取N位", "从右取N位", "按分隔符提取",
        "前后关键字之间提取", "指定字符前提取", "指定字符后提取", "删除前缀", "删除后缀"
    ]
    OUTPUT_MODES = ["生成新字段", "覆盖源字段"]
    UNMATCHED_MODES = ["留空", "保留原值", "填写固定值", "跳过该行"]
    FORMAT_PARSE_TYPES = ["日期", "时间", "日期时间"]
    FORMAT_INPUT_STRUCTURES = ["固定位置", "分隔符", "自动识别常见格式"]
    FORMAT_YEAR_RULES = ["20xx", "19xx", "自动窗口", "不补全"]
    FORMAT_DATE_ORDERS = ["年-月-日", "月-日-年", "日-月-年"]
    FORMAT_OUTPUT_MODES = ["生成新字段", "覆盖源字段", "生成多个字段"]
    CURRENT_DATETIME_OUTPUT_MODES = ["生成新字段", "覆盖已有字段"]
    CURRENT_DATETIME_TIME_MODES = ["整次运行固定同一时间", "逐行实时获取"]
    CURRENT_DATETIME_FORMAT_MODES = ["占位符模板", "Python strftime"]
    NEW_COLUMNS_CONFLICT_MODES = ["自动改名", "跳过已有字段", "覆盖已有字段", "存在则报错"]
    NEW_COLUMNS_VALUE_MODES = ["统一默认值", "按列配置值", "空值"]
    SEPARATOR_OPTIONS = ["空字符", "空格", "换行", "Windows换行", "制表符", "-", "_", "/", "\\", "|", ",", ";", ":", ".", "+", "自定义"]

    def __init__(self, app):
        self.app = app
        self.window = tk.Toplevel(app.root)
        self.window.title("计划 / 工作流处理")
        self.window.geometry("1680x950")
        self.window.minsize(1050, 650)
        self.window.transient(app.root)

        self.nodes = []
        self.preview_headers = list(app.headers)
        self.preview_rows = [list(row) for row in app.rows]
        self.current_config_widgets = {}
        self.separator_widgets = []
        self.field_listbox = None
        self.status_var = tk.StringVar(value="计划窗口已打开。先添加节点，再预览或执行完整计划。")
        self.output_mode_var = tk.StringVar(value="输出到主界面预览区")
        self.output_table_var = tk.StringVar(value=self.make_default_output_table_name())
        self.backup_before_overwrite_var = tk.BooleanVar(value=True)
        self.table_access_policy_var = tk.StringVar(value="只审计")
        self.node_type_var = tk.StringVar(value=self.NODE_TYPES[0])
        self.selected_node_index = None
        self.preview_edit_mode = False
        self.preview_edit_entry = None
        self.preview_edit_btn_text = tk.StringVar(value="修改模式:关")
        self.preview_dirty = False
        self.current_transit_tables = {}
        self.last_workflow_context = {}
        self.last_table_access_logs = []
        self.last_table_access_precheck = []
        # “当前预览结果”独立缓存：结果预览区临时载入 SQLite/中转/主界面表时，
        # 不应覆盖最后一次计划预览/执行得到的结果，否则下拉切换后会丢失原预览结果。
        self.plan_preview_headers = list(self.preview_headers)
        self.plan_preview_rows = [list(row) for row in self.preview_rows]
        self.preview_view_kind = "preview"
        # 结果预览区表格选择：用于快速查看当前预览、主界面表、SQLite表和中转副表。
        self.preview_table_var = tk.StringVar(value="当前预览结果")
        self.preview_table_map = {}
        self.preview_search_var = tk.StringVar(value="")
        self.preview_search_matches = []
        self.preview_search_index = -1

        # 循环单步调试缓存：在“循环判断回跳”节点点击“执行循环一次”时复用。
        # 用于逐次运行循环体，后续预览节点可接着这个 N 次循环后的上下文继续执行。
        self.manual_loop_context = None
        self.manual_loop_headers = None
        self.manual_loop_rows = None
        self.manual_loop_start_idx = None
        self.manual_loop_judge_idx = None
        self.manual_loop_after_index = None
        self.manual_loop_logs = []

        # 后台执行/进度条状态：主界面不直接跑耗时流程，后台线程负责执行，Queue 回传进度。
        # 第一版采用线程 worker，接口按“可迁移到子进程 worker”的消息协议设计。
        self.workflow_worker_thread = None
        self.workflow_worker_queue = queue.Queue()
        self.workflow_worker_cancel = None
        self.workflow_worker_running = False
        self.workflow_progress_var = tk.DoubleVar(value=0)
        self.node_progress_var = tk.DoubleVar(value=0)
        self.workflow_progress_text = tk.StringVar(value="总进度：空闲")
        self.node_progress_text = tk.StringVar(value="当前节点：空闲")
        self.worker_status_text = tk.StringVar(value="执行状态：空闲")
        self.workflow_current_task = None
        self.workflow_widget_state_backup = {}
        self.workflow_cancel_button = None

        # 外部插件节点：启动/打开计划窗口时扫描 plugins 目录并注册。
        self.plugin_registry = {}
        self.plugin_display_map = {}
        self.plugin_load_errors = []
        self.load_plugins(show_status=False)

        # 计划模板库：程序真实目录下的 plan 文件夹。
        # 只识别 template_type == "workflow_plan" 的新版模板。
        self.plan_dir = self.get_plan_dir()
        # 节点组模板库：程序真实目录下的 groups 文件夹。
        self.group_dir = self.get_group_dir()
        self.plan_template_var = tk.StringVar(value="")
        self.plan_template_map = {}

        self.build_ui()
        self.refresh_node_list()
        self.refresh_preview_tree(self.preview_headers, self.preview_rows)
        self.refresh_plan_template_list(show_status=False)

    def make_default_output_table_name(self):
        base = self.app.sanitize_sql_name(self.app.table_name_var.get(), "计划结果")
        return f"{base}_计划结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    def normalize_table_access_policy(self, value=None):
        if value is None:
            value = self.table_access_policy_var.get()
        return TableAccessManager.normalize_permission_policy(value)

    def table_access_policy_display(self, value=None):
        policy = self.normalize_table_access_policy(value)
        return self.TABLE_ACCESS_POLICY_DISPLAY.get(policy, "只审计")

    def set_table_access_policy(self, value):
        self.table_access_policy_var.set(self.table_access_policy_display(value))

    def normalize_table_access_write_mode(self, mode):
        return TableAccessManager.normalize_write_mode(mode)

    def write_mode_permission_set(self, mode, exists=False, read=False, partial=False):
        perms = {key: False for key, _ in self.table_access_permission_items()}
        for key in TableAccessManager.required_permissions_for_write_mode(mode, exists=exists, partial=partial):
            if key in perms:
                perms[key] = True
        if read:
            perms["read_table"] = True
        return perms

    def write_mode_display_text(self, mode):
        standard = self.normalize_table_access_write_mode(mode)
        labels = {
            "": "",
            "current_table_default": "当前表默认",
            "create_new": "新建表写入",
            "append": "追加行",
            "overlay_by_order": "按顺序覆盖",
            "update_by_key": "按键更新",
            "upsert_by_key": "匹配更新或追加",
            "clear_keep_schema": "清空保留结构写入",
            "keep_schema_insert": "保留结构写入",
            "replace_table": "替换整表",
            "timestamp_new": "自动时间戳新表",
            "fail_if_exists": "存在则报错",
            "write_fields_only": "指定字段写入",
            "fill_blank_fields": "字段空缺补齐",
        }
        return labels.get(standard, str(mode or ""))

    def build_ui(self):
        main = ttk.Frame(self.window, padding=8)
        main.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(main)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))

        source_frame = ttk.LabelFrame(left, text="1. 输入数据源", padding=8)
        source_frame.pack(fill=tk.X)
        ttk.Label(source_frame, text=f"当前输入：{len(self.app.rows)} 行 × {len(self.app.headers)} 列").pack(anchor=tk.W)
        ttk.Button(source_frame, text="重新读取主界面当前预览", command=self.reload_from_app_preview).pack(fill=tk.X, pady=(6, 0))

        node_frame = ttk.LabelFrame(left, text="2. 工作流节点", padding=8)
        node_frame.pack(fill=tk.BOTH, expand=True, pady=8)

        add_frame = ttk.Frame(node_frame)
        add_frame.pack(fill=tk.X)
        self.node_type_combo = ttk.Combobox(add_frame, textvariable=self.node_type_var, values=self.get_node_type_values(), width=22, state="readonly")
        self.node_type_combo.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(add_frame, text="添加节点", command=self.add_node).pack(side=tk.LEFT)
        ttk.Button(add_frame, text="刷新插件", command=self.refresh_plugins).pack(side=tk.LEFT, padx=(4, 0))

        node_list_wrap = ttk.Frame(node_frame)
        node_list_wrap.pack(fill=tk.BOTH, expand=True, pady=6)
        self.node_listbox = tk.Listbox(node_list_wrap, width=42, height=24, exportselection=False, selectmode=tk.EXTENDED)
        node_list_scroll = ttk.Scrollbar(node_list_wrap, orient=tk.VERTICAL, command=self.node_listbox.yview)
        self.node_listbox.configure(yscrollcommand=node_list_scroll.set)
        self.node_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        node_list_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.node_listbox.bind("<<ListboxSelect>>", self.on_node_select)

        node_btns1 = ttk.Frame(node_frame)
        node_btns1.pack(fill=tk.X)
        for text_, cmd in [
            ("删除", self.delete_node),
            ("上移", self.move_node_up),
            ("下移", self.move_node_down),
            ("启用/禁用", self.toggle_node_enabled),
        ]:
            ttk.Button(node_btns1, text=text_, command=cmd).pack(side=tk.LEFT, padx=2, pady=2)

        node_btns2 = ttk.Frame(node_frame)
        node_btns2.pack(fill=tk.X)
        ttk.Button(node_btns2, text="复制节点", command=self.copy_node).pack(side=tk.LEFT, padx=2, pady=2)
        ttk.Button(node_btns2, text="合并为组", command=self.merge_selected_nodes_to_group).pack(side=tk.LEFT, padx=2, pady=2)
        ttk.Button(node_btns2, text="展开组", command=self.expand_selected_group).pack(side=tk.LEFT, padx=2, pady=2)
        ttk.Button(node_btns2, text="清空节点", command=self.clear_nodes).pack(side=tk.LEFT, padx=2, pady=2)

        node_btns3 = ttk.Frame(node_frame)
        node_btns3.pack(fill=tk.X)
        ttk.Button(node_btns3, text="字段权限层", command=self.open_table_access_window).pack(side=tk.LEFT, padx=2, pady=2)
        ttk.Button(node_btns3, text="权限预检", command=self.open_table_access_precheck_window).pack(side=tk.LEFT, padx=2, pady=2)
        ttk.Button(node_btns3, text="审计日志", command=self.open_table_access_audit_window).pack(side=tk.LEFT, padx=2, pady=2)
        ttk.Button(node_btns3, text="跳转管理", command=self.open_jump_manager_window).pack(side=tk.LEFT, padx=2, pady=2)

        policy_frame = ttk.Frame(node_frame)
        policy_frame.pack(fill=tk.X)
        ttk.Label(policy_frame, text="权限策略：").pack(side=tk.LEFT, padx=(2, 2), pady=2)
        ttk.Combobox(
            policy_frame,
            textvariable=self.table_access_policy_var,
            values=self.TABLE_ACCESS_POLICY_CHOICES,
            width=10,
            state="readonly",
        ).pack(side=tk.LEFT, padx=2, pady=2)

        tpl_frame = ttk.LabelFrame(left, text="3. 计划模板", padding=8)
        tpl_frame.pack(fill=tk.X)

        tpl_row1 = ttk.Frame(tpl_frame)
        tpl_row1.pack(fill=tk.X, pady=(0, 4))
        ttk.Button(tpl_row1, text="保存计划模板", command=self.save_plan_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(tpl_row1, text="载入计划模板", command=self.load_plan_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(tpl_row1, text="打开plan目录", command=self.open_plan_dir).pack(side=tk.LEFT, padx=2)

        tpl_row2 = ttk.Frame(tpl_frame)
        tpl_row2.pack(fill=tk.X)
        self.plan_template_combo = ttk.Combobox(
            tpl_row2,
            textvariable=self.plan_template_var,
            width=27,
            state="readonly"
        )
        self.plan_template_combo.pack(side=tk.LEFT, padx=2)
        self.plan_template_combo.configure(postcommand=lambda: self.refresh_plan_template_list(show_status=False))
        ttk.Button(tpl_row2, text="载入选中模板", command=self.load_selected_plan_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(tpl_row2, text="刷新模板", command=self.refresh_plan_template_list).pack(side=tk.LEFT, padx=2)

        right = ttk.Frame(main)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 节点配置区内容较多，使用一个固定高度的可滚动区域，避免控件被窗口截断。
        self.config_outer = ttk.LabelFrame(right, text="4. 节点配置", padding=8)
        self.config_outer.pack(fill=tk.X)
        self.config_outer.configure(height=310)
        self.config_outer.pack_propagate(False)

        self.config_canvas = tk.Canvas(self.config_outer, highlightthickness=0)
        self.config_y_scroll = ttk.Scrollbar(self.config_outer, orient=tk.VERTICAL, command=self.config_canvas.yview)
        self.config_x_scroll = ttk.Scrollbar(self.config_outer, orient=tk.HORIZONTAL, command=self.config_canvas.xview)
        self.config_frame = ttk.Frame(self.config_canvas)

        self.config_canvas_window = self.config_canvas.create_window((0, 0), window=self.config_frame, anchor="nw")
        self.config_canvas.configure(
            yscrollcommand=self.config_y_scroll.set,
            xscrollcommand=self.config_x_scroll.set
        )

        self.config_canvas.grid(row=0, column=0, sticky="nsew")
        self.config_y_scroll.grid(row=0, column=1, sticky="ns")
        self.config_x_scroll.grid(row=1, column=0, sticky="ew")
        self.config_outer.rowconfigure(0, weight=1)
        self.config_outer.columnconfigure(0, weight=1)

        self.config_frame.bind("<Configure>", self._on_config_frame_configure)
        self.config_canvas.bind("<Configure>", self._on_config_canvas_configure)
        self.config_canvas.bind("<Enter>", self._bind_config_mousewheel)
        self.config_canvas.bind("<Leave>", self._unbind_config_mousewheel)

        action_frame = ttk.Frame(right)
        action_frame.pack(fill=tk.X, pady=8)
        ttk.Button(action_frame, text="预览到当前节点", command=self.preview_to_selected_node).pack(side=tk.LEFT, padx=4)
        ttk.Button(action_frame, text="预览完整计划", command=self.preview_full_plan).pack(side=tk.LEFT, padx=4)
        ttk.Button(action_frame, text="执行计划", command=self.execute_plan).pack(side=tk.LEFT, padx=4)

        progress_frame = ttk.LabelFrame(right, text="执行进度", padding=8)
        progress_frame.pack(fill=tk.X, pady=(0, 8))
        self.workflow_progress_label = ttk.Label(progress_frame, textvariable=self.workflow_progress_text, anchor=tk.W)
        self.workflow_progress_label.grid(row=0, column=0, sticky="ew", padx=4, pady=(2, 0))
        self.workflow_progress_bar = ttk.Progressbar(progress_frame, variable=self.workflow_progress_var, maximum=100, mode="determinate")
        self.workflow_progress_bar.grid(row=1, column=0, sticky="ew", padx=4, pady=(2, 6))
        self.node_progress_label = ttk.Label(progress_frame, textvariable=self.node_progress_text, anchor=tk.W)
        self.node_progress_label.grid(row=2, column=0, sticky="ew", padx=4, pady=(2, 0))
        self.node_progress_bar = ttk.Progressbar(progress_frame, variable=self.node_progress_var, maximum=100, mode="determinate")
        self.node_progress_bar.grid(row=3, column=0, sticky="ew", padx=4, pady=(2, 6))
        worker_btns = ttk.Frame(progress_frame)
        worker_btns.grid(row=2, column=1, sticky=tk.E, padx=(8, 4), pady=0)
        self.workflow_cancel_button = ttk.Button(worker_btns, text="取消后台任务", command=self.cancel_background_workflow)
        self.workflow_cancel_button.pack(side=tk.LEFT, padx=2)
        self.workflow_cancel_button.configure(state="disabled")
        self.worker_status_label = ttk.Label(progress_frame, textvariable=self.worker_status_text, anchor=tk.W, wraplength=980, justify=tk.LEFT)
        self.worker_status_label.grid(row=4, column=0, columnspan=2, sticky="ew", padx=4, pady=(4, 0))
        progress_frame.columnconfigure(0, weight=1)
        def update_progress_wrap(event, label=self.worker_status_label):
            try:
                label.configure(wraplength=max(320, int(event.width) - 32))
            except Exception:
                pass
        progress_frame.bind("<Configure>", update_progress_wrap)

        output_frame = ttk.LabelFrame(right, text="5. 输出设置", padding=8)
        output_frame.pack(fill=tk.X)
        ttk.Label(output_frame, text="输出方式：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(
            output_frame,
            textvariable=self.output_mode_var,
            values=["输出到主界面预览区", "保存为SQLite新表", "覆盖当前表", "导出为xlsx"],
            width=20,
            state="readonly"
        ).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Label(output_frame, text="输出表名：").grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(output_frame, textvariable=self.output_table_var, width=36).grid(row=0, column=3, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(output_frame, text="覆盖前自动备份旧表", variable=self.backup_before_overwrite_var).grid(row=0, column=4, sticky=tk.W, padx=4, pady=4)

        preview_frame = ttk.LabelFrame(right, text="6. 结果预览", padding=6)
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=8)

        preview_toolbar = ttk.Frame(preview_frame)
        preview_toolbar.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 4))
        ttk.Button(
            preview_toolbar,
            textvariable=self.preview_edit_btn_text,
            command=self.toggle_preview_edit_mode
        ).pack(side=tk.LEFT, padx=4)

        ttk.Label(preview_toolbar, text="查看表：").pack(side=tk.LEFT, padx=(10, 2))
        self.preview_table_combo = ttk.Combobox(
            preview_toolbar,
            textvariable=self.preview_table_var,
            width=34,
            state="readonly"
        )
        self.preview_table_combo.pack(side=tk.LEFT, padx=2)
        self.preview_table_combo.configure(postcommand=lambda: self.refresh_preview_table_choices(show_status=False))
        ttk.Button(
            preview_toolbar,
            text="载入选中表",
            command=self.load_selected_preview_table
        ).pack(side=tk.LEFT, padx=(4, 8))

        ttk.Label(
            preview_toolbar,
            text="开启后可双击下方预览单元格修改；再次预览/重新执行计划会重新生成预览。",
            foreground="gray"
        ).pack(side=tk.LEFT, padx=6)

        preview_search_frame = ttk.Frame(preview_frame)
        preview_search_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 4))
        ttk.Label(preview_search_frame, text="搜索：").pack(side=tk.LEFT, padx=(4, 4))
        preview_search_entry = ttk.Entry(preview_search_frame, textvariable=self.preview_search_var, width=38)
        preview_search_entry.pack(side=tk.LEFT, padx=(4, 4))
        preview_search_entry.bind("<Return>", lambda e: self.search_preview_table(reset=True))
        ttk.Button(
            preview_search_frame,
            text="搜索",
            command=lambda: self.search_preview_table(reset=True)
        ).pack(side=tk.LEFT, padx=(12, 8))
        ttk.Button(
            preview_search_frame,
            text="上一个",
            command=self.search_preview_prev
        ).pack(side=tk.LEFT, padx=(12, 8))
        ttk.Button(
            preview_search_frame,
            text="下一个",
            command=self.search_preview_next
        ).pack(side=tk.LEFT, padx=(12, 8))
        ttk.Button(
            preview_search_frame,
            text="导出为 xlsx",
            command=self.export_preview_to_xlsx
        ).pack(side=tk.LEFT, padx=(4, 8))

        self.preview_tree = ttk.Treeview(preview_frame, show="headings")
        y_scroll = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_tree.yview)
        x_scroll = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview_tree.grid(row=2, column=0, sticky="nsew")
        y_scroll.grid(row=2, column=1, sticky="ns")
        x_scroll.grid(row=3, column=0, sticky="ew")
        self.preview_tree.bind("<Double-1>", self.on_preview_tree_double_click)
        preview_frame.rowconfigure(2, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        ttk.Label(right, textvariable=self.status_var, padding=(0, 4)).pack(fill=tk.X)
        self.show_empty_config()

    def _on_config_frame_configure(self, event=None):
        """更新节点配置区滚动范围。"""
        if hasattr(self, "config_canvas"):
            self.config_canvas.configure(scrollregion=self.config_canvas.bbox("all"))

    def _on_config_canvas_configure(self, event=None):
        """让内部配置区域宽度跟随 Canvas，减少横向截断。"""
        if hasattr(self, "config_canvas") and hasattr(self, "config_canvas_window"):
            try:
                self.config_canvas.itemconfigure(self.config_canvas_window, width=event.width)
            except Exception:
                pass

    def _bind_config_mousewheel(self, event=None):
        if hasattr(self, "config_canvas"):
            self.config_canvas.bind_all("<MouseWheel>", self._on_config_mousewheel)
            self.config_canvas.bind_all("<Shift-MouseWheel>", self._on_config_shift_mousewheel)

    def _unbind_config_mousewheel(self, event=None):
        if hasattr(self, "config_canvas"):
            self.config_canvas.unbind_all("<MouseWheel>")
            self.config_canvas.unbind_all("<Shift-MouseWheel>")

    def _on_config_mousewheel(self, event):
        if hasattr(self, "config_canvas"):
            self.config_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_config_shift_mousewheel(self, event):
        if hasattr(self, "config_canvas"):
            self.config_canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

    def set_plan_preview_result(self, headers, rows, display=True, source_label="当前预览结果"):
        """保存最后一次计划预览/执行结果。

        注意：self.preview_headers/self.preview_rows 表示“当前结果预览区正在显示的表”，
        可能是 SQLite 表、中转副表或主界面预览表；而 self.plan_preview_headers/rows
        专门保存“当前预览结果”，避免用户临时查看其他表后丢失计划预览结果。
        """
        self.plan_preview_headers = list(headers or [])
        self.plan_preview_rows = [list(row) for row in (rows or [])]
        if display:
            self.preview_view_kind = "preview"
            self.preview_table_var.set("当前预览结果")
            self.preview_headers = list(self.plan_preview_headers)
            self.preview_rows = [list(row) for row in self.plan_preview_rows]
            self.refresh_preview_tree(self.preview_headers, self.preview_rows)

    def get_plan_preview_result(self):
        headers = list(getattr(self, "plan_preview_headers", []))
        rows = [list(row) for row in getattr(self, "plan_preview_rows", [])]
        return headers, rows

    def reload_from_app_preview(self):
        headers = list(self.app.headers)
        rows = [list(row) for row in self.app.rows]
        self.set_plan_preview_result(headers, rows, display=True, source_label="主界面当前预览")
        self.status_var.set(f"已重新读取主界面当前预览，并保存为当前预览结果：{len(rows)} 行 × {len(headers)} 列。")
        self.rebuild_current_config()

    def show_empty_config(self):
        self.clear_config_frame()
        ttk.Label(self.config_frame, text="请先添加并选择一个节点。每个节点会接收上一步结果，并输出给下一步。", foreground="gray").pack(anchor=tk.W)

    def clear_config_frame(self):
        for child in self.config_frame.winfo_children():
            child.destroy()
        self.current_config_widgets = {}
        self.separator_widgets = []
        self.field_listbox = None
        if hasattr(self, "config_canvas"):
            self.config_canvas.yview_moveto(0)
            self.config_canvas.xview_moveto(0)
            self.config_canvas.after_idle(lambda: self.config_canvas.configure(scrollregion=self.config_canvas.bbox("all")))

    def get_selected_node_index(self):
        sel = self.node_listbox.curselection()
        if not sel:
            return None
        return sel[0]

    def on_node_select(self, event=None):
        idx = self.get_selected_node_index()
        self.selected_node_index = idx
        self.rebuild_current_config()

    def rebuild_current_config(self):
        idx = self.get_selected_node_index()
        if idx is None or idx < 0 or idx >= len(self.nodes):
            self.show_empty_config()
            return
        self.build_node_config(idx)

    def refresh_node_list(self, select_index=None, reveal=True):
        self.ensure_node_tree_identity(self.nodes)
        selected = self.get_selected_node_index() if select_index is None else select_index
        self.node_listbox.delete(0, tk.END)
        for idx, node in enumerate(self.nodes, start=1):
            mark = "√" if node.get("enabled", True) else "×"
            self.node_listbox.insert(tk.END, f"[{mark}] {idx}. {node.get('type')}：{node.get('name', '')}")
        if selected is not None and self.nodes:
            selected = min(selected, len(self.nodes) - 1)
            self.selected_node_index = selected
            self.node_listbox.selection_clear(0, tk.END)
            self.node_listbox.selection_set(selected)
            self.node_listbox.activate(selected)
            if reveal:
                self.node_listbox.see(selected)
        elif not self.nodes:
            self.selected_node_index = None


    # ------------------------------------------------------------------
    # 外部 Python 插件节点
    # ------------------------------------------------------------------
    def get_plugins_dir(self):
        path = os.path.join(getattr(self.app, "app_dir", get_app_dir()), "plugins")
        os.makedirs(path, exist_ok=True)
        return path

    def get_plugin_data_dir(self, plugin_id=None):
        base = os.path.join(getattr(self.app, "app_dir", get_app_dir()), "plugin_data")
        if plugin_id:
            base = os.path.join(base, self.app.sanitize_sql_name(plugin_id, "plugin"))
        os.makedirs(base, exist_ok=True)
        return base

    def get_plugin_log_dir(self):
        path = os.path.join(getattr(self.app, "app_dir", get_app_dir()), "logs", "plugins")
        os.makedirs(path, exist_ok=True)
        return path

    def get_plugin_env_dir(self, plugin_id=None):
        base = os.path.join(getattr(self.app, "app_dir", get_app_dir()), "plugin_envs")
        if plugin_id:
            base = os.path.join(base, self.app.sanitize_sql_name(plugin_id, "plugin"))
        os.makedirs(base, exist_ok=True)
        return base

    def get_node_type_values(self):
        values = list(self.NODE_TYPES)
        values.extend(sorted(getattr(self, "plugin_display_map", {}).keys()))
        return values

    def refresh_plugins(self):
        self.load_plugins(show_status=True)
        if hasattr(self, "node_type_combo"):
            self.node_type_combo["values"] = self.get_node_type_values()
        if self.node_type_var.get() not in self.get_node_type_values():
            self.node_type_var.set(self.NODE_TYPES[0])
        self.rebuild_current_config()

    def load_plugins(self, show_status=False):
        """扫描 plugins 目录并注册插件。"""
        self.plugin_registry = {}
        self.plugin_display_map = {}
        self.plugin_load_errors = []
        plugins_dir = self.get_plugins_dir()
        registry, errors = scan_plugins(plugins_dir)
        self.plugin_registry = registry
        self.plugin_load_errors = errors

        used_names = {}
        external_only_count = 0
        for plugin_id, item in sorted(self.plugin_registry.items(), key=lambda kv: kv[1]["info"].get("name", kv[0])):
            name = str(item["info"].get("name", plugin_id)).strip() or plugin_id
            suffix = ""
            if item.get("load_status") == "仅独立环境运行":
                external_only_count += 1
                suffix = " [仅独立]"
            display = f"插件 / {name}{suffix}"
            if display in used_names:
                used_names[display] += 1
                display = f"插件 / {name} ({plugin_id}){suffix}"
            else:
                used_names[display] = 1
            self.plugin_display_map[display] = plugin_id

        if show_status:
            msg = f"插件刷新完成：已注册 {len(self.plugin_registry)} 个插件"
            if external_only_count:
                msg += f"，其中仅独立环境 {external_only_count} 个"
            if self.plugin_load_errors:
                msg += f"，加载失败 {len(self.plugin_load_errors)} 个"
                first = self.plugin_load_errors[0]
                msg += f"；示例：{first.get('file')} - {first.get('error')}"
            self.status_var.set(msg)


    def default_config_for_plugin(self, plugin_id):
        item = self.plugin_registry.get(plugin_id, {})
        schema = item.get("schema", [])
        params = {}
        for field in schema:
            if not isinstance(field, dict):
                continue
            name = field.get("name")
            if not name:
                continue
            default = field.get("default", "")
            if field.get("type") == "multi_field_select" and default == "":
                default = []
            params[name] = default
        info = item.get("info", {})
        default_run_mode = info.get("run_mode") or item.get("run_mode_default") or "主程序内置环境"
        if default_run_mode in ("external_python", "独立环境", "插件独立环境"):
            default_run_mode = "插件独立环境"
        else:
            default_run_mode = "主程序内置环境"
        return {
            "plugin_id": plugin_id,
            "params": params,
            "input_tables": [],
            "run_mode": default_run_mode,
            "external_python": "",
            "external_env_dir": self.get_plugin_env_dir(plugin_id),
            "external_entry": item.get("external_entry", item.get("path", "")),
            "external_timeout": "0",
            "output_mode": "使用插件返回结果",
            "save_output_as_transit": False,
            "transit_name": item.get("info", {}).get("name", plugin_id),
            "transit_conflict_mode": "覆盖",
            "save_plugin_log_file": True,
            "save_plugin_log_sqlite": False,
            "save_plugin_log_transit": False,
            "plugin_log_transit_name": f"{item.get('info', {}).get('name', plugin_id)}_日志",
            "plugin_log_in_preview": False,
            "plugin_failure_policy": "停止工作流",
        }

    def get_sqlite_table_names(self):
        db_path = self.app.db_path_var.get().strip()
        if not db_path or not os.path.exists(db_path):
            return []
        try:
            return TableAccessManager(db_path).list_tables()
        except Exception:
            return []

    def get_workflow_snapshot(self, context=None):
        """返回后台任务快照。后台线程优先使用快照，避免直接读取 Tkinter 变量。"""
        if isinstance(context, dict):
            snapshot = context.get("workflow_snapshot") or {}
            if isinstance(snapshot, dict):
                return snapshot
        return {}

    def get_workflow_db_path(self, context=None):
        """执行期统一获取 SQLite 路径：优先读 workflow_snapshot，兜底读主线程 UI 变量。"""
        snapshot = self.get_workflow_snapshot(context)
        db_path = str(snapshot.get("db_path") or "").strip()
        if db_path:
            return db_path
        try:
            return self.app.db_path_var.get().strip()
        except Exception:
            return ""

    def make_node_id(self):
        return "node_" + uuid.uuid4().hex[:12]

    def table_permission_set(self, read=False, write=False, create=False, append=False, update=False,
                             clear=False, replace=False, alter=False, delete=False, drop=False):
        return {
            "read_table": bool(read),
            "write_table": bool(write),
            "create_table": bool(create),
            "append_rows": bool(append),
            "update_rows": bool(update),
            "clear_table": bool(clear),
            "replace_table": bool(replace),
            "alter_schema": bool(alter),
            "delete_rows": bool(delete),
            "drop_table": bool(drop),
        }

    def make_table_access_entry(self, role, table, source_type="SQLite表", is_current_table=False,
                                permissions=None, write_mode="", field_mapping=None, log_only=False,
                                table_pattern="", pattern_type="glob", declared_by=""):
        return {
            "role": role,
            "table": table,
            "table_pattern": str(table_pattern or "").strip(),
            "pattern_type": str(pattern_type or "glob").strip(),
            "declared_by": str(declared_by or "").strip(),
            "source_type": source_type,
            "is_current_table": bool(is_current_table),
            "permissions": permissions or self.table_permission_set(read=True),
            "write_mode": self.normalize_table_access_write_mode(write_mode),
            "field_mapping_mode": "by_name",
            "field_mapping": field_mapping or {},
            "log_only": bool(log_only),
        }

    def get_plugin_table_access_specs(self, config):
        config = config or {}
        plugin_id = str(config.get("plugin_id", "") or "").strip()
        item = self.plugin_registry.get(plugin_id, {}) if hasattr(self, "plugin_registry") else {}
        module = item.get("module")
        params = dict(config.get("params", {}) or {})
        specs = None
        provider = getattr(module, "get_table_access_spec", None) if module is not None else None
        if callable(provider):
            try:
                specs = provider(params, {"plugin_id": plugin_id, "config_probe": True})
            except TypeError:
                specs = provider(params)
            except Exception:
                specs = None
        if specs is None:
            info = item.get("info", {}) or {}
            specs = info.get("table_access") or info.get("table_access_spec") or []
        if isinstance(specs, dict):
            specs = specs.get("tables") or [specs]
        return [spec for spec in (specs or []) if isinstance(spec, dict)]

    def plugin_has_table_access_declaration(self, config):
        config = config or {}
        plugin_id = str(config.get("plugin_id", "") or "").strip()
        item = self.plugin_registry.get(plugin_id, {}) if hasattr(self, "plugin_registry") else {}
        module = item.get("module")
        if callable(getattr(module, "get_table_access_spec", None)):
            return True
        info = item.get("info", {}) or {}
        return bool(info.get("table_access") or info.get("table_access_spec"))

    def plugin_needs_table_access_declaration(self, config):
        config = config or {}
        plugin_id = str(config.get("plugin_id", "") or "").strip()
        item = self.plugin_registry.get(plugin_id, {}) if hasattr(self, "plugin_registry") else {}
        info = item.get("info", {}) or {}
        danger = str(info.get("danger_level", "") or "").strip().lower()
        return danger in {"db_write", "database_write"} or bool(info.get("database_requests"))

    def make_plugin_declared_access_entry(self, plugin_id, spec):
        spec = spec or {}
        permissions = {key: False for key, _ in self.table_access_permission_items()}
        permissions.update({
            key: bool(value)
            for key, value in (spec.get("permissions") or {}).items()
            if key in permissions
        })
        return self.make_table_access_entry(
            spec.get("role") or "plugin_declared",
            spec.get("table") or "",
            source_type=spec.get("source_type") or "SQLite表",
            is_current_table=bool(spec.get("is_current_table")),
            permissions=permissions,
            write_mode=spec.get("write_mode") or "",
            field_mapping=copy.deepcopy(spec.get("field_mapping") or {}),
            log_only=bool(spec.get("log_only")),
            table_pattern=spec.get("table_pattern") or "",
            pattern_type=spec.get("pattern_type") or "glob",
            declared_by=plugin_id,
        )

    def default_table_access_for_node(self, node):
        return workflow_build_default_table_access_for_node(
            node,
            self.make_table_access_entry,
            self.table_permission_set,
            normalize_selected_columns_write_mode=getattr(self, "normalize_selected_columns_write_mode", None),
            normalize_group_transit_conflict_mode=getattr(self, "normalize_group_transit_conflict_mode", None),
            get_plugin_table_access_specs=self.get_plugin_table_access_specs,
            make_plugin_declared_access_entry=self.make_plugin_declared_access_entry,
        )

    def ensure_node_identity(self, node, force_new=False):
        if not isinstance(node, dict):
            return node
        if force_new or not str(node.get("node_id", "")).strip():
            node["node_id"] = self.make_node_id()
        if not isinstance(node.get("table_access"), dict):
            node["table_access"] = self.default_table_access_for_node(node)
        return node

    def ensure_node_tree_identity(self, nodes, force_new=False):
        for node in nodes or []:
            self.ensure_node_identity(node, force_new=force_new)
            cfg = node.get("config", {}) if isinstance(node, dict) else {}
            child_nodes = cfg.get("nodes") if isinstance(cfg, dict) else None
            if isinstance(child_nodes, list):
                self.ensure_node_tree_identity(child_nodes, force_new=force_new)

    def refresh_node_table_access(self, node):
        if isinstance(node, dict) and (
            not isinstance(node.get("table_access"), dict)
            or bool(node.get("table_access", {}).get("auto_generated", True))
        ):
            node["table_access"] = self.default_table_access_for_node(node)
        return node

    def refresh_node_tree_table_access(self, nodes):
        for node in nodes or []:
            self.ensure_node_identity(node)
            self.refresh_node_table_access(node)
            cfg = node.get("config", {}) if isinstance(node, dict) else {}
            child_nodes = cfg.get("nodes") if isinstance(cfg, dict) else None
            if isinstance(child_nodes, list):
                self.refresh_node_tree_table_access(child_nodes)

    def table_access_permission_items(self):
        return [
            ("read_table", "读表"),
            ("write_table", "写表"),
            ("create_table", "新建表"),
            ("append_rows", "追加行"),
            ("update_rows", "更新行"),
            ("clear_table", "清空表"),
            ("replace_table", "替换表"),
            ("alter_schema", "改结构"),
            ("delete_rows", "删行"),
            ("drop_table", "删表"),
        ]

    def field_permission_items(self):
        return [
            ("read_field", "可读"),
            ("write_field", "可写"),
            ("create_field", "可创建"),
            ("protect_field", "保护"),
        ]

    def get_node_table_access(self, node):
        self.ensure_node_identity(node)
        access = node.get("table_access")
        if not isinstance(access, dict):
            access = self.default_table_access_for_node(node)
            node["table_access"] = access
        access.setdefault("version", 1)
        tables = access.get("tables")
        if not isinstance(tables, list):
            access["tables"] = []
        return access

    def mark_node_table_access_manual(self, node):
        access = self.get_node_table_access(node)
        access["auto_generated"] = False
        return access

    def table_access_table_choices(self, node=None):
        values = ["__CURRENT_TABLE__"]
        try:
            values.extend(self.app.get_table_names())
        except Exception:
            pass
        if isinstance(node, dict):
            for entry in self.get_node_table_access(node).get("tables", []):
                table = str((entry or {}).get("table", "") or "").strip()
                if table:
                    values.append(table)
        result = []
        for value in values:
            if value not in result:
                result.append(value)
        return result

    def table_permission_summary(self, entry):
        perms = (entry or {}).get("permissions") or {}
        labels = []
        label_map = dict(self.table_access_permission_items())
        for key, _ in self.table_access_permission_items():
            if perms.get(key):
                labels.append(label_map.get(key, key))
        if not labels:
            return "无权限"
        return "/".join(labels[:4]) + ("..." if len(labels) > 4 else "")

    def table_access_entry_table_label(self, entry):
        return workflow_table_access_entry_table_label(entry)

    def table_access_operation_summary(self, entry):
        return workflow_table_access_operation_summary(
            entry,
            write_mode_formatter=self.write_mode_display_text,
        )

    def table_access_entry_status(self, entry):
        return workflow_table_access_entry_status(entry)

    def table_access_node_status(self, node):
        access = self.get_node_table_access(node)
        tables = access.get("tables", [])
        if not tables:
            return "未配置"
        statuses = [self.table_access_entry_status(entry) for entry in tables]
        if any(s in ("未绑定", "未授权") for s in statuses):
            return "待配置"
        if any(s == "危险写入" for s in statuses):
            return "需确认"
        if any(s == "已授权" for s in statuses):
            return "已授权"
        if all(s in ("只读", "只记录", "当前表") for s in statuses):
            return "只读/记录"
        return "OK"

    def table_access_field_items(self, entry):
        return workflow_table_access_field_items(entry)

    def find_table_access_field_rule(self, entry, target="", source="", field_index=None):
        return workflow_find_table_access_field_rule(entry, target=target, source=source, field_index=field_index)

    def make_table_access_field_key(self, mapping, source_field, target_field):
        return workflow_make_table_access_field_key(mapping, source_field, target_field)

    def field_permission_status(self, item):
        item = item or {}
        perms = item.get("permissions") or {}
        if perms.get("protect_field"):
            return "保护"
        if perms.get("write_field"):
            return "可写"
        if perms.get("read_field"):
            return "只读"
        return "未授权"

    def field_bool_text(self, value):
        return "是" if bool(value) else "否"

    def get_table_access_field_choices(self, node_index, entry):
        entry = entry or {}
        table = str(entry.get("table", "") or "").strip()
        choices = []
        try:
            headers, _ = self.get_headers_rows_before(node_index)
            choices.extend(headers or [])
        except Exception:
            choices.extend(self.preview_headers or [])
        if table and table != "__CURRENT_TABLE__" and entry.get("source_type", "SQLite表") == "SQLite表":
            try:
                choices.extend(self.app.get_table_columns(table))
            except Exception:
                pass
        for _, item in self.table_access_field_items(entry):
            for key in ("source_field", "target_field", "field", "name"):
                value = str(item.get(key, "") or "").strip()
                if value:
                    choices.append(value)
        result = []
        for value in choices:
            if value and value not in result:
                result.append(value)
        return result

    def auto_match_table_access_fields(self, node_index, entry):
        entry = entry or {}
        source_fields = []
        try:
            source_fields, _ = self.get_headers_rows_before(node_index)
        except Exception:
            source_fields = list(self.preview_headers or [])

        table = str(entry.get("table", "") or "").strip()
        target_fields = []
        if table and table != "__CURRENT_TABLE__" and entry.get("source_type", "SQLite表") == "SQLite表":
            try:
                target_fields = self.app.get_table_columns(table)
            except Exception:
                target_fields = []
        return workflow_apply_auto_field_mapping_by_name(
            entry,
            source_fields,
            target_fields,
            lambda value: self.app.sanitize_sql_name(value, ""),
            make_key=self.make_table_access_field_key,
        )

    def auto_match_table_access_fields_by_order(self, node_index, entry):
        entry = entry or {}
        try:
            source_fields, _ = self.get_headers_rows_before(node_index)
        except Exception:
            source_fields = list(self.preview_headers or [])

        table = str(entry.get("table", "") or "").strip()
        target_fields = []
        if table and table != "__CURRENT_TABLE__" and entry.get("source_type", "SQLite表") == "SQLite表":
            try:
                target_fields = self.app.get_table_columns(table)
            except Exception:
                target_fields = []
        return workflow_apply_auto_field_mapping_by_order(entry, source_fields, target_fields)

    def apply_table_access_preset_to_vars(self, preset, permission_vars, log_only_var=None):
        config = workflow_table_access_preset_config(
            preset,
            [key for key, _ in self.table_access_permission_items()],
        )
        if config is None:
            return
        for key, var in permission_vars.items():
            var.set(bool(config["permissions"].get(key)))
        if log_only_var is not None:
            log_only_var.set(bool(config["log_only"]))

    def table_access_entry_match_score(self, actual, expected):
        return workflow_table_access_entry_match_score(actual, expected)

    def find_matching_table_access_entry(self, actual_tables, expected):
        return workflow_find_matching_table_access_entry(actual_tables, expected)

    def normalize_precheck_transit_name(self, table_name):
        return workflow_normalize_precheck_transit_name(table_name)

    def add_table_access_precheck_issue(self, issues, severity, node_label, node, entry, message,
                                        suggestion="", category="permission", blocking=None):
        issues.append(workflow_make_table_access_precheck_issue(
            severity,
            node_label,
            node,
            entry,
            message,
            suggestion=suggestion,
            category=category,
            blocking=blocking,
            write_mode_formatter=self.write_mode_display_text,
        ))

    def table_access_precheck_sort_key(self, issue):
        return workflow_table_access_precheck_sort_key(issue)

    def table_access_precheck_summary_text(self, issues):
        return workflow_table_access_precheck_summary_text(issues)

    def table_access_precheck_actionable(self, issues):
        return workflow_table_access_precheck_actionable(issues)

    def table_access_precheck_blocking(self, issues):
        return workflow_table_access_precheck_blocking(issues)

    def get_precheck_sqlite_tables(self):
        db_path = self.get_workflow_db_path()
        if not db_path or not os.path.exists(db_path):
            return None
        try:
            return set(TableAccessManager(db_path, node_type="权限预检").list_tables())
        except Exception:
            return None

    def iter_nodes_for_table_access_precheck(self, nodes=None, stop_index=None, prefix=""):
        node_list = nodes if nodes is not None else self.nodes
        yield from workflow_iter_nodes_for_table_access_precheck(node_list, stop_index=stop_index, prefix=prefix)

    def build_table_access_precheck(self, execute_actions=True, stop_index=None, nodes=None):
        """
        执行前权限预检。

        以节点配置重新推导“期望表访问”，再和当前保存的 table_access 对比。
        这样既能发现默认映射遗漏，也能发现用户手动收窄权限后的运行风险。
        """
        node_list = nodes if nodes is not None else self.nodes
        self.ensure_node_tree_identity(node_list)
        self.refresh_node_tree_table_access(node_list)

        issues = []
        db_path = self.get_workflow_db_path()
        sqlite_tables = self.get_precheck_sqlite_tables()
        produced_transit = set((self.current_transit_tables or {}).keys())
        permission_label_map = dict(self.table_access_permission_items())
        db_exists = os.path.exists(db_path) if db_path else None

        if execute_actions:
            output_mode = self.output_mode_var.get()
            output_table = self.output_table_var.get().strip()
            issues.extend(workflow_evaluate_workflow_output_precheck(
                output_mode,
                output_table,
                db_path=db_path,
                write_mode_formatter=self.write_mode_display_text,
            ))

        for node_label, node in self.iter_nodes_for_table_access_precheck(node_list, stop_index=stop_index):
            if not isinstance(node, dict):
                continue
            if not node.get("enabled", True):
                continue

            node_type = node.get("type", "")
            config = node.get("config", {}) or {}
            expected_access = self.default_table_access_for_node(node)
            actual_access = self.get_node_table_access(node)
            node_result = workflow_evaluate_node_table_access_precheck(
                node_label,
                node,
                expected_access,
                actual_access,
                permission_label_map=permission_label_map,
                execute_actions=execute_actions,
                db_path=db_path,
                db_exists=db_exists,
                sqlite_tables=sqlite_tables,
                produced_transit=produced_transit,
                needs_plugin_declaration=node_type == "插件节点" and self.plugin_needs_table_access_declaration(config),
                has_plugin_declaration=node_type == "插件节点" and self.plugin_has_table_access_declaration(config),
                write_mode_formatter=self.write_mode_display_text,
            )
            issues.extend(node_result.get("issues", []))
            for transit_name in node_result.get("produced_transit", []) or []:
                produced_transit.add(transit_name)

        issues.sort(key=self.table_access_precheck_sort_key)
        self.last_table_access_precheck = issues
        return issues

    def show_table_access_precheck_dialog(self, issues, title="权限预检", allow_continue=False):
        issues = list(issues or [])
        result = {"continue": not allow_continue}
        win = tk.Toplevel(self.window)
        win.title(title)
        win.geometry("1280x680")
        win.minsize(980, 520)
        win.transient(self.window)

        main = ttk.Frame(win, padding=8)
        main.pack(fill=tk.BOTH, expand=True)
        summary_var = tk.StringVar(value=self.table_access_precheck_summary_text(issues))
        ttk.Label(main, textvariable=summary_var, font=("TkDefaultFont", 10, "bold")).pack(anchor=tk.W, pady=(0, 6))

        filter_frame = ttk.Frame(main)
        filter_frame.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(filter_frame, text="级别：").pack(side=tk.LEFT, padx=(0, 4))
        severity_var = tk.StringVar(value="全部")
        severity_combo = ttk.Combobox(filter_frame, textvariable=severity_var, values=["全部", "error", "warning", "info"], width=10, state="readonly")
        severity_combo.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(filter_frame, text="搜索：").pack(side=tk.LEFT, padx=(0, 4))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_frame, textvariable=search_var, width=34)
        search_entry.pack(side=tk.LEFT, padx=(0, 8))

        tree_wrap = ttk.Frame(main)
        tree_wrap.pack(fill=tk.BOTH, expand=True)
        columns = ("severity", "category", "blocking", "node", "source", "table", "role", "operation", "message", "suggestion")
        tree = ttk.Treeview(tree_wrap, columns=columns, show="headings", height=18)
        for col, text, width in [
            ("severity", "级别", 72),
            ("category", "类型", 72),
            ("blocking", "阻断", 52),
            ("node", "节点", 180),
            ("source", "来源", 82),
            ("table", "表", 150),
            ("role", "角色", 82),
            ("operation", "操作", 150),
            ("message", "问题", 320),
            ("suggestion", "建议", 260),
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor=tk.W)
        tree.tag_configure("error", foreground="#b00020")
        tree.tag_configure("warning", foreground="#8a5a00")
        tree.tag_configure("info", foreground="#335c99")
        yscroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(tree_wrap, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)

        def row_text(issue):
            return " ".join(str(issue.get(key, "") or "") for key in ["severity", "category", "blocking", "node", "source_type", "table", "role", "operation", "message", "suggestion"])

        def refresh_tree(*_):
            tree.delete(*tree.get_children())
            selected_sev = severity_var.get()
            keyword = search_var.get().strip().lower()
            visible = 0
            for idx, issue in enumerate(issues):
                sev = issue.get("severity", "info")
                if selected_sev != "全部" and sev != selected_sev:
                    continue
                if keyword and keyword not in row_text(issue).lower():
                    continue
                visible += 1
                tree.insert(
                    "",
                    tk.END,
                    iid=str(idx),
                    values=(
                        sev,
                        issue.get("category", ""),
                        "是" if issue.get("blocking") else "否",
                        issue.get("node", ""),
                        issue.get("source_type", ""),
                        issue.get("table", ""),
                        issue.get("role", ""),
                        issue.get("operation", ""),
                        issue.get("message", ""),
                        issue.get("suggestion", ""),
                    ),
                    tags=(sev,),
                )
            summary_var.set(self.table_access_precheck_summary_text(issues) + f" 当前显示 {visible} 项。")

        def show_detail(event=None):
            sel = tree.selection()
            if not sel:
                return
            issue = issues[int(sel[0])]
            detail = (
                f"级别：{issue.get('severity', '')}\n"
                f"类型：{issue.get('category', '')}\n"
                f"阻断执行：{'是' if issue.get('blocking') else '否'}\n"
                f"节点：{issue.get('node', '')}\n"
                f"表：{issue.get('source_type', '')} / {issue.get('table', '')}\n"
                f"角色：{issue.get('role', '')}\n"
                f"操作：{issue.get('operation', '')}\n\n"
                f"问题：{issue.get('message', '')}\n\n"
                f"建议：{issue.get('suggestion', '')}"
            )
            messagebox.showinfo("预检详情", detail, parent=win)

        tree.bind("<Double-1>", show_detail)
        severity_var.trace_add("write", refresh_tree)
        search_var.trace_add("write", refresh_tree)

        bottom = ttk.Frame(win, padding=(8, 0, 8, 8))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="打开字段权限层", command=lambda: (win.destroy(), self.open_table_access_window())).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="详情", command=show_detail).pack(side=tk.LEFT, padx=4)
        if allow_continue:
            def continue_run():
                result["continue"] = True
                win.destroy()
            def cancel_run():
                result["continue"] = False
                win.destroy()
            ttk.Button(bottom, text="继续执行", command=continue_run).pack(side=tk.RIGHT, padx=4)
            ttk.Button(bottom, text="取消执行", command=cancel_run).pack(side=tk.RIGHT, padx=4)
            win.protocol("WM_DELETE_WINDOW", cancel_run)
        else:
            ttk.Button(bottom, text="关闭", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        refresh_tree()
        self.center_toplevel(win, self.window, 1280, 680)
        try:
            win.grab_set()
        except Exception:
            pass
        self.window.wait_window(win)
        return bool(result.get("continue"))

    def confirm_table_access_precheck(self, execute_actions=True, stop_index=None):
        issues = self.build_table_access_precheck(execute_actions=execute_actions, stop_index=stop_index)
        self.last_table_access_precheck = list(issues or [])
        actionable = self.table_access_precheck_actionable(issues)
        if not actionable:
            self.status_var.set(self.table_access_precheck_summary_text(issues))
            return True
        policy = self.normalize_table_access_policy()
        if policy == "audit":
            self.status_var.set(self.table_access_precheck_summary_text(actionable) + " 当前策略为只审计，执行不会因预检提示而中断。")
            return True
        if policy == "strict":
            blocking = self.table_access_precheck_blocking(actionable)
            if not blocking:
                self.status_var.set(self.table_access_precheck_summary_text(actionable) + " 当前仅有风险提醒，强制模式允许继续执行。")
                return True
            self.show_table_access_precheck_dialog(
                blocking,
                title="执行前权限预检 - 强制拦截",
                allow_continue=False,
            )
            self.status_var.set("执行计划已拦截：当前权限策略为强制拦截，请先处理权限预检项。")
            return False
        return self.show_table_access_precheck_dialog(
            actionable,
            title="执行前权限预检",
            allow_continue=True,
        )

    def open_table_access_precheck_window(self):
        issues = self.build_table_access_precheck(execute_actions=True)
        self.show_table_access_precheck_dialog(issues, title="权限预检", allow_continue=False)

    def jump_node_label(self, idx, node):
        node_type = node.get("type", "")
        name = str(node.get("name", "") or "").strip()
        label = f"{idx + 1}.{node_type}"
        if name:
            label += f" / {name}"
        return label

    def collect_jump_anchors(self, nodes=None):
        node_list = nodes if nodes is not None else self.nodes
        anchors = []
        by_id = {}
        for idx, node in enumerate(node_list or []):
            if node.get("type") != "跳转锚点节点":
                continue
            cfg = node.get("config", {}) or {}
            anchor_id = str(cfg.get("anchor_id", "") or "").strip()
            entry = {
                "anchor_id": anchor_id,
                "anchor_name": str(cfg.get("anchor_name", "") or node.get("name", "") or "").strip(),
                "description": str(cfg.get("description", "") or "").strip(),
                "node_index": idx,
                "node_id": node.get("node_id", ""),
                "node_name": node.get("name", ""),
                "enabled": bool(node.get("enabled", True)),
                "node": node,
            }
            anchors.append(entry)
            if anchor_id:
                by_id.setdefault(anchor_id, []).append(entry)
        return {"all": anchors, "by_id": by_id}

    def collect_condition_flag_producers(self, nodes=None):
        node_list = nodes if nodes is not None else self.nodes
        flags = {}
        for idx, node in enumerate(node_list or []):
            if node.get("type") != "条件判断节点" or not node.get("enabled", True):
                continue
            flag_name = str((node.get("config", {}) or {}).get("flag_name", "") or "").strip()
            if not flag_name:
                continue
            flags.setdefault(flag_name, []).append({
                "node_index": idx,
                "node": node,
                "label": self.jump_node_label(idx, node),
            })
        return flags

    def resolve_jump_anchor_index(self, anchor_id, anchors_info=None, nodes=None):
        anchor_id = str(anchor_id or "").strip()
        if not anchor_id:
            return None, "目标锚点未配置"
        anchors_info = anchors_info if isinstance(anchors_info, dict) else self.collect_jump_anchors(nodes=nodes)
        matches = list((anchors_info.get("by_id") or {}).get(anchor_id, []) or [])
        if not matches:
            return None, f"目标锚点不存在：{anchor_id}"
        enabled = [item for item in matches if item.get("enabled")]
        if not enabled:
            return None, f"目标锚点已禁用：{anchor_id}"
        if len(enabled) > 1:
            return None, f"目标锚点重复：{anchor_id}"
        target_idx = int(enabled[0].get("node_index", -1))
        return target_idx, f"有效：节点 {target_idx + 1}"

    def jump_relation_status_text(self, relation, anchors_info=None, nodes=None):
        if not relation.get("enabled", True):
            return "跳转节点已禁用"
        target = str(relation.get("target_anchor_id", "") or "").strip()
        if not target:
            return "未配置目标锚点"
        target_idx, message = self.resolve_jump_anchor_index(target, anchors_info=anchors_info, nodes=nodes)
        if target_idx is None:
            return message
        return f"有效 -> 节点 {target_idx + 1}"

    def collect_jump_relations(self, nodes=None, anchors_info=None):
        node_list = nodes if nodes is not None else self.nodes
        anchors_info = anchors_info if isinstance(anchors_info, dict) else self.collect_jump_anchors(nodes=node_list)
        relations = []
        for idx, node in enumerate(node_list or []):
            node_type = node.get("type", "")
            cfg = node.get("config", {}) or {}
            enabled = bool(node.get("enabled", True))
            if node_type == "无条件跳转节点":
                relation = {
                    "source_index": idx,
                    "source_label": self.jump_node_label(idx, node),
                    "source_type": node_type,
                    "kind": "无条件",
                    "flag_name": "",
                    "condition_value": "始终",
                    "target_anchor_id": str(cfg.get("target_anchor_id", "") or "").strip(),
                    "enabled": enabled,
                    "is_default": False,
                    "node": node,
                }
                relation["status"] = self.jump_relation_status_text(relation, anchors_info=anchors_info, nodes=node_list)
                relations.append(relation)
            elif node_type == "条件跳转节点":
                flag_name = str(cfg.get("flag_name", "") or "").strip()
                rules = cfg.get("jump_rules", [])
                if not isinstance(rules, list):
                    rules = []
                for rule_idx, rule in enumerate(rules):
                    if not isinstance(rule, dict):
                        continue
                    relation = {
                        "source_index": idx,
                        "source_label": self.jump_node_label(idx, node),
                        "source_type": node_type,
                        "kind": "条件",
                        "flag_name": flag_name,
                        "condition_value": str(rule.get("value", "") or "").strip(),
                        "target_anchor_id": str(rule.get("target_anchor_id", "") or "").strip(),
                        "enabled": enabled,
                        "is_default": False,
                        "rule_index": rule_idx,
                        "node": node,
                    }
                    relation["status"] = self.jump_relation_status_text(relation, anchors_info=anchors_info, nodes=node_list)
                    relations.append(relation)
                default_anchor = str(cfg.get("default_anchor_id", "") or "").strip()
                if default_anchor:
                    relation = {
                        "source_index": idx,
                        "source_label": self.jump_node_label(idx, node),
                        "source_type": node_type,
                        "kind": "默认",
                        "flag_name": flag_name,
                        "condition_value": "DEFAULT",
                        "target_anchor_id": default_anchor,
                        "enabled": enabled,
                        "is_default": True,
                        "node": node,
                    }
                    relation["status"] = self.jump_relation_status_text(relation, anchors_info=anchors_info, nodes=node_list)
                    relations.append(relation)
                if not rules and not default_anchor:
                    relation = {
                        "source_index": idx,
                        "source_label": self.jump_node_label(idx, node),
                        "source_type": node_type,
                        "kind": "条件",
                        "flag_name": flag_name,
                        "condition_value": "",
                        "target_anchor_id": "",
                        "enabled": enabled,
                        "is_default": False,
                        "node": node,
                        "status": "未配置跳转规则",
                    }
                    relations.append(relation)
        return relations

    def add_jump_validation_issue(self, issues, severity, item, message, suggestion="", relation=None, anchor=None):
        issues.append({
            "severity": severity,
            "item": item,
            "message": message,
            "suggestion": suggestion,
            "relation": relation,
            "anchor": anchor,
        })

    def next_enabled_node_after_anchor(self, anchor, nodes=None):
        node_list = nodes if nodes is not None else self.nodes
        start = int(anchor.get("node_index", -1)) + 1
        for idx in range(start, len(node_list or [])):
            if (node_list[idx] or {}).get("enabled", True):
                return idx
        return None

    def validate_jump_relations(self, nodes=None):
        node_list = nodes if nodes is not None else self.nodes
        anchors_info = self.collect_jump_anchors(nodes=node_list)
        relations = self.collect_jump_relations(nodes=node_list, anchors_info=anchors_info)
        flag_producers = self.collect_condition_flag_producers(nodes=node_list)
        issues = []

        for anchor in anchors_info.get("all", []):
            anchor_id = anchor.get("anchor_id", "")
            label = f"{anchor.get('node_index', -1) + 1}.锚点"
            if not anchor_id:
                self.add_jump_validation_issue(
                    issues, "error", label, "锚点ID为空，其他跳转节点无法引用它。",
                    "给锚点填写唯一、稳定的锚点ID。", anchor=anchor
                )
            if anchor.get("enabled") and self.next_enabled_node_after_anchor(anchor, nodes=node_list) is None:
                self.add_jump_validation_issue(
                    issues, "warning", anchor_id or label, "锚点后没有可执行节点。",
                    "如果该锚点不是终点，请在锚点后添加处理节点。", anchor=anchor
                )

        for anchor_id, matches in (anchors_info.get("by_id") or {}).items():
            enabled_matches = [m for m in matches if m.get("enabled")]
            if len(matches) > 1:
                self.add_jump_validation_issue(
                    issues, "error", anchor_id, f"锚点ID重复：{len(matches)} 个节点使用同一个ID。",
                    "保留一个锚点ID，其他锚点改名；重复锚点运行时默认不跳转。",
                    anchor=matches[0],
                )
            if matches and not enabled_matches:
                self.add_jump_validation_issue(
                    issues, "warning", anchor_id, "该锚点当前全部处于禁用状态。",
                    "启用目标锚点，或调整跳转节点目标。", anchor=matches[0]
                )

        referenced = {str(rel.get("target_anchor_id", "") or "").strip() for rel in relations if str(rel.get("target_anchor_id", "") or "").strip()}
        for anchor in anchors_info.get("all", []):
            anchor_id = anchor.get("anchor_id", "")
            if anchor_id and anchor.get("enabled") and anchor_id not in referenced:
                self.add_jump_validation_issue(
                    issues, "info", anchor_id, "锚点未被任何跳转节点引用。",
                    "如果只是流程定位标记可以保留；如果希望跳到这里，请在跳转节点中绑定它。", anchor=anchor
                )

        checked_flag_nodes = set()
        for rel in relations:
            if not rel.get("enabled", True):
                continue
            source_idx = int(rel.get("source_index", -1))
            source_label = rel.get("source_label", "")
            target = str(rel.get("target_anchor_id", "") or "").strip()
            if rel.get("source_type") == "条件跳转节点":
                flag_name = str(rel.get("flag_name", "") or "").strip()
                flag_key = (source_idx, flag_name)
                if flag_key not in checked_flag_nodes:
                    checked_flag_nodes.add(flag_key)
                    if not flag_name:
                        self.add_jump_validation_issue(
                            issues, "warning", source_label, "条件跳转节点未填写读取标志。",
                            "填写条件判断节点输出的标志名；未填写时运行默认不跳转。", relation=rel
                        )
                    elif flag_name not in flag_producers:
                        self.add_jump_validation_issue(
                            issues, "warning", source_label, f"未找到条件标志来源：{flag_name}",
                            "在该节点之前添加条件判断节点，或确认标志名完全一致。", relation=rel
                        )
                    elif all(item.get("node_index", 0) > source_idx for item in flag_producers.get(flag_name, [])):
                        self.add_jump_validation_issue(
                            issues, "warning", source_label, f"条件标志 {flag_name} 的生成节点位于跳转节点之后。",
                            "把条件判断节点移到条件跳转节点之前。", relation=rel
                        )
                if not str(rel.get("condition_value", "") or "").strip() and not rel.get("is_default"):
                    self.add_jump_validation_issue(
                        issues, "warning", source_label, "条件规则的条件值为空。",
                        "填写 TRUE/FALSE 或条件判断节点实际输出值；空值规则很容易误判。", relation=rel
                    )

            if not target:
                self.add_jump_validation_issue(
                    issues, "warning", source_label, "跳转目标锚点未配置，运行时默认不跳转。",
                    "选择一个锚点；如果确实希望未命中时继续执行，可以保留默认锚点为空。", relation=rel
                )
                continue

            target_idx, message = self.resolve_jump_anchor_index(target, anchors_info=anchors_info, nodes=node_list)
            if target_idx is None:
                self.add_jump_validation_issue(
                    issues, "error", source_label, message,
                    "检查锚点ID是否存在、是否启用，以及是否重复。", relation=rel
                )
                continue
            if target_idx == source_idx:
                self.add_jump_validation_issue(
                    issues, "error", source_label, "跳转目标指向当前节点，可能形成自跳转。",
                    "改为跳到独立锚点，或删除该规则。", relation=rel
                )
            elif target_idx < source_idx:
                self.add_jump_validation_issue(
                    issues, "warning", source_label, f"目标锚点在当前节点之前：节点 {target_idx + 1}",
                    "这会形成回跳路径，请确认有条件能够退出，避免死循环。", relation=rel
                )

        severity_order = {"error": 0, "warning": 1, "info": 2}
        issues.sort(key=lambda item: (severity_order.get(item.get("severity"), 9), item.get("item", "")))
        return issues

    def jump_validation_summary_text(self, issues):
        issues = list(issues or [])
        if not issues:
            return "跳转校验完成：未发现明显问题。"
        counts = {}
        for issue in issues:
            sev = issue.get("severity", "info")
            counts[sev] = counts.get(sev, 0) + 1
        parts = []
        if counts.get("error"):
            parts.append(f"错误 {counts['error']}")
        if counts.get("warning"):
            parts.append(f"警告 {counts['warning']}")
        if counts.get("info"):
            parts.append(f"提示 {counts['info']}")
        return "跳转校验完成：" + "，".join(parts)

    def jump_issue_detail_text(self, issue):
        if not issue:
            return ""
        lines = [
            f"级别：{issue.get('severity', '')}",
            f"对象：{issue.get('item', '')}",
            f"问题：{issue.get('message', '')}",
        ]
        if issue.get("suggestion"):
            lines.append(f"建议：{issue.get('suggestion')}")
        rel = issue.get("relation") or {}
        if rel:
            lines.extend([
                "",
                "关系：",
                f"来源：{rel.get('source_label', '')}",
                f"类型：{rel.get('kind', '')}",
                f"读取标志：{rel.get('flag_name', '')}",
                f"条件值：{rel.get('condition_value', '')}",
                f"目标锚点：{rel.get('target_anchor_id', '')}",
                f"状态：{rel.get('status', '')}",
            ])
        anchor = issue.get("anchor") or {}
        if anchor:
            lines.extend([
                "",
                "锚点：",
                f"节点：{anchor.get('node_index', -1) + 1}",
                f"锚点ID：{anchor.get('anchor_id', '')}",
                f"名称：{anchor.get('anchor_name', '')}",
                f"启用：{'是' if anchor.get('enabled') else '否'}",
            ])
        return "\n".join(lines)

    def show_jump_precheck_dialog(self, issues, title="跳转校验", allow_continue=False):
        issues = list(issues or [])
        result = {"continue": not allow_continue}
        win = tk.Toplevel(self.window)
        win.title(title)
        win.geometry("1180x620")
        win.minsize(900, 480)
        win.transient(self.window)

        main = ttk.Frame(win, padding=8)
        main.pack(fill=tk.BOTH, expand=True)
        summary_var = tk.StringVar(value=self.jump_validation_summary_text(issues))
        ttk.Label(main, textvariable=summary_var, font=("TkDefaultFont", 10, "bold")).pack(anchor=tk.W, pady=(0, 6))
        ttk.Label(main, text="跳转目标无效时运行会默认不跳转；这里用于提前发现配置风险。", foreground="gray").pack(anchor=tk.W, pady=(0, 6))

        tree_wrap = ttk.Frame(main)
        tree_wrap.pack(fill=tk.BOTH, expand=True)
        columns = ("severity", "item", "message", "suggestion")
        tree = ttk.Treeview(tree_wrap, columns=columns, show="headings", height=18)
        for col, text, width in [
            ("severity", "级别", 70),
            ("item", "对象", 180),
            ("message", "问题", 420),
            ("suggestion", "建议", 360),
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor=tk.W)
        tree.tag_configure("error", foreground="#b00020")
        tree.tag_configure("warning", foreground="#8a5a00")
        tree.tag_configure("info", foreground="#335c99")
        yscroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(tree_wrap, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)

        for idx, issue in enumerate(issues):
            sev = issue.get("severity", "info")
            tree.insert(
                "",
                tk.END,
                iid=str(idx),
                values=(sev, issue.get("item", ""), issue.get("message", ""), issue.get("suggestion", "")),
                tags=(sev,),
            )

        def show_detail(event=None):
            sel = tree.selection()
            if not sel:
                return
            issue = issues[int(sel[0])]
            messagebox.showinfo("跳转校验详情", self.jump_issue_detail_text(issue), parent=win)

        def open_manager():
            result["continue"] = False
            win.destroy()
            self.open_jump_manager_window()

        tree.bind("<Double-1>", show_detail)

        bottom = ttk.Frame(win, padding=(8, 0, 8, 8))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="打开跳转管理", command=open_manager).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="详情", command=show_detail).pack(side=tk.LEFT, padx=4)
        if allow_continue:
            def continue_run():
                result["continue"] = True
                win.destroy()
            def cancel_run():
                result["continue"] = False
                win.destroy()
            ttk.Button(bottom, text="继续运行", command=continue_run).pack(side=tk.RIGHT, padx=4)
            ttk.Button(bottom, text="取消运行", command=cancel_run).pack(side=tk.RIGHT, padx=4)
            win.protocol("WM_DELETE_WINDOW", cancel_run)
        else:
            ttk.Button(bottom, text="关闭", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        self.center_toplevel(win, self.window, 1180, 620)
        try:
            win.grab_set()
        except Exception:
            pass
        self.window.wait_window(win)
        return bool(result.get("continue"))

    def confirm_jump_precheck(self, execute_actions=False, stop_index=None):
        issues = self.validate_jump_relations()
        actionable = [issue for issue in issues if issue.get("severity") in ("error", "warning")]
        self.last_jump_precheck = list(issues or [])
        if not actionable:
            return True
        errors = [issue for issue in actionable if issue.get("severity") == "error"]
        if not execute_actions and not errors:
            self.status_var.set(self.jump_validation_summary_text(actionable) + " 预览继续执行；可在跳转管理中查看。")
            return True
        return self.show_jump_precheck_dialog(
            actionable,
            title="执行前跳转校验" if execute_actions else "预览前跳转校验",
            allow_continue=True,
        )

    def open_jump_manager_window(self):
        self.ensure_node_tree_identity(self.nodes)
        win = tk.Toplevel(self.window)
        win.title("跳转管理")
        win.geometry("1360x740")
        win.minsize(1050, 560)
        win.transient(self.window)

        main = ttk.Frame(win, padding=8)
        main.pack(fill=tk.BOTH, expand=True)
        summary_var = tk.StringVar()
        ttk.Label(
            main,
            text="跳转系统只管理锚点与跳转关系，不管理表映射、字段映射或字段权限。",
            foreground="gray",
        ).pack(anchor=tk.W, pady=(0, 4))
        ttk.Label(main, textvariable=summary_var, font=("TkDefaultFont", 10, "bold")).pack(anchor=tk.W, pady=(0, 6))

        panes = ttk.Panedwindow(main, orient=tk.HORIZONTAL)
        panes.pack(fill=tk.BOTH, expand=True)

        left = ttk.LabelFrame(panes, text="锚点", padding=6)
        middle = ttk.LabelFrame(panes, text="跳转关系", padding=6)
        right = ttk.Frame(panes)
        panes.add(left, weight=1)
        panes.add(middle, weight=2)
        panes.add(right, weight=2)

        anchor_tree = ttk.Treeview(left, columns=("index", "anchor", "name", "refs", "status"), show="headings", height=20)
        for col, text, width in [
            ("index", "#", 45),
            ("anchor", "锚点ID", 150),
            ("name", "名称", 120),
            ("refs", "引用", 55),
            ("status", "状态", 85),
        ]:
            anchor_tree.heading(col, text=text)
            anchor_tree.column(col, width=width, anchor=tk.W)
        anchor_scroll = ttk.Scrollbar(left, orient=tk.VERTICAL, command=anchor_tree.yview)
        anchor_tree.configure(yscrollcommand=anchor_scroll.set)
        anchor_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        anchor_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        anchor_tree.tag_configure("disabled", foreground="#777777")
        anchor_tree.tag_configure("error", foreground="#b00020")

        relation_tree = ttk.Treeview(
            middle,
            columns=("source", "kind", "flag", "value", "target", "status"),
            show="headings",
            height=20,
        )
        for col, text, width in [
            ("source", "来源节点", 190),
            ("kind", "类型", 70),
            ("flag", "标志", 120),
            ("value", "条件值", 90),
            ("target", "目标锚点", 140),
            ("status", "状态", 190),
        ]:
            relation_tree.heading(col, text=text)
            relation_tree.column(col, width=width, anchor=tk.W)
        rel_y = ttk.Scrollbar(middle, orient=tk.VERTICAL, command=relation_tree.yview)
        rel_x = ttk.Scrollbar(middle, orient=tk.HORIZONTAL, command=relation_tree.xview)
        relation_tree.configure(yscrollcommand=rel_y.set, xscrollcommand=rel_x.set)
        relation_tree.grid(row=0, column=0, sticky="nsew")
        rel_y.grid(row=0, column=1, sticky="ns")
        rel_x.grid(row=1, column=0, sticky="ew")
        middle.rowconfigure(0, weight=1)
        middle.columnconfigure(0, weight=1)
        relation_tree.tag_configure("ok", foreground="#1b5e20")
        relation_tree.tag_configure("warning", foreground="#8a5a00")
        relation_tree.tag_configure("error", foreground="#b00020")
        relation_tree.tag_configure("disabled", foreground="#777777")

        detail_frame = ttk.LabelFrame(right, text="详情", padding=6)
        detail_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 6))
        detail_text = tk.Text(detail_frame, height=12, wrap=tk.WORD)
        detail_text.pack(fill=tk.BOTH, expand=True)
        detail_text.configure(state=tk.DISABLED)

        issue_frame = ttk.LabelFrame(right, text="校验结果", padding=6)
        issue_frame.pack(fill=tk.BOTH, expand=True)
        issue_tree = ttk.Treeview(issue_frame, columns=("severity", "item", "message", "suggestion"), show="headings", height=10)
        for col, text, width in [
            ("severity", "级别", 70),
            ("item", "对象", 130),
            ("message", "问题", 220),
            ("suggestion", "建议", 220),
        ]:
            issue_tree.heading(col, text=text)
            issue_tree.column(col, width=width, anchor=tk.W)
        issue_scroll = ttk.Scrollbar(issue_frame, orient=tk.VERTICAL, command=issue_tree.yview)
        issue_tree.configure(yscrollcommand=issue_scroll.set)
        issue_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        issue_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        issue_tree.tag_configure("error", foreground="#b00020")
        issue_tree.tag_configure("warning", foreground="#8a5a00")
        issue_tree.tag_configure("info", foreground="#335c99")

        state = {"anchors": [], "relations": [], "issues": []}

        def set_detail(text):
            detail_text.configure(state=tk.NORMAL)
            detail_text.delete("1.0", tk.END)
            detail_text.insert("1.0", text or "")
            detail_text.configure(state=tk.DISABLED)

        def relation_tag(relation):
            if not relation.get("enabled", True):
                return "disabled"
            status = relation.get("status", "")
            if status.startswith("有效"):
                return "ok"
            if "不存在" in status or "重复" in status:
                return "error"
            return "warning"

        def refresh_all():
            self.ensure_node_tree_identity(self.nodes)
            anchors_info = self.collect_jump_anchors()
            relations = self.collect_jump_relations(anchors_info=anchors_info)
            issues = self.validate_jump_relations()
            state["anchors"] = anchors_info.get("all", [])
            state["relations"] = relations
            state["issues"] = issues

            refs = {}
            for rel in relations:
                target = str(rel.get("target_anchor_id", "") or "").strip()
                if target:
                    refs[target] = refs.get(target, 0) + 1

            anchor_tree.delete(*anchor_tree.get_children())
            for idx, anchor in enumerate(state["anchors"]):
                anchor_id = anchor.get("anchor_id", "")
                status = "启用" if anchor.get("enabled") else "禁用"
                tag = ""
                if not anchor.get("enabled"):
                    tag = "disabled"
                if anchor_id and len((anchors_info.get("by_id") or {}).get(anchor_id, [])) > 1:
                    status = "重复"
                    tag = "error"
                anchor_tree.insert(
                    "",
                    tk.END,
                    iid=str(idx),
                    values=(anchor.get("node_index", -1) + 1, anchor_id, anchor.get("anchor_name", ""), refs.get(anchor_id, 0), status),
                    tags=(tag,),
                )

            relation_tree.delete(*relation_tree.get_children())
            for idx, rel in enumerate(relations):
                relation_tree.insert(
                    "",
                    tk.END,
                    iid=str(idx),
                    values=(
                        rel.get("source_label", ""),
                        rel.get("kind", ""),
                        rel.get("flag_name", ""),
                        rel.get("condition_value", ""),
                        rel.get("target_anchor_id", ""),
                        rel.get("status", ""),
                    ),
                    tags=(relation_tag(rel),),
                )

            issue_tree.delete(*issue_tree.get_children())
            for idx, issue in enumerate(issues):
                issue_tree.insert(
                    "",
                    tk.END,
                    iid=str(idx),
                    values=(issue.get("severity", ""), issue.get("item", ""), issue.get("message", ""), issue.get("suggestion", "")),
                    tags=(issue.get("severity", ""),),
                )
            summary_var.set(
                f"锚点 {len(state['anchors'])} 个，跳转关系 {len(relations)} 条。"
                + self.jump_validation_summary_text(issues)
            )
            if state["anchors"]:
                anchor_tree.selection_set("0")
                anchor_tree.focus("0")
                show_anchor_detail()
            else:
                set_detail("当前工作流还没有跳转锚点节点。")

        def show_anchor_detail(event=None):
            sel = anchor_tree.selection()
            if not sel:
                return
            idx = int(sel[0])
            anchors = state.get("anchors", [])
            if idx < 0 or idx >= len(anchors):
                return
            anchor = anchors[idx]
            anchor_id = anchor.get("anchor_id", "")
            refs = [rel for rel in state.get("relations", []) if rel.get("target_anchor_id") == anchor_id]
            lines = [
                f"锚点节点：{anchor.get('node_index', -1) + 1}",
                f"锚点ID：{anchor_id}",
                f"显示名称：{anchor.get('anchor_name', '')}",
                f"启用：{'是' if anchor.get('enabled') else '否'}",
                f"说明：{anchor.get('description', '') or '-'}",
                "",
                f"引用关系：{len(refs)} 条",
            ]
            for rel in refs[:20]:
                lines.append(f"- {rel.get('source_label', '')} / {rel.get('kind', '')} / {rel.get('condition_value', '')}")
            if len(refs) > 20:
                lines.append(f"... 仅显示前 20 条，共 {len(refs)} 条。")
            set_detail("\n".join(lines))

        def show_relation_detail(event=None):
            sel = relation_tree.selection()
            if not sel:
                return
            idx = int(sel[0])
            relations = state.get("relations", [])
            if idx < 0 or idx >= len(relations):
                return
            rel = relations[idx]
            lines = [
                f"来源节点：{rel.get('source_label', '')}",
                f"跳转类型：{rel.get('kind', '')}",
                f"读取标志：{rel.get('flag_name', '') or '-'}",
                f"条件值：{rel.get('condition_value', '') or '-'}",
                f"目标锚点：{rel.get('target_anchor_id', '') or '-'}",
                f"状态：{rel.get('status', '')}",
                "",
                "运行规则：",
                "目标有效时跳到锚点节点；锚点节点自身不计算，随后继续执行锚点后的节点。",
                "目标缺失、禁用、不存在或重复时，默认不跳转并继续后续节点。",
            ]
            set_detail("\n".join(lines))

        def show_issue_detail(event=None):
            sel = issue_tree.selection()
            if not sel:
                return
            idx = int(sel[0])
            issues = state.get("issues", [])
            if 0 <= idx < len(issues):
                set_detail(self.jump_issue_detail_text(issues[idx]))

        anchor_tree.bind("<<TreeviewSelect>>", show_anchor_detail)
        relation_tree.bind("<<TreeviewSelect>>", show_relation_detail)
        issue_tree.bind("<<TreeviewSelect>>", show_issue_detail)
        relation_tree.bind("<Double-1>", show_relation_detail)
        issue_tree.bind("<Double-1>", show_issue_detail)

        bottom = ttk.Frame(win, padding=(8, 0, 8, 8))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="刷新", command=refresh_all).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="关闭", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        refresh_all()
        self.center_toplevel(win, self.window, 1360, 740)

    def table_access_log_text(self, event):
        try:
            return json.dumps(event, ensure_ascii=False, default=str)
        except Exception:
            return str(event)

    def open_table_access_audit_window(self):
        logs_state = {"logs": list(self.last_table_access_logs or [])}
        win = tk.Toplevel(self.window)
        win.title("表访问权限审计日志")
        win.geometry("1320x700")
        win.minsize(980, 520)
        win.transient(self.window)

        main = ttk.Frame(win, padding=8)
        main.pack(fill=tk.BOTH, expand=True)
        summary_var = tk.StringVar()

        filter_frame = ttk.Frame(main)
        filter_frame.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(filter_frame, textvariable=summary_var, font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=(0, 16))
        ttk.Label(filter_frame, text="状态：").pack(side=tk.LEFT, padx=(0, 4))
        status_var = tk.StringVar(value="全部")
        status_combo = ttk.Combobox(filter_frame, textvariable=status_var, values=["全部", "ok", "warning", "denied", "missing", "compat"], width=10, state="readonly")
        status_combo.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(filter_frame, text="搜索：").pack(side=tk.LEFT, padx=(0, 4))
        search_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=search_var, width=34).pack(side=tk.LEFT, padx=(0, 8))

        tree_wrap = ttk.Frame(main)
        tree_wrap.pack(fill=tk.BOTH, expand=True)
        columns = ("time", "node", "source", "table", "operation", "status", "mode", "policy", "message")
        tree = ttk.Treeview(tree_wrap, columns=columns, show="headings", height=20)
        for col, text, width in [
            ("time", "时间", 145),
            ("node", "节点", 155),
            ("source", "来源", 82),
            ("table", "表", 150),
            ("operation", "操作", 150),
            ("status", "状态", 78),
            ("mode", "模式", 110),
            ("policy", "策略", 70),
            ("message", "信息", 360),
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor=tk.W)
        tree.tag_configure("ok", foreground="#1b5e20")
        tree.tag_configure("warning", foreground="#8a5a00")
        tree.tag_configure("denied", foreground="#b00020")
        tree.tag_configure("missing", foreground="#b00020")
        tree.tag_configure("compat", foreground="#555555")
        yscroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(tree_wrap, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)

        def log_row(event):
            node = event.get("node_name") or event.get("node_type") or event.get("node_id") or ""
            source = event.get("source_type") or event.get("access_source_type") or ""
            mode = event.get("write_mode") or event.get("mode") or ""
            return (
                event.get("time", ""),
                node,
                source,
                event.get("table_name", ""),
                event.get("operation_checked") or event.get("operation", ""),
                event.get("status", ""),
                mode,
                event.get("policy", ""),
                event.get("message", ""),
            )

        def refresh_tree(*_):
            tree.delete(*tree.get_children())
            selected_status = status_var.get()
            keyword = search_var.get().strip().lower()
            visible = 0
            counts = {}
            for idx, event in enumerate(logs_state["logs"]):
                status = str(event.get("status", "") or "")
                counts[status] = counts.get(status, 0) + 1
                if selected_status != "全部" and status != selected_status:
                    continue
                text = self.table_access_log_text(event).lower()
                if keyword and keyword not in text:
                    continue
                visible += 1
                row = log_row(event)
                tag = status if status in ("ok", "warning", "denied", "missing", "compat") else ""
                tree.insert("", tk.END, iid=str(idx), values=row, tags=(tag,))
            count_text = "，".join(f"{k or '无状态'} {v}" for k, v in sorted(counts.items()))
            summary_var.set(f"最近日志 {len(logs_state['logs'])} 条，当前显示 {visible} 条" + (f"（{count_text}）" if count_text else ""))

        def reload_logs():
            logs_state["logs"] = list(self.last_table_access_logs or [])
            refresh_tree()

        def clear_logs():
            self.last_table_access_logs = []
            logs_state["logs"] = []
            refresh_tree()

        def show_log_detail(event=None):
            sel = tree.selection()
            if not sel:
                return
            item = logs_state["logs"][int(sel[0])]
            messagebox.showinfo("审计日志详情", self.table_access_log_text(item), parent=win)

        def export_logs():
            if not logs_state["logs"]:
                messagebox.showwarning("提示", "当前没有可导出的审计日志。", parent=win)
                return
            path = filedialog.asksaveasfilename(
                title="导出表访问审计日志",
                defaultextension=".csv",
                filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")],
                parent=win,
            )
            if not path:
                return
            fieldnames = sorted({key for event in logs_state["logs"] if isinstance(event, dict) for key in event.keys()})
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for event in logs_state["logs"]:
                    row = {}
                    for key in fieldnames:
                        value = event.get(key, "")
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        row[key] = value
                    writer.writerow(row)
            messagebox.showinfo("导出完成", f"已导出审计日志：\n{path}", parent=win)

        tree.bind("<Double-1>", show_log_detail)
        status_var.trace_add("write", refresh_tree)
        search_var.trace_add("write", refresh_tree)

        bottom = ttk.Frame(win, padding=(8, 0, 8, 8))
        bottom.pack(fill=tk.X)
        ttk.Button(bottom, text="刷新最近日志", command=reload_logs).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="导出CSV", command=export_logs).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="清空最近日志", command=clear_logs).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="详情", command=show_log_detail).pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="关闭", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        refresh_tree()
        if not logs_state["logs"]:
            summary_var.set("最近日志 0 条。先预览或执行一次工作流后，这里会显示表访问审计。")
        self.center_toplevel(win, self.window, 1320, 700)

    def build_table_access_window_shell(self):
        win = tk.Toplevel(self.window)
        win.title("字段权限层")
        win.geometry("1180x720")
        win.minsize(900, 560)
        win.transient(self.window)

        main = ttk.Frame(win, padding=8)
        main.pack(fill=tk.BOTH, expand=True)

        panes = ttk.Panedwindow(main, orient=tk.HORIZONTAL)
        panes.pack(fill=tk.BOTH, expand=True)

        left = ttk.LabelFrame(panes, text="节点层", padding=6)
        detail = ttk.Frame(panes)
        panes.add(left, weight=1)
        panes.add(detail, weight=3)

        detail_tabs = ttk.Notebook(detail)
        detail_tabs.pack(fill=tk.BOTH, expand=True)
        middle = ttk.Frame(detail_tabs, padding=6)
        right = ttk.Frame(detail_tabs, padding=6)
        detail_tabs.add(middle, text="表权限层")
        detail_tabs.add(right, text="字段权限层")

        return {
            "win": win,
            "left": left,
            "middle": middle,
            "right": right,
        }

    def build_table_access_list_section(self, parent):
        node_tree = ttk.Treeview(
            parent,
            columns=("index", "type", "name", "status"),
            show="headings",
            height=22,
        )
        for col, text, width in workflow_table_access_node_tree_columns():
            node_tree.heading(col, text=text)
            node_tree.column(col, width=width, anchor=tk.W)
        node_y = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=node_tree.yview)
        node_tree.configure(yscrollcommand=node_y.set)
        node_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        node_y.pack(side=tk.RIGHT, fill=tk.Y)
        return {"node_tree": node_tree}

    def build_table_access_table_form_section(self, parent):
        table_tree_frame = ttk.Frame(parent)
        table_tree_frame.pack(fill=tk.BOTH, expand=True)
        table_tree = ttk.Treeview(
            table_tree_frame,
            columns=("role", "table", "operation", "current", "permissions", "mode", "status"),
            show="headings",
            height=12,
        )
        for col, text, width in workflow_table_access_table_tree_columns():
            table_tree.heading(col, text=text)
            table_tree.column(col, width=width, anchor=tk.W, stretch=False)
        table_y = ttk.Scrollbar(table_tree_frame, orient=tk.VERTICAL, command=table_tree.yview)
        table_x = ttk.Scrollbar(table_tree_frame, orient=tk.HORIZONTAL, command=table_tree.xview)
        table_tree.configure(yscrollcommand=table_y.set, xscrollcommand=table_x.set)
        table_tree.grid(row=0, column=0, sticky="nsew")
        table_y.grid(row=0, column=1, sticky="ns")
        table_x.grid(row=1, column=0, sticky="ew")
        table_tree_frame.rowconfigure(0, weight=1)
        table_tree_frame.columnconfigure(0, weight=1)

        table_form = ttk.LabelFrame(parent, text="表角色设置", padding=6)
        table_form.pack(fill=tk.X, pady=(6, 0))
        role_var = tk.StringVar()
        source_type_var = tk.StringVar(value="SQLite表")
        table_var = tk.StringVar()
        write_mode_var = tk.StringVar()
        preset_var = tk.StringVar(value="自定义")
        is_current_var = tk.BooleanVar(value=False)
        log_only_var = tk.BooleanVar(value=False)
        permission_vars = {key: tk.BooleanVar(value=False) for key, _ in self.table_access_permission_items()}

        ttk.Label(table_form, text="角色").grid(row=0, column=0, sticky=tk.W, padx=3, pady=3)
        ttk.Combobox(table_form, textvariable=role_var, values=workflow_table_access_role_choices(), width=12).grid(row=0, column=1, sticky=tk.W, padx=3, pady=3)
        ttk.Label(table_form, text="来源").grid(row=0, column=2, sticky=tk.W, padx=3, pady=3)
        ttk.Combobox(table_form, textvariable=source_type_var, values=workflow_table_access_source_type_choices(), width=12, state="readonly").grid(row=0, column=3, sticky=tk.W, padx=3, pady=3)
        ttk.Label(table_form, text="实际表").grid(row=1, column=0, sticky=tk.W, padx=3, pady=3)
        table_combo = ttk.Combobox(table_form, textvariable=table_var, values=self.table_access_table_choices(), width=25)
        table_combo.grid(row=1, column=1, columnspan=2, sticky=tk.W, padx=3, pady=3)
        ttk.Label(table_form, text="写入模式").grid(row=1, column=3, sticky=tk.W, padx=3, pady=3)
        ttk.Combobox(
            table_form,
            textvariable=write_mode_var,
            values=self.STANDARD_WRITE_MODE_CHOICES,
            width=19,
        ).grid(row=1, column=4, sticky=tk.W, padx=3, pady=3)
        ttk.Label(table_form, text="预设").grid(row=2, column=0, sticky=tk.W, padx=3, pady=3)
        preset_combo = ttk.Combobox(
            table_form,
            textvariable=preset_var,
            values=workflow_table_access_preset_choices(),
            width=18,
            state="readonly",
        )
        preset_combo.grid(row=2, column=1, sticky=tk.W, padx=3, pady=3)
        ttk.Checkbutton(table_form, text="当前表", variable=is_current_var).grid(row=2, column=2, sticky=tk.W, padx=3, pady=3)
        ttk.Checkbutton(table_form, text="只记录", variable=log_only_var).grid(row=2, column=3, sticky=tk.W, padx=3, pady=3)
        ttk.Label(table_form, text="字段权限范围").grid(row=4, column=0, sticky=tk.W, padx=3, pady=3)
        field_mapping_mode_var = tk.StringVar(value="按字段名")
        ttk.Combobox(
            table_form,
            textvariable=field_mapping_mode_var,
            values=workflow_table_access_field_mapping_mode_choices(),
            width=12,
            state="readonly",
        ).grid(row=4, column=1, sticky=tk.W, padx=3, pady=3)

        perm_frame = ttk.Frame(table_form)
        perm_frame.grid(row=3, column=0, columnspan=5, sticky=tk.W, pady=(4, 0))
        for idx, (key, label) in enumerate(self.table_access_permission_items()):
            ttk.Checkbutton(perm_frame, text=label, variable=permission_vars[key]).grid(row=idx // 5, column=idx % 5, sticky=tk.W, padx=4, pady=2)

        return {
            "table_tree": table_tree,
            "table_form": table_form,
            "role_var": role_var,
            "source_type_var": source_type_var,
            "table_var": table_var,
            "write_mode_var": write_mode_var,
            "preset_var": preset_var,
            "preset_combo": preset_combo,
            "is_current_var": is_current_var,
            "log_only_var": log_only_var,
            "permission_vars": permission_vars,
            "field_mapping_mode_var": field_mapping_mode_var,
            "table_combo": table_combo,
        }

    def build_table_access_field_form_section(self, parent):
        field_tree_frame = ttk.Frame(parent)
        field_tree_frame.pack(fill=tk.BOTH, expand=True)
        field_tree = ttk.Treeview(
            field_tree_frame,
            columns=("source_index", "source", "target_index", "target", "read", "write", "create", "protect", "status"),
            show="headings",
            height=14,
        )
        for col, text, width in workflow_table_access_field_tree_columns():
            field_tree.heading(col, text=text)
            field_tree.column(col, width=width, anchor=tk.W, stretch=False)
        field_y = ttk.Scrollbar(field_tree_frame, orient=tk.VERTICAL, command=field_tree.yview)
        field_x = ttk.Scrollbar(field_tree_frame, orient=tk.HORIZONTAL, command=field_tree.xview)
        field_tree.configure(yscrollcommand=field_y.set, xscrollcommand=field_x.set)
        field_tree.grid(row=0, column=0, sticky="nsew")
        field_y.grid(row=0, column=1, sticky="ns")
        field_x.grid(row=1, column=0, sticky="ew")
        field_tree_frame.rowconfigure(0, weight=1)
        field_tree_frame.columnconfigure(0, weight=1)

        field_form = ttk.LabelFrame(parent, text="字段权限设置", padding=6)
        field_form.pack(fill=tk.X, pady=(6, 0))
        source_field_var = tk.StringVar()
        target_field_var = tk.StringVar()
        source_index_var = tk.StringVar()
        target_index_var = tk.StringVar()
        field_permission_vars = {key: tk.BooleanVar(value=False) for key, _ in self.field_permission_items()}

        ttk.Label(field_form, text="来源字段").grid(row=0, column=0, sticky=tk.W, padx=3, pady=3)
        source_field_combo = ttk.Combobox(field_form, textvariable=source_field_var, width=22)
        source_field_combo.grid(row=0, column=1, sticky=tk.W, padx=3, pady=3)
        ttk.Label(field_form, text="源序号").grid(row=0, column=2, sticky=tk.W, padx=3, pady=3)
        ttk.Entry(field_form, textvariable=source_index_var, width=6).grid(row=0, column=3, sticky=tk.W, padx=3, pady=3)
        ttk.Label(field_form, text="目标字段").grid(row=1, column=0, sticky=tk.W, padx=3, pady=3)
        target_field_combo = ttk.Combobox(field_form, textvariable=target_field_var, width=22)
        target_field_combo.grid(row=1, column=1, sticky=tk.W, padx=3, pady=3)
        ttk.Label(field_form, text="目序号").grid(row=1, column=2, sticky=tk.W, padx=3, pady=3)
        ttk.Entry(field_form, textvariable=target_index_var, width=6).grid(row=1, column=3, sticky=tk.W, padx=3, pady=3)
        fp_frame = ttk.Frame(field_form)
        fp_frame.grid(row=2, column=0, columnspan=4, sticky=tk.W, pady=(4, 0))
        for idx, (key, label) in enumerate(self.field_permission_items()):
            ttk.Checkbutton(fp_frame, text=label, variable=field_permission_vars[key]).grid(row=0, column=idx, sticky=tk.W, padx=4, pady=2)

        return {
            "field_tree": field_tree,
            "field_form": field_form,
            "source_field_var": source_field_var,
            "target_field_var": target_field_var,
            "source_index_var": source_index_var,
            "target_index_var": target_index_var,
            "field_permission_vars": field_permission_vars,
            "source_field_combo": source_field_combo,
            "target_field_combo": target_field_combo,
        }

    def build_table_access_table_action_buttons(self, table_form, commands):
        table_btns = ttk.Frame(table_form)
        table_btns.grid(row=5, column=0, columnspan=5, sticky=tk.W, pady=(6, 0))
        buttons = {
            "add_table_entry": ttk.Button(table_btns, text="新增表角色", command=commands["add_table_entry"]),
            "save_table_entry": ttk.Button(table_btns, text="保存表设置", command=commands["save_table_entry"]),
            "delete_table_entry": ttk.Button(table_btns, text="删除表角色", command=commands["delete_table_entry"]),
            "rebuild_default_access": ttk.Button(table_btns, text="重建默认", command=commands["rebuild_default_access"]),
            "check_all_permissions": ttk.Button(table_btns, text="检查权限", command=commands["check_all_permissions"]),
            "preview_impact": ttk.Button(table_btns, text="预览影响", command=commands["preview_impact"]),
        }
        for button in buttons.values():
            button.pack(side=tk.LEFT, padx=3)
        return buttons

    def build_table_access_field_action_buttons(self, field_form, commands):
        field_btns = ttk.Frame(field_form)
        field_btns.grid(row=3, column=0, columnspan=4, sticky=tk.W, pady=(6, 0))
        buttons = {
            "add_field_entry": ttk.Button(field_btns, text="新增字段", command=commands["add_field_entry"]),
            "save_field_entry": ttk.Button(field_btns, text="保存字段", command=commands["save_field_entry"]),
            "delete_field_entry": ttk.Button(field_btns, text="删除字段", command=commands["delete_field_entry"]),
            "auto_match_fields": ttk.Button(field_btns, text="按字段名生成权限", command=commands["auto_match_fields"]),
            "auto_match_fields_by_order": ttk.Button(field_btns, text="按列顺序生成权限", command=commands["auto_match_fields_by_order"]),
            "clear_fields": ttk.Button(field_btns, text="清空字段", command=commands["clear_fields"]),
        }
        for button in buttons.values():
            button.pack(side=tk.LEFT, padx=3)
        return buttons

    def build_table_access_bottom_buttons(self, win, commands):
        bottom = ttk.Frame(win, padding=(8, 0, 8, 8))
        bottom.pack(fill=tk.X)
        buttons = {
            "refresh": ttk.Button(bottom, text="刷新节点列表", command=commands["refresh"]),
            "precheck": ttk.Button(bottom, text="权限预检", command=commands["precheck"]),
            "audit": ttk.Button(bottom, text="审计日志", command=commands["audit"]),
            "close": ttk.Button(bottom, text="关闭", command=commands["close"]),
        }
        buttons["refresh"].pack(side=tk.LEFT, padx=4)
        buttons["precheck"].pack(side=tk.LEFT, padx=4)
        buttons["audit"].pack(side=tk.LEFT, padx=4)
        buttons["close"].pack(side=tk.RIGHT, padx=4)
        return buttons

    def current_table_access_window_node(self, state):
        idx = state.get("node_index")
        if idx is None or idx < 0 or idx >= len(self.nodes):
            return None
        return self.nodes[idx]

    def refresh_table_access_node_tree(self, node_tree, state):
        state["refreshing_node_tree"] = True
        try:
            node_tree.delete(*node_tree.get_children())
            for idx, node in enumerate(self.nodes):
                mark = "√" if node.get("enabled", True) else "×"
                node_tree.insert(
                    "",
                    tk.END,
                    iid=str(idx),
                    values=(idx + 1, f"{mark} {node.get('type', '')}", node.get("name", ""), self.table_access_node_status(node)),
                )
            selected = state.get("node_index")
            if selected is not None and 0 <= selected < len(self.nodes):
                node_tree.selection_set(str(selected))
                node_tree.focus(str(selected))
        finally:
            state["refreshing_node_tree"] = False

    def load_table_access_table_form(self, table_section, entry, table_choices):
        entry = entry or {}
        table_section["role_var"].set(entry.get("role", "target"))
        table_section["source_type_var"].set(entry.get("source_type", "SQLite表"))
        table_section["table_var"].set(entry.get("table", ""))
        table_section["write_mode_var"].set(self.normalize_table_access_write_mode(entry.get("write_mode", "")))
        table_section["field_mapping_mode_var"].set(workflow_field_mapping_mode_display(entry))
        table_section["is_current_var"].set(bool(entry.get("is_current_table")))
        table_section["log_only_var"].set(bool(entry.get("log_only")))
        perms = entry.get("permissions") or {}
        for key, var in table_section["permission_vars"].items():
            var.set(bool(perms.get(key)))
        table_section["preset_var"].set("自定义")
        table_section["table_combo"].configure(values=table_choices)

    def refresh_table_access_field_choices(self, field_section, choices):
        field_section["source_field_combo"].configure(values=choices)
        field_section["target_field_combo"].configure(values=choices)

    def refresh_table_access_field_tree(self, state, field_tree, entry, field_section, choices):
        state["field_keys"] = workflow_render_field_mapping_tree(
            field_tree,
            entry,
            self.field_bool_text,
            self.field_permission_status,
        )
        self.refresh_table_access_field_choices(field_section, choices)

    def current_table_access_window_access(self, state):
        node = self.current_table_access_window_node(state)
        return self.get_node_table_access(node) if node is not None else {"tables": []}

    def current_table_access_window_table_entry(self, state):
        access = self.current_table_access_window_access(state)
        idx = state.get("table_index")
        tables = access.get("tables", [])
        if idx is None or idx < 0 or idx >= len(tables):
            return None
        return tables[idx]

    def load_table_access_window_table_form(self, state, table_section, entry):
        self.load_table_access_table_form(
            table_section,
            entry,
            self.table_access_table_choices(self.current_table_access_window_node(state)),
        )

    def collect_table_access_window_table_form(self, table_section):
        permission_vars = table_section["permission_vars"]
        return {
            "role": table_section["role_var"].get(),
            "source_type": table_section["source_type_var"].get(),
            "table": table_section["table_var"].get(),
            "is_current_table": table_section["is_current_var"].get(),
            "log_only": table_section["log_only_var"].get(),
            "write_mode": self.normalize_table_access_write_mode(table_section["write_mode_var"].get()),
            "field_mapping_mode": workflow_field_mapping_mode_value(table_section["field_mapping_mode_var"].get()),
            "permissions": {key: bool(var.get()) for key, var in permission_vars.items()},
        }

    def refresh_table_access_window_field_tree(self, state, field_section, field_tree):
        entry = self.current_table_access_window_table_entry(state)
        self.refresh_table_access_field_tree(
            state,
            field_tree,
            entry,
            field_section,
            self.get_table_access_field_choices(state.get("node_index") or 0, entry or {}),
        )

    def refresh_table_access_window_table_tree(
        self,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        select_index=None,
    ):
        access = self.current_table_access_window_access(state)
        tables = access.get("tables", [])
        if tables:
            if select_index is None:
                select_index = state.get("table_index")
            select_index = workflow_render_table_access_tree(
                table_tree,
                tables,
                self.table_access_entry_table_label,
                self.table_access_operation_summary,
                self.table_permission_summary,
                self.write_mode_display_text,
                self.table_access_entry_status,
                select_index=select_index,
            )
            state["table_index"] = select_index
            self.load_table_access_window_table_form(state, table_section, tables[select_index])
        else:
            workflow_render_table_access_tree(
                table_tree,
                tables,
                self.table_access_entry_table_label,
                self.table_access_operation_summary,
                self.table_permission_summary,
                self.write_mode_display_text,
                self.table_access_entry_status,
            )
            state["table_index"] = None
            self.load_table_access_window_table_form(state, table_section, {})
        self.refresh_table_access_window_field_tree(state, field_section, field_tree)
        self.refresh_table_access_node_tree(node_tree, state)

    def on_table_access_window_node_selected(
        self,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
        event=None,
        force=False,
    ):
        if state.get("refreshing_node_tree"):
            return
        sel = node_tree.selection()
        if not sel:
            return
        selected_index = int(sel[0])
        if not force and selected_index == state.get("node_index"):
            return
        state["node_index"] = selected_index
        state["table_index"] = None
        self.refresh_table_access_window_table_tree(
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
        )
        node = self.current_table_access_window_node(state)
        if node:
            status_var.set(f"当前节点：{state['node_index'] + 1}.{node.get('type')} / {node.get('name', '')}")

    def on_table_access_window_table_selected(
        self,
        state,
        table_section,
        field_section,
        table_tree,
        field_tree,
        event=None,
    ):
        sel = table_tree.selection()
        if not sel:
            return
        state["table_index"] = int(sel[0])
        entry = self.current_table_access_window_table_entry(state)
        self.load_table_access_window_table_form(state, table_section, entry)
        self.refresh_table_access_window_field_tree(state, field_section, field_tree)

    def on_table_access_window_field_selected(self, state, field_section, field_tree, event=None):
        sel = field_tree.selection()
        if not sel:
            return
        row_idx = int(sel[0])
        entry = self.current_table_access_window_table_entry(state)
        if not entry or row_idx >= len(state["field_keys"]):
            return
        key = state["field_keys"][row_idx]
        item = workflow_field_mapping_item(entry, key)
        if item is None:
            return
        workflow_load_field_form(
            item,
            field_section["source_field_var"],
            field_section["target_field_var"],
            field_section["source_index_var"],
            field_section["target_index_var"],
            field_section["field_permission_vars"],
        )

    def save_table_access_window_table_entry(
        self,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
    ):
        node = self.current_table_access_window_node(state)
        if node is None:
            return
        access = self.mark_node_table_access_manual(node)
        result = workflow_save_table_access_entry(
            access,
            state.get("table_index"),
            self.collect_table_access_window_table_form(table_section),
            lambda: self.make_table_access_entry("target", ""),
        )
        idx = result["table_index"]
        state["table_index"] = idx
        self.refresh_table_access_window_table_tree(
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
            select_index=idx,
        )
        status_var.set("表角色设置已保存。")

    def add_table_access_window_table_entry(
        self,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
    ):
        node = self.current_table_access_window_node(state)
        if node is None:
            return
        access = self.mark_node_table_access_manual(node)
        result = workflow_add_table_access_entry(
            access,
            self.make_table_access_entry(
                "target",
                "",
                permissions=self.table_permission_set(read=True),
            ),
        )
        state["table_index"] = result["table_index"]
        self.refresh_table_access_window_table_tree(
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
            select_index=state["table_index"],
        )

    def delete_table_access_window_table_entry(
        self,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
    ):
        node = self.current_table_access_window_node(state)
        idx = state.get("table_index")
        if node is None or idx is None:
            return
        access = self.mark_node_table_access_manual(node)
        result = workflow_delete_table_access_entry(access, idx)
        state["table_index"] = result["table_index"]
        self.refresh_table_access_window_table_tree(
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
            select_index=state["table_index"],
        )
        status_var.set("表角色已删除。")

    def rebuild_table_access_window_default_access(
        self,
        win,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
    ):
        node = self.current_table_access_window_node(state)
        if node is None:
            return
        if not messagebox.askyesno("重建默认映射", "将根据当前节点配置重建 table_access，并覆盖手动设置。继续吗？", parent=win):
            return
        workflow_rebuild_table_access(node, self.default_table_access_for_node(node))
        state["table_index"] = None
        self.refresh_table_access_window_table_tree(
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
        )
        status_var.set("已重建默认映射。")

    def check_table_access_window_permissions(self, win, status_var):
        result = workflow_build_table_access_permission_check(
            self.nodes,
            self.get_node_table_access,
            self.table_access_entry_status,
        )
        messagebox.showinfo("权限检查", result["message"], parent=win)
        status_var.set("权限检查完成。")

    def preview_table_access_window_impact(self, win, state):
        node = self.current_table_access_window_node(state)
        entry = self.current_table_access_window_table_entry(state)
        message = workflow_build_table_access_impact_preview(
            state.get("node_index") or 0,
            node,
            entry,
            self.table_access_field_items(entry) if entry is not None else [],
            self.table_access_entry_table_label,
            self.table_access_operation_summary,
            self.table_access_entry_status,
            self.table_permission_summary,
            self.write_mode_display_text,
        )
        if message is None:
            messagebox.showwarning("预览影响", "请先选择节点和表角色。", parent=win)
            return
        messagebox.showinfo("预览影响", message, parent=win)

    def apply_table_access_window_table_preset(self, table_section, event=None):
        preset = table_section["preset_var"].get()
        self.apply_table_access_preset_to_vars(
            preset,
            table_section["permission_vars"],
            table_section["log_only_var"],
        )
        preset_config = workflow_table_access_preset_config(
            preset,
            [key for key, _ in self.table_access_permission_items()],
        )
        if preset_config and preset_config.get("write_mode"):
            table_section["write_mode_var"].set(preset_config["write_mode"])

    def save_table_access_window_field_entry(self, state, table_section, field_section, field_tree, status_var):
        entry = self.current_table_access_window_table_entry(state)
        node = self.current_table_access_window_node(state)
        if entry is None or node is None:
            return
        self.mark_node_table_access_manual(node)
        sel = field_tree.selection()
        key = workflow_selected_field_key(sel, state["field_keys"])
        workflow_upsert_field_mapping_entry(
            entry,
            key,
            field_section["source_field_var"].get(),
            field_section["target_field_var"].get(),
            field_section["source_index_var"].get(),
            field_section["target_index_var"].get(),
            "by_order" if table_section["field_mapping_mode_var"].get() == "按列顺序" else "by_name",
            {pkey: bool(var.get()) for pkey, var in field_section["field_permission_vars"].items()},
            self.make_table_access_field_key,
        )
        self.refresh_table_access_window_field_tree(state, field_section, field_tree)
        status_var.set("字段映射已保存。")

    def add_table_access_window_field_entry(self, table_section, field_section, field_tree):
        workflow_reset_field_form(
            field_section["source_field_var"],
            field_section["target_field_var"],
            field_section["source_index_var"],
            field_section["target_index_var"],
            field_section["field_permission_vars"],
            write_enabled=table_section["permission_vars"]["write_table"].get(),
        )
        field_tree.selection_remove(field_tree.selection())

    def delete_table_access_window_field_entry(self, state, field_section, field_tree, status_var):
        entry = self.current_table_access_window_table_entry(state)
        node = self.current_table_access_window_node(state)
        sel = field_tree.selection()
        if entry is None or node is None or not sel:
            return
        key = workflow_selected_field_key(sel, state["field_keys"])
        if key and workflow_delete_field_mapping_entry(entry, key):
            self.mark_node_table_access_manual(node)
            self.refresh_table_access_window_field_tree(state, field_section, field_tree)
            status_var.set("字段映射已删除。")

    def auto_match_table_access_window_fields(self, state, field_section, field_tree, status_var):
        entry = self.current_table_access_window_table_entry(state)
        node = self.current_table_access_window_node(state)
        if entry is None or node is None:
            return
        self.mark_node_table_access_manual(node)
        count = self.auto_match_table_access_fields(state.get("node_index") or 0, entry)
        self.refresh_table_access_window_field_tree(state, field_section, field_tree)
        status_var.set(f"自动字段匹配完成：{count} 个字段。")

    def auto_match_table_access_window_fields_by_order(self, state, table_section, field_section, field_tree, status_var):
        entry = self.current_table_access_window_table_entry(state)
        node = self.current_table_access_window_node(state)
        if entry is None or node is None:
            return
        self.mark_node_table_access_manual(node)
        count = self.auto_match_table_access_fields_by_order(state.get("node_index") or 0, entry)
        table_section["field_mapping_mode_var"].set("按列顺序")
        self.refresh_table_access_window_field_tree(state, field_section, field_tree)
        status_var.set(f"按列顺序字段匹配完成：{count} 个字段。")

    def clear_table_access_window_fields(self, state, field_section, field_tree, status_var):
        entry = self.current_table_access_window_table_entry(state)
        node = self.current_table_access_window_node(state)
        if entry is None or node is None:
            return
        self.mark_node_table_access_manual(node)
        workflow_clear_field_mapping(entry)
        self.refresh_table_access_window_field_tree(state, field_section, field_tree)
        status_var.set("字段映射已清空。")

    def create_table_access_selection_callbacks(
        self,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
    ):
        def current_node():
            return self.current_table_access_window_node(state)

        def current_table_entry():
            return self.current_table_access_window_table_entry(state)

        def refresh_field_tree():
            self.refresh_table_access_window_field_tree(state, field_section, field_tree)

        def refresh_node_tree():
            self.refresh_table_access_node_tree(node_tree, state)

        def refresh_table_tree(select_index=None):
            self.refresh_table_access_window_table_tree(
                state,
                table_section,
                field_section,
                node_tree,
                table_tree,
                field_tree,
                select_index=select_index,
            )

        def on_node_selected(event=None, force=False):
            self.on_table_access_window_node_selected(
                state,
                table_section,
                field_section,
                node_tree,
                table_tree,
                field_tree,
                status_var,
                event=event,
                force=force,
            )

        def on_table_selected(event=None):
            self.on_table_access_window_table_selected(
                state,
                table_section,
                field_section,
                table_tree,
                field_tree,
                event=event,
            )

        def on_field_selected(event=None):
            self.on_table_access_window_field_selected(state, field_section, field_tree, event=event)

        return {
            "current_node": current_node,
            "current_table_entry": current_table_entry,
            "refresh_node_tree": refresh_node_tree,
            "refresh_table_tree": refresh_table_tree,
            "refresh_field_tree": refresh_field_tree,
            "on_node_selected": on_node_selected,
            "on_table_selected": on_table_selected,
            "on_field_selected": on_field_selected,
        }

    def create_table_access_table_action_callbacks(
        self,
        win,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
    ):
        def save_table_entry():
            self.save_table_access_window_table_entry(
                state,
                table_section,
                field_section,
                node_tree,
                table_tree,
                field_tree,
                status_var,
            )

        def add_table_entry():
            self.add_table_access_window_table_entry(
                state,
                table_section,
                field_section,
                node_tree,
                table_tree,
                field_tree,
            )

        def delete_table_entry():
            self.delete_table_access_window_table_entry(
                state,
                table_section,
                field_section,
                node_tree,
                table_tree,
                field_tree,
                status_var,
            )

        def rebuild_default_access():
            self.rebuild_table_access_window_default_access(
                win,
                state,
                table_section,
                field_section,
                node_tree,
                table_tree,
                field_tree,
                status_var,
            )

        def check_all_permissions():
            self.check_table_access_window_permissions(win, status_var)

        def preview_impact():
            self.preview_table_access_window_impact(win, state)

        def apply_table_preset(event=None):
            self.apply_table_access_window_table_preset(table_section, event=event)

        return {
            "save_table_entry": save_table_entry,
            "add_table_entry": add_table_entry,
            "delete_table_entry": delete_table_entry,
            "rebuild_default_access": rebuild_default_access,
            "check_all_permissions": check_all_permissions,
            "preview_impact": preview_impact,
            "apply_table_preset": apply_table_preset,
        }

    def create_table_access_field_action_callbacks(
        self,
        state,
        table_section,
        field_section,
        field_tree,
        status_var,
    ):
        def save_field_entry():
            self.save_table_access_window_field_entry(state, table_section, field_section, field_tree, status_var)

        def add_field_entry():
            self.add_table_access_window_field_entry(table_section, field_section, field_tree)

        def delete_field_entry():
            self.delete_table_access_window_field_entry(state, field_section, field_tree, status_var)

        def auto_match_fields():
            self.auto_match_table_access_window_fields(state, field_section, field_tree, status_var)

        def auto_match_fields_by_order():
            self.auto_match_table_access_window_fields_by_order(state, table_section, field_section, field_tree, status_var)

        def clear_fields():
            self.clear_table_access_window_fields(state, field_section, field_tree, status_var)

        return {
            "save_field_entry": save_field_entry,
            "add_field_entry": add_field_entry,
            "delete_field_entry": delete_field_entry,
            "auto_match_fields": auto_match_fields,
            "auto_match_fields_by_order": auto_match_fields_by_order,
            "clear_fields": clear_fields,
        }

    def create_table_access_window_callbacks(
        self,
        win,
        state,
        table_section,
        field_section,
        node_tree,
        table_tree,
        field_tree,
        status_var,
    ):
        selection_callbacks = self.create_table_access_selection_callbacks(
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
            status_var,
        )
        table_callbacks = self.create_table_access_table_action_callbacks(
            win,
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
            status_var,
        )
        field_callbacks = self.create_table_access_field_action_callbacks(
            state,
            table_section,
            field_section,
            field_tree,
            status_var,
        )
        callbacks = {}
        callbacks.update(selection_callbacks)
        callbacks.update(table_callbacks)
        callbacks.update(field_callbacks)
        return callbacks

    def open_table_access_window(self, initial_index=None):
        self.ensure_node_tree_identity(self.nodes)
        if initial_index is None:
            initial_index = self.get_selected_node_index()
        if initial_index is None and self.nodes:
            initial_index = 0

        state = {"node_index": initial_index, "table_index": None, "field_keys": [], "refreshing_node_tree": False}

        shell = self.build_table_access_window_shell()
        win = shell["win"]
        left = shell["left"]
        middle = shell["middle"]
        right = shell["right"]

        left_section = self.build_table_access_list_section(left)
        node_tree = left_section["node_tree"]

        table_section = self.build_table_access_table_form_section(middle)
        table_tree = table_section["table_tree"]
        table_form = table_section["table_form"]
        preset_combo = table_section["preset_combo"]

        field_section = self.build_table_access_field_form_section(right)
        field_tree = field_section["field_tree"]
        field_form = field_section["field_form"]

        status_var = tk.StringVar(value="选择节点后可编辑表权限与字段映射。")
        ttk.Label(win, textvariable=status_var, foreground="gray").pack(fill=tk.X, padx=8, pady=(0, 6))

        callbacks = self.create_table_access_window_callbacks(
            win,
            state,
            table_section,
            field_section,
            node_tree,
            table_tree,
            field_tree,
            status_var,
        )

        preset_combo.bind("<<ComboboxSelected>>", callbacks["apply_table_preset"])
        node_tree.bind("<<TreeviewSelect>>", callbacks["on_node_selected"])
        table_tree.bind("<<TreeviewSelect>>", callbacks["on_table_selected"])
        field_tree.bind("<<TreeviewSelect>>", callbacks["on_field_selected"])

        self.build_table_access_table_action_buttons(
            table_form,
            {
                "add_table_entry": callbacks["add_table_entry"],
                "save_table_entry": callbacks["save_table_entry"],
                "delete_table_entry": callbacks["delete_table_entry"],
                "rebuild_default_access": callbacks["rebuild_default_access"],
                "check_all_permissions": callbacks["check_all_permissions"],
                "preview_impact": callbacks["preview_impact"],
            },
        )
        self.build_table_access_field_action_buttons(
            field_form,
            {
                "add_field_entry": callbacks["add_field_entry"],
                "save_field_entry": callbacks["save_field_entry"],
                "delete_field_entry": callbacks["delete_field_entry"],
                "auto_match_fields": callbacks["auto_match_fields"],
                "auto_match_fields_by_order": callbacks["auto_match_fields_by_order"],
                "clear_fields": callbacks["clear_fields"],
            },
        )
        self.build_table_access_bottom_buttons(
            win,
            {
                "refresh": lambda: (callbacks["refresh_node_tree"](), callbacks["refresh_table_tree"](state.get("table_index"))),
                "precheck": self.open_table_access_precheck_window,
                "audit": self.open_table_access_audit_window,
                "close": win.destroy,
            },
        )

        callbacks["refresh_node_tree"]()
        if self.nodes:
            initial_index = max(0, min(int(initial_index or 0), len(self.nodes) - 1))
            node_tree.selection_set(str(initial_index))
            node_tree.focus(str(initial_index))
            callbacks["on_node_selected"](force=True)
        else:
            status_var.set("当前没有节点，请先添加工作流节点。")

    def get_table_manager(self, context=None, node=None, node_type="", node_name=""):
        db_path = self.get_workflow_db_path(context)
        if isinstance(context, dict) and "table_access_policy" not in context:
            snapshot = context.get("workflow_snapshot") or {}
            if isinstance(snapshot, dict) and snapshot.get("table_access_policy") is not None:
                context["table_access_policy"] = TableAccessManager.normalize_permission_policy(snapshot.get("table_access_policy"))
        current = (context or {}).get("current_node_info", {}) if isinstance(context, dict) else {}
        table_access = None
        if isinstance(node, dict):
            self.ensure_node_identity(node)
            node_id = node.get("node_id", "")
            node_name = node.get("name", node_name)
            node_type = node.get("type", node_type)
            table_access = node.get("table_access") if isinstance(node.get("table_access"), dict) else None
        else:
            node_id = current.get("node_id", "")
            node_name = current.get("node_name", node_name)
            node_type = current.get("node_type", node_type)
            table_access = current.get("table_access") if isinstance(current.get("table_access"), dict) else None
        return TableAccessManager(
            db_path,
            node_id=node_id,
            node_name=node_name,
            node_type=node_type,
            context=context,
            table_access=table_access,
        )

    def get_workflow_output_manager(self, table_name, overwrite=False, context=None):
        db_path = self.get_workflow_db_path(context)
        exists = bool(db_path and os.path.exists(db_path) and TableAccessManager(db_path).table_exists(table_name))
        permissions = self.table_permission_set(
            read=bool(overwrite and exists),
            write=True,
            create=True,
            replace=bool(overwrite),
        )
        access = {
            "version": 1,
            "auto_generated": True,
            "system_scope": "workflow_output",
            "tables": [
                self.make_table_access_entry(
                    "workflow_output",
                    table_name,
                    permissions=permissions,
                    write_mode="replace_table" if overwrite else "timestamp_new",
                    declared_by="workflow_output",
                )
            ],
        }
        policy = None
        if isinstance(context, dict):
            snapshot = context.get("workflow_snapshot") or {}
            policy = context.get("table_access_policy") or (snapshot.get("table_access_policy") if isinstance(snapshot, dict) else None)
        return TableAccessManager(
            db_path,
            node_id="__workflow_output__",
            node_name="工作流最终输出",
            node_type="工作流输出",
            context=context if isinstance(context, dict) else None,
            table_access=access,
            permission_policy=policy,
        )

    def transit_write_permissions_for_mode(self, exists=False, write_mode="", partial=False):
        required = TableAccessManager.required_permissions_for_write_mode(write_mode or "replace_table", exists=exists, partial=partial)
        standard = TableAccessManager.normalize_write_mode(write_mode)
        if standard in {"overlay_by_order", "write_fields_only", "fill_blank_fields", "clear_keep_schema"}:
            required.append("alter_schema")
        result = []
        for perm in required:
            if perm not in result:
                result.append(perm)
        return result

    def check_transit_table_permission(self, context, table_name, permissions, operation="transit_table",
                                       fields=None, field_action=None, write_mode="", node_type=""):
        table_name = str(table_name or "").strip()
        if not table_name:
            return None
        manager = self.get_table_manager(context if isinstance(context, dict) else None, node_type=node_type or "中转副表")
        manager.check_table_permission(
            table_name,
            permissions,
            operation=operation,
            fields=fields,
            field_action=field_action,
            write_mode=write_mode,
            source_type="中转副表",
        )
        return manager

    def check_transit_table_write_permission(self, context, table_name, exists=False, write_mode="",
                                             fields=None, partial=False, node_type="", operation="write_transit_table"):
        return self.check_transit_table_permission(
            context,
            table_name,
            self.transit_write_permissions_for_mode(exists=exists, write_mode=write_mode, partial=partial),
            operation=operation,
            fields=fields,
            field_action="write",
            write_mode=write_mode,
            node_type=node_type,
        )

    def log_transit_table_event(self, manager, operation, table_name, headers=None, rows=None, message="", **extra):
        if not isinstance(manager, TableAccessManager):
            return
        headers = list(headers or [])
        row_count = len(rows or [])
        extra.setdefault("source_type", "中转副表")
        extra.setdefault("rows", row_count)
        extra.setdefault("columns", len(headers))
        if not message:
            message = f"{operation} {table_name}：{row_count} 行 × {len(headers)} 列"
        manager._log_event(operation, table_name, message=message, **extra)

    def check_current_table_permission(self, context, headers, write=False, operation="current_table"):
        manager = self.get_table_manager(context if isinstance(context, dict) else None, node_type="当前工作流表")
        manager.check_table_permission(
            "__CURRENT_TABLE__",
            ["write_table", "update_rows"] if write else ["read_table"],
            operation=operation,
            fields=list(headers or []),
            field_action="write" if write else "read",
            write_mode="current_table_default" if write else "",
            source_type="当前工作流表",
        )
        return manager

    def log_current_table_transform(self, manager, before_shape, headers, rows, node_type=""):
        if not isinstance(manager, TableAccessManager):
            return
        after_shape = (len(rows or []), len(headers or []))
        manager._log_event(
            "transform_current_table",
            "__CURRENT_TABLE__",
            source_type="当前工作流表",
            before_rows=before_shape[0],
            before_columns=before_shape[1],
            rows=after_shape[0],
            columns=after_shape[1],
            message=f"当前工作流表处理完成：{before_shape[0]}×{before_shape[1]} -> {after_shape[0]}×{after_shape[1]}，节点 {node_type}",
        )

    def get_workflow_output_mode(self, context=None):
        snapshot = self.get_workflow_snapshot(context)
        value = str(snapshot.get("output_mode") or "").strip()
        if value:
            return value
        try:
            return self.output_mode_var.get()
        except Exception:
            return "输出到主界面预览区"

    def get_workflow_output_table(self, context=None):
        snapshot = self.get_workflow_snapshot(context)
        value = str(snapshot.get("output_table") or snapshot.get("workflow_name") or "").strip()
        if value:
            return value
        try:
            return self.output_table_var.get().strip()
        except Exception:
            return ""

    def get_workflow_backup_before_overwrite(self, context=None):
        snapshot = self.get_workflow_snapshot(context)
        if "backup_before_overwrite" in snapshot:
            return bool(snapshot.get("backup_before_overwrite"))
        try:
            return bool(self.backup_before_overwrite_var.get())
        except Exception:
            return True

    def get_workflow_sqlite_columns(self, table_name, context=None):
        """执行期读取 SQLite 字段，后台线程使用快照中的 db_path。"""
        db_path = self.get_workflow_db_path(context)
        if not db_path:
            raise ValueError("请先设置 SQLite 数据库路径。")
        return self.get_table_manager(context).get_columns(table_name)

    def read_plugin_input_table_source(self, spec, current_headers, current_rows, context=None):
        """按插件节点多表配置读取一张输入表。"""
        spec = spec or {}
        context = context or {"transit_tables": {}}
        source_type = str(spec.get("source_type") or "当前工作流表").strip() or "当前工作流表"
        if source_type == "当前工作流表":
            headers = list(current_headers or [])
            rows = [list(r) for r in self.normalize_rows(current_rows or [], len(headers))]
            return {
                "type": "table",
                "headers": headers,
                "rows": rows,
                "source_name": "workflow_current",
                "meta": {"source_type": source_type},
            }
        if source_type == "SQLite表":
            table = str(spec.get("sqlite_table") or spec.get("table") or "").strip()
            if not table:
                raise ValueError("插件额外输入表未选择 SQLite 表。")
            data = self.get_table_manager(context).read_table(table)
            headers = list(data.get("headers", []))
            rows = [list(row) for row in data.get("rows", [])]
            return {
                "type": "table",
                "headers": headers,
                "rows": rows,
                "source_name": f"SQLite:{table}",
                "meta": {"source_type": source_type, "table_name": table},
            }
        if source_type == "中转副表":
            name = str(spec.get("transit_table") or spec.get("table") or "").strip()
            if not name:
                raise ValueError("插件额外输入表未选择中转副表。")
            manager = self.check_transit_table_permission(
                context,
                name,
                ["read_table"],
                operation="read_transit_table",
                field_action="read",
                node_type="插件节点",
            )
            item = (context.get("transit_tables", {}) or {}).get(name)
            if not item:
                raise ValueError(f"插件额外输入表未找到中转副表：{name}")
            headers = list(item.get("headers", []) or [])
            rows = [list(r) for r in (item.get("rows", []) or [])]
            self.log_transit_table_event(manager, "read_transit_table", name, headers, rows, message=f"读取中转副表 {name}：{len(rows)} 行 × {len(headers)} 列")
            return {
                "type": "table",
                "headers": headers,
                "rows": rows,
                "source_name": f"中转:{name}",
                "meta": {"source_type": source_type, "table_name": name},
            }
        raise ValueError(f"未知插件输入表来源类型：{source_type}")

    def build_plugin_input_tables(self, config, current_headers, current_rows, context=None):
        """构建插件可用的多输入表字典，兼容旧版单表 input_data。"""
        primary = self.read_plugin_input_table_source(
            {"source_type": "当前工作流表"},
            current_headers,
            current_rows,
            context,
        )
        tables = {
            "当前表": primary,
            "workflow_current": primary,
            "primary": primary,
        }
        for index, spec in enumerate(config.get("input_tables", []) or [], start=1):
            if not isinstance(spec, dict):
                continue
            if spec.get("enabled", True) is False:
                continue
            table_data = self.read_plugin_input_table_source(spec, current_headers, current_rows, context)
            alias = str(spec.get("alias") or "").strip() or f"输入表{index}"
            table_data = dict(table_data)
            meta = dict(table_data.get("meta") or {})
            meta.update({"alias": alias, "input_index": index})
            table_data["meta"] = meta
            tables[alias] = table_data
        return tables

    def read_plugin_input_table_headers(self, spec, current_headers, context=None):
        """仅读取插件输入表字段，用于节点配置下拉菜单，避免 UI 阶段整表加载。"""
        spec = spec or {}
        context = context or {"transit_tables": {}}
        source_type = str(spec.get("source_type") or "当前工作流表").strip() or "当前工作流表"
        if source_type == "当前工作流表":
            return list(current_headers or [])
        if source_type == "SQLite表":
            table = str(spec.get("sqlite_table") or spec.get("table") or "").strip()
            if not table:
                return []
            return list(self.get_workflow_sqlite_columns(table, context))
        if source_type == "中转副表":
            name = str(spec.get("transit_table") or spec.get("table") or "").strip()
            item = (context.get("transit_tables", {}) or {}).get(name) if name else None
            return list((item or {}).get("headers", []) or [])
        return []

    def build_plugin_input_table_headers(self, config, current_headers, context=None):
        """构建插件可用输入表的字段映射，供动态参数控件使用。"""
        table_headers = {
            "当前表": list(current_headers or []),
            "workflow_current": list(current_headers or []),
            "primary": list(current_headers or []),
        }
        for index, spec in enumerate(config.get("input_tables", []) or [], start=1):
            if not isinstance(spec, dict) or spec.get("enabled", True) is False:
                continue
            alias = str(spec.get("alias") or "").strip() or f"输入表{index}"
            try:
                table_headers[alias] = self.read_plugin_input_table_headers(spec, current_headers, context)
            except Exception:
                table_headers.setdefault(alias, [])
        return table_headers

    def center_toplevel(self, win, parent=None, width=None, height=None):
        """把 Toplevel 放到父窗口中心；没有父窗口时放到屏幕中心。"""
        try:
            parent = parent or self.window
            win.update_idletasks()
            w = int(width or win.winfo_width() or win.winfo_reqwidth() or 600)
            h = int(height or win.winfo_height() or win.winfo_reqheight() or 400)
            if parent is not None and parent.winfo_exists():
                parent.update_idletasks()
                px = parent.winfo_rootx()
                py = parent.winfo_rooty()
                pw = parent.winfo_width()
                ph = parent.winfo_height()
                if pw <= 1 or ph <= 1:
                    px = py = 0
                    pw = win.winfo_screenwidth()
                    ph = win.winfo_screenheight()
            else:
                px = py = 0
                pw = win.winfo_screenwidth()
                ph = win.winfo_screenheight()
            x = max(0, px + (pw - w) // 2)
            y = max(0, py + (ph - h) // 2)
            win.geometry(f"{w}x{h}+{x}+{y}" if width or height else f"+{x}+{y}")
        except Exception:
            pass

    def show_centered_toplevel(self, win, parent=None, width=None, height=None):
        self.center_toplevel(win, parent, width, height)
        try:
            win.deiconify()
        except Exception:
            pass
        try:
            win.lift()
        except Exception:
            pass
        try:
            win.focus_set()
        except Exception:
            pass

    def plugin_config_context_with_live_transit(self, transit_context=None, include_rows=False):
        """插件配置期复用上次真实预览生成的中转副表。

        include_rows=False 时只补表名和字段，避免点选插件节点时复制大表；
        打开插件自带设置窗口时再传入真实行数据。
        """
        config_context = copy.deepcopy(transit_context or {"transit_tables": {}})
        transit_tables = config_context.setdefault("transit_tables", {})
        reused = []

        live_tables = {}
        if isinstance(getattr(self, "last_workflow_context", None), dict):
            live_tables.update(self.last_workflow_context.get("transit_tables", {}) or {})
        live_tables.update(getattr(self, "current_transit_tables", {}) or {})

        for name, live_item in live_tables.items():
            if not isinstance(live_item, dict):
                continue
            live_rows = list(live_item.get("rows", []) or [])
            if not live_rows:
                continue
            existing = transit_tables.get(name)
            existing_rows = []
            if isinstance(existing, dict):
                existing_rows = list(existing.get("rows", []) or [])
            if existing_rows:
                continue
            if include_rows:
                transit_tables[name] = copy.deepcopy(live_item)
            else:
                headers = list(live_item.get("headers", []) or [])
                source = live_item.get("source", "上次真实预览")
                if isinstance(existing, dict):
                    merged = copy.deepcopy(existing)
                    if not merged.get("headers") and headers:
                        merged["headers"] = headers
                    merged.setdefault("rows", [])
                    merged.setdefault("source", source)
                    transit_tables[name] = merged
                else:
                    transit_tables[name] = {"headers": headers, "rows": [], "source": source}
            if name not in reused:
                reused.append(name)

        if reused:
            config_context["_reused_preview_transit_tables"] = reused
        return config_context

    def plugin_config_transit_reuse_note(self, transit_context=None):
        return workflow_plugin_config_transit_reuse_note(transit_context)

    def get_plugin_dynamic_parameter_choices_for_config(
        self,
        item,
        config,
        params,
        spec,
        key,
        headers,
        current_rows=None,
        transit_context=None,
        input_table_headers=None,
    ):
        choices = workflow_get_plugin_static_parameter_choices(spec)
        provider = getattr(item.get("module"), "get_dynamic_parameter_options", None)
        if not callable(provider):
            return choices
        try:
            context = transit_context or {}
            plugin_context = self.make_plugin_context(config, context, execute_actions=False)
            plugin_context["input_table_headers"] = input_table_headers or self.build_plugin_input_table_headers(config, headers, context)
            plugin_context["plugin_input_table_specs"] = copy.deepcopy(config.get("input_tables", []))
            try:
                plugin_context["input_tables"] = self.build_plugin_input_tables(config, headers, current_rows or [], context)
            except Exception as table_exc:
                plugin_context["input_tables_error"] = str(table_exc)
            dynamic = provider(key, dict(params), plugin_context)
            return workflow_normalize_plugin_dynamic_parameter_choices(choices, dynamic)
        except Exception:
            return choices

    def run_plugin_custom_config_window(
        self,
        item,
        config,
        params,
        headers,
        current_rows=None,
        transit_context=None,
        dynamic_param_controls=None,
        refresh_dynamic_controls=None,
    ):
        window_transit_context = self.plugin_config_context_with_live_transit(transit_context, include_rows=True)
        plugin_context = self.make_plugin_context(config, window_transit_context or {}, execute_actions=False)
        try:
            input_tables = self.build_plugin_input_tables(config, headers, current_rows or [], window_transit_context or {})
            plugin_context["input_tables"] = input_tables
            plugin_context["plugin_input_table_specs"] = copy.deepcopy(config.get("input_tables", []))
            reuse_note_for_window = self.plugin_config_transit_reuse_note(window_transit_context)
            if reuse_note_for_window:
                plugin_context["plugin_config_data_note"] = reuse_note_for_window
                self.status_var.set(reuse_note_for_window)
        except Exception as table_exc:
            plugin_context["input_tables_error"] = str(table_exc)
            plugin_context["plugin_input_table_specs"] = copy.deepcopy(config.get("input_tables", []))
        result = item["module"].open_config_window(self.window, dict(params), plugin_context)
        if not workflow_apply_plugin_custom_config_result(config, params, result):
            return False
        for control in dynamic_param_controls or []:
            key = control.get("key", "")
            var = control.get("var")
            if var is not None and key in params:
                var.set(params.get(key, ""))
        if callable(refresh_dynamic_controls):
            refresh_dynamic_controls()
        return True

    def refresh_plugin_dynamic_config_controls(self, controls, set_param, get_choices):
        for control in controls or []:
            combo = control.get("combo")
            var = control.get("var")
            spec = control.get("spec", {})
            key = control.get("key", "")
            typ = control.get("type", "")
            if combo is None or var is None:
                continue
            current = str(var.get() or "")
            state = workflow_build_plugin_dynamic_control_state(
                typ,
                spec,
                current,
                get_choices(control) if callable(get_choices) else [],
            )
            try:
                combo.configure(values=state["choices"])
            except Exception:
                pass
            desired = state["value"]
            if desired != current:
                var.set(desired)
            set_param(key, var.get())

    def get_group_config_source_headers(self, source_type, headers, transit_context=None, transit_name="", sqlite_table=""):
        sqlite_columns = []
        if source_type == "SQLite表" and sqlite_table:
            try:
                sqlite_columns = self.get_workflow_sqlite_columns(
                    sqlite_table,
                    context=transit_context if isinstance(transit_context, dict) else None,
                )
            except Exception:
                sqlite_columns = []
        return workflow_group_source_headers_for_mapping(
            source_type,
            headers,
            (transit_context or {}).get("transit_tables", {}) if isinstance(transit_context, dict) else {},
            transit_name,
            sqlite_columns,
        )

    def save_group_inner_node_json_text(self, config, index, text):
        nodes = config.setdefault("nodes", [])
        if index is None or index < 0 or index >= len(nodes):
            raise ValueError("请先选择一个组内节点。")
        nodes[index] = workflow_parse_group_inner_node_json(text)
        return index

    def load_group_template_into_config(self, config):
        data = self.load_group_template_dialog()
        if data is None:
            return False
        workflow_apply_group_template_config(config, self.group_config_from_template_data(data))
        self.rebuild_current_config()
        return True

    def get_group_inner_node_type_values(self):
        return workflow_group_inner_node_type_values(self.get_node_type_values())

    def make_group_inner_node(self, node_type):
        return workflow_make_group_inner_node(
            node_type,
            plugin_display_map=getattr(self, "plugin_display_map", {}),
            plugin_registry=getattr(self, "plugin_registry", {}),
            plugin_config_factory=self.default_config_for_plugin,
            default_name_factory=self.default_name_for_node,
            default_config_factory=self.default_config_for_type,
        )

    def add_group_inner_node_to_config(self, config, node_type):
        _, index = workflow_add_group_inner_node(
            config,
            node_type,
            plugin_display_map=getattr(self, "plugin_display_map", {}),
            plugin_registry=getattr(self, "plugin_registry", {}),
            plugin_config_factory=self.default_config_for_plugin,
            default_name_factory=self.default_name_for_node,
            default_config_factory=self.default_config_for_type,
        )
        return index

    def apply_group_inner_node_list_action_to_config(self, config, index, action, delta=0):
        nodes, select_idx = workflow_apply_group_inner_node_list_action(
            config.setdefault("nodes", []),
            index,
            action,
            delta=delta,
        )
        config["nodes"] = nodes
        return select_idx

    def build_group_input_source_controls(self, input_frame, config, transit_context=None):
        return workflow_group_config_ui.build_group_input_source_controls(self, input_frame, config, transit_context=transit_context)

    def build_group_input_fields_controls(self, input_frame, config, refresh_mapping):
        return workflow_group_config_ui.build_group_input_fields_controls(self, input_frame, config, refresh_mapping)

    def build_group_mapping_tree_control(self, input_frame):
        return workflow_group_config_ui.build_group_mapping_tree_control(input_frame)

    def build_group_mapping_edit_controls(self, input_frame):
        return workflow_group_config_ui.build_group_mapping_edit_controls(input_frame)

    def build_group_mapping_action_buttons(
        self,
        map_edit,
        apply_mapping,
        auto_mapping,
        use_source_headers,
        infer_inputs,
    ):
        return workflow_group_config_ui.build_group_mapping_action_buttons(
            map_edit,
            apply_mapping,
            auto_mapping,
            use_source_headers,
            infer_inputs,
        )

    def build_group_input_mapping_section(self, frame, config, headers, transit_context=None, row=2):
        return workflow_group_config_ui.build_group_input_mapping_section(
            self,
            frame,
            config,
            headers,
            transit_context=transit_context,
            row=row,
        )

    def build_group_output_section(self, frame, config, row=3):
        return workflow_group_config_ui.build_group_output_section(self, frame, config, row=row)

    def refresh_group_source_field_combo(self, source_field_combo, source_field_var, source_headers):
        return workflow_group_config_ui.refresh_group_source_field_combo(source_field_combo, source_field_var, source_headers)

    def sync_group_mapping_edit_from_selected(
        self,
        config,
        mapping_tree,
        selected_input_var,
        source_field_var,
        default_value_var,
        refresh_source_fields,
    ):
        return workflow_group_config_ui.sync_group_mapping_edit_from_selected(
            config,
            mapping_tree,
            selected_input_var,
            source_field_var,
            default_value_var,
            refresh_source_fields,
        )

    def refresh_group_selected_input_combo(self, config, selected_input_combo, selected_input_var, sync_detail=None):
        return workflow_group_config_ui.refresh_group_selected_input_combo(config, selected_input_combo, selected_input_var, sync_detail=sync_detail)

    def refresh_group_mapping_tree(self, config, mapping_tree, refresh_selected_inputs):
        return workflow_group_config_ui.refresh_group_mapping_tree(config, mapping_tree, refresh_selected_inputs)

    def apply_group_mapping_from_controls(self, config, selected_input_var, source_field_var, default_value_var, refresh_mapping):
        return workflow_group_config_ui.apply_group_mapping_from_controls(
            config,
            selected_input_var,
            source_field_var,
            default_value_var,
            refresh_mapping,
            messagebox_module=messagebox,
        )

    def auto_group_mapping_by_name_from_source(self, config, get_source_headers, refresh_mapping):
        return workflow_group_config_ui.auto_group_mapping_by_name_from_source(config, get_source_headers, refresh_mapping)

    def use_group_source_headers_as_inputs(self, config, get_source_headers, set_input_fields_text, refresh_mapping):
        return workflow_group_config_ui.use_group_source_headers_as_inputs(
            config,
            get_source_headers,
            set_input_fields_text,
            refresh_mapping,
        )

    def create_group_input_mapping_callbacks(
        self,
        config,
        transit_context,
        get_source_headers,
        mapping_tree,
        selected_input_combo,
        selected_input_var,
        source_field_combo,
        source_field_var,
        default_value_var,
        input_fields_var,
    ):
        return workflow_group_config_ui.create_group_input_mapping_callbacks(
            self,
            config,
            transit_context,
            get_source_headers,
            mapping_tree,
            selected_input_combo,
            selected_input_var,
            source_field_combo,
            source_field_var,
            default_value_var,
            input_fields_var,
            messagebox_module=messagebox,
        )

    def infer_and_apply_group_input_fields_for_config(
        self,
        config,
        transit_context,
        get_source_headers,
        set_input_fields_text,
        refresh_mapping,
    ):
        return workflow_group_config_ui.infer_and_apply_group_input_fields_for_config(
            self,
            config,
            transit_context,
            get_source_headers,
            set_input_fields_text,
            refresh_mapping,
            messagebox_module=messagebox,
        )

    def build_group_inner_nodes_section(self, frame, config, row):
        return workflow_group_config_ui.build_group_inner_nodes_section(
            self,
            frame,
            config,
            row,
            messagebox_module=messagebox,
        )
    def build_plugin_run_environment_section(self, frame, config, item, plugin_id, start_row=3):
        available_run_modes = item.get("available_run_modes") or ["主程序内置环境", "插件独立环境"]
        status_state = workflow_build_plugin_load_status_state(
            item.get("load_status", "可内置运行"),
            item.get("metadata_source", ""),
            item.get("import_error", ""),
        )
        row = start_row
        ttk.Label(frame, text=status_state["text"], foreground=status_state["foreground"], wraplength=1050).grid(row=row, column=0, columnspan=4, sticky=tk.W, padx=4, pady=2)
        row += 1
        if status_state["import_error_text"]:
            ttk.Label(frame, text=status_state["import_error_text"], foreground="#b26a00", wraplength=1050).grid(row=row, column=0, columnspan=4, sticky=tk.W, padx=4, pady=(0, 6))
            row += 1

        normalized_run_mode = workflow_normalize_plugin_run_mode(
            config.get("run_mode", item.get("run_mode_default", "主程序内置环境")),
            available_run_modes,
        )
        config["run_mode"] = normalized_run_mode
        run_mode_var = tk.StringVar(value=normalized_run_mode)
        ttk.Label(frame, text="运行环境：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        run_mode_combo = ttk.Combobox(frame, textvariable=run_mode_var, values=available_run_modes, state="readonly", width=18)
        run_mode_combo.grid(row=row, column=1, sticky=tk.W, padx=4, pady=4)
        run_mode_var.trace_add("write", lambda *_, v=run_mode_var: config.__setitem__("run_mode", v.get()))
        ttk.Label(frame, text="独立环境适合插件依赖未打包进主程序的情况", foreground="gray").grid(row=row, column=2, columnspan=2, sticky=tk.W, padx=4, pady=4)
        row += 1

        external_python_var = tk.StringVar(value=config.get("external_python", ""))
        ttk.Label(frame, text="独立Python：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(frame, textvariable=external_python_var, width=58).grid(row=row, column=1, columnspan=2, sticky=tk.W, padx=4, pady=4)

        def choose_external_python(v=external_python_var):
            path = filedialog.askopenfilename(title="选择插件独立环境 python.exe", filetypes=[("Python", "python.exe;python"), ("所有文件", "*.*")])
            if path:
                v.set(path)

        ttk.Button(frame, text="选择", command=choose_external_python).grid(row=row, column=3, sticky=tk.W, padx=4, pady=4)
        external_python_var.trace_add("write", lambda *_, v=external_python_var: config.__setitem__("external_python", v.get()))
        row += 1

        env_dir_var = tk.StringVar(value=config.get("external_env_dir", self.get_plugin_env_dir(plugin_id)))
        ttk.Label(frame, text="环境目录：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(frame, textvariable=env_dir_var, width=58).grid(row=row, column=1, columnspan=2, sticky=tk.W, padx=4, pady=4)

        def open_env_dir(v=env_dir_var):
            path = v.get().strip() or self.get_plugin_env_dir(plugin_id)
            os.makedirs(path, exist_ok=True)
            try:
                os.startfile(path)
            except Exception as e:
                messagebox.showerror("打开失败", f"无法打开环境目录：\n{path}\n\n{e}")

        ttk.Button(frame, text="打开", command=open_env_dir).grid(row=row, column=3, sticky=tk.W, padx=4, pady=4)
        env_dir_var.trace_add("write", lambda *_, v=env_dir_var: config.__setitem__("external_env_dir", v.get()))
        row += 1

        entry_var = tk.StringVar(value=config.get("external_entry", item.get("external_entry", item.get("path", ""))))
        ttk.Label(frame, text="外部入口：").grid(row=row, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(frame, textvariable=entry_var, width=58).grid(row=row, column=1, columnspan=2, sticky=tk.W, padx=4, pady=4)

        def test_external_python(v=external_python_var):
            py = v.get().strip() or self.find_external_python(config, item, allow_current=True)
            try:
                out = subprocess.check_output([py, "--version"], stderr=subprocess.STDOUT, text=True, timeout=10)
                messagebox.showinfo("测试成功", out.strip())
            except Exception as e:
                messagebox.showerror("测试失败", str(e))

        ttk.Button(frame, text="测试环境", command=test_external_python).grid(row=row, column=3, sticky=tk.W, padx=4, pady=4)
        entry_var.trace_add("write", lambda *_, v=entry_var: config.__setitem__("external_entry", v.get()))
        return row + 1

    def refresh_plugin_input_listbox(self, input_lb, config):
        input_lb.delete(0, tk.END)
        for spec in config.get("input_tables", []) or []:
            input_lb.insert(tk.END, workflow_format_plugin_input_spec(spec))

    def open_plugin_input_spec_editor(
        self,
        config,
        index,
        sqlite_tables,
        transit_names,
        refresh_input_lb,
        refresh_plugin_dynamic_controls,
    ):
        specs = config.setdefault("input_tables", [])
        editing = index is not None and 0 <= index < len(specs)
        source_spec = copy.deepcopy(specs[index]) if editing else workflow_default_plugin_input_spec(len(specs), sqlite_tables, transit_names)
        win = tk.Toplevel(self.window)
        try:
            win.withdraw()
        except Exception:
            pass
        win.title("插件输入表设置")
        win.transient(self.window)
        body = ttk.Frame(win, padding=10)
        body.pack(fill=tk.BOTH, expand=True)

        alias_var = tk.StringVar(value=source_spec.get("alias", ""))
        source_type_var = tk.StringVar(value=source_spec.get("source_type", "SQLite表"))
        sqlite_var = tk.StringVar(value=source_spec.get("sqlite_table", source_spec.get("table", "")))
        transit_var = tk.StringVar(value=source_spec.get("transit_table", source_spec.get("table", "")))
        enabled_var = tk.BooleanVar(value=bool(source_spec.get("enabled", True)))

        ttk.Label(body, text="别名：").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(body, textvariable=alias_var, width=30).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(body, text="启用", variable=enabled_var).grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)

        ttk.Label(body, text="来源类型：").grid(row=1, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(
            body,
            textvariable=source_type_var,
            values=["当前工作流表", "SQLite表", "中转副表"],
            state="readonly",
            width=18,
        ).grid(row=1, column=1, sticky=tk.W, padx=4, pady=4)

        ttk.Label(body, text="SQLite表：").grid(row=2, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(body, textvariable=sqlite_var, values=sqlite_tables, width=34, state="normal").grid(row=2, column=1, columnspan=2, sticky=tk.W, padx=4, pady=4)
        ttk.Label(body, text="中转副表：").grid(row=3, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Combobox(body, textvariable=transit_var, values=transit_names, width=34, state="normal").grid(row=3, column=1, columnspan=2, sticky=tk.W, padx=4, pady=4)
        ttk.Label(body, text="建议别名示例：文档读取表、新内容表。别名是插件读取多表时的键名。", foreground="gray", wraplength=520).grid(row=4, column=0, columnspan=3, sticky=tk.W, padx=4, pady=(4, 8))

        btns = ttk.Frame(body)
        btns.grid(row=5, column=0, columnspan=3, sticky=tk.E, padx=4, pady=4)

        def on_ok():
            new_spec = workflow_build_plugin_input_spec(
                alias_var.get(),
                source_type_var.get(),
                sqlite_var.get(),
                transit_var.get(),
                enabled_var.get(),
                fallback_index=len(specs),
            )
            if editing:
                specs[index] = new_spec
            else:
                specs.append(new_spec)
            config["input_tables"] = specs
            refresh_input_lb()
            win.destroy()
            refresh_plugin_dynamic_controls()

        ttk.Button(btns, text="确定", command=on_ok).pack(side=tk.RIGHT, padx=4)
        ttk.Button(btns, text="取消", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        def show_input_window():
            self.show_centered_toplevel(win, self.window)
            win.grab_set()

        win.after_idle(show_input_window)

    def build_plugin_input_tables_section(
        self,
        frame,
        config,
        row,
        sqlite_tables,
        transit_names,
        refresh_plugin_dynamic_controls,
    ):
        input_specs = workflow_ensure_plugin_input_specs(config)
        input_frame = ttk.LabelFrame(frame, text="插件多表输入（可选）", padding=6)
        input_frame.grid(row=row, column=0, columnspan=4, sticky="ew", padx=4, pady=(4, 8))
        ttk.Label(
            input_frame,
            text="默认会传入当前工作流表；这里可额外传入 SQLite 表或中转副表，插件可从 input_data['tables'] / context['input_tables'] 按别名读取。",
            foreground="gray",
            wraplength=1050,
        ).grid(row=0, column=0, columnspan=5, sticky=tk.W, padx=4, pady=(0, 4))
        input_lb = tk.Listbox(input_frame, height=4, width=88, exportselection=False)
        input_lb.grid(row=1, column=0, columnspan=4, sticky="ew", padx=4, pady=4)

        def refresh_input_lb():
            self.refresh_plugin_input_listbox(input_lb, config)

        def selected_input_index():
            sel = input_lb.curselection()
            return int(sel[0]) if sel else None

        def edit_input_spec(index=None):
            self.open_plugin_input_spec_editor(
                config,
                index,
                sqlite_tables,
                transit_names,
                refresh_input_lb,
                refresh_plugin_dynamic_controls,
            )

        def edit_selected_input():
            idx = selected_input_index()
            if idx is not None:
                edit_input_spec(idx)

        def delete_selected_input():
            idx = selected_input_index()
            specs = config.setdefault("input_tables", [])
            if idx is not None and 0 <= idx < len(specs):
                del specs[idx]
                refresh_input_lb()
                refresh_plugin_dynamic_controls()

        input_btns = ttk.Frame(input_frame)
        input_btns.grid(row=1, column=4, sticky=tk.NW, padx=4, pady=4)
        ttk.Button(input_btns, text="增加", command=lambda: edit_input_spec(None)).pack(fill=tk.X, pady=2)
        ttk.Button(input_btns, text="编辑", command=edit_selected_input).pack(fill=tk.X, pady=2)
        ttk.Button(input_btns, text="删除", command=delete_selected_input).pack(fill=tk.X, pady=2)
        ttk.Button(input_btns, text="刷新", command=lambda: (refresh_input_lb(), refresh_plugin_dynamic_controls())).pack(fill=tk.X, pady=2)
        refresh_input_lb()
        return {
            "input_specs": input_specs,
            "input_listbox": input_lb,
            "refresh_input_lb": refresh_input_lb,
            "next_row": row + 1,
        }

    def create_plugin_dynamic_config_context(self, item, config, params, headers, transit_context, current_rows, dynamic_param_controls):
        state = {"refreshing_dynamic_controls": False}

        def set_param(key, value):
            params[key] = value
            config["params"] = params

        def get_input_table_header_map():
            return self.build_plugin_input_table_headers(config, headers, transit_context or {})

        def get_input_table_alias_choices():
            return workflow_get_plugin_input_table_alias_choices(
                get_input_table_header_map(),
                config.get("input_tables", []) or [],
            )

        def get_field_choices_for_table_param(spec):
            return workflow_get_plugin_field_choices_for_table_param(
                spec,
                params,
                get_input_table_header_map(),
            )

        def get_dynamic_parameter_choices(spec, key):
            return self.get_plugin_dynamic_parameter_choices_for_config(
                item,
                config,
                params,
                spec,
                key,
                headers,
                current_rows=current_rows,
                transit_context=transit_context or {},
                input_table_headers=get_input_table_header_map(),
            )

        def dynamic_choices_for_control(control):
            typ = control.get("type", "")
            spec = control.get("spec", {})
            key = control.get("key", "")
            if typ == "dynamic_select":
                return get_dynamic_parameter_choices(spec, key)
            if typ == "input_table_select":
                return get_input_table_alias_choices()
            if typ == "input_table_field_select":
                return get_field_choices_for_table_param(spec)
            return []

        def refresh_plugin_dynamic_controls():
            state["refreshing_dynamic_controls"] = True
            try:
                self.refresh_plugin_dynamic_config_controls(
                    dynamic_param_controls,
                    set_param,
                    dynamic_choices_for_control,
                )
            finally:
                state["refreshing_dynamic_controls"] = False

        return {
            "set_param": set_param,
            "get_input_table_alias_choices": get_input_table_alias_choices,
            "get_field_choices_for_table_param": get_field_choices_for_table_param,
            "get_dynamic_parameter_choices": get_dynamic_parameter_choices,
            "refresh_plugin_dynamic_controls": refresh_plugin_dynamic_controls,
            "is_refreshing_dynamic_controls": lambda: state["refreshing_dynamic_controls"],
        }

    def build_plugin_schema_parameter_controls(
        self,
        frame,
        schema,
        config,
        params,
        headers,
        row,
        dynamic_param_controls,
        dynamic_context,
    ):
        return workflow_build_plugin_schema_parameter_controls_ui(
            self,
            frame,
            schema,
            config,
            params,
            headers,
            row,
            dynamic_param_controls,
            dynamic_context,
        )

    def build_plugin_output_and_log_section(
        self,
        frame,
        config,
        item,
        params,
        headers,
        current_rows,
        transit_context,
        dynamic_param_controls,
        refresh_plugin_dynamic_controls,
        row,
    ):
        return workflow_build_plugin_output_and_log_section_ui(
            self,
            frame,
            config,
            item,
            params,
            headers,
            current_rows,
            transit_context,
            dynamic_param_controls,
            refresh_plugin_dynamic_controls,
            row,
        )

    def build_plugin_node_config(self, config, headers, transit_context=None, current_rows=None):
        return workflow_build_plugin_node_config_ui(self, config, headers, transit_context=transit_context, current_rows=current_rows)

    def normalize_plugin_logs(self, logs, plugin_id="", node_name="插件节点"):
        return workflow_normalize_plugin_logs(logs, plugin_id=plugin_id, node_name=node_name)

    def save_plugin_logs_to_file(self, plugin_id, log_items):
        if not log_items:
            return ""
        log_dir = self.get_plugin_log_dir()
        os.makedirs(log_dir, exist_ok=True)
        safe_id = re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fff]+", "_", str(plugin_id or "plugin"))
        path = os.path.join(log_dir, f"{safe_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
        with open(path, "w", encoding="utf-8") as f:
            for it in log_items:
                f.write(f"[{it.get('time','')}] [{it.get('level','INFO')}] [{it.get('plugin_id','')}] {it.get('object','')} {it.get('message','')}\n")
                tb = it.get("traceback") or ""
                if tb:
                    f.write(str(tb).rstrip() + "\n")
        return path

    def save_plugin_logs_to_sqlite(self, log_items, db_path=None, context=None):
        if not log_items:
            return 0
        db_path = str(db_path or "").strip()
        if not db_path:
            db_path = self.get_workflow_db_path(context)
        if not db_path:
            return 0
        if isinstance(context, dict):
            manager = self.get_table_manager(context, node_type="插件日志")
            if not manager.db_path:
                current = context.get("current_node_info", {}) if isinstance(context.get("current_node_info"), dict) else {}
                manager = TableAccessManager(
                    db_path,
                    node_id=current.get("node_id", ""),
                    node_name=current.get("node_name", ""),
                    node_type=current.get("node_type", "插件日志"),
                    context=context,
                    table_access=current.get("table_access") if isinstance(current.get("table_access"), dict) else None,
                )
            return manager.write_plugin_logs(log_items)
        return TableAccessManager(db_path, node_type="插件日志").write_plugin_logs(log_items)

    def plugin_log_items_to_table(self, log_items):
        return workflow_plugin_log_items_to_table(log_items)

    def save_plugin_output_to_transit(self, context, name, headers, rows, conflict_mode="覆盖", source="插件输出"):
        if context is None:
            return "未保存：无上下文"
        transit_tables = context.setdefault("transit_tables", {})
        base_name = str(name or "插件输出").strip() or "插件输出"
        headers = list(headers or [])
        rows = [list(r) for r in (rows or [])]
        exists_before = base_name in transit_tables
        if conflict_mode == "自动加时间戳":
            manager = self.check_transit_table_write_permission(
                context,
                base_name,
                exists=exists_before,
                write_mode=conflict_mode,
                fields=headers,
                node_type="插件节点",
            )
            final_name = self.make_unique_transit_name(base_name, transit_tables)
            transit_tables[final_name] = {"headers": headers, "rows": rows, "source": source}
            self.log_transit_table_event(manager, "write_transit_table", final_name, headers, rows, write_mode=conflict_mode, message=f"写入中转副表 {final_name}：{len(rows)} 行 × {len(headers)} 列")
            return f"中转副表：{final_name}"
        if conflict_mode == "追加" and base_name in transit_tables:
            manager = self.check_transit_table_write_permission(
                context,
                base_name,
                exists=True,
                write_mode=conflict_mode,
                fields=headers,
                node_type="插件节点",
            )
            old = transit_tables.get(base_name, {}) or {}
            mh, mr = self.append_headers_rows(old.get("headers", []), old.get("rows", []), headers, rows)
            transit_tables[base_name] = {"headers": mh, "rows": mr, "source": f"{source}:追加"}
            self.log_transit_table_event(manager, "append_transit_table", base_name, mh, mr, write_mode=conflict_mode, appended_rows=len(rows), message=f"追加中转副表 {base_name}：新增 {len(rows)} 行，累计 {len(mr)} 行")
            return f"中转副表追加：{base_name}（新增 {len(rows)} 行，累计 {len(mr)} 行）"
        manager = self.check_transit_table_write_permission(
            context,
            base_name,
            exists=exists_before,
            write_mode=conflict_mode or "覆盖",
            fields=headers,
            node_type="插件节点",
        )
        transit_tables[base_name] = {"headers": headers, "rows": rows, "source": source}
        self.log_transit_table_event(manager, "write_transit_table", base_name, headers, rows, write_mode=conflict_mode or "覆盖", message=f"写入中转副表 {base_name}：{len(rows)} 行 × {len(headers)} 列")
        return f"中转副表：{base_name}"

    def save_plugin_log_outputs(self, plugin_id, plugin_name, config, log_items, plugin_context=None, context=None, execute_actions=False, include_transit=True, suppress_errors=False):
        log_saved_parts = []
        plugin_context = plugin_context or {}
        should_save_persistent = execute_actions or config.get("plugin_log_in_preview", False)
        if config.get("save_plugin_log_file", True) and should_save_persistent:
            try:
                path = self.save_plugin_logs_to_file(plugin_id, log_items)
                if path:
                    log_saved_parts.append(f"日志文件：{path}")
            except Exception as e:
                if not suppress_errors:
                    log_saved_parts.append(f"日志文件保存失败：{e}")
        if config.get("save_plugin_log_sqlite", False) and should_save_persistent:
            try:
                cnt = self.save_plugin_logs_to_sqlite(log_items, db_path=plugin_context.get("db_path"), context=context)
                if cnt:
                    log_saved_parts.append(f"SQLite日志：{cnt}条")
            except Exception as e:
                if not suppress_errors:
                    log_saved_parts.append(f"SQLite日志保存失败：{e}")
        if include_transit and config.get("save_plugin_log_transit", False):
            try:
                lh, lr = self.plugin_log_items_to_table(log_items)
                log_name = config.get("plugin_log_transit_name") or f"{plugin_name or plugin_id}_日志"
                part = self.save_plugin_output_to_transit(context, log_name, lh, lr, config.get("transit_conflict_mode", "覆盖"), source=f"插件日志:{plugin_id}")
                log_saved_parts.append(part)
            except Exception as e:
                if not suppress_errors:
                    log_saved_parts.append(f"日志中转保存失败：{e}")
        return log_saved_parts

    def save_plugin_result_transit_output(self, config, item, plugin_id, context, headers, rows, source_prefix="插件"):
        if not workflow_should_save_plugin_output_as_transit(config):
            return []
        name = config.get("transit_name") or item.get("info", {}).get("name", plugin_id)
        part = self.save_plugin_output_to_transit(
            context,
            name,
            headers,
            rows,
            config.get("transit_conflict_mode", "覆盖"),
            source=f"{source_prefix}:{plugin_id}",
        )
        return [part]

    def merge_plugin_output_fields_to_current(self, cur_headers, cur_rows, out_headers, out_rows):
        return workflow_merge_plugin_output_fields_to_current(cur_headers, cur_rows, out_headers, out_rows)

    def is_external_plugin_mode(self, config, item=None):
        return workflow_is_external_plugin_mode(config, item)

    def find_external_python(self, config, item=None, allow_current=False, return_info=False):
        """查找外部插件 Python。

        return_info=True 时返回 (python_exe, used_current_fallback, warning_message)，
        便于后台线程在日志/状态中明确提示是否回退到了当前 Python。
        """
        py = str(config.get("external_python", "")).strip()
        if py and os.path.exists(py):
            return (py, False, "") if return_info else py
        env_dir = str(config.get("external_env_dir", "")).strip()
        if env_dir:
            candidates = [
                os.path.join(env_dir, "Scripts", "python.exe"),
                os.path.join(env_dir, "bin", "python"),
                os.path.join(env_dir, "python.exe"),
                os.path.join(env_dir, "python"),
            ]
            for c in candidates:
                if os.path.exists(c):
                    return (c, False, "") if return_info else c
        if allow_current and not getattr(sys, "frozen", False):
            warn = "未找到插件独立 Python，当前处于源码运行模式，已回退使用主程序 Python。正式使用独立环境插件时建议配置 plugin_envs/插件ID/Scripts/python.exe。"
            return (sys.executable, True, warn) if return_info else sys.executable
        raise FileNotFoundError("未找到插件独立 Python。请在插件节点中选择 plugin_envs/插件ID/Scripts/python.exe，或先创建插件独立环境。")

    def make_external_plugin_json_context(self, config, context=None, execute_actions=False):
        plugin_id = config.get("plugin_id", "")
        context = context or {}
        snapshot = context.get("workflow_snapshot", {}) if isinstance(context, dict) else {}
        db_path = str(snapshot.get("db_path", "")).strip()
        workflow_name = str(snapshot.get("workflow_name", "")).strip()
        app_dir = snapshot.get("app_dir") or getattr(self.app, "app_dir", get_app_dir())
        if not db_path:
            # 兼容非后台/旧入口调用；后台线程应优先走 snapshot。
            db_path = self.get_workflow_db_path(context)
        if not workflow_name:
            workflow_name = self.get_workflow_output_table(context)
        return {
            "app_dir": app_dir,
            # 独立进程不接收真实数据库路径。需要落库时返回 database_requests，
            # 由主程序在当前节点权限上下文中统一执行。
            "db_path": "",
            "database_access": "managed_requests",
            "database_available": bool(db_path),
            "plugins_dir": self.get_plugins_dir(),
            "plugin_data_dir": self.get_plugin_data_dir(plugin_id),
            "log_dir": self.get_plugin_log_dir(),
            "is_preview": not bool(execute_actions),
            "execute_actions": bool(execute_actions),
            "is_config_probe": bool(context.get("is_config_probe")),
            "workflow_name": workflow_name,
            "node_name": config.get("name") or config.get("node_name") or "插件节点",
            "plugin_id": plugin_id,
            "transit_tables": context.get("transit_tables", {}),
            "input_tables": context.get("input_tables", {}),
            "plugin_input_table_specs": copy.deepcopy(config.get("input_tables", [])),
        }

    def run_external_plugin_process(self, item, input_data, params, config, context=None, execute_actions=False):
        """使用独立 Python 环境运行插件。

        外部插件入口需支持：
            python plugin.py --input input.json --output output.json
        执行过程中可以向 stdout 输出 JSON 行：
            {"type":"node_progress","current":1,"total":10,"message":"..."}

        本版将 stdout 读取移动到独立线程，主循环短轮询 cancel_event/timeout，
        避免外部插件长时间不输出时 readline() 阻塞导致取消不及时。
        """
        plugin_id = config.get("plugin_id", item.get("id", "plugin"))
        context = context or {}
        logs = []
        progress_callback = context.get("progress_callback")
        cancel_event = context.get("cancel_event")

        python_exe, used_current_fallback, fallback_warning = self.find_external_python(config, item, allow_current=True, return_info=True)
        if fallback_warning:
            logs.append({"level": "WARNING", "message": fallback_warning})
            if callable(progress_callback):
                try:
                    progress_callback({
                        "type": "node_progress",
                        "node_name": config.get("name") or "插件节点",
                        "plugin_id": plugin_id,
                        "message": fallback_warning,
                    })
                except Exception:
                    pass

        entry = str(config.get("external_entry") or item.get("external_entry") or item.get("path") or "").strip()
        if not entry:
            raise FileNotFoundError("未配置外部插件入口文件")
        if not os.path.isabs(entry):
            entry = os.path.join(self.get_plugins_dir(), entry)
        if not os.path.exists(entry):
            raise FileNotFoundError(f"外部插件入口不存在：{entry}")

        run_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        run_dir = os.path.join(self.get_plugin_data_dir(plugin_id), "runs", run_id)
        os.makedirs(run_dir, exist_ok=True)
        input_path = os.path.join(run_dir, "input.json")
        output_path = os.path.join(run_dir, "output.json")
        payload = {
            "input_data": input_data,
            "params": params,
            "context": self.make_external_plugin_json_context(config, context, execute_actions=execute_actions),
        }
        with open(input_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        cmd = [python_exe, entry, "--input", input_path, "--output", output_path]
        timeout_text = str(config.get("external_timeout", "0") or "0").strip()
        try:
            timeout = float(timeout_text)
        except Exception:
            timeout = 0.0
        start_time = time.time()
        stdout_queue = queue.Queue()
        stdout_done = threading.Event()

        def stdout_reader(pipe):
            try:
                if pipe is None:
                    return
                for line in iter(pipe.readline, ""):
                    stdout_queue.put(line)
            except Exception as e:
                stdout_queue.put(json.dumps({"type": "node_log", "level": "WARNING", "message": f"读取外部插件输出失败：{e}"}, ensure_ascii=False))
            finally:
                stdout_done.set()
                try:
                    if pipe is not None:
                        pipe.close()
                except Exception:
                    pass

        def terminate_process(proc, reason, exc):
            try:
                if proc.poll() is None:
                    proc.terminate()
                    try:
                        proc.wait(timeout=3)
                    except subprocess.TimeoutExpired:
                        proc.kill()
                        proc.wait(timeout=3)
            except Exception:
                try:
                    if proc.poll() is None:
                        proc.kill()
                except Exception:
                    pass
            raise exc

        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding="utf-8", errors="replace")
        reader_thread = threading.Thread(target=stdout_reader, args=(proc.stdout,), daemon=True)
        reader_thread.start()
        code = None
        try:
            while True:
                drained = 0
                while drained < 200:
                    try:
                        line = stdout_queue.get_nowait()
                    except queue.Empty:
                        break
                    handle_plugin_stdout_line(
                        line,
                        logs,
                        progress_callback=progress_callback,
                        node_name=config.get("name") or "插件节点",
                        plugin_id=plugin_id,
                    )
                    drained += 1

                if cancel_event is not None and cancel_event.is_set():
                    terminate_process(proc, "cancel", RuntimeError("用户取消外部插件执行"))
                if timeout > 0 and (time.time() - start_time) > timeout:
                    terminate_process(proc, "timeout", TimeoutError(f"外部插件执行超时：{timeout}秒"))

                code = proc.poll()
                if code is not None and stdout_done.is_set() and stdout_queue.empty():
                    break
                time.sleep(0.05)

            code = proc.wait(timeout=1)
        finally:
            if proc.poll() is None:
                try:
                    proc.kill()
                except Exception:
                    pass
            try:
                reader_thread.join(timeout=1)
            except Exception:
                pass

        if not os.path.exists(output_path):
            if code != 0:
                raise RuntimeError(f"外部插件进程返回错误码：{code}，且未生成 output.json。运行目录：{run_dir}")
            raise FileNotFoundError(f"外部插件未生成 output.json：{output_path}")
        with open(output_path, "r", encoding="utf-8") as f:
            result = json.load(f)
        if isinstance(result, dict):
            old_logs = result.get("logs", []) or []
            if logs:
                result["logs"] = old_logs + logs
            if result.get("ok", True):
                self.execute_external_plugin_database_requests(
                    result,
                    config,
                    context,
                    execute_actions=execute_actions,
                )
            if used_current_fallback:
                summary = result.get("summary", {}) or {}
                summary["used_current_python_fallback"] = True
                summary["actual_python"] = python_exe
                result["summary"] = summary
            if code != 0 and result.get("ok", True):
                result["ok"] = False
                result["message"] = result.get("message") or f"外部插件进程返回错误码：{code}"
        return result

    def execute_external_plugin_database_requests(self, result, config, context=None, execute_actions=False):
        if not isinstance(result, dict):
            return []
        requests = [item for item in (result.get("database_requests") or []) if isinstance(item, dict)]
        if not requests:
            return []
        logs = result.setdefault("logs", [])
        if not execute_actions:
            logs.append({
                "level": "INFO",
                "message": f"预览模式未执行外部插件数据库请求：{len(requests)} 项",
            })
            result["database_results"] = [
                {"status": "preview_skipped", "operation": item.get("operation", "")}
                for item in requests
            ]
            return result["database_results"]

        manager = self.get_table_manager(
            context if isinstance(context, dict) else None,
            node_type="插件节点",
            node_name=config.get("name") or config.get("node_name") or "插件节点",
        )
        results = []
        for index, request in enumerate(requests, start=1):
            operation = str(request.get("operation", "") or "").strip()
            if operation != "write_table":
                raise ValueError(f"外部插件数据库请求不支持操作：{operation or '<empty>'}")
            table_name = str(request.get("table_name", "") or "").strip()
            headers = list(request.get("headers") or [])
            rows = [list(row) for row in (request.get("rows") or [])]
            mode = request.get("mode") or "replace"
            info = manager.write_table(table_name, headers, rows, mode=mode)
            results.append({
                "status": "ok",
                "request_index": index,
                "operation": operation,
                **info,
            })
            logs.append({
                "level": "INFO",
                "message": f"主程序已执行外部插件数据库请求 {index}/{len(requests)}：{table_name}",
            })
        result["database_results"] = results
        if isinstance(context, dict) and results:
            context["needs_refresh_table_list"] = True
        return results

    def make_plugin_context(self, config, context=None, execute_actions=False):
        plugin_id = config.get("plugin_id", "")
        context = context or {}
        snapshot = context.get("workflow_snapshot", {}) if isinstance(context, dict) else {}
        db_path = str(snapshot.get("db_path", "")).strip()
        if not db_path:
            # 兼容非后台/旧入口调用；后台线程应优先走 snapshot。
            db_path = self.get_workflow_db_path(context)
        workflow_name = str(snapshot.get("workflow_name", "")).strip()
        if not workflow_name:
            workflow_name = self.get_workflow_output_table(context)
        app_dir = snapshot.get("app_dir") or getattr(self.app, "app_dir", get_app_dir())
        context = context or {}
        node_name = config.get("name") or config.get("node_name") or "插件节点"
        progress_callback = context.get("progress_callback")
        cancel_event = context.get("cancel_event")

        def report_progress(current=None, total=None, message="", **extra):
            """给外部插件使用的轻量进度上报函数。

            插件可以调用：
                context["report_progress"](i, total, "正在处理 xxx")
            主界面会把它显示到“当前节点进度条”。
            """
            if not callable(progress_callback):
                return
            msg = {
                "type": "node_progress",
                "node_name": node_name,
                "plugin_id": plugin_id,
                "current": current,
                "total": total,
                "message": message or "插件处理中",
            }
            msg.update(extra)
            try:
                progress_callback(msg)
            except Exception:
                pass

        return {
            "app_dir": app_dir,
            "db_path": db_path,
            "db": self.get_table_manager(context, node_type="插件节点", node_name=node_name),
            "plugins_dir": self.get_plugins_dir(),
            "plugin_data_dir": self.get_plugin_data_dir(plugin_id),
            "log_dir": self.get_plugin_log_dir(),
            "is_preview": not bool(execute_actions),
            "execute_actions": bool(execute_actions),
            "is_config_probe": bool(context.get("is_config_probe")),
            "workflow_name": workflow_name,
            "node_name": node_name,
            "plugin_id": plugin_id,
            "transit_tables": context.get("transit_tables", {}),
            "input_tables": context.get("input_tables", {}),
            "plugin_input_table_specs": copy.deepcopy(config.get("input_tables", [])),
            # 后台进度 / 取消透传给插件。
            # progress_callback 是底层消息通道；report_progress 是推荐给插件使用的轻量封装。
            "progress_callback": progress_callback,
            "report_progress": report_progress,
            "cancel_event": cancel_event,
        }

    def is_plugin_config_probe(self, context=None, execute_actions=False):
        """配置界面字段探测：只推断字段，不真实执行插件。"""
        return bool((context or {}).get("is_config_probe")) and not bool(execute_actions)

    def build_plugin_probe_input_tables(self, config, current_headers, context=None):
        """构建仅含字段的插件输入表，避免配置阶段加载整表或触发重节点。"""
        table_headers = self.build_plugin_input_table_headers(config, current_headers, context or {})
        tables = {}
        for alias, headers in (table_headers or {}).items():
            tables[alias] = {
                "type": "table",
                "headers": list(headers or []),
                "rows": [],
                "source_name": alias,
                "meta": {"lazy_schema": True, "source_type": "config_probe"},
            }
        primary = tables.get("当前表") or {
            "type": "table",
            "headers": list(current_headers or []),
            "rows": [],
            "source_name": "workflow_current",
            "meta": {"lazy_schema": True, "source_type": "config_probe"},
        }
        tables.setdefault("当前表", primary)
        tables.setdefault("workflow_current", primary)
        tables.setdefault("primary", primary)
        return tables

    def normalize_plugin_output_schema(self, schema, fallback_headers=None):
        return workflow_normalize_plugin_output_schema(schema, fallback_headers=fallback_headers)

    def get_plugin_output_schema_table(self, item, input_data, params, plugin_context, fallback_headers=None):
        return workflow_get_plugin_output_schema_table(item, input_data, params, plugin_context, fallback_headers=fallback_headers)

    def apply_lazy_plugin_probe_node(self, headers, rows, config, item, params, runtime_context):
        """配置阶段插件懒加载：返回字段和空值，不调用插件 run。"""
        plugin_id = config.get("plugin_id", "")
        input_tables = self.build_plugin_probe_input_tables(config, headers, runtime_context)
        runtime_context["input_tables"] = input_tables
        runtime_context["plugin_input_table_specs"] = copy.deepcopy(config.get("input_tables", []))
        input_data = workflow_make_plugin_input_data(plugin_id, headers, [], input_tables, lazy_schema=True)
        plugin_context = self.make_plugin_context(config, runtime_context, execute_actions=False)
        schema_table = self.get_plugin_output_schema_table(item, input_data, params, plugin_context, fallback_headers=headers)
        schema_declared = schema_table is not None
        if schema_table is None:
            schema_table = {
                "type": "table",
                "headers": list(headers),
                "rows": [list(r) for r in rows],
                "meta": {"lazy_schema": True, "schema_fallback": "pass_through"},
            }

        new_headers = list(schema_table.get("headers", headers))
        new_rows = [list(r) for r in schema_table.get("rows", [])]
        output_mode = config.get("output_mode", "使用插件返回结果")
        transit_parts = self.save_plugin_result_transit_output(config, item, plugin_id, runtime_context, new_headers, new_rows, source_prefix="插件字段探测")

        final_headers, final_rows = workflow_build_plugin_probe_final_output(
            headers,
            rows,
            new_headers,
            new_rows,
            output_mode,
            schema_declared,
        )

        plugin_name = item.get("info", {}).get("name", plugin_id)
        stat = workflow_build_plugin_probe_stat(plugin_name, schema_declared, final_headers, transit_parts)
        return final_headers, final_rows, stat

    def run_plugin_node_runtime(self, headers, rows, config, item, params, runtime_context, execute_actions=False):
        plugin_id = config.get("plugin_id", "")
        input_tables = self.build_plugin_input_tables(config, headers, rows, runtime_context)
        runtime_context["input_tables"] = input_tables
        runtime_context["plugin_input_table_specs"] = copy.deepcopy(config.get("input_tables", []))
        input_data = workflow_make_plugin_input_data(plugin_id, headers, rows, input_tables)
        plugin_context = self.make_plugin_context(config, runtime_context, execute_actions=execute_actions)

        if self.is_external_plugin_mode(config, item):
            result = self.run_external_plugin_process(
                item,
                input_data,
                params,
                config,
                runtime_context,
                execute_actions=execute_actions,
            )
        else:
            module = item.get("module")
            if module is None:
                raise RuntimeError("该插件未在主程序环境中导入。请将运行环境设置为“插件独立环境”，或改用单文件内置插件。")
            validate = getattr(module, "validate_params", None)
            if callable(validate):
                ok_msg = validate(params, input_data, plugin_context)
                if isinstance(ok_msg, tuple):
                    ok, msg = ok_msg
                    if not ok:
                        raise ValueError(msg or "插件参数校验失败")
                elif ok_msg is False:
                    raise ValueError("插件参数校验失败")
            result = module.run(input_data, params, plugin_context)

        normalized_result = workflow_normalize_plugin_run_result(result, input_data, headers, rows)
        return normalized_result, plugin_context, input_data

    def apply_plugin_node(self, headers, rows, config, context=None, execute_actions=False):
        plugin_id = config.get("plugin_id", "")
        item = self.plugin_registry.get(plugin_id)
        if not item:
            raise ValueError(f"插件未加载或缺失：{plugin_id}")
        params = dict(config.get("params", {}))
        runtime_context = dict(context or {})
        if isinstance(context, dict):
            runtime_context["table_access_logs"] = context.setdefault("table_access_logs", [])
        if self.is_plugin_config_probe(runtime_context, execute_actions=execute_actions):
            return self.apply_lazy_plugin_probe_node(headers, rows, config, item, params, runtime_context)
        plugin_context = None
        failure_policy = config.get("plugin_failure_policy", "停止工作流")

        try:
            normalized_result, plugin_context, _input_data = self.run_plugin_node_runtime(
                headers,
                rows,
                config,
                item,
                params,
                runtime_context,
                execute_actions=execute_actions,
            )
            message = normalized_result["message"]
            logs = normalized_result["logs"]
            summary = normalized_result["summary"]
            new_headers = normalized_result["headers"]
            new_rows = normalized_result["rows"]
            ok = True
            error_message = ""
        except Exception as e:
            ok = False
            error_message = str(e)
            logs = [{"level": "ERROR", "message": error_message, "traceback": traceback.format_exc()}]
            message = error_message
            summary = {"ok": False, "error": error_message}
            new_headers = list(headers)
            new_rows = [list(r) for r in rows]
            if failure_policy == "停止工作流":
                log_items = self.normalize_plugin_logs(logs, plugin_id=plugin_id, node_name=config.get("name") or "插件节点")
                self.save_plugin_log_outputs(
                    plugin_id,
                    item.get("info", {}).get("name", plugin_id),
                    config,
                    log_items,
                    plugin_context=plugin_context,
                    context=runtime_context,
                    execute_actions=execute_actions,
                    include_transit=False,
                    suppress_errors=True,
                )
                raise
            else:
                new_headers, new_rows = workflow_build_plugin_failure_output(
                    plugin_id,
                    error_message,
                    traceback.format_exc(),
                    headers,
                    rows,
                    failure_policy,
                )

        log_items = self.normalize_plugin_logs(logs, plugin_id=plugin_id, node_name=config.get("name") or "插件节点")
        plugin_name = item.get("info", {}).get("name", plugin_id)
        log_saved_parts = self.save_plugin_log_outputs(
            plugin_id,
            plugin_name,
            config,
            log_items,
            plugin_context=plugin_context,
            context=context,
            execute_actions=execute_actions,
        )

        output_mode = config.get("output_mode", "使用插件返回结果")
        transit_parts = self.save_plugin_result_transit_output(config, item, plugin_id, context, new_headers, new_rows)

        final_headers, final_rows = workflow_build_plugin_final_output(
            headers,
            rows,
            new_headers,
            new_rows,
            output_mode,
        )
        stat = workflow_build_plugin_status_text(
            plugin_name,
            plugin_id,
            ok,
            failure_policy,
            message,
            summary,
            transit_parts,
            log_saved_parts,
            log_items,
        )
        return final_headers, final_rows, stat

    def default_config_for_type(self, node_type):
        table_names = []
        needs_sqlite_defaults = {"匹配值输出列名", "选定列写入指定表", "字段映射写入表"}
        if node_type in needs_sqlite_defaults:
            try:
                table_names = self.app.get_table_names()
            except Exception:
                pass
        table_columns = {}
        for table in table_names[:1]:
            try:
                table_columns[table] = self.app.get_table_columns(table)
            except Exception:
                table_columns[table] = []
        return workflow_default_config_for_type(
            node_type,
            preview_headers=self.preview_headers,
            table_names=table_names,
            table_columns=table_columns,
            app_dir=getattr(self.app, "app_dir", get_app_dir()),
        )

    def default_name_for_node(self, node_type):
        return {
            "节点组 / 子工作流": "节点组 / 子工作流",
            "循环执行起点": "循环执行起点",
            "循环判断回跳": "循环判断回跳",
            "批量替换": "批量替换",
            "数据提取": "数据提取",
            "格式规范化 / 日期时间解析": "格式规范化 / 日期时间解析",
            "新建日期时间列": "新建日期时间列",
            "新建列": "新建列",
            "合并列": "合并列",
            "批量更改列名": "批量更改列名",
            "去重 / 重复数据处理": "去重 / 重复数据处理",
            "列数字运算": "列数字运算",
            "匹配值输出列名": "匹配值输出列名",
            "复制列": "复制列",
            "复制行": "复制行",
            "删除行": "删除行",
            "填充值": "填充值",
            "序列填充": "序列填充",
            "区域填充": "区域填充",
            "行数据映射填充": "行数据映射填充",
            "保存中转数据": "保存中转数据",
            "字段映射写入表": "字段映射写入表",
            "高级筛选": "筛选数据",
            "删除列": "删除列",
            "移动列": "整理列顺序",
        }.get(node_type, node_type)

    def add_node(self):
        node_type = self.node_type_var.get()
        if node_type in getattr(self, "plugin_display_map", {}):
            plugin_id = self.plugin_display_map[node_type]
            plugin_info = self.plugin_registry.get(plugin_id, {}).get("info", {})
            node = {
                "enabled": True,
                "type": "插件节点",
                "name": plugin_info.get("name", plugin_id),
                "config": self.default_config_for_plugin(plugin_id),
            }
        else:
            node = {
                "enabled": True,
                "type": node_type,
                "name": self.default_name_for_node(node_type),
                "config": self.default_config_for_type(node_type),
            }
        self.ensure_node_identity(node)
        selected = self.node_listbox.curselection()
        insert_at = int(selected[0]) + 1 if len(selected) == 1 else len(self.nodes)
        self.nodes.insert(insert_at, node)
        self.refresh_node_list(select_index=insert_at, reveal=True)
        self.build_node_config(insert_at)
        if len(selected) == 1:
            self.status_var.set(f"已在当前节点下方插入：{node.get('name', node.get('type', '节点'))}")
        else:
            self.status_var.set(f"已追加节点：{node.get('name', node.get('type', '节点'))}")

    def delete_node(self):
        idx = self.get_selected_node_index()
        if idx is None:
            return
        del self.nodes[idx]
        self.refresh_node_list()
        self.rebuild_current_config()

    def move_node_up(self):
        idx = self.get_selected_node_index()
        if idx is None or idx <= 0:
            return
        self.nodes[idx - 1], self.nodes[idx] = self.nodes[idx], self.nodes[idx - 1]
        self.refresh_node_list(select_index=idx - 1, reveal=True)
        self.rebuild_current_config()

    def move_node_down(self):
        idx = self.get_selected_node_index()
        if idx is None or idx >= len(self.nodes) - 1:
            return
        self.nodes[idx + 1], self.nodes[idx] = self.nodes[idx], self.nodes[idx + 1]
        self.refresh_node_list(select_index=idx + 1, reveal=True)
        self.rebuild_current_config()

    def toggle_node_enabled(self):
        idx = self.get_selected_node_index()
        if idx is None:
            return
        self.nodes[idx]["enabled"] = not self.nodes[idx].get("enabled", True)
        self.refresh_node_list(select_index=idx, reveal=True)

    def copy_node(self):
        idx = self.get_selected_node_index()
        if idx is None:
            return
        import copy
        new_node = copy.deepcopy(self.nodes[idx])
        new_node["name"] = f"{new_node.get('name', new_node.get('type'))}_复制"
        self.ensure_node_tree_identity([new_node], force_new=True)
        self.nodes.insert(idx + 1, new_node)
        self.refresh_node_list(select_index=idx + 1, reveal=True)
        self.rebuild_current_config()

    def clear_nodes(self):
        if self.nodes and not messagebox.askyesno("确认", "是否清空所有计划节点？"):
            return
        self.nodes.clear()
        self.refresh_node_list()
        self.show_empty_config()

    def update_node_name(self, idx, name_var):
        if 0 <= idx < len(self.nodes):
            self.nodes[idx]["name"] = name_var.get().strip() or self.nodes[idx]["type"]
            self.refresh_node_list(select_index=idx, reveal=True)

    def make_config_preview_context(self):
        """
        配置界面专用的预运行上下文。

        用途：刷新某个节点配置时，会临时运行它前面的节点，以便拿到“到当前节点为止”的字段列表和中转副表。
        这里允许“选定列写入指定表”在配置预运行时写入【当前工作表】和【中转副表】，
        这样后续高级筛选、匹配值输出列名、插件节点等配置界面才能看到这些临时字段。

        注意：selected_columns_config_preview_only 会在该节点内部拦截 SQLite 写入，
        防止只是切换/刷新配置界面时误改真实数据库。
        """
        return {
            "transit_tables": {},
            "loop_states": {},
            "loop_results": {},
            "is_config_probe": True,
            "allow_selected_columns_write_in_preview": True,
            "selected_columns_config_preview_only": True,
        }

    def get_headers_rows_before(self, idx):
        return self.run_plan(
            stop_index=idx - 1,
            raise_error=True,
            initial_context=self.make_config_preview_context(),
        )[:2]

    def get_transit_context_before(self, idx):
        """运行到指定节点之前，取得已经保存的内存中转副表。配置界面用于列出可引用的中转表。"""
        if idx is None or idx <= 0:
            return self.make_config_preview_context()
        try:
            _, _, _, context = self.run_plan(
                stop_index=idx - 1,
                raise_error=False,
                return_context=True,
                initial_context=self.make_config_preview_context(),
            )
            return context
        except Exception:
            return self.make_config_preview_context()

    def build_node_config(self, idx):
        self.clear_config_frame()
        node = self.nodes[idx]
        config = node.setdefault("config", {})
        try:
            available_headers, available_rows = self.get_headers_rows_before(idx)
        except Exception:
            available_headers = list(self.preview_headers)
            available_rows = [list(r) for r in self.preview_rows]

        title = ttk.Frame(self.config_frame)
        title.pack(fill=tk.X)
        ttk.Label(title, text=f"节点类型：{node.get('type')}   ").pack(side=tk.LEFT)
        ttk.Label(title, text="节点名称：").pack(side=tk.LEFT)
        name_var = tk.StringVar(value=node.get("name", node.get("type", "")))
        ttk.Entry(title, textvariable=name_var, width=28).pack(side=tk.LEFT, padx=4)
        ttk.Button(title, text="更新名称", command=lambda: self.update_node_name(idx, name_var)).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(title, text="启用", variable=self.make_node_enabled_var(idx)).pack(side=tk.LEFT, padx=8)
        ttk.Button(title, text="字段权限层", command=lambda idx=idx: self.open_table_access_window(initial_index=idx)).pack(side=tk.LEFT, padx=4)

        node_type = node.get("type")
        if node_type == "节点组 / 子工作流":
            transit_context = self.get_transit_context_before(idx)
            self.build_group_node_config(config, available_headers, transit_context)
        elif node_type == "循环执行起点":
            transit_context = self.get_transit_context_before(idx)
            self.build_loop_start_config(config, available_headers, transit_context)
        elif node_type == "循环判断回跳":
            self.build_loop_judge_config(config, available_headers)
        elif node_type == "跳转锚点节点":
            self.build_jump_anchor_config(config)
        elif node_type == "无条件跳转节点":
            self.build_unconditional_jump_config(config)
        elif node_type == "条件判断节点":
            self.build_condition_check_config(config, available_headers)
        elif node_type == "条件跳转节点":
            self.build_conditional_jump_config(config)
        elif node_type == "批量替换":
            self.build_replace_config(config, available_headers)
        elif node_type == "数据提取":
            self.build_extract_config(config, available_headers)
        elif node_type == "格式规范化 / 日期时间解析":
            self.build_format_datetime_config(config, available_headers)
        elif node_type == "新建日期时间列":
            self.build_current_datetime_column_config(config, available_headers)
        elif node_type == "新建列":
            self.build_new_columns_config(config, available_headers)
        elif node_type == "合并列":
            self.build_merge_config(config, available_headers)
        elif node_type == "批量更改列名":
            self.build_rename_columns_config(config, available_headers)
        elif node_type == "去重 / 重复数据处理":
            self.build_dedupe_config(config, available_headers)
        elif node_type == "列数字运算":
            self.build_numeric_column_config(config, available_headers)
        elif node_type == "匹配值输出列名":
            transit_context = self.get_transit_context_before(idx)
            self.build_match_value_output_field_name_config(config, available_headers, transit_context)
        elif node_type == "插件节点":
            transit_context = self.get_transit_context_before(idx)
            self.build_plugin_node_config(config, available_headers, transit_context, available_rows)
        elif node_type == "复制列":
            self.build_copy_column_config(config, available_headers)
        elif node_type == "复制行":
            self.build_copy_row_config(config, available_headers)
        elif node_type == "删除行":
            self.build_delete_rows_config(config, available_headers)
        elif node_type == "填充值":
            self.build_fill_value_config(config, available_headers)
        elif node_type == "序列填充":
            self.build_sequence_fill_config(config, available_headers)
        elif node_type == "区域填充":
            self.build_area_fill_config(config, available_headers)
        elif node_type == "行数据映射填充":
            self.build_row_data_mapping_config(config, available_headers)
        elif node_type == "保存中转数据":
            self.build_save_transit_config(config, available_headers)
        elif node_type == "选定列写入指定表":
            transit_context = self.get_transit_context_before(idx)
            self.build_selected_columns_write_config(config, available_headers, idx, transit_context)
        elif node_type == "字段映射写入表":
            self.build_writeback_config(config, available_headers)
        elif node_type == "高级筛选":
            transit_context = self.get_transit_context_before(idx)
            self.build_filter_config(config, available_headers, transit_context)
        elif node_type == "删除列":
            self.build_delete_columns_config(config, available_headers)
        elif node_type == "移动列":
            self.build_move_columns_config(config, available_headers)
        elif node_type == "获取文件列表":
            self.build_file_list_config(config)
        elif node_type == "批量重命名":
            self.build_batch_rename_config(config, available_headers)
        else:
            ttk.Label(self.config_frame, text="未知节点类型。", foreground="red").pack(anchor=tk.W)

    def make_node_enabled_var(self, idx):
        var = tk.BooleanVar(value=self.nodes[idx].get("enabled", True))
        def on_change(*_):
            if 0 <= idx < len(self.nodes):
                self.nodes[idx]["enabled"] = bool(var.get())
                self.refresh_node_list(select_index=idx, reveal=True)
        var.trace_add("write", on_change)
        return var

    def add_labeled_entry(self, parent, label, value, row, col, width=20):
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky=tk.W, padx=4, pady=4)
        var = tk.StringVar(value=value)
        ttk.Entry(parent, textvariable=var, width=width).grid(row=row, column=col + 1, sticky=tk.W, padx=4, pady=4)
        return var

    def add_labeled_combo(self, parent, label, value, values, row, col, width=20, readonly=True):
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky=tk.W, padx=4, pady=4)
        var = tk.StringVar(value=value if value in values or not readonly else (values[0] if values else value))
        state = "readonly" if readonly else "normal"
        ttk.Combobox(parent, textvariable=var, values=values, width=width, state=state).grid(row=row, column=col + 1, sticky=tk.W, padx=4, pady=4)
        return var

    def add_labeled_combo_control(self, parent, label, value, values, row, col, width=20, readonly=True):
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky=tk.W, padx=4, pady=4)
        var = tk.StringVar(value=value if value in values or not readonly else (values[0] if values else value))
        state = "readonly" if readonly else "normal"
        combo = ttk.Combobox(parent, textvariable=var, values=values, width=width, state=state)
        combo.grid(row=row, column=col + 1, sticky=tk.W, padx=4, pady=4)
        return var, combo

    def refresh_combo_values(self, combo, var, values, keep_custom=True, fallback=""):
        values = [str(v) for v in (values or [])]
        current = str(var.get() or "")
        display_values = list(values)
        if current and current not in display_values and keep_custom:
            display_values = [current] + display_values
        combo.configure(values=display_values)
        if not current:
            var.set(fallback if fallback in values else (values[0] if values else fallback))
        elif current not in values and not keep_custom:
            var.set(fallback if fallback in values else (values[0] if values else fallback))

    def refresh_listbox_values(self, listbox, values, selected_values=None):
        selected_values = set(selected_values or [])
        listbox.delete(0, tk.END)
        selected_indices = []
        for i, value in enumerate(values or []):
            listbox.insert(tk.END, value)
            if value in selected_values:
                selected_indices.append(i)
        for i in selected_indices:
            listbox.selection_set(i)
        return selected_indices

    def sync_var_to_config(self, var, config, key, cast=str):
        def on_change(*_):
            try:
                config[key] = cast(var.get())
            except Exception:
                config[key] = var.get()
        var.trace_add("write", on_change)
        return var

    def sync_bool_to_config(self, var, config, key):
        def on_change(*_):
            config[key] = bool(var.get())
        var.trace_add("write", on_change)
        return var


    # ------------------------------
    # 节点组 / 子工作流
    # ------------------------------
    def build_group_node_config(self, config, headers, transit_context=None):
        return workflow_group_config_ui.build_group_node_config(self, config, headers, transit_context=transit_context)

    def merge_selected_nodes_to_group(self):
        return workflow_group_template_ui.merge_selected_nodes_to_group(
            self,
            messagebox_module=messagebox,
            simpledialog_module=simpledialog,
        )

    def expand_selected_group(self):
        return workflow_group_template_ui.expand_selected_group(self, messagebox_module=messagebox)

    def get_group_dir(self):
        return workflow_group_template_ui.get_group_dir(self, get_app_dir)

    def validate_group_template_data(self, data):
        return workflow_group_template_ui.validate_group_template_data(data)

    def build_group_template_data(self, config, group_name=None):
        return workflow_group_template_ui.build_group_template_data(config, group_name=group_name)

    def group_config_from_template_data(self, data):
        return workflow_group_template_ui.group_config_from_template_data(data)

    def save_group_template_from_config(self, config):
        return workflow_group_template_ui.save_group_template_from_config(
            self,
            config,
            atomic_write_json,
            messagebox_module=messagebox,
            filedialog_module=filedialog,
        )

    def load_group_template_dialog(self):
        return workflow_group_template_ui.load_group_template_dialog(
            self,
            load_json_file_with_recovery,
            messagebox_module=messagebox,
            filedialog_module=filedialog,
        )

    def open_group_dir(self):
        return workflow_group_template_ui.open_group_dir(self, messagebox_module=messagebox)
    def build_loop_start_config(self, config, headers, transit_context=None):
        return workflow_build_loop_start_config_ui(self, config, headers, transit_context=transit_context)

    def build_loop_judge_config(self, config, headers):
        return workflow_build_loop_judge_config_ui(self, config, headers)

    def jump_anchor_choices(self):
        return workflow_jump_anchor_choices_ui(self.nodes)

    def anchor_id_from_choice(self, value):
        return workflow_anchor_id_from_choice_ui(value)

    def set_anchor_var_to_config(self, var, config, key):
        return workflow_set_anchor_var_to_config_ui(var, config, key)

    def build_jump_anchor_config(self, config):
        return workflow_build_jump_anchor_config_ui(self, config)

    def build_unconditional_jump_config(self, config):
        return workflow_build_unconditional_jump_config_ui(self, config)

    def build_condition_check_config(self, config, headers):
        return workflow_build_condition_check_config_ui(self, config, headers)

    def build_conditional_jump_config(self, config):
        return workflow_build_conditional_jump_config_ui(self, config)

    def get_loop_source_table_data(self, headers, rows, config, context=None):
        source_type = config.get("source_type", "当前表")
        if source_type == "当前表":
            return list(headers), [list(r) for r in rows], "当前表"
        if source_type == "SQLite表":
            table_name = config.get("source_table", "")
            if not table_name:
                raise ValueError("循环执行起点未选择 SQLite 来源表。")
            db = self.get_table_manager(context if isinstance(context, dict) else None, node_type="循环执行起点")
            data = db.read_table(table_name)
            return list(data.get("headers", [])), [list(r) for r in data.get("rows", [])], f"SQLite:{table_name}"
        if source_type == "中转副表":
            name = config.get("transit_table", "")
            manager = self.check_transit_table_permission(
                context,
                name,
                ["read_table"],
                operation="read_transit_table",
                field_action="read",
                node_type="循环执行起点",
            )
            tables = (context or {}).get("transit_tables", {})
            if name not in tables:
                raise ValueError(f"未找到中转副表：{name}")
            data = tables[name]
            source_headers = list(data.get("headers", []))
            source_rows = [list(r) for r in data.get("rows", [])]
            self.log_transit_table_event(manager, "read_transit_table", name, source_headers, source_rows, message=f"循环执行起点读取中转副表 {name}：{len(source_rows)} 行 × {len(source_headers)} 列")
            return source_headers, source_rows, f"中转:{name}"
        return list(headers), [list(r) for r in rows], "当前表"

    def loop_last_non_empty_row_index(self, headers, rows, field):
        return workflow_loop_last_non_empty_row_index(headers, rows, field)

    def init_loop_state(self, headers, rows, config, context=None):
        source_headers, source_rows, source_name = self.get_loop_source_table_data(headers, rows, config, context=context)
        return workflow_init_loop_state_from_source(source_headers, source_rows, source_name, config)

    def apply_loop_start_node(self, headers, rows, config, context=None):
        context = context if context is not None else {}
        states = context.setdefault("loop_states", {})
        loop_id = config.get("loop_id", "loop") or "loop"
        state = states.get(loop_id)
        if state is None:
            state = self.init_loop_state(headers, rows, config, context=context)
            states[loop_id] = state
        start_result = workflow_take_next_loop_item(state)
        table_name = start_result["table_name"]
        current_headers = start_result["current_headers"]
        transit_rows = start_result["transit_rows"]
        transit_tables = context.setdefault("transit_tables", {})
        manager = self.check_transit_table_write_permission(
            context,
            table_name,
            exists=table_name in transit_tables,
            write_mode="覆盖当前循环项",
            fields=current_headers,
            node_type="循环执行起点",
        )
        transit_tables[table_name] = {
            "headers": list(current_headers),
            "rows": [list(r) for r in transit_rows],
            "source": start_result["transit_source"],
        }
        if start_result.get("no_pending"):
            message = f"循环执行起点写入空当前项中转副表 {table_name}"
        else:
            message = f"循环执行起点写入当前项中转副表 {table_name}：1 行 × {len(current_headers)} 列"
        self.log_transit_table_event(manager, "write_transit_table", table_name, current_headers, transit_rows, write_mode="覆盖当前循环项", message=message)
        return workflow_build_loop_start_output(headers, rows, start_result, output_current_as_table=config.get("output_current_as_table", True))

    def evaluate_loop_condition(self, headers, rows, config, context=None, loop_state=None):
        return workflow_evaluate_loop_condition(headers, rows, config, loop_state=loop_state)

    def find_loop_start_index(self, loop_id, current_idx, nodes=None):
        node_list = nodes if nodes is not None else self.nodes
        return workflow_find_loop_start_index(loop_id, current_idx, node_list)

    def find_loop_judge_index(self, loop_id, start_idx, end_idx, nodes=None):
        node_list = nodes if nodes is not None else self.nodes
        return workflow_find_loop_judge_index(loop_id, start_idx, end_idx, node_list)

    def apply_loop_judge_node(self, headers, rows, config, context=None):
        context = context if context is not None else {}
        loop_id = config.get("loop_id", "")
        if not loop_id:
            raise ValueError("循环判断回跳节点未绑定循环执行起点。")
        state = context.setdefault("loop_states", {}).get(loop_id)
        if not state:
            raise ValueError(f"未找到循环状态：{loop_id}。请确认循环执行起点在本节点之前。")
        judge_result = workflow_apply_loop_judge_to_state(headers, rows, config, state)
        if judge_result.get("no_current"):
            return headers, rows, judge_result["stat"], judge_result["ctrl"]
        result_headers = judge_result["result_headers"]
        result_row = judge_result["result_row"]
        results = context.setdefault("loop_results", {}).setdefault(loop_id, {"headers": result_headers, "rows": []})
        results["rows"].append(result_row)
        result_name = config.get("result_table_name", "循环结果") or "循环结果"
        transit_tables = context.setdefault("transit_tables", {})
        result_rows = [list(r) for r in results["rows"]]
        result_manager = self.check_transit_table_write_permission(
            context,
            result_name,
            exists=result_name in transit_tables,
            write_mode="覆盖循环结果",
            fields=result_headers,
            node_type="循环判断回跳",
        )
        transit_tables[result_name] = {"headers": result_headers, "rows": result_rows, "source": f"循环:{loop_id}:结果"}
        self.log_transit_table_event(result_manager, "write_transit_table", result_name, result_headers, result_rows, write_mode="覆盖循环结果", message=f"循环判断回跳写入结果中转副表 {result_name}：{len(result_rows)} 行 × {len(result_headers)} 列")
        queue_name = judge_result["queue_name"]
        queue_rows = judge_result["queue_rows"]
        queue_headers = judge_result["queue_headers"]
        queue_manager = self.check_transit_table_write_permission(
            context,
            queue_name,
            exists=queue_name in transit_tables,
            write_mode="覆盖循环队列",
            fields=queue_headers,
            node_type="循环判断回跳",
        )
        transit_tables[queue_name] = {"headers": list(queue_headers), "rows": queue_rows, "source": f"循环:{loop_id}:队列"}
        self.log_transit_table_event(queue_manager, "write_transit_table", queue_name, queue_headers, queue_rows, write_mode="覆盖循环队列", message=f"循环判断回跳写入队列中转副表 {queue_name}：{len(queue_rows)} 行 × {len(queue_headers)} 列")
        return workflow_build_loop_judge_output(headers, rows, config, state, judge_result, results["rows"])

    def build_file_list_config(self, config):
        return workflow_build_file_list_config_ui(self, config)

    def build_batch_rename_config(self, config, headers):
        return workflow_build_batch_rename_config_ui(self, config, headers)

    def build_replace_config(self, config, headers):
        return workflow_build_replace_config_ui(self, config, headers)

    def build_extract_config(self, config, headers):
        return workflow_build_extract_config_ui(self, config, headers)

    def build_format_datetime_config(self, config, headers):
        return workflow_build_format_datetime_config_ui(self, config, headers)

    def build_current_datetime_column_config(self, config, headers):
        return workflow_build_current_datetime_column_config_ui(self, config, headers)

    def build_new_columns_config(self, config, headers):
        return workflow_build_new_columns_config_ui(self, config, headers)

    def build_merge_config(self, config, headers):
        return workflow_build_merge_config_ui(self, config, headers)

    def build_match_value_output_field_name_config(self, config, headers, transit_context=None):
        return workflow_build_match_value_output_field_name_config_ui(self, config, headers, transit_context)

    def build_numeric_column_config(self, config, headers):
        return workflow_build_numeric_column_config_ui(self, config, headers)

    def build_rename_columns_config(self, config, headers):
        return workflow_build_rename_columns_config_ui(self, config, headers)

    def ensure_separator_count(self, config):
        return workflow_ensure_separator_count_ui(config)

    def parse_separator_text(self, text):
        return workflow_parse_separator_text(text)

    def separator_to_input_text(self, text):
        return workflow_separator_to_input_text(text)

    def sep_value_to_display(self, sep):
        return workflow_sep_value_to_display(sep, self.SEPARATOR_OPTIONS)

    def display_to_sep_value(self, display, custom):
        return workflow_display_to_sep_value(display, custom)

    def preview_plan_separator(self, parent, left_name, right_name, combo_var, custom_var):
        return workflow_preview_plan_separator_ui(self, left_name, right_name, combo_var, custom_var)

    def refresh_merge_separator_ui(self, parent, config):
        return workflow_refresh_merge_separator_ui(self, parent, config)


    def build_row_data_mapping_config(self, config, headers):
        return workflow_build_row_data_mapping_config_ui(self, config, headers)

    def build_save_transit_config(self, config, headers):
        return workflow_build_save_transit_config_ui(self, config, headers)



    # ------------------------------
    # 选定列写入指定表
    # ------------------------------
    def build_selected_columns_write_config(self, config, headers, idx=None, transit_context=None):
        return workflow_build_selected_columns_write_config_ui(self, config, headers, idx, transit_context)

    def get_selected_columns_write_selected_fields(self, config, source_headers):
        return workflow_get_selected_columns_write_selected_fields(config, source_headers)

    def make_selected_columns_target_fields(self, config, selected_fields):
        return workflow_make_selected_columns_target_fields(config, selected_fields)

    def read_selected_columns_source_table(self, config, current_headers, current_rows, context=None):
        """读取选定列写入节点的来源表。"""
        source_type = config.get("source_type", "当前工作流表")
        context = context or {"transit_tables": {}}
        if source_type == "当前工作流表":
            return list(current_headers), [list(r) for r in self.normalize_rows(current_rows, len(current_headers))], "当前工作流表"
        if source_type == "SQLite表":
            table = str(config.get("source_sqlite_table", "")).strip()
            if not table:
                raise ValueError("请选择 SQLite 来源表。")
            data = self.get_table_manager(context, node_type="选定列写入指定表").read_table(table)
            headers = list(data.get("headers", []))
            rows = [list(row) for row in data.get("rows", [])]
            return headers, rows, f"SQLite:{table}"
        if source_type == "中转副表":
            name = str(config.get("source_transit_table", "")).strip()
            if not name:
                raise ValueError("请选择中转来源表。")
            manager = self.check_transit_table_permission(
                context,
                name,
                ["read_table"],
                operation="read_transit_table",
                field_action="read",
                node_type="选定列写入指定表",
            )
            item = (context.get("transit_tables", {}) or {}).get(name)
            if not item:
                raise ValueError(f"未找到中转来源表：{name}")
            headers = list(item.get("headers", []) or [])
            rows = [list(r) for r in (item.get("rows", []) or [])]
            self.log_transit_table_event(manager, "read_transit_table", name, headers, rows, message=f"读取中转来源表 {name}：{len(rows)} 行 × {len(headers)} 列")
            return headers, rows, f"中转:{name}"
        raise ValueError(f"未知来源类型：{source_type}")

    def read_selected_columns_target_table(self, config, context=None, current_headers=None, current_rows=None):
        """读取选定列写入节点的目标表。目标不存在时返回空表。"""
        target_type = config.get("target_type", "SQLite表")
        context = context or {"transit_tables": {}}
        if target_type == "当前工作表":
            headers = list(current_headers or [])
            rows = [list(r) for r in self.normalize_rows(current_rows or [], len(headers))]
            return headers, rows, "当前工作表"
        if target_type == "SQLite表":
            table = str(config.get("target_table", "")).strip()
            if not table:
                raise ValueError("请输入 SQLite 目标表。")
            try:
                if not self.sqlite_table_exists_by_name(self.app.sanitize_sql_name(table, "选定列结果"), context=context):
                    return [], [], f"SQLite:{table}"
                real_table = self.app.sanitize_sql_name(table, "选定列结果")
                data = self.get_table_manager(context, node_type="选定列写入指定表").read_table(real_table)
                headers = list(data.get("headers", []))
                rows = [list(row) for row in data.get("rows", [])]
                return headers, rows, f"SQLite:{real_table}"
            except Exception:
                return [], [], f"SQLite:{table}"
        if target_type == "中转副表":
            name = str(config.get("target_transit_table", "")).strip() or "选定列结果"
            manager = self.check_transit_table_permission(
                context,
                name,
                ["read_table"],
                operation="read_transit_table",
                field_action="read",
                node_type="选定列写入指定表",
            )
            item = (context.get("transit_tables", {}) or {}).get(name)
            if not item:
                self.log_transit_table_event(manager, "read_transit_table", name, [], [], message=f"读取中转目标表 {name}：目标尚不存在")
                return [], [], f"中转:{name}"
            headers = list(item.get("headers", []) or [])
            rows = [list(r) for r in (item.get("rows", []) or [])]
            self.log_transit_table_event(manager, "read_transit_table", name, headers, rows, message=f"读取中转目标表 {name}：{len(rows)} 行 × {len(headers)} 列")
            return headers, rows, f"中转:{name}"
        raise ValueError(f"未知目标类型：{target_type}")

    def selected_columns_should_write(self, old_value, new_value, overwrite_rule):
        return workflow_selected_columns_should_write(old_value, new_value, overwrite_rule)

    def normalize_selected_columns_write_mode(self, write_mode):
        return workflow_normalize_selected_columns_write_mode(write_mode)

    def build_selected_columns_write_preview(self, config, current_headers, current_rows, context=None):
        source_headers, source_rows, source_name = self.read_selected_columns_source_table(config, current_headers, current_rows, context)
        target_headers, target_rows, target_name = self.read_selected_columns_target_table(config, context, current_headers, current_rows)
        return workflow_build_selected_columns_write_preview_rows(
            config,
            source_headers,
            source_rows,
            source_name,
            target_headers,
            target_rows,
            target_name,
        )

    def apply_selected_columns_to_memory_table(self, target_headers, target_rows, selected_target_headers, selected_rows, config):
        return workflow_apply_selected_columns_to_memory_table(
            target_headers,
            target_rows,
            selected_target_headers,
            selected_rows,
            config,
        )

    def get_selected_columns_write_payload(self, config, current_headers, current_rows, context=None):
        source_headers, source_rows, source_name = self.read_selected_columns_source_table(config, current_headers, current_rows, context)
        selected_fields, target_fields, selected_rows = workflow_build_selected_columns_write_payload(
            config,
            source_headers,
            source_rows,
        )
        return selected_fields, target_fields, selected_rows, source_name

    def apply_selected_columns_write_current_table(self, headers, rows, config, target_fields, selected_rows):
        new_headers, new_rows = self.apply_selected_columns_to_memory_table(headers, rows, target_fields, selected_rows, config)
        return new_headers, new_rows, f"已写入当前工作表：{len(new_rows)} 行 × {len(new_headers)} 列，结果继续传给后续节点"

    def apply_selected_columns_write_transit_table(self, headers, rows, config, context, target_name, target_fields, selected_rows):
        mode = self.normalize_selected_columns_write_mode(config.get("write_mode", "局部覆盖，保留目标原行数"))
        exists_before = target_name in context["transit_tables"]
        manager = self.check_transit_table_write_permission(
            context,
            target_name,
            exists=exists_before,
            write_mode=mode,
            fields=target_fields,
            partial=mode in ("局部覆盖，保留目标原行数", "清空目标字段后覆盖，保留目标原行数"),
            node_type="选定列写入指定表",
        )
        old = context["transit_tables"].get(target_name, {}) or {}
        old_headers = list(old.get("headers", []) or [])
        old_rows = [list(r) for r in (old.get("rows", []) or [])]
        new_headers, new_rows = self.apply_selected_columns_to_memory_table(old_headers, old_rows, target_fields, selected_rows, config)
        context["transit_tables"][target_name] = {
            "headers": new_headers,
            "rows": [list(r) for r in new_rows],
            "source": "选定列写入指定表"
        }
        self.log_transit_table_event(
            manager,
            "write_transit_table",
            target_name,
            new_headers,
            new_rows,
            write_mode=mode,
            message=f"写入中转副表 {target_name}：{len(new_rows)} 行 × {len(new_headers)} 列，模式 {mode}",
        )
        return headers, rows, f"已写入中转副表：{target_name}（{len(new_rows)} 行 × {len(new_headers)} 列），主流程数据透传"

    def apply_selected_columns_write_sqlite_table(self, headers, rows, config, context, target_name, target_fields, selected_rows):
        sqlite_name = self.app.sanitize_sql_name(target_name, "选定列结果")
        mode = self.normalize_selected_columns_write_mode(config.get("write_mode", "复制列到目标表新建字段"))
        if mode == "覆盖重建目标表":
            saved = self.save_result_to_sqlite(
                target_fields,
                selected_rows,
                sqlite_name,
                overwrite=True,
                backup=bool(config.get("backup_before_write", True)),
                context=context,
            )
            return headers, rows, f"已覆盖重建 SQLite 表：{saved}（{len(selected_rows)} 行 × {len(target_fields)} 列），主流程数据透传"
        target_headers, target_rows, _target_label = self.read_selected_columns_target_table(
            {**config, "target_type": "SQLite表", "target_table": sqlite_name},
            context,
        )
        new_headers, new_rows = self.apply_selected_columns_to_memory_table(target_headers, target_rows, target_fields, selected_rows, config)
        saved = self.save_result_to_sqlite(
            new_headers,
            new_rows,
            sqlite_name,
            overwrite=True,
            backup=bool(config.get("backup_before_write", True)),
            context=context,
        )
        return headers, rows, f"已复制选定列到 SQLite 表字段：{saved}（{len(new_rows)} 行 × {len(new_headers)} 列），主流程数据透传"

    def apply_selected_columns_write_node(self, headers, rows, config, context=None, execute_actions=False):
        """执行“选定列写入指定表”。

        默认透传当前数据；勾选 enable_write 后：
        - 执行计划会写入目标表；
        - 预览完整计划/预览到当前节点时，只有本节点允许预览写入，避免误触发其他副作用节点；
        - 目标类型为“当前工作表”时，写入结果会作为后续节点输入。
        """
        context = context if context is not None else {"transit_tables": {}}
        context.setdefault("transit_tables", {})
        selected_fields, target_fields, selected_rows, source_name = self.get_selected_columns_write_payload(config, headers, rows, context)
        target_type, target_name = workflow_resolve_selected_columns_write_target(config)
        allow_preview_write = bool(context.get("allow_selected_columns_write_in_preview", False))
        # 配置界面刷新/切换节点时也会临时运行前置节点。
        # 这个场景只允许生成当前工作表字段和内存中转副表，严禁写真实 SQLite，避免误改数据库。
        config_preview_only = bool(context.get("selected_columns_config_preview_only", False))
        skip_stat = workflow_get_selected_columns_write_skip_stat(
            config,
            source_name,
            selected_fields,
            selected_rows,
            execute_actions=execute_actions,
            allow_preview_write=allow_preview_write,
            config_preview_only=config_preview_only,
        )
        if skip_stat:
            return headers, rows, skip_stat

        if target_type == "当前工作表":
            return self.apply_selected_columns_write_current_table(headers, rows, config, target_fields, selected_rows)

        if target_type == "中转副表":
            return self.apply_selected_columns_write_transit_table(
                headers,
                rows,
                config,
                context,
                target_name,
                target_fields,
                selected_rows,
            )

        if target_type == "SQLite表":
            return self.apply_selected_columns_write_sqlite_table(
                headers,
                rows,
                config,
                context,
                target_name,
                target_fields,
                selected_rows,
            )

        raise ValueError(f"未知目标类型：{target_type}")

    def build_writeback_config(self, config, headers):
        return workflow_build_writeback_config_ui(self, config, headers)

    def build_filter_header_risk_section(self, frame, start_row=0):
        return workflow_build_filter_header_risk_section_ui(self, frame, start_row=start_row)

    def build_filter_source_table_section(self, frame, config, headers, selected_tables, transit_context, sync_extra_tables, start_row=2):
        return workflow_build_filter_source_table_section_ui(
            self,
            frame,
            config,
            headers,
            selected_tables,
            transit_context,
            sync_extra_tables,
            start_row=start_row,
        )

    def build_filter_condition_section(self, frame, config, all_fields, start_row=3):
        return workflow_build_filter_condition_section_ui(self, frame, config, all_fields, start_row=start_row)

    def build_filter_join_section(self, frame, config, all_fields, current_fields, start_row=4):
        return workflow_build_filter_join_section_ui(self, frame, config, all_fields, current_fields, start_row=start_row)

    def build_filter_output_section(self, frame, config, all_fields, start_row=5):
        return workflow_build_filter_output_section_ui(self, frame, config, all_fields, start_row=start_row)

    def refresh_filter_risk_text(self, headers, config, risk_var, risk_label):
        return workflow_refresh_filter_risk_text_ui(self, headers, config, risk_var, risk_label)

    def refresh_filter_condition_value_input(self, field_state, value_source_var, value_var, value_combo):
        return workflow_refresh_filter_condition_value_input_ui(field_state, value_source_var, value_var, value_combo)

    def filter_tree_rows(self, tree):
        return workflow_filter_tree_rows_ui(tree)

    def replace_filter_tree_rows(self, tree, rows):
        return workflow_replace_filter_tree_rows_ui(tree, rows)

    def edit_filter_condition_cell(self, event, cond_tree, cond_edit_mode, sync_conditions_from_tree):
        return workflow_edit_filter_condition_cell_ui(event, cond_tree, cond_edit_mode, sync_conditions_from_tree)

    def build_filter_condition_action_buttons(self, condition_section, config, refresh_filter_risk_text):
        return workflow_build_filter_condition_action_buttons_ui(self, condition_section, config, refresh_filter_risk_text)

    def build_filter_join_action_buttons(self, join_section, config, refresh_filter_risk_text):
        return workflow_build_filter_join_action_buttons_ui(self, join_section, config, refresh_filter_risk_text)

    def refresh_filter_actual_output_text(self, out_list, actual_output_var, headers, field_state, config):
        return workflow_refresh_filter_actual_output_text_ui(out_list, actual_output_var, headers, field_state, config)

    def sync_filter_output_fields(self, out_list, actual_output_var, headers, field_state, config):
        return workflow_sync_filter_output_fields_ui(out_list, actual_output_var, headers, field_state, config)

    def build_filter_output_action_buttons(self, output_section, config, headers, field_state):
        return workflow_build_filter_output_action_buttons_ui(self, output_section, config, headers, field_state)

    def refresh_filter_field_sources(
        self,
        headers,
        config,
        transit_context,
        field_state,
        source_section,
        condition_section,
        join_section,
        output_section,
        sync_output_fields,
        refresh_condition_value_input,
        refresh_filter_risk_text,
    ):
        return workflow_refresh_filter_field_sources_ui(
            self,
            headers,
            config,
            transit_context,
            field_state,
            source_section,
            condition_section,
            join_section,
            output_section,
            sync_output_fields,
            refresh_condition_value_input,
            refresh_filter_risk_text,
        )
    def build_filter_config(self, config, headers, transit_context=None):
        return workflow_build_filter_config_ui(self, config, headers, transit_context=transit_context)

    def select_all_output_fields(self, listbox, config):
        return workflow_select_all_output_fields_ui(listbox, config)

    def invert_output_fields(self, listbox, config):
        return workflow_invert_output_fields_ui(listbox, config)

    def select_current_table_output_fields(self, listbox, config):
        return workflow_select_current_table_output_fields_ui(listbox, config)

    def build_copy_column_config(self, config, headers):
        return workflow_build_copy_column_config_ui(self, config, headers)

    def build_copy_row_config(self, config, headers):
        return workflow_build_copy_row_config_ui(self, config, headers)

    def build_delete_rows_config(self, config, headers):
        return workflow_build_delete_rows_config_ui(self, config, headers)

    def build_fill_value_config(self, config, headers):
        return workflow_build_fill_value_config_ui(self, config, headers)

    def build_sequence_fill_config(self, config, headers):
        return workflow_build_sequence_fill_config_ui(self, config, headers)

    def build_area_fill_config(self, config, headers):
        return workflow_build_area_fill_config_ui(self, config, headers)


    def build_dedupe_config(self, config, headers):
        return workflow_build_dedupe_config_ui(self, config, headers)

    def build_delete_columns_config(self, config, headers):
        return workflow_build_delete_columns_config_ui(self, config, headers)

    def build_move_columns_config(self, config, headers):
        return workflow_build_move_columns_config_ui(self, config, headers)

    def toggle_preview_edit_mode(self):
        self.preview_edit_mode = not self.preview_edit_mode
        if self.preview_edit_mode:
            self.preview_edit_btn_text.set("修改模式:开")
            self.status_var.set("计划预览修改模式已开启：双击结果预览表格中的单元格即可修改当前预览数据。")
        else:
            self.preview_edit_btn_text.set("修改模式:关")
            self.status_var.set("计划预览修改模式已关闭。")
            if self.preview_edit_entry is not None:
                self.preview_edit_entry.destroy()
                self.preview_edit_entry = None

    def on_preview_tree_double_click(self, event):
        if not self.preview_edit_mode:
            return
        region = self.preview_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.preview_tree.identify_row(event.y)
        col_id = self.preview_tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        try:
            col_index = int(col_id.replace("#", "")) - 1
            row_index = self.preview_tree.index(row_id)
        except Exception:
            return
        if row_index < 0 or row_index >= len(self.preview_rows):
            return
        if col_index < 0 or col_index >= len(self.preview_headers):
            return
        bbox = self.preview_tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, width, height = bbox
        old_value = ""
        if col_index < len(self.preview_rows[row_index]):
            old_value = self.preview_rows[row_index][col_index]
        if self.preview_edit_entry is not None:
            self.preview_edit_entry.destroy()
            self.preview_edit_entry = None
        entry = ttk.Entry(self.preview_tree)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, old_value)
        entry.select_range(0, tk.END)
        entry.focus()
        closed = {"done": False}

        def close_editor(save=True):
            if closed["done"]:
                return
            closed["done"] = True
            if save:
                new_value = entry.get()
                while len(self.preview_rows[row_index]) < len(self.preview_headers):
                    self.preview_rows[row_index].append("")
                self.preview_rows[row_index][col_index] = new_value
                values = list(self.preview_tree.item(row_id, "values"))
                while len(values) < len(self.preview_headers):
                    values.append("")
                values[col_index] = new_value
                self.preview_tree.item(row_id, values=values)
                if getattr(self, "preview_view_kind", "preview") == "preview":
                    self.plan_preview_headers = list(self.preview_headers)
                    self.plan_preview_rows = [list(r) for r in self.preview_rows]
                self.preview_dirty = True
                self.status_var.set(f"已修改计划预览：第 {row_index + 1} 行，第 {col_index + 1} 列。")
            entry.destroy()
            self.preview_edit_entry = None

        entry.bind("<Return>", lambda e: close_editor(save=True))
        entry.bind("<FocusOut>", lambda e: close_editor(save=True))
        entry.bind("<Escape>", lambda e: close_editor(save=False))
        self.preview_edit_entry = entry

    def refresh_preview_tree(self, headers, rows, limit=1000):
        if self.preview_edit_entry is not None:
            self.preview_edit_entry.destroy()
            self.preview_edit_entry = None
        self.preview_search_matches = []
        self.preview_search_index = -1
        self.preview_dirty = False
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = headers
        for h in headers:
            self.preview_tree.heading(h, text=h)
            self.preview_tree.column(h, width=140, minwidth=80, anchor=tk.W, stretch=False)
        self.preview_tree.tag_configure("search_match", background="#fff7cc")
        self.preview_tree.tag_configure("search_current", background="#ffd580")
        for row in rows[:limit]:
            fixed = list(row)
            if len(fixed) < len(headers):
                fixed += [""] * (len(headers) - len(fixed))
            if len(fixed) > len(headers):
                fixed = fixed[:len(headers)]
            self.preview_tree.insert("", tk.END, values=fixed)
        try:
            self.refresh_preview_table_choices(show_status=False)
        except Exception:
            pass

    def clear_preview_search_marks(self):
        for iid in self.preview_tree.get_children():
            self.preview_tree.item(iid, tags=())
        self.preview_search_matches = []
        self.preview_search_index = -1

    def search_preview_table(self, reset=True):
        keyword = self.preview_search_var.get().strip()
        if not keyword:
            messagebox.showwarning("提示", "请输入搜索关键词。")
            return

        keyword_lower = keyword.lower()
        self.clear_preview_search_marks()

        for iid in self.preview_tree.get_children():
            values = self.preview_tree.item(iid, "values")
            row_text = "\t".join(str(v) for v in values)
            if keyword_lower in row_text.lower():
                self.preview_search_matches.append(iid)
                self.preview_tree.item(iid, tags=("search_match",))

        if not self.preview_search_matches:
            self.status_var.set(f"搜索完成：未找到包含『{keyword}』的结果预览行。")
            return

        self.preview_search_index = 0 if reset else max(self.preview_search_index, 0)
        self.goto_preview_search_result()
        self.status_var.set(f"搜索完成：找到 {len(self.preview_search_matches)} 行匹配『{keyword}』。")

    def goto_preview_search_result(self):
        if not self.preview_search_matches:
            return
        self.preview_search_index %= len(self.preview_search_matches)
        current_iid = self.preview_search_matches[self.preview_search_index]
        for iid in self.preview_search_matches:
            self.preview_tree.item(iid, tags=("search_match",))
        self.preview_tree.item(current_iid, tags=("search_current",))
        self.preview_tree.selection_set(current_iid)
        self.preview_tree.focus(current_iid)
        self.preview_tree.see(current_iid)
        self.status_var.set(f"当前搜索结果：{self.preview_search_index + 1}/{len(self.preview_search_matches)}")

    def search_preview_next(self):
        if not self.preview_search_matches:
            self.search_preview_table(reset=True)
            return
        self.preview_search_index += 1
        self.goto_preview_search_result()

    def search_preview_prev(self):
        if not self.preview_search_matches:
            self.search_preview_table(reset=True)
            return
        self.preview_search_index -= 1
        self.goto_preview_search_result()

    def refresh_preview_table_choices(self, show_status=False):
        """刷新结果预览区的表格下拉菜单。

        用于快速查看：当前预览结果、主界面当前表、已生成的中转副表、SQLite 数据库表。
        这样调试循环/高级筛选时，不需要临时用高级筛选节点查看表内容。
        """
        choices = []
        mapping = {}

        def add_choice(label, key):
            display = label
            # 理论上前缀已能区分；这里仍做一次重复保护。
            if display in mapping:
                n = 2
                while f"{display} ({n})" in mapping:
                    n += 1
                display = f"{display} ({n})"
            choices.append(display)
            mapping[display] = key

        add_choice("当前预览结果", ("preview", None))
        add_choice("主界面当前预览", ("main_preview", None))

        for name in sorted((self.current_transit_tables or {}).keys()):
            add_choice(f"中转:{name}", ("transit", name))

        try:
            for table in self.get_sqlite_table_names():
                add_choice(f"SQLite:{table}", ("sqlite", table))
        except Exception:
            pass

        self.preview_table_map = mapping
        if hasattr(self, "preview_table_combo"):
            self.preview_table_combo["values"] = choices
        current = self.preview_table_var.get()
        if not current or current not in mapping:
            self.preview_table_var.set("当前预览结果" if "当前预览结果" in mapping else (choices[0] if choices else ""))
        if show_status:
            self.status_var.set(f"已刷新结果预览可查看表：{len(choices)} 个。")
        return choices

    def read_sqlite_table_for_preview(self, table_name):
        """读取 SQLite 表为 headers/rows，用于结果预览区快速查看。"""
        db_path = self.get_workflow_db_path(None)
        if not db_path or not os.path.exists(db_path):
            raise ValueError("当前 SQLite 数据库路径不存在。")
        data = TableAccessManager(db_path, node_type="结果预览").read_table(table_name)
        return list(data.get("headers", [])), [list(row) for row in data.get("rows", [])]

    def load_selected_preview_table(self):
        """把下拉菜单选中的表加载到计划窗口结果预览区。"""
        self.refresh_preview_table_choices(show_status=False)
        selected = self.preview_table_var.get()
        if not selected:
            messagebox.showwarning("提示", "请先选择要查看的表。")
            return
        key = self.preview_table_map.get(selected)
        if not key:
            messagebox.showwarning("提示", "选中的表不存在或已失效，请刷新后重试。")
            return
        kind, name = key
        try:
            if kind == "preview":
                headers, rows = self.get_plan_preview_result()
                label = "当前预览结果"
            elif kind == "main_preview":
                headers = list(self.app.headers)
                rows = [list(r) for r in self.app.rows]
                label = "主界面当前预览"
            elif kind == "transit":
                item = (self.current_transit_tables or {}).get(name)
                if item is None:
                    raise ValueError(f"中转副表不存在或尚未生成：{name}")
                headers = list(item.get("headers", []))
                rows = [list(r) for r in item.get("rows", [])]
                label = f"中转:{name}"
            elif kind == "sqlite":
                headers, rows = self.read_sqlite_table_for_preview(name)
                label = f"SQLite:{name}"
            else:
                raise ValueError(f"未知表类型：{kind}")

            self.preview_view_kind = kind
            self.preview_headers, self.preview_rows = headers, rows
            self.refresh_preview_tree(headers, rows)
            self.preview_table_var.set(selected)
            self.status_var.set(f"已载入表到结果预览：{label}，{len(rows)} 行 × {len(headers)} 列。当前预览结果已独立缓存，不会被临时查看表覆盖。")
        except Exception as e:
            messagebox.showerror("载入表失败", str(e))

    def export_preview_to_xlsx(self):
        """导出结果预览区当前显示的数据，复用主界面的 xlsx 导出流程。"""
        headers = list(self.preview_headers or [])
        rows = [list(row) for row in (self.preview_rows or [])]
        if not headers:
            messagebox.showwarning("提示", "当前结果预览没有可导出的表格字段。")
            return
        table_name = self.preview_table_var.get().strip() or "计划预览结果"
        self.app.export_current_preview_to_xlsx(
            headers=headers,
            rows=rows,
            table_name=table_name,
            title="导出为 xlsx",
        )

    def format_logs(self, logs):
        if not logs:
            return ""
        last = logs[-3:]
        text = "  最近节点：" + "；".join(last)
        return text[:500]

    def reset_manual_loop_context(self):
        self.manual_loop_context = None
        self.manual_loop_headers = None
        self.manual_loop_rows = None
        self.manual_loop_start_idx = None
        self.manual_loop_judge_idx = None
        self.manual_loop_after_index = None
        self.manual_loop_logs = []
        self.status_var.set("已重置单步循环缓存。后续预览将重新从计划开头执行。")

    def execute_loop_once_from_selected_judge(self):
        idx = self.get_selected_node_index()
        if idx is None:
            messagebox.showwarning("提示", "请先选择一个【循环判断回跳】节点。")
            return
        node = self.nodes[idx]
        if node.get("type") != "循环判断回跳":
            messagebox.showwarning("提示", "请先选中【循环判断回跳】节点，再点击执行循环一次。")
            return
        loop_id = node.get("config", {}).get("loop_id", "")
        if not loop_id:
            messagebox.showwarning("提示", "当前循环判断节点没有绑定循环名称。")
            return
        start_idx = self.find_loop_start_index(loop_id, idx)
        if start_idx is None:
            messagebox.showerror("循环错误", f"未找到对应循环执行起点：{loop_id}")
            return
        try:
            # 首次单步，先执行循环起点之前的节点，得到进入循环前的表与上下文。
            # 若继续点击同一个循环判断节点，则复用上一次 context 中的循环队列表和中转结果。
            if (self.manual_loop_context is None or
                self.manual_loop_start_idx != start_idx or
                self.manual_loop_judge_idx != idx):
                if start_idx > 0:
                    base_headers, base_rows, base_logs, base_context = self.run_plan(stop_index=start_idx - 1, raise_error=True, return_context=True)
                else:
                    base_headers = list(self.app.headers)
                    base_rows = [list(r) for r in self.app.rows]
                    base_logs = []
                    base_context = {"transit_tables": {}, "loop_states": {}, "loop_results": {}}
                self.manual_loop_headers = base_headers
                self.manual_loop_rows = base_rows
                self.manual_loop_context = base_context
                self.manual_loop_start_idx = start_idx
                self.manual_loop_judge_idx = idx
                self.manual_loop_after_index = idx + 1
                self.manual_loop_logs = list(base_logs)

            headers, rows, logs, context = self.run_plan(
                start_index=start_idx,
                stop_index=idx,
                raise_error=True,
                return_context=True,
                initial_headers=self.manual_loop_headers,
                initial_rows=self.manual_loop_rows,
                initial_context=self.manual_loop_context,
                suppress_jump_at_stop=True,
            )
            self.manual_loop_headers = headers
            self.manual_loop_rows = rows
            self.manual_loop_context = context
            self.manual_loop_logs.extend(logs)
            self.current_transit_tables = context.get("transit_tables", {})

            # 若判断节点写出了循环结果中转表，则优先显示它，便于观察已执行 N 次的累计结果。
            result_name = node.get("config", {}).get("result_table_name", "循环结果") or "循环结果"
            display_headers, display_rows = headers, rows
            if result_name in self.current_transit_tables:
                item = self.current_transit_tables[result_name]
                display_headers = list(item.get("headers", headers))
                display_rows = [list(r) for r in item.get("rows", rows)]

            self.set_plan_preview_result(display_headers, display_rows, display=True)

            state = context.get("loop_states", {}).get(loop_id, {})
            done = sum(1 for r in state.get("queue_rows", []) if str(r[0]).strip() == "2")
            pending = sum(1 for r in state.get("queue_rows", []) if str(r[0]).strip() == "0")
            failed = sum(1 for r in state.get("queue_rows", []) if str(r[0]).strip() == "3")
            self.status_var.set(
                f"已执行循环一次：{loop_id}，完成 {done}，待执行 {pending}，失败 {failed}。"
                f"后续选择判断节点之后的节点预览时，会基于当前单步循环缓存继续执行。"
                + self.format_logs(logs)
            )
        except Exception as e:
            messagebox.showerror("执行循环一次失败", str(e))

    def run_plan(self, stop_index=None, raise_error=False, execute_actions=False, return_context=False,
                 start_index=0, initial_headers=None, initial_rows=None, initial_context=None, suppress_jump_at_stop=False,
                 progress_callback=None, cancel_event=None, workflow_snapshot=None):
        """执行计划。

        这里从原来的简单 for 循环升级为 PC（程序计数器）模式，用于支持
        “循环执行起点 / 循环判断回跳”这类需要跳转的节点。
        """
        initial_state = workflow_run_plan_context.build_run_plan_initial_state(
            self,
            stop_index=stop_index,
            start_index=start_index,
            initial_headers=initial_headers,
            initial_rows=initial_rows,
            initial_context=initial_context,
            progress_callback=progress_callback,
            cancel_event=cancel_event,
            workflow_snapshot=workflow_snapshot,
            normalize_policy=TableAccessManager.normalize_permission_policy,
        )
        node_list = initial_state["node_list"]
        headers = initial_state["headers"]
        rows = initial_state["rows"]
        logs = initial_state["logs"]
        context = initial_state["context"]
        end = initial_state["end"]
        pc = initial_state["pc"]
        steps = initial_state["steps"]
        max_steps = initial_state["max_steps"]
        anchors_info = initial_state["anchors_info"]

        while pc < len(node_list) and pc <= end:
            if cancel_event is not None and cancel_event.is_set():
                logs.append("用户取消后台执行，工作流已安全停止。")
                break
            steps += 1
            if steps > max_steps:
                raise RuntimeError("工作流执行步数超过安全上限，疑似循环未正确结束。")

            idx = pc
            node = node_list[idx]
            self.ensure_node_identity(node)
            self.refresh_node_table_access(node)
            if not node.get("enabled", True):
                logs.append(f"跳过 {idx+1}.{node.get('type')}")
                pc += 1
                continue

            node_type, config = workflow_run_plan_step.prepare_node_execution(
                context,
                node,
                idx,
                len(node_list),
                steps,
                progress_callback,
            )
            try:
                jump_to = None
                before_shape, current_table_manager = workflow_run_plan_step.begin_node_execution(
                    self,
                    context,
                    headers,
                    rows,
                    node_type,
                )

                headers, rows, stat, jump_to = workflow_run_plan_dispatch.dispatch_run_plan_node(
                    self,
                    headers,
                    rows,
                    node,
                    context,
                    execute_actions=execute_actions,
                    anchors_info=anchors_info,
                    node_list=node_list,
                    idx=idx,
                    end=end,
                )

                pc, should_stop = workflow_run_plan_step.finish_node_execution(
                    self,
                    logs,
                    current_table_manager,
                    before_shape,
                    idx,
                    node_type,
                    config,
                    headers,
                    rows,
                    stat,
                    jump_to,
                    end,
                    len(node_list),
                    steps,
                    progress_callback,
                    suppress_jump_at_stop=suppress_jump_at_stop,
                )
                if should_stop:
                    break

            except Exception as e:
                pc = workflow_run_plan_step.handle_node_execution_error(
                    progress_callback,
                    logs,
                    idx,
                    len(node_list),
                    node_type,
                    e,
                    raise_error=raise_error,
                )

        # 不在后台线程直接写 self.current_transit_tables；由 workflow_done 回到主线程后统一更新。
        if return_context:
            return headers, rows, logs, context
        return headers, rows, logs


    def unique_keep_order(self, values):
        return workflow_unique_keep_order(values)

    def parse_group_input_fields(self, config):
        return workflow_parse_group_input_fields(config)


    def parse_new_column_names_for_group_analysis(self, text, strip_name=True, allow_empty=False):
        """
        节点组入口自动推导辅助：解析“新建列”节点的字段定义。

        支持：
        - 字段A
        - 字段B=默认值
        - 逗号/分号/换行混合分隔
        这里只关心字段名，不关心默认值。
        """
        result = []
        for part in re.split(r"[\n,，;；]+", str(text or "")):
            item = part.strip() if strip_name else str(part)
            if not item and not allow_empty:
                continue
            if "=" in item:
                name = item.split("=", 1)[0]
            else:
                name = item
            name = name.strip() if strip_name else name
            if name or allow_empty:
                result.append(name)
        return self.unique_keep_order(result)

    def add_group_field_ref(self, target, value):
        """节点组入口自动推导辅助：把字段名/字段列表安全加入 target。"""
        if value is None:
            return
        if isinstance(value, str):
            text = value.strip()
            if text:
                target.append(text)
            return
        if isinstance(value, (list, tuple, set)):
            for item in value:
                self.add_group_field_ref(target, item)

    def add_group_field_refs_from_dict_list(self, target, items, keys):
        """从规则列表中按多个可能 key 收集字段名。"""
        if not isinstance(items, list):
            return
        for item in items:
            if not isinstance(item, dict):
                continue
            for key in keys:
                self.add_group_field_ref(target, item.get(key))

    def classify_group_filter_field_reference(self, field, extra_tables=None):
        """
        将高级筛选字段引用转换为节点组静态分析使用的字段名。

        当前表限定名去掉本轮“当前表.”前缀；副表字段保留限定名，但标记为
        external，表示它由高级筛选节点自行读取，不属于节点组入口。
        """
        text = str(field or "").strip()
        if not text:
            return "", ""
        for table in extra_tables or []:
            table_name = str(table or "").strip()
            if table_name and text.startswith(f"{table_name}."):
                return "external", text
        if text.startswith("当前表."):
            return "current", text[len("当前表."):]
        return "current", text

    def get_group_filter_external_output_fields(self, config, context=None):
        """读取无显式投影时高级筛选会输出的副表字段。"""
        fields = []
        unresolved = []
        transit_tables = (context or {}).get("transit_tables", {})
        for table in list((config or {}).get("extra_tables", []) or []):
            table_name = str(table or "").strip()
            if not table_name:
                continue
            try:
                if table_name.startswith("中转:"):
                    transit_name = table_name.split(":", 1)[1]
                    item = transit_tables.get(transit_name)
                    if not isinstance(item, dict):
                        raise ValueError("中转副表尚未生成")
                    columns = list(item.get("headers", []) or [])
                else:
                    columns = list(self.get_workflow_sqlite_columns(table_name, context))
                fields.extend(f"{table_name}.{column}" for column in columns)
            except Exception as exc:
                unresolved.append(f"{table_name}（{exc}）")
        return self.unique_keep_order(fields), unresolved

    def analyze_group_filter_field_io(self, config, context=None):
        """专门分析节点组内高级筛选的条件、匹配规则和投影字段。"""
        cfg = config or {}
        extra_tables = list(cfg.get("extra_tables", []) or [])
        reads = []
        writes = []
        write_prefixes = []

        def add_current_read(field):
            owner, name = self.classify_group_filter_field_reference(field, extra_tables)
            if owner == "current":
                self.add_group_field_ref(reads, name)

        def add_output(field):
            owner, name = self.classify_group_filter_field_reference(field, extra_tables)
            if not name:
                return
            if owner == "current":
                self.add_group_field_ref(reads, name)
            self.add_group_field_ref(writes, name)

        for cond in cfg.get("conditions", []) or []:
            if not isinstance(cond, dict):
                continue
            add_current_read(cond.get("field"))
            if self.normalize_filter_condition_value_source(cond) == "字段值":
                add_current_read(cond.get("value"))

        for rule in cfg.get("join_rules", []) or []:
            if not isinstance(rule, dict):
                continue
            add_current_read(rule.get("left"))
            add_current_read(rule.get("right"))

        for field in cfg.get("output_fields", []) or []:
            add_output(field)

        note = "当前表字段作为组内输入；副表字段由高级筛选自行读取"
        if cfg.get("output_fields"):
            note += "；显式输出字段参与后续节点推导"
        else:
            external_fields, unresolved = self.get_group_filter_external_output_fields(
                cfg,
                context=context,
            )
            writes.extend(external_fields)
            write_prefixes.extend(
                f"{str(table).strip()}."
                for table in extra_tables
                if str(table).strip()
            )
            note += f"；未指定输出字段，已推导副表输出 {len(external_fields)} 个字段"
            if unresolved:
                note += "；结构未解析：" + "、".join(unresolved)
        return {
            "read_fields": self.unique_keep_order(reads),
            "write_fields": self.unique_keep_order(writes),
            "write_field_prefixes": self.unique_keep_order(write_prefixes),
            "note": note,
        }

    def analyze_group_inner_node_field_io(self, node, context=None):
        """
        分析组内单个节点的字段输入/输出。

        返回：
        {
            "read_fields": [...],     # 节点需要读取的字段
            "write_fields": [...],    # 节点会生成/覆盖的字段
            "note": "..."            # 简短说明，用于推导明细
        }

        规则目标不是做到 100% 静态分析，而是覆盖当前工作流里最常见的列处理节点：
        数据提取、格式规范化、新建列、合并列、复制列、填充值、高级筛选等。
        对复杂节点采用“保守读取”：宁可把可能读取的字段列入入口候选，也避免漏掉真正需要外部传入的字段。
        """
        node_type = node.get("type", "")
        cfg = node.get("config", {}) or {}
        reads = []
        writes = []
        note = ""

        if node_type == "批量替换":
            self.add_group_field_ref(reads, cfg.get("target_field"))
            legacy_source = cfg.get("value_source", "手动输入")
            match_source = cfg.get("match_value_source") or legacy_source
            replace_source = cfg.get("replace_value_source") or legacy_source
            if match_source == "列字段":
                self.add_group_field_ref(reads, cfg.get("match_value_field"))
            if replace_source == "列字段":
                self.add_group_field_ref(reads, cfg.get("replace_value_field"))
            self.add_group_field_ref(writes, cfg.get("target_field"))
            note = "读取目标字段及匹配/替换来源字段，覆盖目标字段"

        elif node_type == "数据提取":
            src = cfg.get("source_field")
            self.add_group_field_ref(reads, src)
            if cfg.get("output_mode") == "覆盖源字段":
                self.add_group_field_ref(writes, src)
            else:
                self.add_group_field_ref(writes, cfg.get("new_field"))
            note = "source_field 为输入；新字段/覆盖字段为输出"

        elif node_type == "格式规范化 / 日期时间解析":
            self.add_group_field_ref(reads, cfg.get("source_field"))
            if cfg.get("use_separate_time_field"):
                self.add_group_field_ref(reads, cfg.get("time_source_field"))
            mode = cfg.get("output_mode", "生成新字段")
            parse_type = cfg.get("parse_type", "日期")
            if mode == "覆盖源字段":
                self.add_group_field_ref(writes, cfg.get("source_field"))
            elif mode == "生成多个字段":
                prefix = str(cfg.get("component_prefix") or "解析").strip() or "解析"
                if parse_type in ("日期", "日期时间"):
                    writes.extend([f"{prefix}年", f"{prefix}月", f"{prefix}日"])
                if parse_type in ("时间", "日期时间"):
                    writes.extend([f"{prefix}时", f"{prefix}分", f"{prefix}秒"])
                self.add_group_field_ref(writes, cfg.get("new_field"))
            else:
                self.add_group_field_ref(writes, cfg.get("new_field"))
            if cfg.get("output_status"):
                self.add_group_field_ref(writes, cfg.get("status_field"))
            note = "日期/时间源字段为输入；标准字段/组件/状态为输出"

        elif node_type == "新建日期时间列":
            if cfg.get("output_mode") == "覆盖已有字段":
                self.add_group_field_ref(writes, cfg.get("target_field"))
            else:
                self.add_group_field_ref(writes, cfg.get("new_field"))
            note = "不读取外部字段，只生成日期时间字段"

        elif node_type == "新建列":
            writes.extend(self.parse_new_column_names_for_group_analysis(
                cfg.get("columns_text", ""),
                strip_name=bool(cfg.get("strip_column_name", True)),
                allow_empty=bool(cfg.get("allow_empty_name", False)),
            ))
            note = "不读取外部字段，只新建字段"

        elif node_type == "合并列":
            self.add_group_field_ref(reads, cfg.get("fields"))
            self.add_group_field_ref(writes, cfg.get("output_field"))
            note = "合并字段为输入；合并结果为输出"

        elif node_type == "批量更改列名":
            # 只处理手动映射：旧字段是输入，新字段是输出。
            self.add_group_field_refs_from_dict_list(reads, cfg.get("mappings"), ["old", "old_field", "source", "source_field", "from"])
            self.add_group_field_refs_from_dict_list(writes, cfg.get("mappings"), ["new", "new_field", "target", "target_field", "to"])
            self.add_group_field_ref(reads, cfg.get("scope_fields"))
            note = "按映射读取旧字段并输出新字段"

        elif node_type == "去重 / 重复数据处理":
            self.add_group_field_ref(reads, cfg.get("key_fields"))
            if cfg.get("add_marker_columns"):
                for key in ["duplicate_group_field", "duplicate_status_field", "duplicate_index_field", "duplicate_count_field", "keep_flag_field"]:
                    self.add_group_field_ref(writes, cfg.get(key))
            note = "去重键字段为输入；标记列为输出"

        elif node_type == "列数字运算":
            self.add_group_field_ref(reads, cfg.get("target_field"))
            if cfg.get("operand_source") == "另一列字段":
                self.add_group_field_ref(reads, cfg.get("operand_field"))
            self.add_group_field_ref(reads, cfg.get("reference_field"))
            if cfg.get("output_mode") == "覆盖原列":
                self.add_group_field_ref(writes, cfg.get("target_field"))
            else:
                self.add_group_field_ref(writes, cfg.get("output_field"))
            note = "目标字段/操作数字段为输入；计算结果为输出"

        elif node_type == "匹配值输出列名":
            self.add_group_field_ref(reads, cfg.get("source_field"))
            for key in ["output_field", "match_value_field", "match_row_field", "status_field"]:
                # 是否实际输出由开关控制，但加入输出集合不会影响入口推导。
                self.add_group_field_ref(writes, cfg.get(key))
            note = "source_field 为输入；匹配结果字段为输出"

        elif node_type == "复制列":
            src = cfg.get("source_field")
            self.add_group_field_ref(reads, src)
            if cfg.get("output_mode") == "覆盖已有字段":
                self.add_group_field_ref(writes, cfg.get("target_field"))
            else:
                self.add_group_field_ref(writes, cfg.get("new_field"))
            note = "源字段为输入；复制目标为输出"

        elif node_type == "删除行":
            if str(cfg.get("delete_mode", "")).startswith("按条件") or cfg.get("condition_field"):
                self.add_group_field_ref(reads, cfg.get("condition_field"))
            self.add_group_field_ref(reads, cfg.get("empty_field"))
            note = "条件/空值判断字段为输入"

        elif node_type == "填充值":
            self.add_group_field_ref(writes, cfg.get("target_field"))
            if cfg.get("value_source") != "手动输入值":
                for key in ["source_field", "source_end_field"]:
                    self.add_group_field_ref(reads, cfg.get(key))
            for key in ["end_field", "reference_field"]:
                self.add_group_field_ref(reads, cfg.get(key))
            note = "来源字段/边界字段为输入；目标字段为输出"

        elif node_type == "序列填充":
            self.add_group_field_ref(writes, cfg.get("target_field"))
            for key in ["end_field", "reference_field"]:
                self.add_group_field_ref(reads, cfg.get(key))
            note = "边界字段为输入；目标字段为输出"

        elif node_type == "区域填充":
            for key in ["start_field", "end_field"]:
                self.add_group_field_ref(writes, cfg.get(key))
            if cfg.get("value_source") != "手动输入值":
                for key in ["source_field", "source_end_field"]:
                    self.add_group_field_ref(reads, cfg.get(key))
            self.add_group_field_ref(reads, cfg.get("reference_field"))
            note = "来源/边界字段为输入；区域字段为输出"

        elif node_type == "行数据映射填充":
            self.add_group_field_ref(reads, cfg.get("value_fields"))
            self.add_group_field_ref(reads, cfg.get("keep_fields"))
            for key in ["output_value_field", "source_field_name", "original_row_field", "status_field"]:
                self.add_group_field_ref(writes, cfg.get(key))
            note = "展开取值字段/保留字段为输入；输出字段为输出"

        elif node_type == "保存中转数据":
            note = "保存当前组内表，不新增入口字段"

        elif node_type == "选定列写入指定表":
            self.add_group_field_ref(reads, cfg.get("selected_fields"))
            # field_mappings 里可能有 source/target 字段。
            self.add_group_field_refs_from_dict_list(reads, cfg.get("field_mappings"), ["source", "source_field", "源字段", "from"])
            self.add_group_field_refs_from_dict_list(writes, cfg.get("field_mappings"), ["target", "target_field", "目标字段", "to"])
            note = "选定来源字段为输入；写入目标字段为副作用输出"

        elif node_type == "字段映射写入表":
            self.add_group_field_refs_from_dict_list(reads, cfg.get("match_rules"), ["source_field", "left_field", "field", "当前表字段"])
            self.add_group_field_refs_from_dict_list(reads, cfg.get("field_mappings"), ["source_field", "source", "当前表字段", "from"])
            note = "匹配规则/字段映射中的当前表字段为输入"

        elif node_type == "高级筛选":
            return self.analyze_group_filter_field_io(cfg, context=context)

        elif node_type == "删除列":
            self.add_group_field_ref(reads, cfg.get("fields"))
            note = "待删除字段为输入"

        elif node_type == "移动列":
            self.add_group_field_ref(reads, cfg.get("order"))
            note = "列顺序字段为输入"

        elif node_type == "批量重命名":
            for key in ["path_field", "new_name_field", "new_path_field", "status_field"]:
                if key in ("status_field", "new_path_field"):
                    self.add_group_field_ref(writes, cfg.get(key))
                else:
                    self.add_group_field_ref(reads, cfg.get(key))
            note = "路径字段/新文件名字段为输入；状态字段为输出"

        elif node_type == "插件节点":
            # 插件参数没有统一字段 schema，第一版只扫描常见字段键。
            self.collect_group_fields_from_nested_config(
                reads,
                cfg,
                field_keys={"source_field", "target_field", "field", "path_field", "file_field", "input_field"},
            )
            self.collect_group_fields_from_nested_config(
                writes,
                cfg,
                field_keys={"output_field", "new_field", "status_field", "result_field"},
            )
            note = "插件节点按常见字段参数保守推导"

        else:
            # 未识别节点：保守扫描常见输入/输出键，避免完全失效。
            self.collect_group_fields_from_nested_config(
                reads,
                cfg,
                field_keys={"source_field", "target_field", "field", "fields", "key_fields", "reference_field"},
            )
            self.collect_group_fields_from_nested_config(
                writes,
                cfg,
                field_keys={"new_field", "output_field", "status_field", "target_field"},
            )
            note = "未知节点，按常见字段键保守推导"

        return {
            "read_fields": self.unique_keep_order(reads),
            "write_fields": self.unique_keep_order(writes),
            "note": note,
        }

    def collect_group_fields_from_nested_config(self, target, value, field_keys=None):
        """递归扫描复杂配置中的字段名。仅在 key 命中 field_keys 时收集。"""
        field_keys = set(field_keys or [])
        if isinstance(value, dict):
            for k, v in value.items():
                if k in field_keys:
                    self.add_group_field_ref(target, v)
                elif isinstance(v, (dict, list, tuple)):
                    self.collect_group_fields_from_nested_config(target, v, field_keys=field_keys)
        elif isinstance(value, (list, tuple)):
            for item in value:
                self.collect_group_fields_from_nested_config(target, item, field_keys=field_keys)

    def infer_group_input_fields_from_nodes(self, nodes, context=None):
        """
        从组内节点顺序自动推导“真正需要从组外传入”的入口字段。

        核心规则：
        - read_fields 中如果字段尚未由前序节点生成，则视为组入口字段。
        - write_fields 加入 produced_fields，后续节点读取它时不再作为入口。
        - 同一个字段既读又写时，如果它之前没有生成，仍然需要作为入口。
        """
        required = []
        produced = set()
        produced_prefixes = []
        details = []
        for idx, node in enumerate(nodes or [], start=1):
            if not node.get("enabled", True):
                details.append({
                    "index": idx,
                    "type": node.get("type", ""),
                    "reads": [],
                    "writes": [],
                    "write_prefixes": [],
                    "required": [],
                    "note": "节点已禁用，跳过推导",
                })
                continue
            info = self.analyze_group_inner_node_field_io(node, context=context)
            reads = info.get("read_fields", [])
            writes = info.get("write_fields", [])
            write_prefixes = info.get("write_field_prefixes", [])
            req_this = []
            for f in reads:
                if f not in produced and not any(str(f).startswith(prefix) for prefix in produced_prefixes):
                    req_this.append(f)
                    required.append(f)
            for f in writes:
                produced.add(f)
            produced_prefixes.extend(
                prefix
                for prefix in write_prefixes
                if prefix and prefix not in produced_prefixes
            )
            details.append({
                "index": idx,
                "type": node.get("type", ""),
                "reads": reads,
                "writes": writes,
                "write_prefixes": write_prefixes,
                "required": self.unique_keep_order(req_this),
                "note": info.get("note", ""),
            })
        return self.unique_keep_order(required), details

    def format_group_input_infer_details(self, inferred, details, limit=20):
        """把入口推导明细整理成弹窗文本。"""
        lines = [f"推导入口字段：{', '.join(inferred) if inferred else '无'}", ""]
        for item in details[:limit]:
            lines.append(f"{item.get('index')}. {item.get('type')}")
            lines.append(f"  读取：{', '.join(item.get('reads') or []) or '-'}")
            lines.append(f"  输出：{', '.join(item.get('writes') or []) or '-'}")
            if item.get("write_prefixes"):
                lines.append(f"  动态输出前缀：{', '.join(item.get('write_prefixes') or [])}")
            lines.append(f"  需要入口：{', '.join(item.get('required') or []) or '-'}")
            if item.get("note"):
                lines.append(f"  说明：{item.get('note')}")
        if len(details) > limit:
            lines.append(f"... 仅显示前 {limit} 个节点，共 {len(details)} 个节点。")
        return "\n".join(lines)

    def normalize_group_transit_conflict_mode(self, mode):
        return workflow_normalize_group_transit_conflict_mode(mode)

    def normalize_group_sqlite_mode(self, mode):
        return workflow_normalize_group_sqlite_mode(mode)

    def get_group_source_table_data(self, headers, rows, config, context=None):
        return workflow_group_runtime.get_group_source_table_data(self, headers, rows, config, context=context)

    def build_group_input_table(self, source_headers, source_rows, config):
        return workflow_build_group_input_table(source_headers, source_rows, config)

    def make_group_child_context(self, parent_context, config):
        return workflow_make_group_child_context(parent_context, config)

    def write_group_outputs(self, result_headers, result_rows, config, parent_context, execute_actions=False):
        return workflow_group_runtime.write_group_outputs(self, result_headers, result_rows, config, parent_context, execute_actions=execute_actions)

    def prepare_group_inner_node_execution(self, child_context, node, node_type, node_index, cur_headers):
        return workflow_group_runtime.prepare_group_inner_node_execution(self, child_context, node, node_type, node_index, cur_headers)

    def run_group_inner_nodes(self, cur_headers, cur_rows, nodes, child_context, execute_actions=False):
        return workflow_group_runtime.run_group_inner_nodes(self, cur_headers, cur_rows, nodes, child_context, execute_actions=execute_actions)

    def apply_group_node(self, headers, rows, config, execute_actions=False, context=None):
        return workflow_group_runtime.apply_group_node(self, headers, rows, config, execute_actions=execute_actions, context=context)

    def append_jump_runtime_log(self, context, event):
        return workflow_jump_runtime.append_jump_runtime_log(context, event)

    def apply_jump_anchor_node(self, headers, rows, config, context=None):
        return workflow_jump_runtime.apply_jump_anchor_node(self, headers, rows, config, context=context)

    def resolve_jump_target_control(self, anchor_id, context=None, anchors_info=None, nodes=None, source="跳转"):
        return workflow_jump_runtime.resolve_jump_target_control(
            self,
            anchor_id,
            context=context,
            anchors_info=anchors_info,
            nodes=nodes,
            source=source,
        )

    def apply_unconditional_jump_node(self, headers, rows, config, context=None, anchors_info=None, nodes=None):
        return workflow_jump_runtime.apply_unconditional_jump_node(
            self,
            headers,
            rows,
            config,
            context=context,
            anchors_info=anchors_info,
            nodes=nodes,
        )

    def condition_count_empty_cells(self, headers, rows, field):
        return workflow_jump_runtime.condition_count_empty_cells(self, headers, rows, field)

    def condition_count_contains_cells(self, headers, rows, field, value, case_sensitive=True):
        return workflow_jump_runtime.condition_count_contains_cells(self, headers, rows, field, value, case_sensitive=case_sensitive)

    def evaluate_condition_check_node(self, headers, rows, config, context=None):
        return workflow_jump_runtime.evaluate_condition_check_node(self, headers, rows, config, context=context)

    def apply_condition_check_node(self, headers, rows, config, context=None):
        return workflow_jump_runtime.apply_condition_check_node(self, headers, rows, config, context=context)

    def find_conditional_jump_target(self, flag_value, config):
        return workflow_jump_runtime.find_conditional_jump_target(flag_value, config)

    def apply_conditional_jump_node(self, headers, rows, config, context=None, anchors_info=None, nodes=None):
        return workflow_jump_runtime.apply_conditional_jump_node(
            self,
            headers,
            rows,
            config,
            context=context,
            anchors_info=anchors_info,
            nodes=nodes,
        )

    def apply_node(self, headers, rows, node, execute_actions=False, context=None):
        node_type = node.get("type")
        config = node.get("config", {})
        if node_type == "节点组 / 子工作流":
            return self.apply_group_node(headers, rows, config, execute_actions=execute_actions, context=context)
        if node_type == "循环执行起点":
            h, r, stat, _ctrl = self.apply_loop_start_node(headers, rows, config, context=context)
            return h, r, stat
        if node_type == "循环判断回跳":
            h, r, stat, _ctrl = self.apply_loop_judge_node(headers, rows, config, context=context)
            return h, r, stat
        if node_type == "获取文件列表":
            return self.apply_file_list_node(headers, rows, config, context=context)
        if node_type == "批量替换":
            return self.apply_replace_node(headers, rows, config, context=context)
        if node_type == "数据提取":
            return self.apply_extract_node(headers, rows, config)
        if node_type == "格式规范化 / 日期时间解析":
            return self.apply_format_datetime_node(headers, rows, config)
        if node_type == "新建日期时间列":
            return self.apply_current_datetime_column_node(headers, rows, config)
        if node_type == "新建列":
            return self.apply_new_columns_node(headers, rows, config)
        if node_type == "合并列":
            return self.apply_merge_node(headers, rows, config, context=context)
        if node_type == "批量更改列名":
            return self.apply_rename_columns_node(headers, rows, config)
        if node_type == "去重 / 重复数据处理":
            return self.apply_dedupe_node(headers, rows, config, context=context)
        if node_type == "列数字运算":
            return self.apply_numeric_column_node(headers, rows, config, context=context)
        if node_type == "匹配值输出列名":
            return self.apply_match_value_output_field_name_node(headers, rows, config, context=context)
        if node_type == "插件节点":
            return self.apply_plugin_node(headers, rows, config, context=context, execute_actions=execute_actions)
        if node_type == "复制列":
            return self.apply_copy_column_node(headers, rows, config)
        if node_type == "复制行":
            return self.apply_copy_row_node(headers, rows, config)
        if node_type == "删除行":
            return self.apply_delete_rows_node(headers, rows, config)
        if node_type == "填充值":
            return self.apply_fill_value_node(headers, rows, config, context=context)
        if node_type == "序列填充":
            return self.apply_sequence_fill_node(headers, rows, config, context=context)
        if node_type == "区域填充":
            return self.apply_area_fill_node(headers, rows, config, context=context)
        if node_type == "行数据映射填充":
            return self.apply_row_data_mapping_node(headers, rows, config)
        if node_type == "保存中转数据":
            return self.apply_save_transit_node(headers, rows, config, context=context, execute_actions=execute_actions)
        if node_type == "选定列写入指定表":
            return self.apply_selected_columns_write_node(headers, rows, config, context=context, execute_actions=execute_actions)
        if node_type == "字段映射写入表":
            return self.apply_writeback_node(headers, rows, config, execute_actions=execute_actions, context=context)
        if node_type == "高级筛选":
            return self.apply_filter_node(headers, rows, config, context=context)
        if node_type == "删除列":
            return self.apply_delete_columns_node(headers, rows, config)
        if node_type == "移动列":
            return self.apply_move_columns_node(headers, rows, config)
        if node_type == "批量重命名":
            return self.apply_batch_rename_node(headers, rows, config, execute_actions=execute_actions, context=context)
        raise ValueError(f"未知节点类型：{node_type}")

    def field_index(self, headers, field):
        if field not in headers:
            raise ValueError(f"字段不存在：{field}")
        return headers.index(field)

    def safe_cell(self, row, idx):
        return core_safe_cell(row, idx)

    def normalize_rows(self, rows, col_count):
        return core_normalize_rows(rows, col_count)

    def compare_values(self, text, op, value, case_sensitive=True):
        text = "" if text is None else str(text)
        value = "" if value is None else str(value)
        t = text if case_sensitive else text.lower()
        v = value if case_sensitive else value.lower()
        if op == "等于" or op == "完全相等":
            return t == v
        if op == "不等于":
            return t != v
        if op == "包含":
            return v in t
        if op == "不包含":
            return v not in t
        if op == "开头是":
            return t.startswith(v)
        if op == "结尾是":
            return t.endswith(v)
        if op == "为空":
            return text == ""
        if op == "不为空":
            return text != ""
        if op == "正则匹配":
            flags = 0 if case_sensitive else re.IGNORECASE
            return re.search(value, text, flags) is not None
        if op in ["大于", "小于", "大于等于", "小于等于"]:
            try:
                a = float(text)
                b = float(value)
            except Exception:
                return False
            if op == "大于":
                return a > b
            if op == "小于":
                return a < b
            if op == "大于等于":
                return a >= b
            if op == "小于等于":
                return a <= b
        return False

    def parse_extensions_filter(self, text_value):
        return workflow_parse_extensions_filter(text_value)

    def is_hidden_path(self, path):
        return workflow_is_hidden_path(path)

    def check_workflow_cancelled(self, context=None):
        """长循环节点内部调用：用户点击取消后，在安全检查点停止。"""
        cancel_event = (context or {}).get("cancel_event")
        if cancel_event is not None and cancel_event.is_set():
            raise RuntimeError("用户取消后台执行")

    def check_workflow_cancelled_periodically(self, context, index, interval=500):
        if index == 0 or index % max(1, int(interval)) == 0:
            self.check_workflow_cancelled(context)

    def report_workflow_node_progress(self, context=None, current=None, total=None, message="", node_name=""):
        """长循环节点内部调用：通过后台 Queue 回传节点内行级/项目级进度。"""
        callback = (context or {}).get("progress_callback")
        if not callable(callback):
            return
        try:
            callback({
                "type": "node_progress",
                "node_name": node_name or "当前节点",
                "current": current,
                "total": total,
                "message": message or "处理中",
            })
        except Exception:
            pass

    def apply_file_list_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context.setdefault("default_directory", getattr(self.app, "app_dir", get_app_dir()))
        node_context["check_cancelled"] = lambda index=None: self.check_workflow_cancelled(context)
        node_context["report_progress"] = (
            lambda current=None, total=None, message="", node_name="获取文件列表": self.report_workflow_node_progress(
                context,
                current=current,
                total=total,
                message=message,
                node_name=node_name,
            )
        )
        return workflow_apply_file_list_node(headers, rows, config, context=node_context)

    def get_or_add_column_index(self, headers, rows, column_name):
        column_name = str(column_name or "").strip()
        if not column_name:
            column_name = "结果"
        if column_name in headers:
            return headers.index(column_name), headers, rows
        headers = list(headers)
        rows = [list(row) for row in rows]
        headers.append(column_name)
        for row in rows:
            row.append("")
        return len(headers) - 1, headers, rows

    def make_numbered_path(self, path):
        return workflow_make_numbered_path(path)

    def apply_batch_rename_node(self, headers, rows, config, execute_actions=False, context=None):
        node_context = dict(context or {})
        node_context.update({
            "check_cancelled": lambda index=None: self.check_workflow_cancelled(context),
            "report_progress": lambda current=None, total=None, message="", node_name="批量重命名": self.report_workflow_node_progress(
                context,
                current=current,
                total=total,
                message=message,
                node_name=node_name,
            ),
            "path_exists": os.path.exists,
            "path_is_dir": os.path.isdir,
            "make_dirs": lambda path: os.makedirs(path, exist_ok=True),
            "rename_file": os.rename,
            "replace_file": os.replace,
            "make_numbered_path": self.make_numbered_path,
        })
        headers, rows, message = workflow_apply_batch_rename_node(
            headers,
            rows,
            config,
            execute_actions=execute_actions,
            context=node_context,
        )

        if node_context.get("batch_rename_do_rename") and bool(config.get("write_log", True)):
            log_path = config.get("log_path") or os.path.abspath("rename_log.csv")
            try:
                os.makedirs(os.path.dirname(os.path.abspath(log_path)), exist_ok=True)
                with open(log_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(BATCH_RENAME_LOG_HEADERS)
                    writer.writerows(node_context.get("batch_rename_log_rows", []))
            except Exception as e:
                return headers, rows, f"重命名完成 {node_context.get('batch_rename_changed', 0)} 项，但日志写入失败：{e}"

        return headers, rows, message

    def apply_replace_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context["check_cancelled"] = lambda index: self.check_workflow_cancelled_periodically(context, index)
        return workflow_apply_replace_node(headers, rows, config, context=node_context)

    def parse_int(self, value, name):
        return workflow_parse_int(value, name)

    def safe_int(self, value, default=0):
        try:
            return int(str(value).strip())
        except Exception:
            return default

    def apply_unmatched_extract(self, text, status, config):
        return workflow_apply_unmatched_extract(text, status, config)

    def post_extract_result(self, result, config):
        return workflow_post_extract_result(result, config)

    def extract_one_value(self, original, config):
        return workflow_extract_one_value(original, config)

    def get_unique_header(self, base_name, headers):
        name = str(base_name or "新字段").strip() or "新字段"
        if name not in headers:
            return name
        counter = 2
        while f"{name}_{counter}" in headers:
            counter += 1
        return f"{name}_{counter}"

    def normalize_datetime_source_text(self, value):
        return workflow_normalize_datetime_source_text(value)

    def parse_format_int(self, value, name, allow_zero=False):
        return workflow_parse_format_int(value, name, allow_zero=allow_zero)

    def slice_by_position(self, text, start, length, base, name):
        return workflow_slice_by_position(text, start, length, base, name)

    def complete_format_year(self, value, config):
        return workflow_complete_format_year(value, config)

    def build_date_parts(self, year, month, day, config):
        return workflow_build_date_parts(year, month, day, config)

    def build_time_parts(self, hour, minute="0", second="0"):
        return workflow_build_time_parts(hour, minute, second)

    def parse_date_fixed(self, text, config):
        return workflow_parse_date_fixed(text, config)

    def parse_time_fixed(self, text, config):
        return workflow_parse_time_fixed(text, config)

    def split_by_config_delimiter(self, text, kind, config):
        return workflow_split_by_config_delimiter(text, kind, config)

    def parse_date_delimited(self, text, config):
        return workflow_parse_date_delimited(text, config)

    def parse_time_delimited(self, text, config):
        return workflow_parse_time_delimited(text, config)

    def parse_date_auto_common(self, text, config):
        return workflow_parse_date_auto_common(text, config)

    def parse_time_auto_common(self, text, config):
        return workflow_parse_time_auto_common(text, config)

    def parse_format_datetime_value(self, date_text, time_text, config):
        return workflow_parse_format_datetime_value(date_text, time_text, config)

    def render_format_template(self, parts, template):
        return workflow_render_format_template(parts, template)

    def format_output_value(self, parts, config):
        return workflow_format_output_value(parts, config)

    def apply_unmatched_format_value(self, original, status, config):
        return workflow_apply_unmatched_format_value(original, status, config)

    def build_format_component_columns(self, parts, parse_type, prefix):
        return workflow_build_format_component_columns(parts, parse_type, prefix)

    def apply_format_datetime_node(self, headers, rows, config):
        return workflow_apply_format_datetime_node(headers, rows, config)

    def get_datetime_parse_warning(self, original, config, parts):
        return workflow_get_datetime_parse_warning(original, config, parts)

    def render_current_datetime_template(self, dt, config):
        return workflow_render_current_datetime_template(dt, config)

    def parse_new_columns_specs(self, config):
        return workflow_parse_new_columns_specs(config)

    def apply_new_columns_node(self, headers, rows, config):
        return workflow_apply_new_columns_node(headers, rows, config)

    def apply_current_datetime_column_node(self, headers, rows, config):
        return workflow_apply_current_datetime_column_node(headers, rows, config)

    def apply_extract_node(self, headers, rows, config):
        return workflow_apply_extract_node(headers, rows, config)

    def apply_merge_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context["check_cancelled"] = lambda index: self.check_workflow_cancelled_periodically(context, index)
        return workflow_apply_merge_node(headers, rows, config, context=node_context)

    def ensure_field_exists(self, headers, rows, field_name):
        return workflow_ensure_field_exists(headers, rows, field_name)

    def ensure_row_count(self, rows, row_count, col_count):
        return workflow_ensure_row_count(
            rows,
            row_count,
            col_count,
            max_expanded_rows=self.MAX_EXPANDED_ROWS,
        )

    def ensure_target_cell_limit(self, row_count, col_count):
        return workflow_ensure_target_cell_limit(
            row_count,
            col_count,
            max_target_cells=self.MAX_TARGET_CELLS,
        )

    def ensure_column_count(self, headers, rows, col_count, base_name="区域复制列"):
        return workflow_ensure_column_count(headers, rows, col_count, base_name)

    def parse_row_number(self, value, name="行号"):
        n = self.parse_int(value, name)
        if n < 1:
            raise ValueError(f"{name} 必须大于等于 1。")
        return n

    def get_config_cell_value(self, headers, rows, config, target_row_idx=None):
        return workflow_get_config_cell_value(headers, rows, config, target_row_idx=target_row_idx)

    def resolve_start_row_index_by_mode(self, headers, rows, target_field, config):
        return workflow_resolve_start_row_index_by_mode(headers, rows, target_field, config)

    def get_source_column_values_by_config(self, headers, rows, config):
        return workflow_get_source_column_values_by_config(headers, rows, config)

    def get_cycle_source_values_by_config(self, headers, rows, config, multi_field=False):
        return workflow_get_cycle_source_values_by_config(headers, rows, config, multi_field=multi_field)

    def get_source_row_multi_field_values_by_config(self, headers, rows, config):
        return workflow_get_source_row_multi_field_values_by_config(headers, rows, config)

    def get_source_area_values_by_config(self, headers, rows, config):
        return workflow_get_source_area_values_by_config(headers, rows, config)

    def resolve_sequence_count_by_source(self, headers, rows, config):
        return workflow_resolve_sequence_count_by_source(headers, rows, config)

    def row_is_empty(self, row, col_count):
        return workflow_row_is_empty(row, col_count)

    def last_non_empty_row_index_by_field(self, headers, rows, field_name):
        return workflow_last_non_empty_row_index_by_field(headers, rows, field_name)

    def resolve_area_end_row_index(self, headers, rows, config):
        return workflow_resolve_area_end_row_index(headers, rows, config)

    def get_fill_targets(self, headers, rows, target_field, start_row_value, direction, end_mode, count_value, end_row_value, end_field_value, reference_field_value="", allow_expand_rows=True, allow_expand_cols=False):
        return workflow_get_fill_targets(
            headers,
            rows,
            target_field,
            start_row_value,
            direction,
            end_mode,
            count_value,
            end_row_value,
            end_field_value,
            reference_field_value=reference_field_value,
            allow_expand_rows=allow_expand_rows,
            allow_expand_cols=allow_expand_cols,
            max_expanded_rows=self.MAX_EXPANDED_ROWS,
            max_target_cells=self.MAX_TARGET_CELLS,
        )

    def should_write_cell(self, current_value, overwrite_rule):
        return workflow_should_write_cell(current_value, overwrite_rule)

    def apply_copy_column_node(self, headers, rows, config):
        return workflow_apply_copy_column_node(headers, rows, config)

    def apply_copy_row_node(self, headers, rows, config):
        return workflow_apply_copy_row_node(headers, rows, config)

    def parse_row_spec_to_indexes(self, spec, max_rows):
        """解析 1,3,5-8 这样的行号列表，返回 0 基下标集合。"""
        return workflow_parse_row_spec_to_indexes(spec, max_rows)

    def apply_delete_rows_node(self, headers, rows, config):
        return workflow_apply_delete_rows_node(headers, rows, config)

    def apply_fill_value_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context.update({
            "check_cancelled": lambda index: self.check_workflow_cancelled_periodically(context, index),
            "max_expanded_rows": self.MAX_EXPANDED_ROWS,
            "max_target_cells": self.MAX_TARGET_CELLS,
        })
        return workflow_apply_fill_value_node(headers, rows, config, context=node_context)

    def format_sequence_value(self, value, config):
        return workflow_format_sequence_value(value, config)

    def apply_sequence_fill_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context.update({
            "check_cancelled": lambda index: self.check_workflow_cancelled_periodically(context, index),
            "max_expanded_rows": self.MAX_EXPANDED_ROWS,
            "max_target_cells": self.MAX_TARGET_CELLS,
        })
        return workflow_apply_sequence_fill_node(headers, rows, config, context=node_context)

    def apply_area_fill_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context.update({
            "check_cancelled": lambda index: self.check_workflow_cancelled_periodically(context, index),
            "max_expanded_rows": self.MAX_EXPANDED_ROWS,
            "max_target_cells": self.MAX_TARGET_CELLS,
        })
        return workflow_apply_area_fill_node(headers, rows, config, context=node_context)

    def get_positive_int(self, value, default_value):
        try:
            n = int(str(value).strip())
            return n if n > 0 else default_value
        except Exception:
            return default_value

    def get_plan_filter_available_fields(self, headers, extra_tables, context=None):
        fields = [f"当前表.{h}" for h in headers]
        transit_tables = (context or {}).get("transit_tables", {})
        for table in extra_tables:
            try:
                if str(table).startswith("中转:"):
                    name = str(table).split(":", 1)[1]
                    item = transit_tables.get(name, {})
                    for col in item.get("headers", []):
                        fields.append(f"{table}.{col}")
                else:
                    for col in self.get_workflow_sqlite_columns(table, context):
                        fields.append(f"{table}.{col}")
            except Exception:
                continue
        return fields

    def normalize_plan_filter_field_reference(self, field, headers, extra_tables=None):
        return workflow_normalize_plan_filter_field_reference(field, headers, extra_tables)

    def normalize_plan_filter_config_field_references(self, config, headers, extra_tables=None):
        return workflow_normalize_plan_filter_config_field_references(config, headers, extra_tables)

    def get_plan_filter_output_base_headers(self, lookup_fields, headers):
        return workflow_get_plan_filter_output_base_headers(lookup_fields, headers)

    def get_plan_filter_output_headers(self, lookup_fields, headers):
        return workflow_get_plan_filter_output_headers(lookup_fields, headers)

    def get_plan_filter_output_header_conflicts(self, lookup_fields, headers):
        return workflow_get_plan_filter_output_header_conflicts(lookup_fields, headers)

    def plan_filter_field_belongs_to_table(self, field, table_name):
        return workflow_plan_filter_field_belongs_to_table(field, table_name)

    def get_plan_filter_field_owner(self, field, headers, extra_tables):
        return workflow_get_plan_filter_field_owner(field, headers, extra_tables)

    def get_plan_filter_hash_join_availability(self, headers, extra_tables, join_rules, join_logic):
        return workflow_get_plan_filter_hash_join_availability(headers, extra_tables, join_rules, join_logic)

    def get_plan_filter_config_warnings(self, headers, extra_tables, conditions, join_rules, join_logic):
        return workflow_get_plan_filter_config_warnings(headers, extra_tables, conditions, join_rules, join_logic)

    def add_plan_filter_required_field(self, field, headers, extra_tables, current_headers, table_fields):
        return workflow_add_plan_filter_required_field(field, headers, extra_tables, current_headers, table_fields)

    def collect_plan_filter_required_fields(self, headers, extra_tables, conditions, join_rules, output_fields, final_fields):
        return workflow_collect_plan_filter_required_fields(
            headers, extra_tables, conditions, join_rules, output_fields, final_fields
        )

    def get_required_columns_for_plan_table(self, table_name, columns, required_fields):
        return workflow_get_required_columns_for_plan_table(table_name, columns, required_fields)

    def make_current_table_records(self, headers, rows, required_headers=None):
        return workflow_make_current_table_records(headers, rows, required_headers)

    def load_plan_table_records(self, table_name, context=None, required_fields=None):
        if str(table_name).startswith("中转:"):
            name = str(table_name).split(":", 1)[1]
            transit_tables = (context or {}).get("transit_tables", {})
            if name not in transit_tables:
                raise ValueError(f"中转副表不存在或尚未生成：{name}")
            item = transit_tables[name]
            all_columns = list(item.get("headers", []))
            columns = self.get_required_columns_for_plan_table(table_name, all_columns, required_fields)
            manager = self.check_transit_table_permission(
                context,
                table_name,
                ["read_table"],
                operation="read_transit_table",
                fields=columns,
                field_action="read",
                node_type="高级筛选",
            )
            column_indexes = [(all_columns.index(col), col) for col in columns]
            db_rows = self.normalize_rows(item.get("rows", []), len(all_columns))
            records = []
            for row in db_rows:
                record = {}
                for i, col in column_indexes:
                    record[f"{table_name}.{col}"] = self.safe_cell(row, i)
                records.append(record)
            self.log_transit_table_event(manager, "read_transit_table", table_name, columns, db_rows, message=f"高级筛选读取中转副表 {table_name}：{len(db_rows)} 行 × {len(columns)} 列")
            return records

        db_path = self.get_workflow_db_path(context)
        if not db_path or not os.path.exists(db_path):
            raise ValueError("当前 SQLite 数据库路径不存在，无法读取副表。")
        all_columns = self.get_workflow_sqlite_columns(table_name, context)
        columns = self.get_required_columns_for_plan_table(table_name, all_columns, required_fields)
        data = self.get_table_manager(context, node_type="高级筛选").read_table(table_name, fields=columns)
        db_rows = [list(row) for row in data.get("rows", [])]
        records = []
        for row in db_rows:
            record = {}
            for i, col in enumerate(columns):
                value = row[i] if i < len(row) else ""
                record[f"{table_name}.{col}"] = value
            records.append(record)
        return records

    def normalize_filter_condition_value_source(self, cond):
        return workflow_normalize_filter_condition_value_source(cond)

    def resolve_plan_condition_value(self, record, cond):
        return workflow_resolve_plan_condition_value(record, cond)

    def eval_plan_condition_record(self, record, cond):
        return workflow_eval_plan_condition_record(record, cond)

    def eval_plan_join_rule_record(self, record, rule):
        return workflow_eval_plan_join_rule_record(record, rule)

    def record_passes_plan_conditions(self, record, conditions, logic):
        return workflow_record_passes_plan_conditions(record, conditions, logic)

    def plan_filter_condition_dependencies(self, cond):
        return workflow_plan_filter_condition_dependencies(cond)

    def record_survives_available_plan_conditions(self, record, conditions, logic):
        return workflow_record_survives_available_plan_conditions(record, conditions, logic)

    def record_passes_plan_join_rules(self, record, join_rules, logic="AND"):
        return workflow_record_passes_plan_join_rules(record, join_rules, logic)

    def get_plan_filter_hash_join_rules(self, table_name, join_rules, join_logic, right_records):
        return workflow_get_plan_filter_hash_join_rules(table_name, join_rules, join_logic, right_records)

    def build_plan_filter_right_index(self, right_records, hash_rules):
        return workflow_build_plan_filter_right_index(right_records, hash_rules)

    def iter_plan_filter_join_candidates(self, left_record, right_records, hash_rules, right_index, missing_key_records):
        return workflow_iter_plan_filter_join_candidates(
            left_record, right_records, hash_rules, right_index, missing_key_records
        )


    def get_row_mapping_end_index(self, rows, start_idx, config, col_count):
        return workflow_get_row_mapping_end_index(rows, start_idx, config, col_count)

    def apply_row_data_mapping_node(self, headers, rows, config):
        return workflow_apply_row_data_mapping_node(headers, rows, config)

    def make_unique_transit_name(self, base_name, transit_tables):
        return workflow_make_unique_transit_name(base_name, transit_tables)

    def append_headers_rows(self, old_headers, old_rows, new_headers, new_rows):
        return workflow_append_headers_rows(old_headers, old_rows, new_headers, new_rows)

    def save_result_to_sqlite_append(self, headers, rows, table_name_raw, context=None):
        """追加写入 SQLite 表；表不存在则创建，字段不足则自动 ADD COLUMN。"""
        table_name = self.app.sanitize_sql_name(table_name_raw, "中转数据")
        sql_columns = self.app.make_sql_columns(headers)
        if not sql_columns:
            raise ValueError("没有可写入的字段。")
        normalized_rows = self.normalize_rows(rows, len(sql_columns))
        info = self.get_table_manager(context, node_type="保存中转数据").write_table(table_name, sql_columns, normalized_rows, mode="append")
        return info.get("table_name", table_name)

    def export_headers_rows_to_xlsx_file(self, headers, rows, path):
        """把指定 headers / rows 导出为 xlsx 文件，复用主程序现有导出逻辑。"""
        if not path:
            raise ValueError("xlsx 导出路径为空。")
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        old_headers = self.app.headers
        old_rows = self.app.rows
        old_raw = self.app.raw_data
        try:
            self.app.headers = list(headers)
            self.app.rows = [list(row) for row in rows]
            self.app.raw_data = ""
            try:
                self.app.export_xlsx_with_openpyxl(path)
            except Exception:
                self.app.export_xlsx_minimal(path)
        finally:
            self.app.headers = old_headers
            self.app.rows = old_rows
            self.app.raw_data = old_raw

    def sqlite_table_exists_by_name(self, table_name, context=None):
        db_path = self.get_workflow_db_path(context)
        if not db_path or not os.path.exists(db_path):
            return False
        try:
            return self.get_table_manager(context).table_exists(table_name)
        except Exception:
            return False

    def apply_save_transit_memory_plan(self, context, memory_plan, headers_copy, rows_copy):
        if not memory_plan:
            return
        manager = self.check_transit_table_write_permission(
            context,
            memory_plan["table_name"],
            exists=bool(memory_plan.get("exists_before")),
            write_mode=memory_plan.get("write_mode", ""),
            fields=memory_plan.get("headers", headers_copy),
            node_type="保存中转数据",
        )
        extra = {
            "write_mode": memory_plan.get("write_mode", ""),
            "message": memory_plan.get("log_message", ""),
        }
        if memory_plan.get("operation") == "append_transit_table":
            extra["appended_rows"] = memory_plan.get("appended_rows", 0)
        context["transit_tables"][memory_plan["table_name"]] = {
            "headers": list(memory_plan.get("headers", headers_copy)),
            "rows": [list(r) for r in memory_plan.get("rows", rows_copy)],
            "source": memory_plan.get("source", "保存中转数据:覆盖"),
        }
        self.log_transit_table_event(
            manager,
            memory_plan.get("operation", "write_transit_table"),
            memory_plan["table_name"],
            memory_plan.get("headers", headers_copy),
            memory_plan.get("rows", rows_copy),
            **extra,
        )

    def execute_save_transit_sqlite(self, options, headers_copy, rows_copy, context=None):
        table_raw = options.get("sqlite_table_raw", options.get("base_name", "中转数据"))
        table_name = self.app.sanitize_sql_name(table_raw, "中转数据")
        mode = options.get("sqlite_mode", "自动加时间戳")
        if mode == "覆盖同名表":
            saved_name = self.save_result_to_sqlite(headers_copy, rows_copy, table_name, overwrite=True, backup=True, context=context)
        elif mode == "追加写入":
            saved_name = self.save_result_to_sqlite_append(headers_copy, rows_copy, table_name, context=context)
        elif mode == "报错停止":
            if self.sqlite_table_exists_by_name(table_name, context=context):
                raise ValueError(f"SQLite 表已存在，按设置停止：{table_name}")
            saved_name = self.save_result_to_sqlite(headers_copy, rows_copy, table_name, overwrite=False, backup=False, context=context)
        else:
            saved_name = self.save_result_to_sqlite(headers_copy, rows_copy, table_name, overwrite=False, backup=False, context=context)
        return f"SQLite表：{saved_name}" + ("（追加写入）" if mode == "追加写入" else "")

    def execute_save_transit_xlsx(self, options, headers_copy, rows_copy):
        xlsx_path = str(options.get("xlsx_path", "")).strip()
        if not xlsx_path:
            export_dir = os.path.join(getattr(self.app, "app_dir", get_app_dir()), "export")
            xlsx_path = os.path.join(export_dir, f"{options.get('base_name', '中转数据')}.xlsx")
        self.export_headers_rows_to_xlsx_file(headers_copy, rows_copy, xlsx_path)
        return f"xlsx：{xlsx_path}"

    def apply_save_transit_node(self, headers, rows, config, context=None, execute_actions=False):
        """保存中转数据：保存当前数据副本，默认不改变主流程数据。"""
        context = context if context is not None else {"transit_tables": {}}
        context.setdefault("transit_tables", {})
        result_headers, result_rows, message = workflow_apply_save_transit_node(
            headers,
            rows,
            config,
            context=context,
            execute_actions=execute_actions,
        )
        options = context.get("save_transit_options", {}) or {}
        headers_copy = context.get("save_transit_headers")
        if headers_copy is None:
            headers_copy = list(headers)
        rows_copy = context.get("save_transit_rows")
        if rows_copy is None:
            rows_copy = [list(row) for row in self.normalize_rows(rows, len(headers_copy))]
        saved_parts = message.split("；") if message else []

        self.apply_save_transit_memory_plan(
            context,
            context.get("save_transit_memory_plan"),
            headers_copy,
            rows_copy,
        )
        if execute_actions and options.get("save_sqlite"):
            saved_parts.append(self.execute_save_transit_sqlite(options, headers_copy, rows_copy, context=context))
        if execute_actions and options.get("save_xlsx"):
            saved_parts.append(self.execute_save_transit_xlsx(options, headers_copy, rows_copy))

        if not saved_parts:
            saved_parts.append("未选择保存位置，仅透传数据")

        return result_headers, result_rows, "；".join(saved_parts)


    def compare_writeback_values(self, left, op, right):
        return workflow_compare_writeback_values(left, op, right)

    def load_target_table_rows_for_writeback(self, table_name, context=None):
        db_path = self.get_workflow_db_path(context)
        if not db_path or not os.path.exists(db_path):
            raise ValueError("SQLite 数据库路径不存在，请先选择数据库。")
        columns, records = self.get_table_manager(context, node_type="字段映射写入表").read_records(
            table_name,
            include_rowid=True,
            include_row_index=True,
        )
        return columns, records

    def backup_sqlite_table_for_writeback(self, table_name, context=None):
        return self.get_table_manager(context, node_type="字段映射写入表").backup_table(table_name)

    def apply_writeback_updates_to_sqlite(self, table_name, actions, context=None):
        db_path = self.get_workflow_db_path(context)
        if not db_path or not os.path.exists(db_path):
            raise ValueError("SQLite 数据库路径不存在，请先选择数据库。")
        return self.get_table_manager(context, node_type="字段映射写入表").apply_cell_actions(
            table_name,
            actions,
            cancel_event=(context or {}).get("cancel_event"),
        )

    def apply_writeback_transaction_to_sqlite(self, table_name, actions, target_fields, context=None):
        db_path = self.get_workflow_db_path(context)
        if not db_path or not os.path.exists(db_path):
            raise ValueError("SQLite 数据库路径不存在，请先选择数据库。")
        return self.get_table_manager(
            context,
            node_type="字段映射写入表",
        ).apply_writeback_transaction(
            table_name,
            actions,
            clear_fields=target_fields,
            cancel_event=(context or {}).get("cancel_event"),
        )

    def clear_writeback_target_fields_in_sqlite(self, table_name, target_fields, context=None):
        """清空 SQLite 目标表中指定字段的全部旧值，返回清空字段数量。"""
        fields = []
        existing = set(self.get_workflow_sqlite_columns(table_name, context))
        for field in target_fields or []:
            field = str(field or "").strip()
            if field and field in existing and field not in fields:
                fields.append(field)
        if not fields:
            return 0
        return self.get_table_manager(context, node_type="字段映射写入表").clear_fields(table_name, fields)

    def build_writeback_full_structure_rows_for_sqlite(self, headers, rows, config, target_columns):
        return workflow_build_writeback_full_structure_rows_for_sqlite(headers, rows, config, target_columns)

    def build_writeback_actions(self, headers, rows, config, context=None):
        table_name = str(config.get("target_table", "")).strip()
        if not table_name:
            raise ValueError("请选择目标表。")
        use_match_rules = bool(config.get("use_match_rules", True))
        match_rules = list(config.get("match_rules", []))
        mappings = list(config.get("field_mappings", []))
        if use_match_rules and not match_rules:
            raise ValueError("已启用匹配规则定位目标行，请至少添加一条匹配规则；如果想按行号顺序写入，请关闭该选项。")
        if not mappings:
            raise ValueError("请至少添加一条字段映射规则。")
        target_columns, target_records = self.load_target_table_rows_for_writeback(table_name, context=context)
        actions = workflow_build_writeback_actions(headers, rows, config, target_columns, target_records)
        return actions, table_name

    def apply_external_table_to_current_node(self, headers, rows, config, context=None):
        source_table = str(config.get("source_table", "")).strip()
        if not source_table:
            raise ValueError("请选择来源表。")
        use_match_rules = bool(config.get("use_match_rules", True))
        match_rules = list(config.get("match_rules", []))
        mappings = list(config.get("field_mappings", []))
        if use_match_rules and not match_rules:
            raise ValueError("已启用匹配规则定位对应行，请至少添加一条匹配规则；如果想按行号顺序写入，请关闭该选项。")
        if not mappings:
            raise ValueError("请至少添加一条字段映射规则。")
        source_columns, source_records = self.load_target_table_rows_for_writeback(source_table, context=context)
        return workflow_apply_external_table_to_current_node(headers, rows, config, source_columns, source_records)

    def apply_writeback_node(self, headers, rows, config, execute_actions=False, context=None):
        if config.get("writeback_direction", "当前表写入SQLite目标表") == "其他表写入当前表":
            return self.apply_external_table_to_current_node(headers, rows, config, context=context)

        table_name = str(config.get("target_table", "")).strip()
        if not table_name:
            raise ValueError("请选择目标表。")
        write_range_mode = config.get("write_range_mode", "局部覆盖，保留目标原行数")
        enable_write = bool(config.get("enable_write", False))
        backup_before_write = bool(config.get("backup_before_write", True))
        output_preview = bool(config.get("output_preview_table", True))

        if write_range_mode == "按来源完整结构覆盖":
            target_columns, _target_records = self.load_target_table_rows_for_writeback(table_name, context=context)
            actions, full_rows = self.build_writeback_full_structure_rows_for_sqlite(headers, rows, config, target_columns)
            stat = workflow_build_writeback_preview_stat(
                write_range_mode,
                actions,
                full_rows=full_rows,
                target_columns=target_columns,
            )
            if execute_actions and enable_write:
                saved = self.save_result_to_sqlite(target_columns, full_rows, table_name, overwrite=True, backup=backup_before_write, context=context)
                stat = workflow_build_writeback_full_structure_execute_stat(saved, full_rows, target_columns)
            else:
                stat += workflow_get_writeback_non_execute_suffix(execute_actions, enable_write)
        else:
            actions, table_name = self.build_writeback_actions(headers, rows, config, context=context)
            action_counts = workflow_count_writeback_actions(actions)
            target_fields = workflow_get_writeback_target_fields(config)
            stat = workflow_build_writeback_preview_stat(write_range_mode, actions, target_fields=target_fields)

            if workflow_should_execute_writeback_update(execute_actions, enable_write, action_counts, write_range_mode):
                backup_name = ""
                if backup_before_write:
                    backup_name = self.backup_sqlite_table_for_writeback(table_name, context=context)
                cleared = 0
                if write_range_mode == "清空目标字段后覆盖，保留目标原行数":
                    result = self.apply_writeback_transaction_to_sqlite(
                        table_name,
                        actions,
                        target_fields,
                        context=context,
                    )
                    cleared = result.get("cleared_fields", 0)
                    actual = result.get("cells", 0)
                else:
                    actual = self.apply_writeback_updates_to_sqlite(
                        table_name,
                        actions,
                        context=context,
                    ) if action_counts["write_count"] > 0 else 0
                stat = workflow_build_writeback_execute_stat(table_name, actual, cleared=cleared, backup_name=backup_name)
            else:
                stat += workflow_get_writeback_non_execute_suffix(execute_actions, enable_write)

        return workflow_finish_writeback_node_output(headers, rows, actions, stat, output_preview)

    def apply_filter_node(self, headers, rows, config, context=None):
        extra_tables = list(config.get("extra_tables", []))
        available_fields = self.get_plan_filter_available_fields(headers, extra_tables, context) if extra_tables else None
        runtime_plan = workflow_build_filter_runtime_plan(headers, config, available_fields=available_fields)

        if (context or {}).get("is_config_probe") and extra_tables:
            return workflow_build_filter_config_probe_result(runtime_plan["output_headers"])

        table_records = {}
        for table in runtime_plan["extra_tables"]:
            table_records[table] = self.load_plan_table_records(
                table,
                context=context,
                required_fields=runtime_plan["table_required"].get(table),
            )

        node_context = {
            "lookup_fields": runtime_plan["lookup_fields"],
            "output_headers": runtime_plan["output_headers"],
            "current_required": runtime_plan["current_required"],
            "table_required": runtime_plan["table_required"],
            "table_records": table_records,
        }
        return workflow_apply_filter_node(headers, rows, runtime_plan["runtime_config"], context=node_context)

    def match_value_output_column_match(self, source_value, lookup_value, mode):
        return workflow_match_value_output_column_match(source_value, lookup_value, mode)

    def load_lookup_table_for_match_value_output(self, config, context=None):
        """读取匹配值输出列名节点使用的匹配表，支持 SQLite 表与内存中转副表。"""
        lookup_source_type = str(config.get("lookup_source_type", "SQLite表")).strip() or "SQLite表"
        lookup_table = str(config.get("lookup_table", "")).strip()
        if not lookup_table:
            raise ValueError("请选择匹配表或中转副表。")
        if lookup_source_type == "中转副表":
            transit_tables = (context or {}).get("transit_tables", {})
            if lookup_table not in transit_tables:
                raise ValueError(f"中转副表不存在或尚未生成：{lookup_table}。请确认保存中转数据节点在当前节点之前执行。")
            item = transit_tables[lookup_table]
            columns = list(item.get("headers", []))
            manager = self.check_transit_table_permission(
                context,
                lookup_table,
                ["read_table"],
                operation="read_transit_table",
                fields=config.get("lookup_fields", []),
                field_action="read",
                node_type="匹配值输出列名",
            )
            raw_rows = self.normalize_rows(item.get("rows", []), len(columns))
            records = []
            for index, row in enumerate(raw_rows, start=1):
                record = {"__rowid__": "", "__row_index__": index}
                for i, col in enumerate(columns):
                    record[col] = self.safe_cell(row, i)
                records.append(record)
            self.log_transit_table_event(manager, "read_transit_table", lookup_table, columns, raw_rows, message=f"匹配值输出列名读取中转副表 {lookup_table}：{len(raw_rows)} 行 × {len(columns)} 列")
            return columns, records
        return self.load_target_table_rows_for_writeback(lookup_table, context=context)

    def apply_match_value_output_field_name_node(self, headers, rows, config, context=None):
        lookup_columns, lookup_records = self.load_lookup_table_for_match_value_output(config, context=context)
        node_context = dict(context or {})
        node_context["lookup_columns"] = lookup_columns
        node_context["lookup_records"] = lookup_records
        node_context["check_cancelled"] = lambda index: self.check_workflow_cancelled_periodically(context, index)
        return workflow_apply_match_value_output_field_name_node(headers, rows, config, context=node_context)

    def apply_rename_columns_node(self, headers, rows, config):
        return workflow_apply_rename_columns_node(headers, rows, config)

    def make_unique_plan_headers(self, headers):
        """字段名去重：重复字段自动追加 _2、_3。"""
        result = []
        counts = {}
        for i, h in enumerate(headers, start=1):
            base = str(h).strip() or f"列{i}"
            if base not in counts:
                counts[base] = 1
                result.append(base)
            else:
                counts[base] += 1
                candidate = f"{base}_{counts[base]}"
                while candidate in counts:
                    counts[base] += 1
                    candidate = f"{base}_{counts[base]}"
                counts[candidate] = 1
                result.append(candidate)
        return result


    def parse_numeric_value_for_column_op(self, value):
        return workflow_parse_numeric_value_for_column_op(value)

    def format_numeric_column_result(self, value, config):
        return workflow_format_numeric_column_result(value, config)

    def get_numeric_node_row_indexes(self, headers, rows, config):
        return workflow_get_numeric_node_row_indexes(headers, rows, config)

    def numeric_node_fallback_value(self, original_value, policy, fixed_value, fail_text):
        return workflow_numeric_node_fallback_value(original_value, policy, fixed_value, fail_text)

    def apply_numeric_column_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context["check_cancelled"] = lambda index: self.check_workflow_cancelled_periodically(context, index)
        node_context["max_expanded_rows"] = self.MAX_EXPANDED_ROWS
        return workflow_apply_numeric_column_node(headers, rows, config, context=node_context)

    def apply_dedupe_node(self, headers, rows, config, context=None):
        node_context = dict(context or {})
        node_context["check_cancelled"] = lambda index: self.check_workflow_cancelled_periodically(context, index)
        return workflow_apply_dedupe_node(headers, rows, config, context=node_context)

    def make_unique_headers_for_append(self, existing_headers, new_headers):
        """给追加字段生成不重复字段名。"""
        return core_make_unique_headers_for_append(existing_headers, new_headers)

    def apply_delete_columns_node(self, headers, rows, config):
        return workflow_apply_delete_columns_node(headers, rows, config)

    def apply_move_columns_node(self, headers, rows, config):
        return workflow_apply_move_columns_node(headers, rows, config)

    def save_result_to_sqlite(self, headers, rows, table_name_raw, overwrite=False, backup=True, context=None):
        db_path = self.get_workflow_db_path(context)
        if not db_path:
            raise ValueError("请先设置 SQLite 数据库路径。")
        table_name = self.app.sanitize_sql_name(table_name_raw, "计划结果")
        sql_columns = self.app.make_sql_columns(headers)
        if not sql_columns:
            raise ValueError("没有可写入的字段。")
        current = (context or {}).get("current_node_info", {}) if isinstance(context, dict) else {}
        if isinstance(current, dict) and current.get("node_id"):
            manager = self.get_table_manager(context, node_type="工作流输出")
        else:
            manager = self.get_workflow_output_manager(table_name, overwrite=overwrite, context=context)
        if overwrite and backup and manager.table_exists(table_name):
            manager.backup_table(table_name)
        mode = "replace" if overwrite else "timestamp"
        info = manager.write_table(table_name, sql_columns, self.normalize_rows(rows, len(sql_columns)), mode=mode)
        return info.get("table_name", table_name)

    def get_plan_dir(self):
        """返回程序真实目录下的 plan 模板目录，并确保目录存在。"""
        base_dir = getattr(self.app, "app_dir", get_app_dir())
        plan_dir = os.path.join(base_dir, "plan")
        os.makedirs(plan_dir, exist_ok=True)
        return plan_dir

    def sanitize_plan_file_name(self, name):
        """生成适合作为文件名的计划模板名称。"""
        name = str(name or "工作流计划").strip()
        name = re.sub(r'[\\/:*?"<>|]+', "_", name)
        name = re.sub(r"\s+", "_", name)
        return name or "工作流计划"

    def build_plan_template_data(self, plan_name=None):
        """
        收集当前计划模板数据。新版模板必须带 template_type。

        plan_name 优先由保存时选择的 JSON 文件名传入，
        这样模板下拉菜单中的计划名会和实际保存文件名保持一致。
        """
        plan_name = str(plan_name or "").strip()
        if not plan_name:
            plan_name = self.output_table_var.get().strip() or "工作流计划"

        self.refresh_node_tree_table_access(self.nodes)
        return {
            "template_type": "workflow_plan",
            "version": "1.0",
            "plan_name": plan_name,
            "nodes": self.nodes,
            "output_mode": self.output_mode_var.get(),
            "output_table": self.output_table_var.get(),
            "backup_before_overwrite": self.backup_before_overwrite_var.get(),
            "table_access_policy": self.normalize_table_access_policy(),
        }

    def validate_plan_template_data(self, data):
        """
        只识别新版计划模板：
        - 必须是 dict
        - template_type 必须等于 workflow_plan
        - nodes 必须是 list
        """
        if not isinstance(data, dict):
            return False, "模板内容不是 JSON 对象。"
        if data.get("template_type") != "workflow_plan":
            return False, "template_type 不是 workflow_plan。"
        if not isinstance(data.get("nodes"), list):
            return False, "nodes 字段不存在或不是列表。"
        return True, ""

    def apply_plan_template_data(self, data, source_path=""):
        """把已验证的计划模板应用到当前计划窗口。"""
        ok, reason = self.validate_plan_template_data(data)
        if not ok:
            raise ValueError(reason)

        self.nodes = data.get("nodes", [])
        self.ensure_node_tree_identity(self.nodes)
        self.output_mode_var.set(data.get("output_mode", "输出到主界面预览区"))
        self.output_table_var.set(data.get("output_table", self.make_default_output_table_name()))
        self.backup_before_overwrite_var.set(bool(data.get("backup_before_overwrite", True)))
        self.set_table_access_policy(data.get("table_access_policy", "audit"))
        self.refresh_node_list()
        self.rebuild_current_config()

        if source_path:
            self.status_var.set(f"计划模板已载入：{source_path}")
        else:
            self.status_var.set("计划模板已载入。")

    def refresh_plan_template_list(self, show_status=True):
        """扫描 plan 目录，只列出能正常读取的新版 workflow_plan JSON。"""
        os.makedirs(self.plan_dir, exist_ok=True)

        template_map = {}
        valid_count = 0
        skipped_count = 0

        try:
            files = sorted(
                f for f in os.listdir(self.plan_dir)
                if f.lower().endswith(".json")
            )
        except Exception as e:
            if hasattr(self, "plan_template_combo"):
                self.plan_template_combo["values"] = []
            self.plan_template_map = {}
            if show_status:
                self.status_var.set(f"读取 plan 目录失败：{e}")
            return

        for file_name in files:
            path = os.path.join(self.plan_dir, file_name)
            try:
                data, _load_info = load_json_with_backup(path)
                ok, _ = self.validate_plan_template_data(data)
                if not ok:
                    skipped_count += 1
                    continue

                # 下拉菜单只显示 JSON 内部的 plan_name，不再显示文件名。
                # 如果 plan_name 为空，则显示“未命名计划”；仍然不显示文件名。
                plan_name = str(data.get("plan_name") or "").strip() or "未命名计划"
                display = plan_name

                # 避免多个模板 plan_name 相同导致映射冲突。
                # 重名时只追加序号，不显示 json 文件名。
                original_display = display
                i = 2
                while display in template_map:
                    display = f"{original_display} ({i})"
                    i += 1

                template_map[display] = path
                valid_count += 1
            except Exception:
                skipped_count += 1
                continue

        self.plan_template_map = template_map
        values = list(template_map.keys())
        if hasattr(self, "plan_template_combo"):
            self.plan_template_combo["values"] = values

        current = self.plan_template_var.get()
        if current not in template_map:
            self.plan_template_var.set(values[0] if values else "")

        if show_status:
            self.status_var.set(
                f"模板刷新完成：可用 {valid_count} 个，跳过 {skipped_count} 个。目录：{self.plan_dir}"
            )

    def open_plan_dir(self):
        """打开程序真实目录下的 plan 模板目录。"""
        os.makedirs(self.plan_dir, exist_ok=True)
        try:
            if hasattr(os, "startfile"):
                os.startfile(self.plan_dir)
            else:
                messagebox.showinfo("plan目录", self.plan_dir)
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开 plan 目录：\n{self.plan_dir}\n\n{e}")

    def save_plan_template(self):
        os.makedirs(self.plan_dir, exist_ok=True)
        default_name = self.sanitize_plan_file_name(self.output_table_var.get() or "工作流计划") + ".json"
        path = filedialog.asksaveasfilename(
            title="保存计划模板",
            initialdir=self.plan_dir,
            initialfile=default_name,
            defaultextension=".json",
            filetypes=[("JSON模板", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return

        # 使用用户实际保存的 JSON 文件名作为 plan_name。
        # 例如保存为“PDF批量重命名.json”，则 JSON 内部写入：
        # "plan_name": "PDF批量重命名"。
        saved_file_name = os.path.basename(path)
        saved_plan_name = os.path.splitext(saved_file_name)[0].strip() or "工作流计划"

        data = self.build_plan_template_data(plan_name=saved_plan_name)
        try:
            atomic_write_json(path, data)
            self.status_var.set(f"计划模板已保存：{path}；plan_name 已同步为：{saved_plan_name}")
            self.refresh_plan_template_list(show_status=False)

            # 保存后尽量自动选中刚保存的模板，便于确认和后续快速载入。
            if hasattr(self, "plan_template_combo") and hasattr(self, "plan_template_map"):
                abs_saved_path = os.path.abspath(path)
                for display_name, template_path in self.plan_template_map.items():
                    if os.path.abspath(template_path) == abs_saved_path:
                        self.plan_template_var.set(display_name)
                        break
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def load_plan_template_from_path(self, path):
        if not path:
            return
        if self.nodes:
            ok = messagebox.askyesno(
                "确认载入模板",
                "当前计划已有节点，载入模板会覆盖当前计划。\n是否继续？"
            )
            if not ok:
                return
        try:
            data = load_json_file_with_recovery(path, parent=self.window)
            self.apply_plan_template_data(data, source_path=path)
        except Exception as e:
            messagebox.showerror("载入失败", str(e))

    def load_plan_template(self):
        path = filedialog.askopenfilename(
            title="载入计划模板",
            initialdir=self.plan_dir,
            filetypes=[("JSON模板", "*.json"), ("所有文件", "*.*")]
        )
        if not path:
            return
        self.load_plan_template_from_path(path)

    def load_selected_plan_template(self):
        display = self.plan_template_var.get()
        if not display:
            messagebox.showwarning("提示", "请先从下拉菜单选择一个计划模板。")
            return

        path = self.plan_template_map.get(display)
        if not path:
            self.refresh_plan_template_list(show_status=False)
            path = self.plan_template_map.get(display)

        if not path:
            messagebox.showwarning("提示", "选中的计划模板不存在或已失效，请刷新模板列表。")
            return

        self.load_plan_template_from_path(path)


    # ==================== 后台执行 / 进度条管理 ====================
    def get_workflow_log_dir(self):
        log_dir = os.path.join(getattr(self.app, "app_dir", get_app_dir()), "logs", "workflow")
        os.makedirs(log_dir, exist_ok=True)
        return log_dir

    def write_workflow_error_log(self, mode, message, traceback_text="", logs=None, snapshot=None):
        """后台线程错误日志。只写文件，不直接操作 Tkinter。"""
        try:
            log_dir = self.get_workflow_log_dir()
            path = os.path.join(log_dir, f"workflow_error_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.log")
            snapshot = snapshot or {}
            node_count = len(snapshot.get("nodes", self.nodes) or [])
            with open(path, "w", encoding="utf-8") as f:
                f.write(f"任务模式：{mode}\n")
                f.write(f"时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"节点数量：{node_count}\n")
                if snapshot.get("db_path"):
                    f.write(f"数据库：{snapshot.get('db_path')}\n")
                if snapshot.get("workflow_name"):
                    f.write(f"工作流/输出名：{snapshot.get('workflow_name')}\n")
                f.write(f"错误信息：{message}\n\n")
                if logs:
                    f.write("执行日志：\n")
                    for item in logs:
                        f.write(f"- {item}\n")
                    f.write("\n")
                if traceback_text:
                    f.write("Traceback：\n")
                    f.write(traceback_text)
            return path
        except Exception:
            return ""

    def build_workflow_task_snapshot(self, mode, stop_index=None, execute_actions=False):
        """在 Tk 主线程创建后台任务快照，后台线程只读普通 Python 数据，避免直接访问 Tk 变量。"""
        return {
            "mode": mode,
            "stop_index": stop_index,
            "execute_actions": bool(execute_actions),
            "app_dir": getattr(self.app, "app_dir", get_app_dir()),
            "db_path": self.app.db_path_var.get().strip(),
            "workflow_name": self.output_table_var.get().strip(),
            "output_table": self.output_table_var.get().strip(),
            "output_mode": self.output_mode_var.get(),
            "backup_before_overwrite": bool(self.backup_before_overwrite_var.get()),
            "table_access_policy": self.normalize_table_access_policy(),
            "headers": copy.deepcopy(self.app.headers),
            "rows": copy.deepcopy(self.app.rows),
            "nodes": copy.deepcopy(self.nodes),
            "manual_loop_context": copy.deepcopy(self.manual_loop_context) if self.manual_loop_context is not None else None,
            "manual_loop_after_index": self.manual_loop_after_index,
            "manual_loop_headers": copy.deepcopy(self.manual_loop_headers) if self.manual_loop_headers is not None else None,
            "manual_loop_rows": copy.deepcopy(self.manual_loop_rows) if self.manual_loop_rows is not None else None,
        }

    def start_workflow_task(self, task_type, title=None, stop_index=None, execute_actions=False):
        """统一后台任务入口。保留 _start_background_workflow 作为底层实现。"""
        title = title or task_type
        return self._start_background_workflow(task_type, title, stop_index=stop_index, execute_actions=execute_actions)

    def _iter_workflow_child_widgets(self, parent):
        for child in parent.winfo_children():
            yield child
            yield from self._iter_workflow_child_widgets(child)

    def _set_workflow_cancel_enabled(self, enabled):
        try:
            if self.workflow_cancel_button is not None and self.workflow_cancel_button.winfo_exists():
                self.workflow_cancel_button.configure(state="normal" if enabled else "disabled")
        except Exception:
            pass

    def _set_workflow_controls_enabled(self, enabled):
        """后台运行期间锁定配置/执行控件，避免节点列表在执行中被改动。"""
        classes = {"TButton", "TCombobox", "TEntry", "TCheckbutton", "TRadiobutton", "Entry", "Text", "Listbox", "Button", "Checkbutton", "Radiobutton", "Spinbox", "TSpinbox"}
        if not enabled:
            self.workflow_widget_state_backup = {}
            for widget in self._iter_workflow_child_widgets(self.window):
                if widget is self.workflow_cancel_button:
                    continue
                try:
                    if widget.winfo_class() not in classes:
                        continue
                    old_state = widget.cget("state")
                    self.workflow_widget_state_backup[widget] = old_state
                    widget.configure(state="disabled")
                except Exception:
                    continue
            self._set_workflow_cancel_enabled(True)
            return

        for widget, old_state in list(self.workflow_widget_state_backup.items()):
            try:
                if widget.winfo_exists():
                    widget.configure(state=old_state)
            except Exception:
                pass
        self.workflow_widget_state_backup = {}
        self._set_workflow_cancel_enabled(False)

    def is_background_workflow_running(self):
        return bool(self.workflow_worker_running and self.workflow_worker_thread and self.workflow_worker_thread.is_alive())

    def cancel_background_workflow(self):
        if not self.is_background_workflow_running():
            self.worker_status_text.set("执行状态：当前没有后台任务。")
            return
        if self.workflow_worker_cancel is not None:
            self.workflow_worker_cancel.set()
        self.worker_status_text.set("执行状态：正在请求取消，当前节点会在安全检查点停止。")

    def _set_background_workflow_state(self, running, title=""):
        self.workflow_worker_running = bool(running)
        if running:
            self.workflow_current_task = title
            self._set_workflow_controls_enabled(False)
            self.worker_status_text.set(f"执行状态：后台运行中 - {title}")
            self.workflow_progress_var.set(0)
            self.node_progress_var.set(0)
            self.workflow_progress_text.set("总进度：准备开始")
            self.node_progress_text.set("当前节点：等待执行")
        else:
            self.workflow_worker_running = False
            self.workflow_current_task = None
            self._set_workflow_controls_enabled(True)

    def _start_background_workflow(self, mode, title, stop_index=None, execute_actions=False):
        if self.is_background_workflow_running():
            messagebox.showwarning("后台任务运行中", "当前已有工作流正在后台执行，请等待完成或先取消。")
            return
        if not self.confirm_jump_precheck(execute_actions=execute_actions, stop_index=stop_index):
            self.status_var.set("工作流已取消：跳转校验未继续。")
            return
        if execute_actions and not self.confirm_table_access_precheck(execute_actions=True, stop_index=stop_index):
            self.status_var.set("执行计划已取消：权限预检未继续。")
            return
        snapshot = self.build_workflow_task_snapshot(mode, stop_index=stop_index, execute_actions=execute_actions)
        self.workflow_worker_queue = queue.Queue()
        self.workflow_worker_cancel = threading.Event()
        self._set_background_workflow_state(True, title)
        self.workflow_worker_thread = threading.Thread(
            target=self._background_workflow_worker,
            args=(mode, stop_index, execute_actions, snapshot),
            daemon=True
        )
        self.workflow_worker_thread.start()
        self.window.after(80, self._poll_background_workflow_queue)

    def _background_progress_callback(self, message):
        try:
            self.workflow_worker_queue.put(message)
        except Exception:
            pass

    def _background_workflow_worker(self, mode, stop_index=None, execute_actions=False, snapshot=None):
        logs = []
        snapshot = snapshot or {}
        try:
            self.workflow_worker_queue.put({"type": "workflow_start", "message": mode})
            if mode == "preview_to":
                idx = int(stop_index)
                manual_loop_context = snapshot.get("manual_loop_context")
                manual_loop_after_index = snapshot.get("manual_loop_after_index")
                if manual_loop_context is not None and manual_loop_after_index is not None and idx >= manual_loop_after_index:
                    preview_context = copy.deepcopy(manual_loop_context)
                    preview_context["allow_selected_columns_write_in_preview"] = True
                    headers, rows, logs, context = self.run_plan(
                        start_index=manual_loop_after_index,
                        stop_index=idx,
                        raise_error=True,
                        return_context=True,
                        initial_headers=snapshot.get("manual_loop_headers"),
                        initial_rows=snapshot.get("manual_loop_rows"),
                        initial_context=preview_context,
                        progress_callback=self._background_progress_callback,
                        cancel_event=self.workflow_worker_cancel,
                        workflow_snapshot=snapshot,
                    )
                    prefix = f"已基于单步循环缓存预览到节点 {idx + 1}"
                else:
                    preview_context = {"transit_tables": {}, "loop_states": {}, "loop_results": {}, "allow_selected_columns_write_in_preview": True}
                    headers, rows, logs, context = self.run_plan(
                        stop_index=idx,
                        raise_error=True,
                        return_context=True,
                        initial_context=preview_context,
                        progress_callback=self._background_progress_callback,
                        cancel_event=self.workflow_worker_cancel,
                        workflow_snapshot=snapshot,
                    )
                    prefix = f"已预览到节点 {idx + 1}"
            elif mode == "preview_full":
                manual_loop_context = snapshot.get("manual_loop_context")
                manual_loop_after_index = snapshot.get("manual_loop_after_index")
                if manual_loop_context is not None and manual_loop_after_index is not None:
                    preview_context = copy.deepcopy(manual_loop_context)
                    preview_context["allow_selected_columns_write_in_preview"] = True
                    headers, rows, logs, context = self.run_plan(
                        start_index=manual_loop_after_index,
                        stop_index=None,
                        raise_error=True,
                        return_context=True,
                        initial_headers=snapshot.get("manual_loop_headers"),
                        initial_rows=snapshot.get("manual_loop_rows"),
                        initial_context=preview_context,
                        progress_callback=self._background_progress_callback,
                        cancel_event=self.workflow_worker_cancel,
                        workflow_snapshot=snapshot,
                    )
                    prefix = "已基于单步循环缓存完成后续计划预览"
                else:
                    preview_context = {"transit_tables": {}, "loop_states": {}, "loop_results": {}, "allow_selected_columns_write_in_preview": True}
                    headers, rows, logs, context = self.run_plan(
                        stop_index=None,
                        raise_error=True,
                        return_context=True,
                        initial_context=preview_context,
                        progress_callback=self._background_progress_callback,
                        cancel_event=self.workflow_worker_cancel,
                        workflow_snapshot=snapshot,
                    )
                    prefix = "完整计划预览完成"
            elif mode == "execute_plan":
                headers, rows, logs, context = self.run_plan(
                    stop_index=None,
                    raise_error=True,
                    execute_actions=execute_actions,
                    return_context=True,
                    progress_callback=self._background_progress_callback,
                    cancel_event=self.workflow_worker_cancel,
                    workflow_snapshot=snapshot,
                )
                prefix = "计划执行完成"
            else:
                raise ValueError(f"未知后台任务模式：{mode}")

            if self.workflow_worker_cancel is not None and self.workflow_worker_cancel.is_set():
                self.workflow_worker_queue.put({"type": "workflow_cancelled", "logs": logs})
                return
            self.workflow_worker_queue.put({
                "type": "workflow_done",
                "mode": mode,
                "prefix": prefix,
                "headers": headers,
                "rows": rows,
                "logs": logs,
                "context": context,
                "snapshot": snapshot,
            })
        except Exception as e:
            if self.workflow_worker_cancel is not None and self.workflow_worker_cancel.is_set():
                logs.append(f"用户取消后台任务：{e}")
                self.workflow_worker_queue.put({"type": "workflow_cancelled", "logs": logs})
                return
            tb = traceback.format_exc()
            log_path = self.write_workflow_error_log(mode, str(e), tb, logs=logs, snapshot=snapshot)
            self.workflow_worker_queue.put({
                "type": "workflow_error",
                "message": str(e),
                "traceback": tb,
                "log_path": log_path,
            })

    def _poll_background_workflow_queue(self):
        try:
            while True:
                msg = self.workflow_worker_queue.get_nowait()
                self._handle_background_workflow_message(msg)
        except queue.Empty:
            pass
        if self.is_background_workflow_running():
            self.window.after(80, self._poll_background_workflow_queue)
        else:
            # 线程已经结束但可能还有最后几条消息，稍后再扫一次。
            if self.workflow_worker_running:
                self.workflow_worker_running = False
                self.window.after(120, self._poll_background_workflow_queue)

    def _handle_background_workflow_message(self, msg):
        mtype = msg.get("type")
        if mtype == "workflow_start":
            self.workflow_progress_var.set(0)
            self.node_progress_var.set(0)
            self.workflow_progress_text.set("总进度：已启动后台执行")
            self.node_progress_text.set("当前节点：等待执行")
            return
        if mtype == "node_start":
            idx = int(msg.get("node_index", 0))
            total = max(1, int(msg.get("node_total", len(self.nodes) or 1)))
            percent = max(0, min(100, idx / total * 100))
            self.workflow_progress_var.set(percent)
            self.node_progress_var.set(0)
            self.workflow_progress_text.set(f"总进度：节点 {idx + 1} / {total}")
            self.node_progress_text.set(f"当前节点：{msg.get('node_name', '')} - 开始")
            self.worker_status_text.set(msg.get("message", "节点开始"))
            return
        if mtype == "node_progress":
            current = msg.get("current")
            total = msg.get("total")
            node_name = msg.get("node_name", "")
            message = msg.get("message", "节点处理中")
            detail_message = msg.get("detail_message") or msg.get("detail") or message
            try:
                current_f = float(current)
                total_f = float(total)
                if total_f > 0:
                    percent = max(0, min(100, current_f / total_f * 100))
                    self.node_progress_var.set(percent)
                    if int(total_f) == total_f and int(current_f) == current_f:
                        self.node_progress_text.set(f"当前节点：{node_name} - {int(current_f)} / {int(total_f)}")
                    else:
                        self.node_progress_text.set(f"当前节点：{node_name} - {current_f:g} / {total_f:g}")
                else:
                    self.node_progress_text.set(f"当前节点：{node_name} - 处理中")
            except Exception:
                self.node_progress_text.set(f"当前节点：{node_name} - 处理中")
            self.worker_status_text.set(detail_message)
            return
        if mtype == "node_done":
            idx = int(msg.get("node_index", 0))
            total = max(1, int(msg.get("node_total", len(self.nodes) or 1)))
            percent = max(0, min(100, (idx + 1) / total * 100))
            self.workflow_progress_var.set(percent)
            self.node_progress_var.set(100)
            self.workflow_progress_text.set(f"总进度：节点 {idx + 1} / {total}")
            self.node_progress_text.set(f"当前节点：{msg.get('node_name', '')} - 完成，{msg.get('rows', 0)} 行 × {msg.get('cols', 0)} 列")
            self.worker_status_text.set(msg.get("message", "节点完成"))
            return
        if mtype == "node_error":
            self.node_progress_text.set(f"当前节点错误：{msg.get('node_name', '')}")
            self.worker_status_text.set(msg.get("message", "节点执行失败"))
            return
        if mtype == "workflow_cancelled":
            self._set_background_workflow_state(False)
            self.workflow_progress_text.set("总进度：已取消")
            self.node_progress_text.set("当前节点：已停止")
            self.status_var.set("后台工作流已取消。" + self.format_logs(msg.get("logs", [])))
            return
        if mtype == "workflow_error":
            self._set_background_workflow_state(False)
            self.workflow_progress_text.set("总进度：执行失败")
            self.node_progress_text.set("当前节点：失败")
            log_path = msg.get("log_path", "")
            if log_path:
                self.worker_status_text.set(f"执行状态：失败，错误日志：{log_path}")
                self.status_var.set(f"后台执行失败：{msg.get('message', '未知错误')}；错误日志：{log_path}")
            else:
                self.worker_status_text.set("执行状态：失败")
                self.status_var.set(f"后台执行失败：{msg.get('message', '未知错误')}")
            messagebox.showerror("后台执行失败", msg.get("message", "未知错误") + (f"\n\n错误日志：{log_path}" if log_path else ""))
            return
        if mtype == "workflow_done":
            self._set_background_workflow_state(False)
            headers = msg.get("headers", [])
            rows = msg.get("rows", [])
            logs = msg.get("logs", [])
            context = msg.get("context", {}) or {}
            snapshot = msg.get("snapshot") or context.get("workflow_snapshot", {}) or {}
            self.current_transit_tables = context.get("transit_tables", {})
            self.last_workflow_context = context
            self.last_table_access_logs = list(context.get("table_access_logs", []) or [])
            # 后台节点如果写入了 SQLite 表，不直接刷新 UI，只在这里回到主线程后统一刷新表列表。
            refresh_requests = context.get("ui_refresh_requests", []) or []
            if context.get("needs_refresh_table_list") or "table_list" in refresh_requests:
                try:
                    self.app.refresh_table_list()
                except Exception:
                    pass
            self.workflow_progress_var.set(100)
            self.node_progress_var.set(100)
            self.workflow_progress_text.set("总进度：完成")
            self.node_progress_text.set("当前节点：完成")
            mode = msg.get("mode")
            if mode in ("preview_full", "preview_to"):
                self.set_plan_preview_result(headers, rows, display=True)
                self.status_var.set(f"{msg.get('prefix', '预览完成')}：{len(rows)} 行 × {len(headers)} 列。" + self.format_logs(logs))
            elif mode == "execute_plan":
                self._finish_execute_plan_output(headers, rows, logs, context=context, snapshot=snapshot)
            return

    def _finish_execute_plan_output(self, headers, rows, logs, context=None, snapshot=None):
        context = context or {}
        snapshot_context = {"workflow_snapshot": snapshot or context.get("workflow_snapshot", {}) or {}}
        mode = self.get_workflow_output_mode(snapshot_context)
        if mode == "输出到主界面预览区":
            self.app.headers = list(headers)
            self.app.rows = [list(row) for row in rows]
            self.app.raw_data = ""
            self.app.refresh_tree()
            self.set_plan_preview_result(headers, rows, display=True)
            self.app.info_var.set(f"计划执行完成，已输出到主界面：{len(rows)} 行 × {len(headers)} 列。")
            self.status_var.set("计划执行完成，已输出到主界面。" + self.format_logs(logs))
            return

        if mode in ["保存为SQLite新表", "覆盖当前表"]:
            table_name = self.get_workflow_output_table(snapshot_context)
            if not table_name:
                messagebox.showwarning("提示", "请填写输出表名。")
                return
            overwrite = mode == "覆盖当前表"
            if overwrite:
                ok = messagebox.askyesno("确认覆盖", f"即将覆盖 SQLite 表：{table_name}\n覆盖前会按设置自动备份。是否继续？")
                if not ok:
                    return
            try:
                saved_name = self.save_result_to_sqlite(
                    headers,
                    rows,
                    table_name,
                    overwrite=overwrite,
                    backup=self.get_workflow_backup_before_overwrite(snapshot_context),
                    context=context,
                )
                self.last_table_access_logs = list(context.get("table_access_logs", []) or [])
                self.app.refresh_table_list()
                self.status_var.set(f"计划执行完成，已保存到 SQLite 表：{saved_name}。" + self.format_logs(logs))
                messagebox.showinfo("保存成功", f"已保存计划结果。\n\n表名：{saved_name}\n行数：{len(rows)}\n列数：{len(headers)}")
            except Exception as e:
                messagebox.showerror("保存失败", str(e))
            return

        if mode == "导出为xlsx":
            path = filedialog.asksaveasfilename(
                title="导出计划结果为 xlsx",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            if not path:
                return
            try:
                self.export_result_to_xlsx(headers, rows, path)
                self.status_var.set(f"计划执行完成，已导出：{path}。" + self.format_logs(logs))
                messagebox.showinfo("导出成功", f"已导出计划结果：\n{path}")
            except Exception as e:
                messagebox.showerror("导出失败", str(e))

    # 覆盖原同步预览/执行入口：改为后台执行，避免长节点阻塞 Tkinter 主事件循环。
    def preview_to_selected_node(self):
        idx = self.get_selected_node_index()
        if idx is None:
            messagebox.showwarning("提示", "请先选择一个节点。")
            return
        self.start_workflow_task("preview_to", f"预览到节点 {idx + 1}", stop_index=idx, execute_actions=False)

    def preview_full_plan(self):
        self.start_workflow_task("preview_full", "预览完整计划", stop_index=None, execute_actions=False)

    def execute_plan(self):
        if self.is_background_workflow_running():
            messagebox.showwarning("后台任务运行中", "当前已有工作流正在后台执行。")
            return
        has_actual_rename = any(
            node.get("enabled", True) and node.get("type") == "批量重命名" and node.get("config", {}).get("actual_rename")
            for node in self.nodes
        )
        if self.preview_dirty and self.preview_headers and self.preview_rows and not has_actual_rename:
            use_current_preview = messagebox.askyesno(
                "使用已修改的计划预览？",
                "检测到结果预览区存在手动修改。\n\n"
                "选择【是】：使用当前预览数据作为输出，不重新执行计划。\n"
                "选择【否】：重新执行计划，当前预览修改会被覆盖。"
            )
            if use_current_preview:
                self._finish_execute_plan_output(list(self.preview_headers), [list(row) for row in self.preview_rows], ["使用手动修改后的当前计划预览结果输出"], snapshot=self.build_workflow_task_snapshot("execute_plan", execute_actions=True))
                return
        if has_actual_rename:
            ok = messagebox.askyesno(
                "确认执行批量重命名",
                "当前计划中存在已勾选【实际执行重命名】的节点。\n\n"
                "执行后会修改磁盘上的文件/文件夹名称。建议先使用【预览完整计划】确认结果无误。\n\n是否继续执行？"
            )
            if not ok:
                return
        self.start_workflow_task("execute_plan", "执行计划", stop_index=None, execute_actions=True)


if __name__ == "__main__":
    # 预留给后续子进程 Worker / PyInstaller 打包使用。当前版本后台执行采用线程 Worker。
    try:
        import multiprocessing
        multiprocessing.freeze_support()
    except Exception:
        pass
    root = tk.Tk()
    app = ClipboardTableApp(root)
    root.mainloop()
