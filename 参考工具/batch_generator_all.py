# -*- coding: utf-8 -*-
import os
import re
import openpyxl
import sys
from dataclasses import dataclass, field
from typing import List, Dict, Any

# 强制设置输出编码，防范 Windows 终端中的字符编码崩溃问题
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

@dataclass
class ControllerConfig:
    """控制器数据承载实体类"""
    huayan_code: str = ""        # 华研编码 (如 1.01.01.01003)
    xiabao_code: str = ""        # 夏宝编码 (如 1.04.01.03047CF)
    raw_params: str = ""         # 原始参数长文本
    
    # 提取后的结构化属性
    project_name: str = ""       # 项目名称
    cabinet_type: str = ""       # 内机箱体
    sales_region: str = "外销"     # 销售地区
    motor_type: str = "直流"      # 内电机类型 (直流 / 交流)
    hardware_version: str = ""   # 内机硬件型号 (如 HYBP2537GK [V1.3])
    is_cooling_only: bool = False # 是否为单冷型号 (True: 单冷, False: 冷暖)
    signature_name: str = "覃文涛" # 签署人姓名
    confirm_type: str = "再次确认。" # 确认类型
    change_reason: str = "按夏宝要求，直流电机板自检电机驱动时间按要求加长至8秒。" # 变更原因
    
    # WiFi功能
    wifi_pid: str = ""           # WiFi 模块 PID
    wifi_client: str = ""        # WiFi 客户名称
    
    # 掉电记忆
    power_failure_memory: str = "开启"
    power_failure_desc: str = "外销型号默认开启"
    
    # 电机转速确认
    cooling_speeds: List[int] = field(default_factory=list) # 制冷转速列表
    heating_speeds: List[int] = field(default_factory=list) # 制热转速列表
    
    # 转速特殊逻辑 (行10)
    test_mode_speed: int = 0
    test_mode_desc: str = ""     # 测试模式转速描述 (包含多转速详情)
    special_speeds: Dict[str, Any] = field(default_factory=dict)
    
    # 附加功能特性
    features: List[str] = field(default_factory=list) # 三档节能、静音、iFEEL、EMC等
    
    # 变更/软件信息 (行12)
    hex_name: str = ""           # 完整的hex文件名
    hex_version: str = "V1.2"    # 版本
    hex_checksum: str = "0xCBF8" # 校验码
    hex_date: str = "251216"     # 日期字符 (如 260526)
    mcu_model: str = "美仁MR88F102CI/032PT_LQFP32"

# ==============================================================================
# 解析策略模式 (Extractors)
# ==============================================================================

class BaseExtractor:
    """解析器基类"""
    def extract(self, raw_text: str, config: ControllerConfig) -> None:
        raise NotImplementedError

class ProjectExtractor(BaseExtractor):
    """提取项目名称、箱体、电机"""
    def extract(self, raw_text: str, config: ControllerConfig) -> None:
        # 1. 区域提取
        region = ""
        for r in ["欧洲", "北美", "外销"]:
            if r in raw_text:
                region = r
                break
        
        # 2. 电压提取
        v_115 = "115V" if "115V" in raw_text else ""
        
        # 3. 容量提取 (优先从 hex_name 捕获，其次从 raw_text 提取)
        capacity = ""
        if config.hex_name:
            # 匹配板子型号 HYBP2612GK- 或 HYDP2533GK- 后面的容量代码
            hex_cap_match = re.search(r'HY[BD]P\d+GK-([^-]+)', config.hex_name)
            if hex_cap_match:
                cap_block = hex_cap_match.group(1)
                num_match = re.match(r'^(\d+)', cap_block)
                if num_match:
                    num = num_match.group(1)
                    if "k" in cap_block.lower():
                        capacity = num + "K"
                    else:
                        capacity = num
                        
        if not capacity:
            cap_match = re.search(r'(\d+K)', raw_text)
            if cap_match:
                capacity = cap_match.group(1)
            else:
                # 兜底匹配国内机型容量数字，如：36冷暖变频, 高效35双冷, 50冷暖, 26单冷
                cap_match2 = re.search(r'(?:高效|外销|电商新能效|正三级)?(\d+)(?:冷暖|单冷|双冷|变频|RD)', raw_text)
                if cap_match2:
                    capacity = cap_match2.group(1)
            
        # 3.5 销售地区与掉电记忆自适应提取 (根据容量是否有K)
        if capacity:
            if "K" in capacity.upper():
                config.sales_region = "外销"
                config.power_failure_memory = "开启"
                config.power_failure_desc = "外销型号默认开启"
            else:
                config.sales_region = "内销"
                config.power_failure_memory = "关闭"
                config.power_failure_desc = "默认关闭"
        else:
            if any(x in raw_text for x in ["北美", "欧洲", "外销"]):
                config.sales_region = "外销"
                config.power_failure_memory = "开启"
                config.power_failure_desc = "外销型号默认开启"
            else:
                config.sales_region = "内销"
                config.power_failure_memory = "关闭"
                config.power_failure_desc = "默认关闭"
            
        # 4. 款式提取 (智能识别款式：K、L、N、E、A、B、F、G、P、J、Q, 均为单个英文大写字母, 兼容中英文及混搭括号)
        style_letter = ""
        
        # 优先在括号内查找款式 (支持中英文及混搭括号，且有且仅有一个大写字母以杜绝 WiFi PID/电机等任何干扰)
        paren_matches = re.finditer(r'[\(（]([^)）]+)[\)）]', raw_text)
        for match in paren_matches:
            paren_content = match.group(1).strip()
            
            # 找出括号中所有的英文字母
            all_letters = re.findall(r'[A-Za-z]', paren_content)
            # 转换为大写
            uppercase_letters = [char.upper() for char in all_letters if char.isupper()]
            
            # 有且仅有一个大写字母，且该字母必须在款式范围内
            if len(uppercase_letters) == 1:
                candidate = uppercase_letters[0]
                if candidate in 'KLNEABFGPJQ':
                    style_letter = candidate
                    break
                
        # 备用方案：在全文查找 款 字样前面的款式名
        if not style_letter:
            let_match = re.search(r'([KLNEABFGPJQ])款', raw_text, re.IGNORECASE)
            if let_match:
                style_letter = let_match.group(1).upper()
            
        # 5. 单冷与冷暖判定
        config.is_cooling_only = "单冷" in raw_text
        mode_str = "单冷变频挂机" if config.is_cooling_only else "冷暖变频挂机"
            
        # 6. 组合项目名称 (统一格式)
        style_part = f"({style_letter}款)" if style_letter else ""
        parts = [region, v_115, capacity, style_part, mode_str]
        parts = [p for p in parts if p]
        config.project_name = "".join(parts)
            
        # 7. 内机箱体自适应
        if capacity:
            if style_letter:
                config.cabinet_type = f"{style_letter}款{capacity}变频挂机"
            else:
                config.cabinet_type = f"{capacity}变频挂机"
        else:
            config.cabinet_type = "变频挂机"
            
        # 8. 电机类型提取
        if "直流" in raw_text:
            config.motor_type = "直流"
        elif "交流" in raw_text:
            config.motor_type = "交流"

class WifiExtractor(BaseExtractor):
    """提取 WiFi 功能与 PID"""
    def extract(self, raw_text: str, config: ControllerConfig) -> None:
        # 1. 寻找 PID，兼容中英文冒号和逗号
        pid_match = re.search(r'(?:PID|pid)[：\s:，]([a-zA-Z0-9]{16})', raw_text)
        if pid_match:
            config.wifi_pid = pid_match.group(1).strip()
            
            # 2. 精准匹配 PID 后面跟着的型号/标识/客户/商标 (支持 + 或 - 或空格分隔)
            client_match = re.search(r'(?:PID|pid)[：\s:，]' + re.escape(config.wifi_pid) + r'\s*[\+\-]\s*([^)\)）\+]+)', raw_text, re.IGNORECASE)
            if client_match:
                config.wifi_client = client_match.group(1).strip()
            else:
                # 备用从 raw_text 提取“X客户”
                backup_match = re.search(r'\+([^+（(]*?客户[^+）)]*)', raw_text)
                if backup_match:
                    config.wifi_client = backup_match.group(1).strip()
                elif "无牌客户" in raw_text:
                    config.wifi_client = "无牌客户-商标空白"

class SpeedExtractor(BaseExtractor):
    """提取电机转速与特殊逻辑"""
    def extract(self, raw_text: str, config: ControllerConfig) -> None:
        # 1. 提取制冷转速 (支持冒号、分号、中文分号或空格)
        cooling_match = re.search(r'制冷[：:;；\s]*([\d/]+)', raw_text)
        if cooling_match:
            config.cooling_speeds = [int(x) for x in cooling_match.group(1).split('/')]
        else:
            # 备用方案：如果在固件信息中直接列出数字且紧随制热（通常是 WIFI PID 商标后直接写制冷转速）
            backup_match = re.search(r'(?:[;；\s\+])([\d/]+)\+制热', raw_text)
            if backup_match:
                config.cooling_speeds = [int(x) for x in backup_match.group(1).split('/')]
            
        # 如果是单冷，不需要提取制热转速
        if not config.is_cooling_only:
            heating_match = re.search(r'制热[：:;；\s]*([\d/]+)', raw_text)
            if heating_match:
                config.heating_speeds = [int(x) for x in heating_match.group(1).split('/')]

            
        # 2. 高精度捕获完整的测试模式转速描述 (包含多转速、文字描述等)
        speed_details = []
        
        # 方式 1: 测试模式转速-最小制冷900，其他1200
        m_c = re.search(r'(测试模式转速-最小制冷\d+，其他\d+)', raw_text)
        if m_c:
            speed_details.append(m_c.group(1))
            
        # 方式 2: 测试模式转速：制冷1600转，制热1500转 或 测试模式转速：制冷1600转
        m_a = re.search(r'测试模式转速[：:\s]*(制冷\d+转(?:[，,]\s*制热\d+转)?)', raw_text)
        if m_a:
            speed_details.append(f"测试模式转速：{m_a.group(1)}")
            
        # 方式 3: 低温制热/最小制热/制热暖气候D/最小制冷/其他/其它测试模式转速等
        m_bd = re.findall(r'((?:低温制热|最小制热|制热暖气候D|最小制冷|其他|其它)测试模式转速)[：:\s]*(\d+rpm|\d+)', raw_text)
        for label, val in m_bd:
            speed_details.append(f"{label}：{val}")
            
        # 方式 4: 通用常规测试模式匹配 (仅在未发现上述复杂描述时)
        if not speed_details:
            m_e = re.search(r'测试模式(?:转速)?[：:\s]*(\d+rpm|\d+)', raw_text)
            if m_e:
                speed_val = m_e.group(1).replace('rpm', '')
                speed_details.append(f"测试模式制冷/制热:{speed_val}")
                
        config.test_mode_desc = "\n".join(speed_details)
        
        # 同时提取单个 test_mode_speed 数字作为数值后备 (三阶段鲁棒匹配)
        test_match = re.search(r'测试模式[：\s:，,]*([0-9]+)', raw_text)
        if test_match:
            config.test_mode_speed = int(test_match.group(1))
        else:
            other_speed_match = re.search(r'其他[^\d]*(\d+)', raw_text)
            if other_speed_match:
                config.test_mode_speed = int(other_speed_match.group(1))
            else:
                test_match2 = re.search(r'测试(?:模式)?转速[：\s:，,-]*.*?(\d+)', raw_text)
                if test_match2:
                    config.test_mode_speed = int(test_match2.group(1))
                    
        if not config.test_mode_speed and config.cooling_speeds:
            config.test_mode_speed = config.cooling_speeds[0]
            
        # 3. 提取特殊转速描述 (动态捕捉制冷D转速与制热D转速的完整标签与数值)
        cooling_d_match = re.search(r'(制冷D[^：:，,+]*)[：:]([0-9]+)', raw_text)
        if cooling_d_match:
            config.special_speeds["cooling_d_label"] = cooling_d_match.group(1).strip()
            config.special_speeds["cooling_d_val"] = int(cooling_d_match.group(2))
            
        heating_d_match = re.search(r'(制热D[^：:，,+]*)[：:]([0-9]+)', raw_text)
        if heating_d_match:
            config.special_speeds["heating_d_label"] = heating_d_match.group(1).strip()
            config.special_speeds["heating_d_val"] = int(heating_d_match.group(2))
            
        temp_stop_match = re.search(r'(达温停压机后[^+\s]+)', raw_text)
        if temp_stop_match:
            config.special_speeds["temp_stop_logic"] = temp_stop_match.group(1)

class FunctionExtractor(BaseExtractor):
    """提取附加功能特征"""
    def extract(self, raw_text: str, config: ControllerConfig) -> None:
        def add_feature(feat):
            if feat not in config.features:
                config.features.append(feat)
                
        # 常见功能检测
        if "三档节能" in raw_text:
            add_feature("三档节能")
        if "ifeel" in raw_text.lower():
            add_feature("iFEEL")
        if "emc" in raw_text.lower():
            add_feature("EMC")
            
        # 左右摆风 (支持多种异形写法，如：左右摆风，左右风，默认左右/摆风，左右摆动，摆风)
        if any(x in raw_text for x in ["左右摆风", "左右风", "摆风", "摆动"]):
            add_feature("左右摆风")
            
        # 静音转速提取
        silent_match = re.search(r'静音[（(](\d+)[）)]', raw_text)
        if silent_match:
            add_feature(f"静音（{silent_match.group(1)}转）")
        elif "静音" in raw_text:
            add_feature("静音")
            
        # 掉电记忆已经在 ProjectExtractor 中根据有无 K 进行了自适应设定，无需在此处强行覆盖
        pass

# ==============================================================================
# 解析通道 (Pipeline)
# ==============================================================================

class ExtractionPipeline:
    def __init__(self):
        self.extractors: List[BaseExtractor] = []
        
    def register(self, extractor: BaseExtractor):
        self.extractors.append(extractor)
        return self
        
    def run(self, raw_text: str, huayan_code: str, xiabao_code: str, hex_name: str = "") -> ControllerConfig:
        config = ControllerConfig(
            huayan_code=huayan_code,
            xiabao_code=xiabao_code,
            raw_params=raw_text,
            hex_name=hex_name
        )
        for ext in self.extractors:
            try:
                ext.extract(raw_text, config)
            except Exception as e:
                print(f"[Warning] {ext.__class__.__name__} 解析异常: {e}")
        return config

# ==============================================================================
# Excel 填充改写器 (Writer)
# ==============================================================================

class ConfirmationWriter:
    def __init__(self, template_path: str):
        self.template_path = template_path
        
    def fill_and_save(self, config: ControllerConfig, output_path: str):
        """加载技术表确认模板，智能改写并输出"""
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # 根据硬件型号确定是定频还是变频 (BP = 变频, DP = 定频)
        is_inverter = "BP" in config.hardware_version
        if not is_inverter:
            config.project_name = config.project_name.replace("变频", "定频")
            config.cabinet_type = config.cabinet_type.replace("变频", "定频")
        
        # 1. 项目名称拼接 (E2:I2 合并)
        feature_suffix = []
        if config.wifi_pid:
            feature_suffix.append("WIFI")
        if "R32" in config.raw_params:
            feature_suffix.append("R32")
            
        # 全局新要求：所有确认表项目名称后缀加入 “遥控器全兼容”
        feature_suffix.append("遥控器全兼容")
        
        for f in config.features:
            if "静音" in f:
                sil_match = re.search(r'静音[（(](\d+)[转）]', f)
                if sil_match:
                    feature_suffix.append(f"静音（{sil_match.group(1)}）")
                else:
                    feature_suffix.append("静音")
            else:
                feature_suffix.append(f)
        
        suffix_str = "+".join(feature_suffix)
        ws["E2"] = f"{config.project_name}（{suffix_str}）"
            
        # 2. 销售地区 (E3) 与 电源规格 (H3)
        ws["E3"] = config.sales_region
        ws["H3"] = "115V" if "115V" in config.raw_params else "220V"
        
        # 3. 料号与硬件信息
        ws["C4"] = config.xiabao_code      # 夏宝电器料号
        ws["H4"] = config.huayan_code      # 夏宝华研板料号
        ws["B5"] = config.cabinet_type     # 内机箱体
        ws["F5"] = config.motor_type       # 内电机类型
        ws["B6"] = config.hardware_version # 内机硬件型号
        
        # 4. 软件校验码 (E6)
        ws["E6"] = config.hex_checksum
        
        # 5. 电机转速确认 (H6)
        cooling_str = "/".join(map(str, config.cooling_speeds))
        if config.is_cooling_only:
            ws["H6"] = f"制冷：{cooling_str}"
        else:
            heating_str = "/".join(map(str, config.heating_speeds))
            ws["H6"] = f"制冷：{cooling_str}\n制热：{heating_str}"
        
        # 6. 主要功能与要求表格 sequential writing (顺序填充有效行，智能消除空行)
        @dataclass
        class FeatureRow:
            name: str
            desc: str
            detail: str
            state: str
            
        active_rows = []
        
        # (1) WiFi 功能行
        if config.wifi_pid:
            client_desc = config.wifi_client if config.wifi_client else "夏宝客户"
            active_rows.append(FeatureRow(
                name="WiFi功能",
                desc="配套显示板带WiFi功能，可使用专用APP控制空调",
                detail=f"模块PID:{config.wifi_pid}-{client_desc}",
                state="主控板带WiFi功能"
            ))
            
        # (2) 掉电记忆行
        active_rows.append(FeatureRow(
            name="掉电记忆",
            desc="运行中断电可自动恢复掉电前的运行状态",
            detail=config.power_failure_desc,
            state=config.power_failure_memory
        ))
        
        # (3) 转速行 (仅变频 Inverter 写入，定频不占位，彻底消除定频转速行空隙)
        if is_inverter:
            speed_details = []
            if config.test_mode_desc:
                for line in config.test_mode_desc.split('\n'):
                    speed_details.append(line)
            elif config.test_mode_speed:
                speed_details.append(f"测试模式制冷/制热:{config.test_mode_speed}")
                
            if "cooling_d_val" in config.special_speeds:
                label = config.special_speeds.get("cooling_d_label", "制冷D转速")
                val = config.special_speeds["cooling_d_val"]
                speed_details.append(f"{label}：{val}")
            if "heating_d_val" in config.special_speeds:
                label = config.special_speeds.get("heating_d_label", "制热D转速")
                val = config.special_speeds["heating_d_val"]
                speed_details.append(f"{label}：{val}")
            if "temp_stop_logic" in config.special_speeds:
                speed_details.append(config.special_speeds['temp_stop_logic'])
                
            active_rows.append(FeatureRow(
                name="转速",
                desc="\n".join(speed_details),
                detail="",
                state="开启"
            ))
            
        # (4) 附加功能行
        if config.features:
            active_rows.append(FeatureRow(
                name="\n+".join(config.features),
                desc="",
                detail="",
                state=f"主控板带{'+'.join(config.features)}功能"
            ))
            
        # (4.5) 遥控器行 (按照用户新需求，全部表格强制加入)
        active_rows.append(FeatureRow(
            name="遥控器",
            desc="遥控器兼容14/15/20字节码",
            detail="",
            state="开启"
        ))
            
        # 强力清除 Row 8, 9, 10, 11, 12 的内容
        for r in [8, 9, 10, 11, 12]:
            ws[f"B{r}"] = ""
            ws[f"C{r}"] = ""
            ws[f"D{r}"] = ""
            ws[f"F{r}"] = ""
            
        # 顺序填充有效行，从 Row 8 开始递增，完美消灭空行！
        for idx, row in enumerate(active_rows):
            r = 8 + idx
            if r > 12:
                break # 防止越界，表格最多支持5行主要功能
                
            ws[f"B{r}"] = row.name
            ws[f"C{r}"] = row.desc
            ws[f"D{r}"] = row.detail
            ws[f"F{r}"] = row.state
            
        # 9. 变更记录与程序名称 (B13:I13 合并)
        desc_parts = []
        if config.confirm_type:
            desc_parts.append(config.confirm_type)
        if config.change_reason:
            desc_parts.append(config.change_reason)
        desc_parts.append(f"程序名称：{config.hex_name}")
        desc_parts.append(f"MCU：{config.mcu_model}")
        change_desc = "\n".join(desc_parts)
        ws["B13"] = change_desc
        
        # 10. 签章日期修改 (I16)
        if len(config.hex_date) == 6:
            fmt_date = f"20{config.hex_date[0:2]}.{config.hex_date[2:4]}.{config.hex_date[4:6]}"
        else:
            fmt_date = "2026.05.26"
        ws["I16"] = f"{config.signature_name}\n{fmt_date}"
        
        # 保存文件
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
        wb.save(output_path)
        wb.close()

# ==============================================================================
# 单芯片族生成控制核心
# ==============================================================================

def process_chip_series(series_name: str, prog_file: str, drv_file: str, hardware_ver: str, signature_name: str, pipeline: ExtractionPipeline, writer: ConfirmationWriter, base_out_dir: str, mcu_model: str = "美仁MR88F102CI/032PT_LQFP32", confirm_type: str = "再次确认。", change_reason: str = "按夏宝要求，直流电机板自检电机驱动时间按要求加长至8秒。", motor_type_override: str = ""):
    """提取并生成某一个控制器系列的所有技术表"""
    print(f"\n--- [开始处理 {series_name} 系列] ---")
    print(f"   程序映射文件: {prog_file}")
    print(f"   送样信息文件: {drv_file}")
    print(f"   内机硬件型号: {hardware_ver}")
    print(f"   签署签章姓名: {signature_name}")
    
    if not os.path.exists(prog_file) or not os.path.exists(drv_file):
        print(f"   ❌ 错误: 找不到对应数据文件，跳过本系列")
        return 0, 0
        
    # 1. 建立送样信息参数索引
    drv_wb = openpyxl.load_workbook(drv_file)
    drv_ws = drv_wb.active
    
    # 动态寻找参数所在的列 (自适应第三列或第四列，寻找含有 '+' 或最符合特征的单元格)
    param_col = 3
    for c in range(3, drv_ws.max_column + 1):
        text_count = 0
        for r in range(1, min(10, drv_ws.max_row + 1)):
            val = str(drv_ws.cell(r, c).value or "")
            if "+" in val or "制冷" in val or len(val) > 40:
                text_count += 1
        if text_count >= 2:
            param_col = c
            break
            
    print(f"   [检测结果] 送样参数处于第 {param_col} 列")
    
    drv_map = {}
    for r in range(1, drv_ws.max_row + 1):
        h_val = drv_ws.cell(r, 1).value
        x_val = drv_ws.cell(r, 2).value
        p_val = drv_ws.cell(r, param_col).value
        
        if not h_val or not p_val:
            continue
            
        h_str = str(h_val).strip()
        x_str = str(x_val).strip() if x_val else ""
        p_str = str(p_val).strip()
        
        # 跳过表头行
        if "编码" in h_str or "参数" in p_str or "华研" in h_str:
            continue
            
        if h_str:
            drv_map[h_str] = p_str
        if x_str:
            drv_map[x_str] = p_str
            
    drv_wb.close()
    print(f"   已成功建立 {len(drv_map)} 条参数索引项")
    
    # 创建系列专有的输出子目录
    out_dir = os.path.join(base_out_dir, f"{series_name}技术确认函")
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
        
    prog_wb = openpyxl.load_workbook(prog_file)
    prog_ws = prog_wb.active
    
    success = 0
    failed = 0
    
    for r in range(2, prog_ws.max_row + 1):
        huayan = str(prog_ws.cell(r, 1).value or "").strip()
        xiabao = str(prog_ws.cell(r, 2).value or "").strip()
        hex_name = str(prog_ws.cell(r, 3).value or "").strip()
        
        if not huayan or not xiabao or not hex_name:
            continue
            
        param_text = drv_map.get(huayan) or drv_map.get(xiabao)
        if not param_text:
            print(f"   ❌ [跳过] 编码 {huayan} / {xiabao} 在送样表 {drv_file} 中检索失败")
            failed += 1
            continue
            
        config = pipeline.run(param_text, huayan, xiabao, hex_name)
        config.hardware_version = hardware_ver
        config.signature_name = signature_name
        config.hex_name = hex_name
        config.mcu_model = mcu_model
        config.confirm_type = confirm_type
        config.change_reason = change_reason
        if motor_type_override:
            config.motor_type = motor_type_override
        
        # 解析 HEX 文件名
        hex_match = re.search(r'_(V\d+\.\d+)_(0x[0-9A-Fa-f]{4})_(\d{6})\.hex$', hex_name)
        if hex_match:
            config.hex_version = hex_match.group(1)
            config.hex_checksum = hex_match.group(2)
            config.hex_date = hex_match.group(3)
            
        output_path = os.path.join(out_dir, f"{xiabao}_确认函.xlsx")
        
        try:
            writer.fill_and_save(config, output_path)
            success += 1
        except Exception as e:
            print(f"   ❌ [错误] {xiabao} 写入失败: {e}")
            failed += 1
            
    prog_wb.close()
    print(f"   完成: 成功 {success} 份，跳过/失败 {failed} 份. 保存于 {out_dir}")
    return success, failed

# ==============================================================================
# 全局运行主入口
# ==============================================================================

def main():
    template_excel = "1.04.01.03047CF_确认函.xlsx"
    base_output_dir = "批量生成确认函"
    
    if not os.path.exists(template_excel):
        print(f"❌ 错误: 找不到全局确认表模板: {template_excel}")
        return
        
    print("==================================================")
    print("🚀 华研-夏宝 2533 & 2537 & 2545 确认函全自动高精度生成系统")
    print("==================================================")
    
    # 1. 组装解析通道
    pipeline = ExtractionPipeline()
    pipeline.register(ProjectExtractor())
    pipeline.register(WifiExtractor())
    pipeline.register(SpeedExtractor())
    pipeline.register(FunctionExtractor())
    
    writer = ConfirmationWriter(template_excel)
    
    # 2. 批量处理 2537 型号族 (覃文涛 签署, 硬件 HYBP2537GK [V1.3])
    succ_2537, fail_2537 = process_chip_series(
        series_name="2537",
        prog_file="2537程序名称.xlsx",
        drv_file="2537送样信息副本.xlsx",
        hardware_ver="HYBP2537GK [V1.3]",
        signature_name="覃文涛",
        pipeline=pipeline,
        writer=writer,
        base_out_dir=base_output_dir,
        mcu_model="美仁MR88FP05DI/032PA_LQFP32",
        confirm_type="首次确认。",
        change_reason="更改为可控硅方案。"
    )
    
    # 3. 批量处理 2533 型号族 (周世诚 签署, 硬件 HYDP2533GK [V1.1])
    succ_2533, fail_2533 = process_chip_series(
        series_name="2533",
        prog_file="2533程序名称.xlsx",
        drv_file="2533送样信息副本.xlsx",
        hardware_ver="HYDP2533GK [V1.1]",
        signature_name="周世诚",
        pipeline=pipeline,
        writer=writer,
        base_out_dir=base_output_dir
    )
    
    # 4. 批量处理 2545 型号族 (周世诚 签署, 硬件 HYBP2545GK [V1.2])
    succ_2545, fail_2545 = process_chip_series(
        series_name="2545",
        prog_file="2545程序名称.xlsx",
        drv_file="2545送样信息副本.xlsx",
        hardware_ver="HYBP2545GK [V1.2]",
        signature_name="周世诚",
        pipeline=pipeline,
        writer=writer,
        base_out_dir=base_output_dir
    )
    
    # 5. 批量处理 2612 型号族 (何章彬 签署, 硬件 HYBP2612GK [V1.0])
    succ_2612, fail_2612 = process_chip_series(
        series_name="2612",
        prog_file="2612程序名称.xlsx",
        drv_file="2612送样信息副本.xlsx",
        hardware_ver="HYBP2612GK [V1.0]",
        signature_name="何章彬",
        pipeline=pipeline,
        writer=writer,
        base_out_dir=base_output_dir,
        mcu_model="美仁MR88FP05DI/032PA_LQFP32",
        confirm_type="再次确认。",
        change_reason="更改为可控硅方案。",
        motor_type_override="PG电机"
    )
    
    # 6. 批量处理 2613 型号族 (何章彬 签署, 硬件 HYBP2613GK [V1.0])
    succ_2613, fail_2613 = process_chip_series(
        series_name="2613",
        prog_file="2613程序名称.xlsx",
        drv_file="2613送样信息副本.xlsx",
        hardware_ver="HYBP2613GK [V1.0]",
        signature_name="何章彬",
        pipeline=pipeline,
        writer=writer,
        base_out_dir=base_output_dir,
        mcu_model="美仁MR88FP05DI/032PA_LQFP32",
        confirm_type="",
        change_reason="对应HYBP2527GK电脑板,PG电机控制由固态继电器方案改为贴片可控硅方案",
        motor_type_override="PG电机"
    )
    
    print("\n" + "=" * 50)
    print("🌟 华研-夏宝 批量生成总结报告")
    print("-" * 50)
    print(f"   2537 系列: 成功 {succ_2537} 份 | 失败 {fail_2537} 份")
    print(f"   2533 系列: 成功 {succ_2533} 份 | 失败 {fail_2533} 份")
    print(f"   2545 系列: 成功 {succ_2545} 份 | 失败 {fail_2545} 份")
    print(f"   2612 系列: 成功 {succ_2612} 份 | 失败 {fail_2612} 份")
    print(f"   2613 系列: 成功 {succ_2613} 份 | 失败 {fail_2613} 份")
    print(f"   总计产出 : {succ_2537 + succ_2533 + succ_2545 + succ_2612 + succ_2613} 份高精度样式无损技术确认函")
    print(f"   结果目录 : '{base_output_dir}/'")
    print("=" * 50)

if __name__ == "__main__":
    main()
